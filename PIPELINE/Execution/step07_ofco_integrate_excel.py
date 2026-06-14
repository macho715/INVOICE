#!/usr/bin/env python3
"""
OFCO Invoice to Excel Integration Script with Price Center Auto-Mapping

This script parses OFCO invoice PDFs using patch.py and inserts the data
into the existing Excel file (ofco.xlsx) following the exact structure
and calculation logic with automatic Price Center to column mapping.
"""

import pandas as pd
import numpy as np
import os
import re
from ofco_parse_pdf_patched import parse_ofco_invoice
from ofco_map_price_center import PriceCenterMapper
from datetime import datetime
import argparse
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# PRICE CENTER to Excel Column Mapping
PRICE_CENTER_TO_COLUMNS = {
    'AGENCY FEE FOR CARGO CLEARANCE': {
        'qty_col': 'AGENCY_FEE_FOR_CARGO_CLEARANCE_QTY',
        'amt_col': 'AGENCY_FEE_FOR_CARGO_CLEARANCE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]  # EA1에 수량과 금액
    },
    'AGENCY FEE FOR BERTHING ARRANGEMENT': {
        'qty_col': 'AGENCY_FEE_FOR_BERTHING_ARRANGEMENT_QTY',
        'amt_col': 'AGENCY_FEE_FOR_BERTHING_ARRANGEMENT_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'AGENCY FEE FOR FW SUPPLY ARRANGEMENT': {
        'qty_col': 'AGENCY_FEE_FOR_FW_SUPPLY_ARRANGEMENT_QTY',
        'amt_col': 'AGENCY_FEE_FOR_FW_SUPPLY_ARRANGEMENT_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'AGENCY FOR ARRANGEMENT PTW': {
        'qty_col': 'AGENCY_FOR_ARRANGEMENT_PTW_QTY',
        'amt_col': 'AGENCY_FOR_ARRANGEMENT_PTW_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'YARD': {
        'qty_col': 'YARD_QTY',
        'amt_col': 'YARD_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'HANDLING FEE': {
        'qty_col': 'OFCO_HANDLING_FEE_PORT_HANDLING_CHARGE_10_QTY',
        'amt_col': 'OFCO_HANDLING_FEE_PORT_HANDLING_CHARGE_10_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'SUPPLY WATER 5000IG': {
        'qty_col': 'SUPPLY_WATER_5001IG_QTY',
        'amt_col': 'SUPPLY_WATER_5001IG_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'FORKLIFT': {
        'qty_col': 'FORKLIFT_HIRE_CHARGE_QTY',
        'amt_col': 'FORKLIFT_HIRE_CHARGE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'FORKLIFT MOB/DE-MOB': {
        'qty_col': 'FORKLIFT_MOB_DE_MOB_CHARGE_QTY',
        'amt_col': 'FORKLIFT_MOB_DE_MOB_CHARGE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'CONSUMABLES': {
        'qty_col': 'CONSUMABLES_QTY',
        'amt_col': 'CONSUMABLES_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'DIESEL VESSEL': {
        'qty_col': 'DIESEL_VESSEL_QTY',
        'amt_col': 'DIESEL_VESSEL_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'DOCUMENT PROCESSING CHARGE': {
        'qty_col': 'DOCUMENT_PROCESSING_CHARGE_QTY',
        'amt_col': 'DOCUMENT_PROCESSING_CHARGE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BULK MATERIAL': {
        'multi_columns': [
            ('DOCUMENT_PROCESSING_CHARGE_QTY', 'DOCUMENT_PROCESSING_CHARGE_AMOUNT', 1),
            ('BULK_MATERIAL_SOLIDS_A_PARCEL_SIZE_0_10_001_TONS_DIRECT_DELIVERY_QTY', 'BULK_MATERIAL_SOLIDS_A_PARCEL_SIZE_0_10_001_TONS_DIRECT_DELIVERY_AMOUNT', 2)
        ]
    },
    'BULK MATERIAL BAGGED CARGO 14': {
        'qty_col': 'BULK_MATERIAL_BAGGED_CARGO_14_QTY',
        'amt_col': 'BULK_MATERIAL_BAGGED_CARGO_14_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BULK MATERIAL BAGGED CARGO 15': {
        'qty_col': 'BULK_MATERIAL_BAGGED_CARGO_15_QTY',
        'amt_col': 'BULK_MATERIAL_BAGGED_CARGO_15_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BULK MATERIAL BAGGED CARGO 16': {
        'qty_col': 'BULK_MATERIAL_BAGGED_CARGO_16_QTY',
        'amt_col': 'BULK_MATERIAL_BAGGED_CARGO_16_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BULK MATERIAL BAGGED CARGO 19': {
        'qty_col': 'BULK_MATERIAL_BAGGED_CARGO_19_QTY',
        'amt_col': 'BULK_MATERIAL_BAGGED_CARGO_19_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PEC CHANGES': {
        'qty_col': 'PEC_CHANGES_QTY',
        'amt_col': 'PEC_CHANGES_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'ANCHORAGE FEES FOR NON-CARGO VESSELS': {
        'qty_col': 'ANCHORAGE_FEES_FOR_NON_CARGO_VESSELS_QTY',
        'amt_col': 'ANCHORAGE_FEES_FOR_NON_CARGO_VESSELS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'CHANNEL TRANSIT CHARGES': {
        'multi_columns': [
            ('PEC_CHANGES_QTY', 'PEC_CHANGES_AMOUNT', 1),  # EA1
            ('CHANNEL_TRANSIT_CROSSING_REQUEST_QTY', 'CHANNEL_TRANSIT_CROSSING_REQUEST_AMOUNT', 2),  # EA2
            ('CHANNEL_CROSSING_CHARGES_FOR_VESSELS_WITH_1000_TO_3_001_GT_QTY', 'CHANNEL_CROSSING_CHARGES_FOR_VESSELS_WITH_1000_TO_3_001_GT_AMOUNT', 3)  # EA3
        ]
    },
    'GENERAL WASTE SERVICE': {
        'qty_col': 'GENERAL_WASTE_SERVICE_QTY',
        'amt_col': 'GENERAL_WASTE_SERVICE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PORT DUES': {
        'multi_columns': [
            ('GENERAL_WASTE_SERVICE_QTY', 'GENERAL_WASTE_SERVICE_AMOUNT', 1),
            ('PORT_DUES_FOR_VESSELS_WITH_ABOVE_1_000_UP_TO_3_001GT_QTY', 'PORT_DUES_FOR_VESSELS_WITH_ABOVE_1_000_UP_TO_3_001GT_AMOUNT', 2)
        ]
    },
    'BUNKERING LUBE OIL SLUDGE SULPHUR BILGE TRANSFER BUNKERING PERMIT': {
        'qty_col': 'BUNKERING_LUBE_OIL_SLUDGE_SULPHUR_BILGE_TRANSFER_BUNKERING_PERMIT_QTY',
        'amt_col': 'BUNKERING_LUBE_OIL_SLUDGE_SULPHUR_BILGE_TRANSFER_BUNKERING_PERMIT_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PORT DUE CARGO': {
        'qty_col': 'PORT_DUE_CARGO_QTY',
        'amt_col': 'PORT_DUE_CARGO_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'WASTE HANDLING': {
        'qty_col': 'WASTE_HANDLING_QTY',
        'amt_col': 'WASTE_HANDLING_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BERTHING/SHIFTING': {
        'qty_col': 'BERTHING_SHIFTING_QTY',
        'amt_col': 'BERTHING_SHIFTING_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PILOTAGE': {
        'qty_col': 'PILOTAGE_QTY',
        'amt_col': 'PILOTAGE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PILOT LAUNCH': {
        'qty_col': 'PILOT_LAUNCH_QTY',
        'amt_col': 'PILOT_LAUNCH_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'GENERAL CARGO': {
        'qty_col': 'GENERAL_CARGO_QTY',
        'amt_col': 'GENERAL_CARGO_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'GATE PASS': {
        'qty_col': 'GATE_PASS_QTY',
        'amt_col': 'GATE_PASS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PORT DUE': {
        'qty_col': 'PORT_DUE_QTY',
        'amt_col': 'PORT_DUE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'WASTE HANDLING2': {
        'qty_col': 'WASTE_HANDLING2_QTY',
        'amt_col': 'WASTE_HANDLING2_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'BUNKERING': {
        'qty_col': 'BUNKERING_QTY',
        'amt_col': 'BUNKERING_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'HEAVY LIFT 20FRT PROJECT': {
        'qty_col': 'HEAVY_LIFT_20FRT_PROJECT_QTY',
        'amt_col': 'HEAVY_LIFT_20FRT_PROJECT_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'MANPOWER RIGGER': {
        'qty_col': 'MANPOWER_RIGGER_QTY',
        'amt_col': 'MANPOWER_RIGGER_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'MANPOWER BANKSMAN': {
        'qty_col': 'MANPOWER_BANKSMAN_QTY',
        'amt_col': 'MANPOWER_BANKSMAN_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'MANPOWER FOREMAN': {
        'qty_col': 'MANPOWER_FOREMAN_QTY',
        'amt_col': 'MANPOWER_FOREMAN_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'MANPOWER SUPERVISOR': {
        'qty_col': 'MANPOWER_SUPERVISOR_QTY',
        'amt_col': 'MANPOWER_SUPERVISOR_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'EQUIPMENT 75TON CRANE': {
        'qty_col': 'EQUIPMENT_75TON_CRANE_QTY',
        'amt_col': 'EQUIPMENT_75TON_CRANE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'EQUIPMENT 100TON CRAN': {
        'qty_col': 'EQUIPMENT_100TON_CRAN_QTY',
        'amt_col': 'EQUIPMENT_100TON_CRAN_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'EQUIPMENT 150TON CRANE': {
        'qty_col': 'EQUIPMENT_150TON_CRANE_QTY',
        'amt_col': 'EQUIPMENT_150TON_CRANE_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'EQUIPMENT SPREAD': {
        'qty_col': 'EQUIPMENT_SPREAD_QTY',
        'amt_col': 'EQUIPMENT_SPREAD_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'PASS ARRANGEMENT': {
        'qty_col': 'OTHERS_QTY',  # Fallback
        'amt_col': 'OTHERS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'DECK MAINTENANCE': {
        'qty_col': 'OTHERS_QTY',
        'amt_col': 'OTHERS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'SHORT TERM PASS': {
        'qty_col': 'OTHERS_QTY',
        'amt_col': 'OTHERS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    'EQUIPMENT PASS': {
        'qty_col': 'OTHERS_QTY',
        'amt_col': 'OTHERS_AMOUNT',
        'ea_mapping': [(1, 'qty', 'amt')]
    },
    # Add more mappings as needed
}

def get_next_version_filename(base_path, invoice_number, date_str):
    """
    Generate filename with auto-incrementing version number.

    Pattern: {invoice_number}_{date}_VER{nn}.xlsx
    Example: OFCO-INV-0001178_20250129_VER01.xlsx
    """
    results_dir = os.path.dirname(base_path) or '.'

    # Find existing versions
    existing_versions = []
    if os.path.exists(results_dir):
        pattern = re.compile(rf"{re.escape(invoice_number)}_{date_str}_VER(\d{{2}})\.xlsx")
        for filename in os.listdir(results_dir):
            match = pattern.match(filename)
            if match:
                version_num = int(match.group(1))
                existing_versions.append(version_num)

    # Determine next version
    next_version = max(existing_versions, default=0) + 1
    if next_version > 99:
        raise ValueError(f"Maximum version (VER99) exceeded for {invoice_number}_{date_str}")

    version_str = f"VER{next_version:02d}"
    filename = f"{invoice_number}_{date_str}_{version_str}.xlsx"

    return os.path.join(results_dir, filename)


def create_empty_row(excel_path='ofco.xlsx'):
    """127개 컬럼을 가진 빈 행 생성"""
    # Excel의 모든 컬럼명 로드
    template_df = pd.read_excel(excel_path, nrows=0)
    return {col: None for col in template_df.columns}

def convert_payload_to_excel_rows(payload, excel_path='ofco.xlsx'):
    """OFCOPayload를 Excel 행 리스트로 변환 (Price Center Auto-Mapping 포함)"""
    rows = []

    # PriceCenterMapper 초기화
    mapper = PriceCenterMapper(excel_path=excel_path)

    for idx, item in enumerate(payload.items):
        row = create_empty_row(excel_path)

        # 기본 정보
        row['Voyage No'] = payload.meta.voyage_no or ''
        row['SUBJECT'] = item.subject or ''
        row['INVOICE NUMBER'] = payload.meta.invoice_number or ''
        row['INVOICE DATE'] = payload.meta.invoice_date or ''

        # 연월 추출
        if payload.meta.invoice_date:
            try:
                date_obj = pd.to_datetime(payload.meta.invoice_date)
                row['INVOICE DATE_YEAR_MONTH'] = date_obj.strftime('%Y-%m-%d')
            except:
                row['INVOICE DATE_YEAR_MONTH'] = payload.meta.invoice_date

        # 비용 센터
        row['COST MAIN'] = item.cost_main or ''
        row['COST CENTER A'] = item.cost_center_a or ''
        row['COST CENTER B'] = item.cost_center_b or ''
        row['PRICE CENTER'] = item.price_center or ''

        # EA 쌍 처리 (최대 4개)
        ea_totals = []
        for i, (qty, rate) in enumerate(item.ea_rates[:4], 1):
            row[f'EA_{i}_Qty'] = qty
            row[f'EA_{i}_Rate'] = rate
            total = qty * rate
            row[f'EA{i}_Total'] = total
            ea_totals.append(total)

        # 나머지 EA 슬롯을 0으로 채움
        for i in range(len(item.ea_rates) + 1, 5):
            row[f'EA_{i}_Qty'] = None
            row[f'EA_{i}_Rate'] = None
            row[f'EA{i}_Total'] = 0.0
            ea_totals.append(0.0)

        # Amount_A_AED 계산
        row['Amount_A_AED'] = sum(ea_totals)

        # Price Center Auto-Mapping 적용
        mapping_result = mapper.map_row({
            'Price Center': item.price_center,
            'EA': item.ea_rates[0][0] if item.ea_rates else 0,
            'Amount': item.amount
        })

        # 매핑 결과를 행에 적용
        for col_name, value in mapping_result.items():
            if col_name in row:
                row[col_name] = value

        # 재무 정보
        row['Total_Amount_AED'] = item.amount

        rows.append(row)

    return pd.DataFrame(rows)

def calculate_sum_columns(df):
    """Sum_B_K_to_BA 등의 계산 컬럼 업데이트"""
    # K부터 BA까지의 AMOUNT 컬럼들을 합산 (K=idx 10, BA=idx 62)
    k_to_ba_cols = []
    bb_to_bi_cols = []

    for i, col in enumerate(df.columns):
        if isinstance(col, str) and '_AMOUNT' in col:
            if 10 <= i <= 62:  # K ~ BA 범위
                k_to_ba_cols.append(col)
            elif 63 <= i <= 78:  # BB ~ BI 범위
                bb_to_bi_cols.append(col)

    for idx, row in df.iterrows():
        # Sum_B_K_to_BA 계산 (K~BA 범위의 AMOUNT 컬럼 합)
        sum_b_k_to_ba = sum([row[col] for col in k_to_ba_cols if pd.notna(row[col])])
        df.at[idx, 'Sum_B_K_to_BA'] = sum_b_k_to_ba

        # Sum_C_BB_to_BI 계산 (BB~BI 범위의 AMOUNT 컬럼 합)
        sum_c_bb_to_bi = sum([row[col] for col in bb_to_bi_cols if pd.notna(row[col])])
        df.at[idx, 'Sum_C_BB_to_BI'] = sum_c_bb_to_bi

        # TotalCost_K 계산 (K열의 AMOUNT 값)
        k_col_idx = 10  # K열 인덱스
        if k_col_idx < len(df.columns):
            k_col_name = df.columns[k_col_idx]
            if isinstance(k_col_name, str) and '_AMOUNT' in k_col_name:
                df.at[idx, 'TotalCost_K'] = row[k_col_name] if pd.notna(row[k_col_name]) else 0.0
            else:
                df.at[idx, 'TotalCost_K'] = 0.0
        else:
            df.at[idx, 'TotalCost_K'] = 0.0

        # Diff 계산
        df.at[idx, 'Diff_A_B'] = row['Amount_A_AED'] - sum_b_k_to_ba
        df.at[idx, 'Diff_A_C'] = row['Amount_A_AED'] - sum_c_bb_to_bi
        df.at[idx, 'Diff_B_C'] = sum_b_k_to_ba - sum_c_bb_to_bi
        df.at[idx, 'Remaining_After_K'] = df.at[idx, 'Diff_A_B']

    return df


def apply_month_color_formatting(excel_path: str, df: pd.DataFrame):
    """
    인보이스 날짜 기반 월별 색상 적용
    
    - openpyxl을 사용하여 Excel 파일 로드
    - INVOICE DATE_YEAR_MONTH 컬럼 분석
    - 월별로 다른 색상 할당
    - 각 행의 배경색 적용
    """
    try:
        # Excel 파일 로드
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 월별 색상 매핑 (12개 월)
        month_colors = {
            1: 'FFFFCC',   # 연한 노랑 (January)
            2: 'CCE5FF',   # 연한 파랑 (February)
            3: 'CCFFCC',   # 연한 초록 (March)
            4: 'FFE5CC',   # 연한 주황 (April)
            5: 'E5CCFF',   # 연한 보라 (May)
            6: 'FFCCE5',   # 연한 분홍 (June)
            7: 'CCFFFF',   # 연한 청록 (July)
            8: 'FFFFE5',   # 연한 아이보리 (August)
            9: 'E5FFCC',   # 연한 라임 (September)
            10: 'FFE5E5',  # 연한 연분홍 (October)
            11: 'CCE5E5',  # 연한 민트 (November)
            12: 'E5E5FF',  # 연한 라벤더 (December)
        }
        
        # INVOICE DATE_YEAR_MONTH 컬럼 위치 찾기
        header_row = 1
        date_column_idx = None
        
        # 헤더 행의 모든 셀 확인
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell_value = cell.value
            if cell_value and str(cell_value).strip() == 'INVOICE DATE_YEAR_MONTH':
                date_column_idx = col_idx
                break
        
        if date_column_idx is None:
            # 대체로 INVOICE DATE 컬럼 사용 시도
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell_value = cell.value
                if cell_value and 'INVOICE DATE' in str(cell_value).upper():
                    date_column_idx = col_idx
                    break
        
        if date_column_idx is None:
            print("Warning: INVOICE DATE_YEAR_MONTH or INVOICE DATE column not found. Skipping color formatting.")
            wb.save(excel_path)
            return
        
        # 각 행에 색상 적용 (헤더 다음부터)
        for row_idx in range(2, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=date_column_idx)
            date_value = date_cell.value
            
            if date_value:
                try:
                    # 날짜 파싱 (다양한 형식 지원)
                    if isinstance(date_value, str):
                        # 문자열 날짜 파싱
                        date_str = str(date_value).strip()
                        # "15 October, 2025" 형식 처리
                        if ',' in date_str:
                            date_obj = pd.to_datetime(date_str, format='%d %B, %Y', errors='coerce')
                        else:
                            date_obj = pd.to_datetime(date_str, errors='coerce')
                    else:
                        date_obj = pd.to_datetime(date_value)
                    
                    if pd.notna(date_obj):
                        month = date_obj.month
                        color_code = month_colors.get(month, 'FFFFFF')  # 기본값: 흰색
                        
                        # 배경색 패턴 생성
                        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
                        
                        # 전체 행에 색상 적용
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = fill
                except Exception as e:
                    # 날짜 파싱 실패 시 해당 행 스킵
                    continue
        
        # 저장
        wb.save(excel_path)
        print(f"Applied month-based color formatting to {ws.max_row - 1} rows")
        
    except Exception as e:
        print(f"Warning: Could not apply color formatting: {e}")


def convert_csv_to_excel_with_mapping(csv_path, excel_path='ofco.xlsx'):
    """CSV 파일을 직접 읽어서 Price Center 매핑을 적용하여 Excel 형식으로 변환"""
    print(f"CSV 파일 읽기: {csv_path}")

    # CSV 파일 읽기
    df = pd.read_csv(csv_path)
    print(f"CSV 로드 완료: {len(df)} rows")

    # PriceCenterMapper 초기화
    mapper = PriceCenterMapper(excel_path=excel_path)

    # Excel 템플릿 로드
    template_df = pd.read_excel(excel_path, nrows=0)
    template_columns = template_df.columns.tolist()

    # 결과 DataFrame 생성
    result_rows = []

    for idx, row in df.iterrows():
        # 기본 행 생성
        new_row = {col: None for col in template_columns}

        # 기본 정보 매핑
        new_row['Voyage No'] = row.get('Voyage No', '')
        new_row['SUBJECT'] = row.get('Subject', '')
        new_row['INVOICE NUMBER'] = row.get('Invoice Number', '')
        new_row['INVOICE DATE'] = row.get('Date', '')
        new_row['INVOICE DATE_YEAR_MONTH'] = row.get('Date', '')

        # 비용 센터
        new_row['COST MAIN'] = row.get('Cost Main', '')
        new_row['COST CENTER A'] = row.get('Cost Center A', '')
        new_row['COST CENTER B'] = row.get('Cost Center B', '')
        new_row['PRICE CENTER'] = row.get('Price Center', '')

        # EA 정보
        ea = float(row.get('EA', 0) or 0)
        rate = float(row.get('Rate', 0) or 0)
        amount = float(row.get('Amount', 0) or 0)

        new_row['EA_1_Qty'] = ea
        new_row['EA_1_Rate'] = rate
        new_row['EA1_Total'] = ea * rate
        new_row['Total_Amount_AED'] = amount
        new_row['Amount_A_AED'] = amount

        # Price Center Auto-Mapping 적용
        mapping_result = mapper.map_row({
            'Price Center': row.get('Price Center', ''),
            'EA': ea,
            'Amount': amount
        })

        # 매핑 결과를 행에 적용
        for col_name, value in mapping_result.items():
            if col_name in new_row:
                new_row[col_name] = value

        result_rows.append(new_row)

    result_df = pd.DataFrame(result_rows)

    # 매핑 통계 출력
    print(f"Price Center 매핑 완료: {len(result_df)} rows")
    mapper._print_mapping_stats(result_df)

    return result_df

def main():
    parser = argparse.ArgumentParser(
        description='OFCO Invoice Excel Integration with Standard Output Naming'
    )
    parser.add_argument('pdf', help='PDF invoice file path')
    parser.add_argument('--excel', default='../templates/ofco.xlsx',
                       help='Excel template path (default: ../templates/ofco.xlsx)')
    parser.add_argument('--output', default=None,
                       help='Output Excel path (optional, overrides standard naming)')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be inserted without saving')
    args = parser.parse_args()

    try:
        # Parse PDF
        print(f"Parsing PDF: {args.pdf}")
        payload = parse_ofco_invoice(args.pdf)

        # Generate standard output filename
        invoice_number = payload.meta.invoice_number or 'UNKNOWN'
        invoice_date = payload.meta.invoice_date or datetime.now().strftime('%Y%m%d')

        # Convert date format: "2025-01-29" -> "20250129"
        if isinstance(invoice_date, str) and '-' in invoice_date:
            date_str = invoice_date.replace('-', '')
        else:
            date_str = str(invoice_date)

        # Create results directory if not exists
        results_dir = '../results'
        os.makedirs(results_dir, exist_ok=True)

        # Generate filename with version
        if args.output:
            output_path = args.output
        else:
            base_path = os.path.join(results_dir, f"{invoice_number}_{date_str}")
            output_path = get_next_version_filename(base_path, invoice_number, date_str)

        print(f"Standard output: {output_path}")

        print(f"Invoice: {payload.meta.invoice_number}")
        print(f"Total items: {len(payload.items)}")

        # Convert to Excel format
        print("Converting to Excel format...")
        new_rows_df = convert_payload_to_excel_rows(payload, args.excel)

        # Calculate sum columns
        print("Calculating sum columns...")
        new_rows_df = calculate_sum_columns(new_rows_df)

        if args.dry_run:
            print("\n=== DRY RUN - Preview of data to be inserted ===")
            print(f"Rows to add: {len(new_rows_df)}")
            print("\nSample data:")
            print(new_rows_df[['Voyage No', 'SUBJECT', 'INVOICE NUMBER', 'PRICE CENTER', 'EA_1_Qty', 'EA_1_Rate', 'EA1_Total', 'Amount_A_AED']].head().to_string())
            return

        # Load existing Excel
        print(f"Loading existing Excel: {args.excel}")
        existing_df = pd.read_excel(args.excel)

        # Get next NO
        last_no = existing_df['NO'].max() if not existing_df.empty else 0
        new_rows_df['NO'] = range(last_no + 1, last_no + 1 + len(new_rows_df))

        # Append
        result_df = pd.concat([existing_df, new_rows_df], ignore_index=True)

        # Save
        print(f"Saving to: {output_path}")
        result_df.to_excel(output_path, index=False, engine='openpyxl')

        # 색상 포맷팅 적용
        print("Applying month-based color formatting...")
        apply_month_color_formatting(output_path, result_df)

        print(f"Successfully added {len(new_rows_df)} rows to {output_path}")
        print(f"Invoice: {payload.meta.invoice_number}")
        print(f"Total items: {len(payload.items)}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
