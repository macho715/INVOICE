#!/usr/bin/env python3
"""
OFCO Pipeline Complete_Pipeline/EA_Slots → LIST 시트 통합 스크립트

파이프라인 실행 결과(Complete_Pipeline 또는 EA_Slots 시트)를 OFCO INVOICE.xlsx의 LIST 시트에 통합.

Usage:
    python ofco_integrate_pipeline_to_excel.py

Options:
    --source: 소스 파일 경로 (기본값: ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx)
    --source-sheet: 소스 시트명 (기본값: EA_Slots)
    --target: 대상 파일 경로 (기본값: OFCO INVOICE.xlsx)
    --backup: 백업 파일 생성 여부 (기본값: True)
    --no-backup: 백업 파일 생성 안 함
"""

import pandas as pd
import sys
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, Tuple, List
import argparse
import re
import os

# HeaderCore 통합 (표준 경로)
try:
    sys.path.insert(0, str(Path(__file__).parent.parent / "shared" / "scripts"))
    from ofco_header_core import HeaderCore, get_header_core
    HEADER_CORE_AVAILABLE = True
except ImportError:
    try:
        # fallback: 현재 디렉토리
        from ofco_header_core import HeaderCore, get_header_core
        HEADER_CORE_AVAILABLE = True
    except ImportError:
        HeaderCore = None
        get_header_core = None
        HEADER_CORE_AVAILABLE = False
        print("[WARNING] HeaderCore를 import할 수 없습니다. 컬럼 순서 정렬이 제한될 수 있습니다.")

# Price Center 매핑 규칙 import (참고용)
try:
    from ofco_integrate_excel import PRICE_CENTER_TO_COLUMNS
except ImportError:
    PRICE_CENTER_TO_COLUMNS = {}


def convert_currency(
    amount: float,
    from_currency: str,
    to_currency: str,
    rate: Optional[float] = None
) -> float:
    """
    USD ↔ AED 통화 변환 (표준화된 버전)

    Args:
        amount: 변환할 금액
        from_currency: 원본 통화 ('USD' 또는 'AED')
        to_currency: 대상 통화 ('USD' 또는 'AED')
        rate: 환율 (기본값: 환경 변수 USD_TO_AED_RATE 또는 3.6725)

    Returns:
        변환된 금액 (2소수점 반올림)

    Note:
        - 표준 환율: 3.6725 AED/USD (2024-2025 평균)
        - 환경 변수 USD_TO_AED_RATE로 오버라이드 가능
    """
    # 입력 검증
    if amount is None or pd.isna(amount):
        return 0.0

    if from_currency == to_currency:
        return round(float(amount), 2)

    # 환율 가져오기 (환경 변수 또는 기본값)
    if rate is None:
        rate = float(os.getenv('USD_TO_AED_RATE', '3.6725'))  # 표준 환율

    # 통화 정규화 (대소문자 처리)
    from_curr = str(from_currency).upper().strip()
    to_curr = str(to_currency).upper().strip()

    try:
        if from_curr == 'USD' and to_curr == 'AED':
            return round(float(amount) * rate, 2)
        elif from_curr == 'AED' and to_curr == 'USD':
            return round(float(amount) / rate, 2)
        else:
            # 지원하지 않는 통화 쌍
            print(f"[WARNING] 지원하지 않는 통화 변환: {from_curr} -> {to_curr}")
            return round(float(amount), 2)
    except (ValueError, TypeError) as e:
        print(f"[ERROR] 통화 변환 실패: {e}")
        return 0.0


def normalize_price_center(price_center: str) -> str:
    """
    Price Center 이름을 LIST 시트 컬럼 형식으로 변환

    예: "AGENCY FEE FOR CARGO CLEARANCE" → "AGENCY_FEE_FOR_CARGO_CLEARANCE"
    """
    if not price_center:
        return ""
    # 대문자로 변환하고 공백/특수문자를 언더스코어로 변경
    normalized = re.sub(r'[^\w\s]', '', str(price_center).upper())
    normalized = re.sub(r'\s+', '_', normalized.strip())
    return normalized


def convert_date(date_str: Any) -> Optional[datetime]:
    """
    날짜 문자열을 datetime 객체로 변환

    지원 형식:
    - "15 October, 2025"
    - "2025-10-15"
    - datetime 객체는 그대로 반환
    """
    if date_str is None or pd.isna(date_str):
        return None

    if isinstance(date_str, datetime):
        return date_str

    if isinstance(date_str, str):
        # "15 October, 2025" 형식
        try:
            return datetime.strptime(date_str, "%d %B, %Y")
        except ValueError:
            pass

        # "2025-10-15" 형식
        try:
            return datetime.strptime(date_str.split()[0], "%Y-%m-%d")
        except ValueError:
            pass

        # 다른 형식 시도
        try:
            return pd.to_datetime(date_str)
        except:
            pass

    return None


def find_price_center_columns(price_center: str, target_df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    """
    Price Center에 해당하는 QTY/AMOUNT 컬럼 찾기

    Returns:
        (qty_col, amt_col) 튜플 또는 (None, None)
    """
    if not price_center:
        return None, None

    # 정규화
    normalized = normalize_price_center(price_center)

    # 대상 파일에서 매칭되는 컬럼 찾기
    qty_col = f"{normalized}_QTY"
    amt_col = f"{normalized}_AMOUNT"

    # 컬럼명을 문자열로 변환하여 비교 (정수 컬럼명 처리)
    target_cols_str = [str(c) for c in target_df.columns]

    if qty_col in target_cols_str and amt_col in target_cols_str:
        # 실제 컬럼명 찾기 (대소문자/공백 차이 고려)
        qty_idx = target_cols_str.index(qty_col)
        amt_idx = target_cols_str.index(amt_col)
        return target_df.columns[qty_idx], target_df.columns[amt_idx]

    # PRICE_CENTER_TO_COLUMNS 규칙 사용
    if PRICE_CENTER_TO_COLUMNS and price_center in PRICE_CENTER_TO_COLUMNS:
        mapping = PRICE_CENTER_TO_COLUMNS[price_center]
        qty_col = mapping.get('qty_col')
        amt_col = mapping.get('amt_col')
        if qty_col and amt_col:
            target_cols_str = [str(c) for c in target_df.columns]
            if qty_col in target_cols_str and amt_col in target_cols_str:
                qty_idx = target_cols_str.index(qty_col)
                amt_idx = target_cols_str.index(amt_col)
                return target_df.columns[qty_idx], target_df.columns[amt_idx]

    return None, None


def transform_row(src_row: pd.Series, target_template: pd.DataFrame) -> Dict[str, Any]:
    """
    Complete_Pipeline 행을 LIST 시트 행 형식으로 변환

    Args:
        src_row: 소스 DataFrame의 행 (Series)
        target_template: 대상 DataFrame (컬럼 구조 참조용)

    Returns:
        변환된 행 데이터 (dict)
    """
    target_row = {}

    # 1. 기본 정보 매핑
    target_row['INVOICE NUMBER'] = src_row.get('invoice_no', '')
    invoice_date = convert_date(src_row.get('invoice_date'))
    if invoice_date:
        target_row['INVOICE DATE'] = invoice_date
        # INVOICE DATE_YEAR_MONTH는 "YYYY-MM-01" 형식 (연-월만)
        target_row['INVOICE DATE_YEAR_MONTH'] = invoice_date.strftime('%Y-%m-01')
    else:
        target_row['INVOICE DATE'] = None
        target_row['INVOICE DATE_YEAR_MONTH'] = None

    target_row['Voyage No'] = src_row.get('voyage_no') if pd.notna(src_row.get('voyage_no')) else None
    target_row['SUBJECT'] = src_row.get('description', '')

    # line_no는 나중에 처리 (Invoice별 순번으로 재계산)
    line_no_raw = src_row.get('line_no')
    if pd.notna(line_no_raw):
        target_row['NO'] = int(line_no_raw) if isinstance(line_no_raw, (int, float)) else None
    else:
        target_row['NO'] = None

    # 2. Cost/Price Center 매핑
    target_row['COST MAIN'] = src_row.get('cost_main', '') if pd.notna(src_row.get('cost_main', '')) else ''
    target_row['COST CENTER A'] = src_row.get('cost_center_a', '') if pd.notna(src_row.get('cost_center_a', '')) else ''
    target_row['COST CENTER B'] = src_row.get('cost_center_b', '') if pd.notna(src_row.get('cost_center_b', '')) else ''

    # PRICE CENTER 우선순위: price_center_name > price_center_a > price_center_b
    price_center = (
        src_row.get('price_center_name') or
        src_row.get('price_center_a') or
        src_row.get('price_center_b') or
        ''
    )
    if price_center and pd.notna(price_center):
        target_row['PRICE CENTER'] = str(price_center)
    else:
        target_row['PRICE CENTER'] = ''

    # 3. EA 슬롯 매핑
    for i in range(1, 5):
        ea_name = src_row.get(f'EA_{i}_Name')
        ea_qty = src_row.get(f'EA_{i}_Qty')
        ea_rate = src_row.get(f'EA_{i}_Rate')
        ea_amount = src_row.get(f'EA_{i}_Amount')

        target_row[f'EA_{i}_Name'] = ea_name if pd.notna(ea_name) else None
        target_row[f'EA_{i}_Qty'] = ea_qty if pd.notna(ea_qty) else None
        target_row[f'EA_{i}_Rate'] = ea_rate if pd.notna(ea_rate) else None
        target_row[f'EA_{i}_Amount'] = ea_amount if pd.notna(ea_amount) else None

    # 3-1. EA_Total_Amount 자동 계산
    ea_amounts = [
        target_row.get('EA_1_Amount') or 0,
        target_row.get('EA_2_Amount') or 0,
        target_row.get('EA_3_Amount') or 0,
        target_row.get('EA_4_Amount') or 0
    ]
    # NaN 값 처리
    ea_amounts = [float(amt) if pd.notna(amt) and amt is not None else 0.0 for amt in ea_amounts]
    target_row['EA_Total_Amount'] = sum(ea_amounts)

    # 4. Price Center별 QTY/AMOUNT 매핑
    price_center_val = target_row.get('PRICE CENTER', '')
    if price_center_val:
        pc_qty_col, pc_amt_col = find_price_center_columns(price_center_val, target_template)
        if pc_qty_col and pc_amt_col:
            # EA_1에서 QTY/AMOUNT 추출 (또는 소스의 Price Center별 컬럼에서)
            # 먼저 소스 파일의 Price Center별 컬럼 확인
            pc_normalized = normalize_price_center(price_center_val)
            src_qty_col = f"{pc_normalized}_QTY"
            src_amt_col = f"{pc_normalized}_AMOUNT"

            # 소스 컬럼명을 문자열로 변환하여 확인
            src_cols_str = [str(c) for c in src_row.index]

            if src_qty_col in src_cols_str and src_amt_col in src_cols_str:
                src_qty_idx = src_cols_str.index(src_qty_col)
                src_amt_idx = src_cols_str.index(src_amt_col)
                qty_val = src_row.iloc[src_qty_idx] if pd.notna(src_row.iloc[src_qty_idx]) else (target_row.get('EA_1_Qty') or 0)
                amt_val = src_row.iloc[src_amt_idx] if pd.notna(src_row.iloc[src_amt_idx]) else (target_row.get('EA_1_Amount') or 0)
            else:
                # EA_1에서 추출
                qty_val = target_row.get('EA_1_Qty') or 0
                amt_val = target_row.get('EA_1_Amount') or 0

            target_row[pc_qty_col] = qty_val if pd.notna(qty_val) else 0
            target_row[pc_amt_col] = amt_val if pd.notna(amt_val) else 0

    # 5. 금액 정보 매핑 (통화 변환 포함)
    # Lines 정보 우선 사용 (amount_excl_tax_lines)
    amount_excl_tax = src_row.get('amount_excl_tax') if pd.notna(src_row.get('amount_excl_tax')) else 0
    # Standard_Lines에는 tax_amount만 있음, tax_amount_lines는 Complete_Pipeline에만 있음
    tax_amount_lines = src_row.get('tax_amount_lines') if pd.notna(src_row.get('tax_amount_lines')) else (
        src_row.get('tax_amount') if pd.notna(src_row.get('tax_amount')) else 0
    )
    total_incl_tax = src_row.get('total_incl_tax') if pd.notna(src_row.get('total_incl_tax')) else (amount_excl_tax + tax_amount_lines)

    # 통화 정보 확인 (소스에서 currency 컬럼 또는 메타데이터)
    source_currency = src_row.get('currency', 'AED')  # 기본값 AED
    if pd.isna(source_currency) or source_currency == '':
        source_currency = 'AED'  # 기본값

    target_row['currency'] = source_currency

    # AED 컬럼 (원본 통화가 AED인 경우 직접 사용, USD인 경우 변환)
    if source_currency == 'USD':
        amount_excl_tax_aed = convert_currency(float(amount_excl_tax), 'USD', 'AED')
        tax_amount_aed = convert_currency(float(tax_amount_lines), 'USD', 'AED')
        total_incl_tax_aed = convert_currency(float(total_incl_tax), 'USD', 'AED')
    else:
        amount_excl_tax_aed = float(amount_excl_tax)
        tax_amount_aed = float(tax_amount_lines)
        total_incl_tax_aed = float(total_incl_tax)

    # USD 컬럼 (원본 통화가 USD인 경우 직접 사용, AED인 경우 변환)
    if source_currency == 'AED':
        amount_excl_tax_usd = convert_currency(float(amount_excl_tax), 'AED', 'USD')
        tax_amount_usd = convert_currency(float(tax_amount_lines), 'AED', 'USD')
        total_incl_tax_usd = convert_currency(float(total_incl_tax), 'AED', 'USD')
    else:
        amount_excl_tax_usd = float(amount_excl_tax)
        tax_amount_usd = float(tax_amount_lines)
        total_incl_tax_usd = float(total_incl_tax)

    target_row['Total_Amount_AED'] = amount_excl_tax_aed
    target_row['Amount_A_AED'] = amount_excl_tax_aed
    target_row['VAT_USD'] = tax_amount_usd
    target_row['Total_Amount_USD'] = total_incl_tax_usd
    target_row['Amount_USD'] = amount_excl_tax_usd

    # 6. 검증 컬럼 매핑
    target_row['calc_check'] = src_row.get('calc_check', False) if pd.notna(src_row.get('calc_check', False)) else False
    target_row['calc_status'] = src_row.get('calc_status', '') if pd.notna(src_row.get('calc_status', '')) else ''
    target_row['vat_check'] = src_row.get('vat_check', False) if pd.notna(src_row.get('vat_check', False)) else False
    target_row['total_check'] = src_row.get('total_check', False) if pd.notna(src_row.get('total_check', False)) else False

    # line_no 컬럼도 매핑
    target_row['line_no'] = target_row.get('NO')

    # 7. 기타 컬럼 초기화 (대상 파일에 있는 모든 컬럼)
    for col in target_template.columns:
        if col not in target_row:
            target_row[col] = None

    return target_row


def handle_existing_data(target_df: pd.DataFrame, invoice_no: str) -> Tuple[pd.DataFrame, int]:
    """
    기존 데이터 처리 (삭제)

    Args:
        target_df: 대상 DataFrame
        invoice_no: 삭제할 Invoice 번호

    Returns:
        (삭제 후 DataFrame, 삭제된 행 수)
    """
    if 'INVOICE NUMBER' not in target_df.columns:
        print(f"[WARNING] 'INVOICE NUMBER' 컬럼이 없습니다. 기존 데이터 삭제를 건너뜁니다.")
        return target_df, 0

    # 기존 데이터 필터링
    mask = target_df['INVOICE NUMBER'] == invoice_no
    existing_count = mask.sum()

    if existing_count > 0:
        print(f"[INFO] 기존 {invoice_no} 데이터 {existing_count}행 삭제 중...")
        target_df = target_df[~mask].copy()
        print(f"[OK] {existing_count}행 삭제 완료")
    else:
        print(f"[INFO] 기존 {invoice_no} 데이터가 없습니다.")

    return target_df, existing_count


def recalculate_sequence_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """
    순번 재계산 (전체 순번, 청구 회차, NO)
    - INVOICE DATE_YEAR_MONTH, INVOICE NUMBER, NO 순으로 정렬 후 전체 순번 할당
    - 청구 회차: INVOICE DATE_YEAR_MONTH별 그룹 내 INVOICE NUMBER별 첫 항목에 회차 할당
    - NO: INVOICE NUMBER별 그룹 내에서 1부터 순차 할당

    Args:
        df: 대상 DataFrame

    Returns:
        순번이 재계산된 DataFrame
    """
    df = df.copy()

    # 정렬: INVOICE DATE_YEAR_MONTH, INVOICE NUMBER, NO 순으로
    sort_cols = []
    if 'INVOICE DATE_YEAR_MONTH' in df.columns:
        sort_cols.append('INVOICE DATE_YEAR_MONTH')
    if 'INVOICE NUMBER' in df.columns:
        sort_cols.append('INVOICE NUMBER')
    if 'NO' in df.columns:
        sort_cols.append('NO')

    if sort_cols:
        df = df.sort_values(by=sort_cols, na_position='last').reset_index(drop=True)
        print(f"[INFO] 데이터 정렬 완료: {', '.join(sort_cols)}")

    # 전체 순번 재계산 (1부터 시작)
    if '전체 순번' in df.columns:
        df['전체 순번'] = range(1, len(df) + 1)
        print(f"[INFO] 전체 순번 재계산 완료: 1 ~ {len(df)}")

    # 청구 회차 계산 (INVOICE DATE_YEAR_MONTH별 그룹 내 INVOICE NUMBER별 첫 항목에 회차 할당)
    if '청구 회차' in df.columns and 'INVOICE NUMBER' in df.columns:
        if 'INVOICE DATE_YEAR_MONTH' in df.columns:
            # INVOICE DATE_YEAR_MONTH별로 그룹화하여 각 그룹 내에서 INVOICE NUMBER별 첫 항목에 회차 할당
            df['청구 회차'] = 0
            for year_month in df['INVOICE DATE_YEAR_MONTH'].dropna().unique():
                mask = df['INVOICE DATE_YEAR_MONTH'] == year_month
                group_df = df[mask].copy()
                # INVOICE NUMBER별로 첫 번째 행만 회차 할당
                first_invoice_mask = ~group_df['INVOICE NUMBER'].duplicated()
                group_df.loc[first_invoice_mask, '청구 회차'] = range(1, first_invoice_mask.sum() + 1)
                df.loc[mask, '청구 회차'] = group_df['청구 회차']
        else:
            # INVOICE NUMBER별로 그룹핑하여 1부터 시작
            df['청구 회차'] = df.groupby('INVOICE NUMBER').cumcount() + 1
        print(f"[INFO] 청구 회차 재계산 완료")

    # NO 계산 (Invoice 내 순번)
    if 'NO' in df.columns and 'INVOICE NUMBER' in df.columns:
        # Invoice별로 그룹핑하여 1부터 시작
        df['NO'] = df.groupby('INVOICE NUMBER').cumcount() + 1
        print(f"[INFO] NO (Invoice 내 순번) 재계산 완료: {df['NO'].min()} ~ {df['NO'].max()}")

    return df


def validate_data_integrity(df: pd.DataFrame, invoice_no: str) -> List[str]:
    """
    데이터 무결성 검증

    Args:
        df: 검증할 DataFrame
        invoice_no: Invoice 번호

    Returns:
        오류 메시지 리스트
    """
    errors = []
    warnings = []

    # Invoice Number 일치 확인
    if 'INVOICE NUMBER' in df.columns:
        mismatched = df[df['INVOICE NUMBER'] != invoice_no]
        if len(mismatched) > 0:
            errors.append(f"Invoice Number 불일치: {len(mismatched)}행")
    else:
        errors.append("필수 컬럼 누락: INVOICE NUMBER")

    # 필수 컬럼 누락 확인
    required_cols = ['INVOICE NUMBER', 'SUBJECT', 'PRICE CENTER']
    for col in required_cols:
        if col not in df.columns:
            errors.append(f"필수 컬럼 누락: {col}")
        else:
            missing = df[col].isna().sum()
            if missing > 0:
                errors.append(f"{col} 누락: {missing}행")

    # EA 정보 누락 확인
    ea_columns = []
    for i in range(1, 5):
        ea_columns.extend([f'EA_{i}_Name', f'EA_{i}_Qty', f'EA_{i}_Rate', f'EA_{i}_Amount'])
    
    missing_ea_cols = [col for col in ea_columns if col not in df.columns]
    if missing_ea_cols:
        warnings.append(f"EA 정보 컬럼 누락: {len(missing_ea_cols)}개 ({', '.join(missing_ea_cols[:5])}...)")
    else:
        # EA 정보가 있으면 값이 있는지 확인
        ea_with_data = 0
        for i in range(1, 5):
            if f'EA_{i}_Amount' in df.columns:
                ea_with_data += (df[f'EA_{i}_Amount'].notna() & (df[f'EA_{i}_Amount'] != 0)).sum()
        if ea_with_data == 0:
            warnings.append("EA 정보가 모두 비어있습니다.")

    # 날짜 정보 확인
    if 'INVOICE DATE_YEAR_MONTH' not in df.columns:
        warnings.append("INVOICE DATE_YEAR_MONTH 컬럼이 없습니다.")
    else:
        nat_count = df['INVOICE DATE_YEAR_MONTH'].isna().sum()
        if nat_count > 0:
            warnings.append(f"INVOICE DATE_YEAR_MONTH NaT 값: {nat_count}행")

    # 통화 정보 확인
    if 'currency' not in df.columns and 'Currency' not in df.columns:
        warnings.append("통화 정보 컬럼이 없습니다. 기본값 AED를 사용합니다.")

    return errors + warnings


def validate_amounts(df: pd.DataFrame) -> List[str]:
    """
    금액 검증 (개선된 버전)

    Args:
        df: 검증할 DataFrame

    Returns:
        오류 메시지 리스트
    """
    errors = []
    warnings = []

    # EA_Total_Amount = Σ(EA_1~4_Amount) 확인
    if all(col in df.columns for col in ['EA_1_Amount', 'EA_2_Amount', 'EA_3_Amount', 'EA_4_Amount', 'EA_Total_Amount']):
        mismatch_count = 0
        for idx, row in df.iterrows():
            ea_amounts = [
                row.get('EA_1_Amount') or 0,
                row.get('EA_2_Amount') or 0,
                row.get('EA_3_Amount') or 0,
                row.get('EA_4_Amount') or 0
            ]
            # NaN 값 처리
            ea_amounts = [float(amt) if pd.notna(amt) and amt is not None else 0.0 for amt in ea_amounts]
            ea_sum = sum(ea_amounts)
            ea_total = float(row.get('EA_Total_Amount') or 0) if pd.notna(row.get('EA_Total_Amount')) else 0.0

            diff = abs(ea_sum - ea_total)
            if diff > 0.01:
                mismatch_count += 1
                invoice_no = row.get('INVOICE NUMBER', 'N/A')
                line_no = row.get('NO', 'N/A')
                errors.append(
                    f"행 {idx} (Invoice: {invoice_no}, Line: {line_no}): "
                    f"EA 합계 불일치 (계산: {ea_sum:.2f}, EA_Total_Amount: {ea_total:.2f}, 차이: {diff:.2f})"
                )

        if mismatch_count == 0:
            warnings.append(f"[OK] EA_Total_Amount 검증 통과: {len(df)}행 모두 정상")
        else:
            warnings.append(f"[WARNING] EA_Total_Amount 불일치: {mismatch_count}/{len(df)}행")

    # calc_check, vat_check, total_check 확인 (상세 정보 포함)
    if 'calc_check' in df.columns:
        calc_failures = df[~df['calc_check']]
        if len(calc_failures) > 0:
            failure_details = []
            for idx, row in calc_failures.iterrows():
                invoice_no = row.get('INVOICE NUMBER', 'N/A')
                line_no = row.get('NO', 'N/A')
                calc_status = row.get('calc_status', 'N/A')
                failure_details.append(f"행 {idx} (Invoice: {invoice_no}, Line: {line_no}, Status: {calc_status})")

            errors.append(f"calc_check 실패: {len(calc_failures)}행")
            # 처음 5개만 상세 정보 추가
            for detail in failure_details[:5]:
                errors.append(f"  - {detail}")
            if len(failure_details) > 5:
                errors.append(f"  ... 및 {len(failure_details) - 5}개 더")
        else:
            warnings.append(f"[OK] calc_check 검증 통과: {len(df)}행 모두 정상")

    if 'vat_check' in df.columns:
        vat_failures = df[~df['vat_check']]
        if len(vat_failures) > 0:
            failure_details = []
            for idx, row in vat_failures.iterrows():
                invoice_no = row.get('INVOICE NUMBER', 'N/A')
                line_no = row.get('NO', 'N/A')
                failure_details.append(f"행 {idx} (Invoice: {invoice_no}, Line: {line_no})")

            errors.append(f"vat_check 실패: {len(vat_failures)}행")
            for detail in failure_details[:5]:
                errors.append(f"  - {detail}")
            if len(failure_details) > 5:
                errors.append(f"  ... 및 {len(failure_details) - 5}개 더")
        else:
            warnings.append(f"[OK] vat_check 검증 통과: {len(df)}행 모두 정상")

    if 'total_check' in df.columns:
        total_failures = df[~df['total_check']]
        if len(total_failures) > 0:
            failure_details = []
            for idx, row in total_failures.iterrows():
                invoice_no = row.get('INVOICE NUMBER', 'N/A')
                line_no = row.get('NO', 'N/A')
                failure_details.append(f"행 {idx} (Invoice: {invoice_no}, Line: {line_no})")

            errors.append(f"total_check 실패: {len(total_failures)}행")
            for detail in failure_details[:5]:
                errors.append(f"  - {detail}")
            if len(failure_details) > 5:
                errors.append(f"  ... 및 {len(failure_details) - 5}개 더")
        else:
            warnings.append(f"[OK] total_check 검증 통과: {len(df)}행 모두 정상")

    # warnings를 errors에 추가 (정보성)
    return errors + warnings


def integrate_data(
    source_file: str,
    target_file: str,
    source_sheet: str = 'EA_Slots',
    backup: bool = True
) -> Tuple[bool, Dict[str, Any]]:
    """
    메인 통합 함수

    Args:
        source_file: 소스 파일 경로
        target_file: 대상 파일 경로
        source_sheet: 소스 시트명 (기본값: 'EA_Slots')
        backup: 백업 파일 생성 여부

    Returns:
        (성공 여부, 통계 정보)
    """
    stats = {
        'source_rows': 0,
        'target_rows_before': 0,
        'target_rows_after': 0,
        'deleted_rows': 0,
        'added_rows': 0,
        'invoice_no': '',
        'errors': [],
        'warnings': [],
        'validation_errors': []
    }

    try:
        # 1. 파일 존재 확인
        source_path = Path(source_file)
        target_path = Path(target_file)

        if not source_path.exists():
            raise FileNotFoundError(f"소스 파일을 찾을 수 없습니다: {source_file}")

        if not target_path.exists():
            raise FileNotFoundError(f"대상 파일을 찾을 수 없습니다: {target_file}")

        print(f"[INFO] 소스 파일: {source_path}")
        print(f"[INFO] 대상 파일: {target_path}")

        # 2. 백업 생성
        if backup:
            backup_path = target_path.with_suffix('.xlsx.backup')
            shutil.copy2(target_path, backup_path)
            print(f"[OK] 백업 파일 생성: {backup_path}")

        # 3. 소스 파일 읽기
        print(f"\n[Step 1/6] 소스 파일 읽기...")
        print(f"[INFO] 소스 시트: {source_sheet}")
        src_df = pd.read_excel(source_path, sheet_name=source_sheet)
        stats['source_rows'] = len(src_df)
        
        # Invoice Number 추출 (여러 컬럼명 시도)
        invoice_no = ''
        if 'invoice_no' in src_df.columns and len(src_df) > 0:
            invoice_no = src_df['invoice_no'].iloc[0]
        elif 'INVOICE NUMBER' in src_df.columns and len(src_df) > 0:
            invoice_no = src_df['INVOICE NUMBER'].iloc[0]
        
        stats['invoice_no'] = invoice_no
        invoice_currency = None
        
        # Standard_Lines 사용 시 추가 데이터 병합 (날짜, EA 정보)
        if source_sheet == 'Standard_Lines':
            # 1. 날짜 정보 병합
            if 'invoice_date' not in src_df.columns:
                invoice_date = None
                date_source = None
                
                # Complete_Pipeline에서 날짜 가져오기 시도
                try:
                    complete_df = pd.read_excel(source_path, sheet_name='Complete_Pipeline')
                    if len(complete_df) > 0:
                        if 'invoice_date' in complete_df.columns:
                            invoice_date = complete_df['invoice_date'].iloc[0]
                            if invoice_date is not None and not pd.isna(invoice_date):
                                date_source = 'Complete_Pipeline'
                                print(f"[INFO] Complete_Pipeline에서 날짜 가져옴: {invoice_date}")
                        if 'currency' in complete_df.columns:
                            non_null_currency = complete_df['currency'].dropna()
                            if len(non_null_currency) > 0:
                                invoice_currency = non_null_currency.iloc[0]
                                print(f"[INFO] Complete_Pipeline에서 통화 가져옴: {invoice_currency}")
                except FileNotFoundError:
                    print(f"[INFO] Complete_Pipeline 시트가 없습니다. Invoice 번호로 매핑 시도...")
                except Exception as e:
                    print(f"[WARNING] Complete_Pipeline에서 날짜 가져오기 실패: {e}")
                    print(f"[INFO] Invoice 번호로 매핑 시도...")
                
                # Complete_Pipeline이 없거나 실패하면 Invoice 번호로 날짜 매핑
                if invoice_date is None or pd.isna(invoice_date):
                    invoice_date_map = {
                        'OFCO-INV-0001178': '15 October, 2025'
                    }
                    if invoice_no in invoice_date_map:
                        invoice_date = invoice_date_map[invoice_no]
                        date_source = 'Invoice_Number_Mapping'
                        print(f"[INFO] Invoice 번호로 날짜 매핑: {invoice_date}")
                    else:
                        print(f"[WARNING] Invoice 번호 '{invoice_no}'에 대한 날짜 매핑이 없습니다.")
                        print(f"[WARNING] 날짜 정보 없이 진행합니다. (INVOICE DATE_YEAR_MONTH가 NaT일 수 있음)")
                
                # 날짜를 모든 행에 추가
                if invoice_date is not None and not pd.isna(invoice_date):
                    src_df['invoice_date'] = invoice_date
                    print(f"[OK] invoice_date 컬럼 추가: {invoice_date} (소스: {date_source})")
                else:
                    # 날짜가 없어도 계속 진행 (경고만)
                    stats['warnings'].append("invoice_date 컬럼이 추가되지 않았습니다. INVOICE DATE_YEAR_MONTH가 NaT일 수 있습니다.")
            
            # 2. EA_Slots 정보 병합 (EA_1~EA_4_Name, Qty, Rate, Amount)
            ea_columns = []
            for i in range(1, 5):
                ea_columns.extend([f'EA_{i}_Name', f'EA_{i}_Qty', f'EA_{i}_Rate', f'EA_{i}_Amount'])
            
            # Standard_Lines에 EA 정보가 없으면 EA_Slots에서 가져오기
            missing_ea_cols = [col for col in ea_columns if col not in src_df.columns]
            if missing_ea_cols:
                print(f"[INFO] Standard_Lines에 EA 정보가 없습니다. EA_Slots 시트에서 병합 시도...")
                try:
                    ea_df = pd.read_excel(source_path, sheet_name='EA_Slots')
                    if len(ea_df) > 0:
                        # line_no 또는 invoice_no 기준으로 병합
                        merge_key = 'line_no' if 'line_no' in src_df.columns and 'line_no' in ea_df.columns else None
                        if not merge_key:
                            merge_key = 'invoice_no' if 'invoice_no' in src_df.columns and 'invoice_no' in ea_df.columns else None

                        ea_cols_to_merge = [col for col in ea_columns if col in ea_df.columns]
                        helper_merge_col = None

                        if not ea_cols_to_merge:
                            print(f"[WARNING] EA_Slots에 EA 컬럼이 없습니다.")
                        else:
                            if merge_key:
                                # ⚠️ 중복 방지: EA_Slots에서 각 merge_key의 첫 번째 행만 사용
                                ea_df_unique = ea_df.drop_duplicates(subset=[merge_key], keep='first')
                                print(f"[INFO] EA_Slots 중복 제거: {len(ea_df)}행 → {len(ea_df_unique)}행 ({merge_key} 기준 첫 번째 행만 사용)")
                            else:
                                # line_no / invoice_no가 없으면 행 순서 기반으로 매핑
                                print(f"[WARNING] EA_Slots와 Standard_Lines 병합 키를 찾을 수 없습니다. 행 순서 기반 병합을 시도합니다.")
                                if len(src_df) != len(ea_df):
                                    print(f"[WARNING] 행 수 불일치 (Standard_Lines: {len(src_df)}행, EA_Slots: {len(ea_df)}행). 첫 {len(src_df)}행만 사용합니다.")
                                ea_df_unique = ea_df.reset_index(drop=True).iloc[:len(src_df)].copy()
                                src_df = src_df.reset_index(drop=True)
                                helper_merge_col = '__ea_merge_row_index__'
                                ea_df_unique[helper_merge_col] = range(len(ea_df_unique))
                                src_df[helper_merge_col] = range(len(src_df))
                                merge_key = helper_merge_col

                            if merge_key:
                                src_df = src_df.merge(
                                    ea_df_unique[[merge_key] + ea_cols_to_merge],
                                    on=merge_key,
                                    how='left',
                                    suffixes=('', '_ea')
                                )

                                # _ea 접미사가 붙은 컬럼을 원래 이름으로 변경
                                for col in ea_cols_to_merge:
                                    if f'{col}_ea' in src_df.columns:
                                        src_df[col] = src_df[f'{col}_ea']
                                        src_df = src_df.drop(columns=[f'{col}_ea'])

                                if helper_merge_col and helper_merge_col in src_df.columns:
                                    src_df = src_df.drop(columns=[helper_merge_col])

                                print(f"[OK] EA_Slots에서 {len(ea_cols_to_merge)}개 EA 컬럼 병합 완료")
                    else:
                        print(f"[WARNING] EA_Slots 시트가 비어있습니다.")
                except FileNotFoundError:
                    print(f"[WARNING] EA_Slots 시트가 없습니다. EA 정보 없이 진행합니다.")
                    stats['warnings'].append("EA_Slots 시트가 없어 EA 정보가 누락될 수 있습니다.")
                except Exception as e:
                    print(f"[WARNING] EA_Slots 읽기 실패: {e}. EA 정보 없이 진행합니다.")
                    stats['warnings'].append(f"EA_Slots 읽기 실패: {e}")
            else:
                print(f"[INFO] Standard_Lines에 EA 정보가 이미 포함되어 있습니다.")

            # 3. 통화 정보 보강 (기본값 AED, Complete_Pipeline 우선)
            default_currency = invoice_currency or 'AED'
            if 'currency' not in src_df.columns:
                src_df['currency'] = default_currency
                print(f"[INFO] currency 컬럼이 없어 기본값 '{default_currency}'로 채웠습니다.")
            else:
                currency_series = src_df['currency']
                missing_currency = currency_series.isna() | (currency_series.astype(str).str.strip() == '')
                if missing_currency.any():
                    src_df.loc[missing_currency, 'currency'] = default_currency
                    print(f"[INFO] currency 결측 {missing_currency.sum()}건을 '{default_currency}'로 보완했습니다.")

            stats['source_currency'] = default_currency
        
        print(f"[OK] {stats['source_rows']}행, {len(src_df.columns)}컬럼 로드됨")
        print(f"[INFO] Invoice: {stats['invoice_no']}")

        # 4. 대상 파일 읽기
        print(f"\n[Step 2/6] 대상 파일 읽기...")
        # Sheet1 시트 사용 (원본 파일의 시트명)
        target_df = pd.read_excel(target_path, sheet_name='Sheet1')
        stats['target_rows_before'] = len(target_df)
        print(f"[OK] {stats['target_rows_before']}행, {len(target_df.columns)}컬럼 로드됨")

        # 5. 기존 데이터 삭제
        print(f"\n[Step 3/6] 기존 데이터 처리...")
        target_df, deleted_count = handle_existing_data(target_df, stats['invoice_no'])
        stats['deleted_rows'] = deleted_count

        # 6. 데이터 변환
        print(f"\n[Step 4/6] 데이터 변환 중...")
        new_rows = []
        for idx, src_row in src_df.iterrows():
            try:
                target_row = transform_row(src_row, target_df)
                new_rows.append(target_row)
            except Exception as e:
                error_msg = f"행 {idx} 변환 실패: {e}"
                print(f"[ERROR] {error_msg}")
                stats['errors'].append(error_msg)

        if not new_rows:
            raise ValueError("변환된 데이터가 없습니다.")

        # 7. 새 DataFrame 생성
        new_df = pd.DataFrame(new_rows)
        stats['added_rows'] = len(new_df)
        print(f"[OK] {stats['added_rows']}행 변환 완료")

        # 8. HeaderCore를 사용한 컬럼 정렬 (146개 표준 헤더 순서)
        print(f"\n[Step 5/6] 데이터 병합 및 컬럼 정렬 중...")
        
        # HeaderCore 사용하여 컬럼 정렬
        if HeaderCore is not None:
            try:
                header_core = get_header_core(reference_file=str(target_path))
                print(f"[HeaderCore] 기준 헤더 로드 완료: {len(header_core.base_headers)}개 컬럼")
                
                # 새 데이터를 HeaderCore 기준으로 정렬
                new_df = header_core.reorder_dataframe(new_df, phase=None, fill_missing=True, map_columns=True)
                print(f"[HeaderCore] 새 데이터 컬럼 정렬 완료: {len(new_df.columns)}개 컬럼")
                
                # 기존 데이터도 HeaderCore 기준으로 정렬
                target_df = header_core.reorder_dataframe(target_df, phase=None, fill_missing=True, map_columns=True)
                print(f"[HeaderCore] 기존 데이터 컬럼 정렬 완료: {len(target_df.columns)}개 컬럼")
            except Exception as e:
                print(f"[WARNING] HeaderCore 정렬 실패: {e}")
                print(f"[INFO] 기본 컬럼 순서로 진행합니다.")
        
        # 컬럼 순서 맞추기 (HeaderCore 실패 시 fallback)
        if len(new_df.columns) != len(target_df.columns) or not all(new_df.columns == target_df.columns):
            new_df = new_df.reindex(columns=target_df.columns, fill_value=None)
        
        # 기존 데이터와 병합 (정렬 기준으로 삽입)
        # INVOICE DATE_YEAR_MONTH 기준으로 정렬하여 적절한 위치에 삽입
        if 'INVOICE DATE_YEAR_MONTH' in target_df.columns and 'INVOICE DATE_YEAR_MONTH' in new_df.columns:
            # 정렬 기준 컬럼 준비
            sort_cols = ['INVOICE DATE_YEAR_MONTH']
            if 'INVOICE NUMBER' in target_df.columns and 'INVOICE NUMBER' in new_df.columns:
                sort_cols.append('INVOICE NUMBER')
            if 'NO' in target_df.columns and 'NO' in new_df.columns:
                sort_cols.append('NO')
            
            # 기존 데이터와 새 데이터 병합
            combined_df = pd.concat([target_df, new_df], ignore_index=True)
            
            # 정렬 수행
            combined_df = combined_df.sort_values(by=sort_cols, na_position='last').reset_index(drop=True)
            print(f"[INFO] 데이터 정렬 후 병합 완료: {sort_cols} 기준")
        else:
            # 정렬 기준이 없으면 단순 병합
            combined_df = pd.concat([target_df, new_df], ignore_index=True)
            print(f"[INFO] 단순 병합 완료 (정렬 기준 컬럼 없음)")
        
        stats['target_rows_after'] = len(combined_df)
        print(f"[OK] 병합 완료: {stats['target_rows_after']}행")

        # 9. 순번 재계산
        print(f"\n[Step 6/6] 순번 재계산 중...")
        combined_df = recalculate_sequence_numbers(combined_df)

        # 9-1. 새 데이터 검증
        print(f"\n[INFO] 새 데이터 검증 중...")
        validation_df = new_df.copy()
        if 'currency' not in validation_df.columns and 'Currency' not in validation_df.columns:
            validation_df['currency'] = stats.get('source_currency', 'AED')

        integrity_results = validate_data_integrity(validation_df, stats['invoice_no'])
        amount_results = validate_amounts(validation_df)

        # 오류와 경고 분리
        integrity_errors = [r for r in integrity_results if not r.startswith('[OK]') and not r.startswith('[WARNING]')]
        integrity_warnings = [r for r in integrity_results if r.startswith('[WARNING]') or r.startswith('[OK]')]
        
        amount_errors = [r for r in amount_results if not r.startswith('[OK]') and not r.startswith('[WARNING]')]
        amount_warnings = [r for r in amount_results if r.startswith('[WARNING]') or r.startswith('[OK]')]

        if integrity_errors:
            stats['validation_errors'].extend(integrity_errors)
            print(f"[ERROR] 무결성 검증 오류: {len(integrity_errors)}개")
            for error in integrity_errors[:5]:
                print(f"  - {error}")
            if len(integrity_errors) > 5:
                print(f"  ... 및 {len(integrity_errors) - 5}개 더")

        if integrity_warnings:
            print(f"[INFO] 무결성 검증 정보:")
            for warning in integrity_warnings[:5]:
                print(f"  {warning}")

        if amount_errors:
            stats['validation_errors'].extend(amount_errors)
            print(f"[ERROR] 금액 검증 오류: {len(amount_errors)}개")
            for error in amount_errors[:5]:
                print(f"  - {error}")
            if len(amount_errors) > 5:
                print(f"  ... 및 {len(amount_errors) - 5}개 더")

        if amount_warnings:
            print(f"[INFO] 금액 검증 정보:")
            for warning in amount_warnings[:5]:
                print(f"  {warning}")

        if not integrity_errors and not amount_errors:
            print(f"[OK] 모든 검증 통과")

        # 10. 파일 저장
        print(f"\n[INFO] 파일 저장 중...")
        try:
            # 기존 파일의 다른 시트 유지를 위해 openpyxl로 로드
            from openpyxl import load_workbook
            wb = load_workbook(target_path)

            # Sheet1 시트 업데이트 (원본 파일의 시트명)
            if 'Sheet1' in wb.sheetnames:
                ws = wb['Sheet1']
                # 기존 데이터 삭제
                ws.delete_rows(1, ws.max_row)
            else:
                # Sheet1이 없으면 생성
                ws = wb.create_sheet('Sheet1')

            # DataFrame을 시트에 쓰기
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r in dataframe_to_rows(combined_df, index=False, header=True):
                ws.append(r)

            # 파일 저장 (임시 파일로 저장 후 이동하여 안전성 확보)
            temp_path = target_path.with_suffix('.xlsx.tmp')
            try:
                wb.save(temp_path)
                wb.close()
                
                # 임시 파일을 원본 파일로 이동
                if temp_path.exists():
                    shutil.move(str(temp_path), str(target_path))
                    print(f"[OK] 파일 저장 완료: {target_path}")
                else:
                    raise Exception("임시 파일 저장 후 이동 실패")
            except Exception as save_error:
                # 임시 파일 정리
                if temp_path.exists():
                    try:
                        temp_path.unlink()
                    except:
                        pass
                raise save_error
                
        except Exception as save_error:
            error_msg = f"파일 저장 실패: {save_error}"
            print(f"[ERROR] {error_msg}")
            stats['errors'].append(error_msg)
            
            # 롤백 시도
            if backup and backup_path.exists():
                print(f"[INFO] 저장 실패로 인한 롤백 시도...")
                try:
                    shutil.copy2(backup_path, target_path)
                    print(f"[OK] 롤백 완료")
                except Exception as rollback_error:
                    print(f"[ERROR] 롤백 실패: {rollback_error}")
                    stats['errors'].append(f"롤백 실패: {rollback_error}")
            
            raise  # 저장 실패는 치명적 오류이므로 재발생

        # 11. 통계 출력
        print(f"\n{'='*80}")
        print(f"통합 완료!")
        print(f"{'='*80}")
        print(f"소스 행 수: {stats['source_rows']}")
        print(f"삭제된 행 수: {stats['deleted_rows']}")
        print(f"추가된 행 수: {stats['added_rows']}")
        print(f"대상 행 수 (변경 전): {stats['target_rows_before']}")
        print(f"대상 행 수 (변경 후): {stats['target_rows_after']}")
        print(f"행 수 변화: {stats['target_rows_after'] - stats['target_rows_before']} ({stats['target_rows_after'] - stats['target_rows_before']:+d})")

        if stats['errors']:
            print(f"\n[ERROR] {len(stats['errors'])}개 오류 발생:")
            for error in stats['errors'][:5]:  # 처음 5개만 출력
                print(f"  - {error}")

        if stats['validation_errors']:
            # 오류와 경고 분리
            actual_errors = [e for e in stats['validation_errors'] if not e.startswith('[OK]') and not e.startswith('[WARNING]')]
            warnings = [e for e in stats['validation_errors'] if e.startswith('[OK]') or e.startswith('[WARNING]')]

            if actual_errors:
                print(f"\n[WARNING] {len(actual_errors)}개 검증 오류:")
                for error in actual_errors[:10]:  # 처음 10개 출력
                    print(f"  {error}")
                if len(actual_errors) > 10:
                    print(f"  ... 및 {len(actual_errors) - 10}개 더")

            if warnings:
                print(f"\n[INFO] 검증 결과:")
                for warning in warnings:
                    print(f"  {warning}")

        return True, stats

    except FileNotFoundError as e:
        error_msg = f"파일을 찾을 수 없습니다: {e}"
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)
        return False, stats
    except PermissionError as e:
        error_msg = f"파일 접근 권한 오류: {e}. 파일이 열려있는지 확인하세요."
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)

        if backup and backup_path.exists():
            print(f"\n[INFO] 롤백 시도 중...")
            try:
                shutil.copy2(backup_path, target_path)
                print(f"[OK] 롤백 완료")
            except Exception as rollback_error:
                print(f"[ERROR] 롤백 실패: {rollback_error}")
                stats['errors'].append(f"롤백 실패: {rollback_error}")
        return False, stats
    except pd.errors.EmptyDataError as e:
        error_msg = f"빈 데이터 파일: {e}"
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)
        return False, stats
    except KeyError as e:
        error_msg = f"필수 컬럼 누락: {e}"
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)
        return False, stats
    except ValueError as e:
        error_msg = f"데이터 형식 오류: {e}"
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)
        return False, stats
    except Exception as e:
        error_msg = f"예상치 못한 오류: {type(e).__name__} - {e}"
        print(f"\n[ERROR] {error_msg}")
        stats['errors'].append(error_msg)

        # 롤백 (백업 파일이 있으면)
        if backup and backup_path.exists():
            print(f"\n[INFO] 롤백 시도 중...")
            try:
                # 롤백 전 원본 파일 삭제 시도 (읽기 전용 파일 등 문제 방지)
                if target_path.exists():
                    try:
                        target_path.unlink()
                    except PermissionError:
                        print(f"[WARNING] 원본 파일 삭제 실패 (읽기 전용일 수 있음). 덮어쓰기 시도...")
                    except Exception as del_error:
                        print(f"[WARNING] 원본 파일 삭제 실패: {del_error}. 덮어쓰기 시도...")
                
                # 백업 파일 복원
                shutil.copy2(backup_path, target_path)
                
                # 복원 검증
                if target_path.exists():
                    restored_size = target_path.stat().st_size
                    backup_size = backup_path.stat().st_size
                    if restored_size == backup_size:
                        print(f"[OK] 롤백 완료: {target_path} ({restored_size} bytes)")
                    else:
                        print(f"[WARNING] 롤백 파일 크기 불일치 (복원: {restored_size}, 백업: {backup_size})")
                else:
                    raise Exception("롤백 후 파일이 존재하지 않습니다")
                    
            except PermissionError as perm_error:
                print(f"[ERROR] 롤백 실패: 파일 권한 오류 - {perm_error}")
                print(f"[ERROR] 수동으로 백업 파일을 복원하세요: {backup_path} → {target_path}")
                stats['errors'].append(f"롤백 실패 (권한 오류): {perm_error}. 수동 복원 필요.")
            except Exception as rollback_error:
                print(f"[ERROR] 롤백 실패: {rollback_error}")
                print(f"[ERROR] 수동으로 백업 파일을 복원하세요: {backup_path} → {target_path}")
                stats['errors'].append(f"롤백 실패: {rollback_error}. 수동 복원 필요.")
        elif backup and not backup_path.exists():
            print(f"[WARNING] 백업 파일이 없습니다: {backup_path}")
            print(f"[WARNING] 롤백을 수행할 수 없습니다.")

        return False, stats


def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(description='OFCO Pipeline Standard_Lines → Sheet1 시트 통합 (v2.0)')
    parser.add_argument('--source', type=str, default='ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx',
                        help='소스 파일 경로')
    parser.add_argument('--source-sheet', type=str, default='Standard_Lines',
                        choices=['Complete_Pipeline', 'EA_Slots', 'Standard_Lines'],
                        help='소스 시트명 (기본값: Standard_Lines)')
    parser.add_argument('--target', type=str, default='OFCO INVOICE.xlsx',
                        help='대상 파일 경로')
    parser.add_argument('--backup', action='store_true', default=True,
                        help='백업 파일 생성 (기본값: True)')
    parser.add_argument('--no-backup', action='store_false', dest='backup',
                        help='백업 파일 생성 안 함')

    args = parser.parse_args()

    print("="*80)
    print("OFCO Pipeline Complete_Pipeline/EA_Slots → LIST 시트 통합")
    print("="*80)
    print()

    success, stats = integrate_data(
        source_file=args.source,
        target_file=args.target,
        source_sheet=args.source_sheet,
        backup=args.backup
    )

    if success:
        print(f"\n[SUCCESS] 통합이 성공적으로 완료되었습니다!")
        sys.exit(0)
    else:
        print(f"\n[FAILURE] 통합 중 오류가 발생했습니다.")
        sys.exit(1)


if __name__ == '__main__':
    main()

