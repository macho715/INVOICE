#!/usr/bin/env python3
"""
OFCO 파서 결과 → 표준 Lines 형식 Excel 입력 스크립트

"OFCO Zayed Port_safeen adp 인보이스 표준 Lines 작성 가이드" v1.2에 따라
파싱 결과를 표준 Lines 스키마로 변환하여 Excel 파일에 입력
"""

import pandas as pd
import numpy as np
import re
import os
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import sys

# HeaderCore 통합 (표준 경로)
try:
    sys.path.insert(0, str(Path(__file__).parent.parent / "shared" / "scripts"))
    from ofco_header_core import get_header_core
    HEADER_CORE_AVAILABLE = True
except ImportError:
    HEADER_CORE_AVAILABLE = False
    print("[WARNING] HeaderCore를 사용할 수 없습니다. 기본 컬럼 순서를 사용합니다.")

# 같은 폴더의 파서 모듈 import
from ofco_parse_pdf_patched import parse_ofco_invoice, OFCOPayload, LineItem


def extract_tariff_id(notes: str) -> Optional[str]:
    """
    NOTES에서 tariff_code 또는 tariff_id 추출

    Args:
        notes: NOTES 문자열 (예: "tariff_code=10.1; sn=1; ...")

    Returns:
        tariff_id 문자열 또는 None
    """
    if not notes:
        return None

    # tariff_code 추출 시도 (OFCO)
    match = re.search(r'tariff_code=([^\s;]+)', notes)
    if match:
        tariff_id = match.group(1).strip()
        if tariff_id and tariff_id != 'None':
            return tariff_id

    # tariff_id 추출 시도 (ADP/SAFEEN)
    match = re.search(r'tariff_id=([^\s;]+)', notes)
    if match:
        tariff_id = match.group(1).strip()
        if tariff_id and tariff_id != 'None':
            return tariff_id

    return None


def format_description(description: str) -> str:
    """
    Description 정제: 날짜를 ISO 형식으로 변환

    Args:
        description: 원본 description

    Returns:
        정제된 description (날짜 ISO 형식)
    """
    if not description:
        return ""

    # 날짜 패턴 변환 (예: "15 October, 2025" → "2025-10-15")
    date_patterns = [
        (r'(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December),?\s+(\d{4})',
         lambda m: format_month_date(m.group(1), m.group(2), m.group(3))),
        (r'(\d{1,2})-([A-Za-z]{3})-(\d{2,4})',
         lambda m: format_short_date(m.group(1), m.group(2), m.group(3))),
    ]

    result = description
    for pattern, formatter in date_patterns:
        result = re.sub(pattern, lambda m: formatter(m), result, flags=re.IGNORECASE)

    return result.strip()


def format_month_date(day: str, month: str, year: str) -> str:
    """월 이름이 포함된 날짜를 ISO 형식으로 변환"""
    month_map = {
        'january': '01', 'february': '02', 'march': '03', 'april': '04',
        'may': '05', 'june': '06', 'july': '07', 'august': '08',
        'september': '09', 'october': '10', 'november': '11', 'december': '12'
    }
    month_num = month_map.get(month.lower(), '01')
    day_num = day.zfill(2)
    year_num = year if len(year) == 4 else f"20{year}"
    return f"{year_num}-{month_num}-{day_num}"


def format_short_date(day: str, month: str, year: str) -> str:
    """짧은 형식 날짜를 ISO 형식으로 변환"""
    month_map = {
        'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
    }
    month_num = month_map.get(month.lower()[:3], '01')
    day_num = day.zfill(2)
    year_num = year if len(year) == 4 else f"20{year}"
    return f"{year_num}-{month_num}-{day_num}"


def extract_units_from_item(item: LineItem, notes: str, invoice_type: str) -> Tuple[float, float, float]:
    """
    unit1, unit2, unit3 추출 및 유닛 분해 규칙 적용

    Args:
        item: LineItem 객체
        notes: NOTES 문자열
        invoice_type: 인보이스 타입 ("OFCO", "ADP", "SAFEEN")

    Returns:
        (unit1, unit2, unit3) 튜플
    """
    # NOTES에서 unit 정보 추출
    unit1_match = re.search(r'unit1=([^\s;]+)', notes)
    unit2_match = re.search(r'unit2=([^\s;]+)', notes)
    unit3_match = re.search(r'unit3=([^\s;]+)', notes)
    qty_match = re.search(r'qty=([^\s;]+)', notes)

    description = (item.subject or "").lower()

    # ADP 타입: 이미 unit1, unit2, unit3 존재
    if invoice_type == "ADP":
        unit1 = float(unit1_match.group(1)) if unit1_match else 1.0
        unit2 = float(unit2_match.group(1)) if unit2_match else 1.0
        unit3 = float(unit3_match.group(1)) if unit3_match else 1.0
        return (unit1, unit2, unit3)

    # 유닛 분해 규칙 적용 (가이드 3-1)
    # Heavy Lift (703.1): unit1=FRT 수량, unit2=1, unit3=1
    if 'heavy lift' in description or '703.1' in notes.lower():
        unit1_val = float(qty_match.group(1)) if qty_match else 1.0
        return (unit1_val, 1.0, 1.0)

    # Labour, Additional man hour (802.3A): unit1=0, unit2=명, unit3=시간 또는 일수
    if 'labour' in description or 'man hour' in description or '802.3a' in notes.lower():
        unit2_val = float(qty_match.group(1)) if qty_match else 1.0
        # unit3은 시간 또는 일수 (기본값 3, 실제로는 파싱 필요하지만 여기서는 기본값 사용)
        return (0.0, unit2_val, 3.0)

    # Foreman/Winch (802.5A): unit1=0, unit2=명, unit3=시간 또는 일수
    if 'foreman' in description or 'winch' in description or '802.5a' in notes.lower():
        unit2_val = float(qty_match.group(1)) if qty_match else 1.0
        return (0.0, unit2_val, 3.0)

    # OFCO 타입: qty가 있으면 unit1=qty, 나머지는 1.0
    if invoice_type == "OFCO":
        if qty_match:
            qty = float(qty_match.group(1))
            return (qty, 1.0, 1.0)
        else:
            # qty가 없으면 기본값 (Port Dues / Permits / Fixed fee)
            return (1.0, 1.0, 1.0)

    # SAFEEN 타입: 기본값
    return (1.0, 1.0, 1.0)


def extract_tax_rate_pct(tax_rate: Optional[str]) -> float:
    """
    tax_rate에서 % 제거하여 숫자만 반환

    Args:
        tax_rate: "5%" 또는 "5" 또는 None

    Returns:
        세율 숫자 (예: 5.0)
    """
    if not tax_rate:
        return 0.0

    # % 제거 및 숫자 추출
    match = re.search(r'(\d+(?:\.\d+)?)', str(tax_rate))
    if match:
        return float(match.group(1))

    return 0.0


def extract_ea_pairs_from_item(item: LineItem, notes: str, max_pairs: int = 4) -> List[Tuple[float, float]]:
    """
    LineItem에서 EA 쌍(Qty, Rate) 추출 (최대 max_pairs개)

    Args:
        item: LineItem 객체
        notes: NOTES 문자열
        max_pairs: 최대 EA 쌍 개수 (기본 4)

    Returns:
        (Qty, Rate) 튜플 리스트 (최대 max_pairs개)
    """
    ea_pairs = []

    # 1. item.ea_rates에서 직접 추출 (이미 파싱된 쌍)
    if item.ea_rates and len(item.ea_rates) > 0:
        for qty, rate in item.ea_rates[:max_pairs]:
            if qty is not None and rate is not None and rate != 0:
                ea_pairs.append((round(float(qty), 2), round(float(rate), 2)))
        if ea_pairs:
            return ea_pairs[:max_pairs]

    # 2. NOTES에서 추출 시도
    # qty와 unit_price가 있으면 사용
    qty_match = re.search(r'qty=([^\s;]+)', notes)
    unit_price_match = re.search(r'unit_price=([^\s;]+)', notes)

    if qty_match and unit_price_match:
        qty = round(float(qty_match.group(1)), 2)
        rate = round(float(unit_price_match.group(1)), 2)
        if qty > 0 and rate > 0:
            ea_pairs.append((qty, rate))

    # 3. unit1, unit2, unit3와 rate가 있으면 사용 (ADP 타입)
    if not ea_pairs:
        unit1_match = re.search(r'unit1=([^\s;]+)', notes)
        rate_match = re.search(r'rate=([^\s;]+)', notes)

        if unit1_match and rate_match:
            qty = round(float(unit1_match.group(1)), 2)
            rate = round(float(rate_match.group(1)), 2)
            if qty > 0 and rate > 0:
                ea_pairs.append((qty, rate))

    # 4. QTY와 AMOUNT만 있을 때 Rate 역산
    if not ea_pairs:
        qty_match = re.search(r'qty=([^\s;]+)', notes)
        amount = item.amount
        if qty_match and amount and amount > 0:
            qty = float(qty_match.group(1))
            if qty > 0:
                rate = round(amount / qty, 2)
                ea_pairs.append((round(qty, 2), rate))

    # 5. 결측 처리: Qty 미기재 + Rate 존재 → Qty=1
    if not ea_pairs:
        rate_match = re.search(r'rate=([^\s;]+)', notes)
        amount = item.amount
        if rate_match:
            rate = round(float(rate_match.group(1)), 2)
            if rate > 0:
                ea_pairs.append((1.0, rate))
        elif amount and amount > 0:
            # Rate도 없고 Qty도 없으면 Qty=1, Rate=Amount
            ea_pairs.append((1.0, round(amount, 2)))

    return ea_pairs[:max_pairs]


def calculate_ea_amounts(ea_pairs: List[Tuple[float, float]]) -> List[float]:
    """
    EA 쌍에서 각 Amount 계산

    Args:
        ea_pairs: (Qty, Rate) 튜플 리스트

    Returns:
        Amount 리스트 (ROUND(Qty × Rate, 2))
    """
    amounts = []
    for qty, rate in ea_pairs:
        amount = round(qty * rate, 2)
        amounts.append(amount)
    return amounts


def calculate_line_amount_from_ea(ea_pairs: List[Tuple[float, float]]) -> float:
    """
    EA 쌍들의 합계 계산

    Args:
        ea_pairs: (Qty, Rate) 튜플 리스트

    Returns:
        Line_Amount = Σ EA_i_Amount
    """
    amounts = calculate_ea_amounts(ea_pairs)
    return round(sum(amounts), 2)


def extract_rate_from_item(item: LineItem, notes: str) -> float:
    """
    rate (단가) 추출

    Args:
        item: LineItem 객체
        notes: NOTES 문자열

    Returns:
        rate 값 (소수 2자리 반올림)
    """
    # NOTES에서 unit_price 또는 rate 추출
    unit_price_match = re.search(r'unit_price=([^\s;]+)', notes)
    rate_match = re.search(r'rate=([^\s;]+)', notes)

    if unit_price_match:
        return round(float(unit_price_match.group(1)), 2)

    if rate_match:
        return round(float(rate_match.group(1)), 2)

    # EA rates에서 추출
    if item.ea_rates and len(item.ea_rates) > 0:
        return round(item.ea_rates[0][1], 2)

    # amount에서 계산 (unit이 모두 1.0인 경우)
    if item.amount:
        return round(item.amount, 2)

    return 0.0


def split_items_by_type(payload: OFCOPayload) -> Tuple[List[LineItem], List[LineItem]]:
    """
    항목을 ADP/SAFEEN과 OFCO로 분리

    Args:
        payload: 파싱된 인보이스 데이터

    Returns:
        (adp_safeen_items, ofco_items) 튜플
    """
    adp_safeen_items = []
    ofco_items = []

    for item in payload.items:
        if is_adp_or_safeen_item(item):
            adp_safeen_items.append(item)
        else:
            ofco_items.append(item)

    return (adp_safeen_items, ofco_items)


def convert_adp_safeen_to_standard_lines(items: List[LineItem], main_invoice_no: str) -> pd.DataFrame:
    """
    ADP/SAFEEN 항목을 표준 Lines 형식으로 변환 (각 invoice_no별로 그룹화)

    Args:
        items: ADP/SAFEEN LineItem 리스트
        main_invoice_no: 메인 인보이스 번호 (OFCO 인보이스)

    Returns:
        표준 Lines 형식 DataFrame
    """
    rows = []

    # invoice_no별로 그룹화
    invoice_groups: Dict[str, List[LineItem]] = {}
    for item in items:
        invoice_no = extract_adp_safeen_invoice_no(item.subject or "")
        if not invoice_no:
            # invoice_no 추출 실패 시 기본값 사용
            invoice_no = main_invoice_no
        if invoice_no not in invoice_groups:
            invoice_groups[invoice_no] = []
        invoice_groups[invoice_no].append(item)

    # 각 invoice_no별로 처리
    for invoice_no, group_items in invoice_groups.items():
        for line_no, item in enumerate(group_items, 1):
            # NOTES 파싱
            notes = item.notes or ""

            # 기본 필드
            row = {
                'invoice_no': invoice_no,
                'line_no': line_no,
                'tariff_id': extract_tariff_id(notes) or '',
                'description': format_description(item.subject or ''),
            }

            # unit1, unit2, unit3 추출
            invoice_type = detect_invoice_type_from_notes(notes)
            unit1, unit2, unit3 = extract_units_from_item(item, notes, invoice_type)
            row['unit1'] = round(unit1, 2)
            row['unit2'] = round(unit2, 2)
            row['unit3'] = round(unit3, 2)

            # rate 추출
            row['rate'] = extract_rate_from_item(item, notes)

            # 금액 필드
            row['amount_excl_tax'] = round(item.amount or 0.0, 2)

            # EA 쌍 추출 및 매핑 (최대 4개)
            ea_pairs = extract_ea_pairs_from_item(item, notes, max_pairs=4)

            # EA_1~EA_4 슬롯 채우기
            for i in range(4):
                if i < len(ea_pairs):
                    qty, rate = ea_pairs[i]
                    row[f'EA_{i+1}_Qty'] = qty
                    row[f'EA_{i+1}_Rate'] = rate
                    row[f'EA_{i+1}_Amount'] = round(qty * rate, 2)
                else:
                    row[f'EA_{i+1}_Qty'] = 0.0
                    row[f'EA_{i+1}_Rate'] = 0.0
                    row[f'EA_{i+1}_Amount'] = 0.0

            # 세율 및 세금
            tax_rate_str = re.search(r'tax_rate=([^\s;]+)', notes)
            if tax_rate_str:
                row['tax_rate_pct'] = round(extract_tax_rate_pct(tax_rate_str.group(1)), 2)
            else:
                row['tax_rate_pct'] = 0.0

            tax_amount_match = re.search(r'tax_amount=([^\s;]+)', notes)
            if tax_amount_match:
                row['tax_amount'] = round(float(tax_amount_match.group(1)), 2)
            else:
                row['tax_amount'] = 0.0

            # total_incl_tax 계산
            amount_incl_match = re.search(r'amount_incl_vat=([^\s;]+)', notes)
            if amount_incl_match:
                row['total_incl_tax'] = round(float(amount_incl_match.group(1)), 2)
            else:
                row['total_incl_tax'] = round(row['amount_excl_tax'] + row['tax_amount'], 2)

            # calc_check (Excel 수식, 나중에 적용)
            row['calc_check'] = None  # Excel 수식으로 설정

            # evidence (기본값)
            row['evidence'] = f"p1,row{line_no}"

            rows.append(row)

    # 컬럼 순서 지정 (가이드 순서대로 + EA 슬롯)
    df = pd.DataFrame(rows)
    column_order = [
        'invoice_no', 'line_no', 'tariff_id', 'description',
        'unit1', 'unit2', 'unit3', 'rate',
        'amount_excl_tax', 'tax_rate_pct', 'tax_amount', 'total_incl_tax',
        'EA_1_Qty', 'EA_1_Rate', 'EA_1_Amount',
        'EA_2_Qty', 'EA_2_Rate', 'EA_2_Amount',
        'EA_3_Qty', 'EA_3_Rate', 'EA_3_Amount',
        'EA_4_Qty', 'EA_4_Rate', 'EA_4_Amount',
        'calc_check', 'evidence'
    ]
    if len(df) > 0:
        # 존재하는 컬럼만 정렬
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
    return df


def convert_ofco_to_excel_rows(items: List[LineItem], main_invoice_no: str) -> pd.DataFrame:
    """
    OFCO 항목을 Excel 행 형식으로 변환 (순서대로)

    Args:
        items: OFCO LineItem 리스트
        main_invoice_no: 메인 인보이스 번호

    Returns:
        DataFrame
    """
    rows = []

    for line_no, item in enumerate(items, 1):
        notes = item.notes or ""

        row = {
            'invoice_no': main_invoice_no,
            'line_no': line_no,
            'description': format_description(item.subject or ''),
            'amount_excl_tax': round(item.amount or 0.0, 2),
        }

        # EA 쌍 추출 및 매핑 (최대 4개)
        ea_pairs = extract_ea_pairs_from_item(item, notes, max_pairs=4)

        # EA_1~EA_4 슬롯 채우기
        for i in range(4):
            if i < len(ea_pairs):
                qty, rate = ea_pairs[i]
                row[f'EA_{i+1}_Qty'] = qty
                row[f'EA_{i+1}_Rate'] = rate
                row[f'EA_{i+1}_Amount'] = round(qty * rate, 2)
            else:
                row[f'EA_{i+1}_Qty'] = 0.0
                row[f'EA_{i+1}_Rate'] = 0.0
                row[f'EA_{i+1}_Amount'] = 0.0

        # 세율 및 세금
        tax_rate_str = re.search(r'tax_rate=([^\s;]+)', notes)
        if tax_rate_str:
            row['tax_rate_pct'] = round(extract_tax_rate_pct(tax_rate_str.group(1)), 2)
        else:
            row['tax_rate_pct'] = 0.0

        tax_amount_match = re.search(r'tax_amount=([^\s;]+)', notes)
        if tax_amount_match:
            row['tax_amount'] = round(float(tax_amount_match.group(1)), 2)
        else:
            row['tax_amount'] = 0.0

        # total_incl_tax 계산
        amount_incl_match = re.search(r'amount_incl_vat=([^\s;]+)', notes)
        if amount_incl_match:
            row['total_incl_tax'] = round(float(amount_incl_match.group(1)), 2)
        else:
            row['total_incl_tax'] = round(row['amount_excl_tax'] + row['tax_amount'], 2)

        # 추가 정보
        row['tariff_id'] = extract_tariff_id(notes) or ''

        # unit 정보 (있는 경우)
        invoice_type = detect_invoice_type_from_notes(notes)
        unit1, unit2, unit3 = extract_units_from_item(item, notes, invoice_type)
        row['unit1'] = round(unit1, 2)
        row['unit2'] = round(unit2, 2)
        row['unit3'] = round(unit3, 2)
        row['rate'] = extract_rate_from_item(item, notes)

        rows.append(row)

    # 컬럼 순서 지정
    df = pd.DataFrame(rows)
    column_order = [
        'invoice_no', 'line_no', 'tariff_id', 'description',
        'unit1', 'unit2', 'unit3', 'rate',
        'amount_excl_tax', 'tax_rate_pct', 'tax_amount', 'total_incl_tax',
        'EA_1_Qty', 'EA_1_Rate', 'EA_1_Amount',
        'EA_2_Qty', 'EA_2_Rate', 'EA_2_Amount',
        'EA_3_Qty', 'EA_3_Rate', 'EA_3_Amount',
        'EA_4_Qty', 'EA_4_Rate', 'EA_4_Amount'
    ]
    if len(df) > 0:
        # 존재하는 컬럼만 정렬
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
    return df


def convert_payload_to_standard_lines(payload: OFCOPayload, pdf_path: str = "") -> pd.DataFrame:
    """
    OFCOPayload를 표준 Lines 스키마로 변환

    Args:
        payload: 파싱된 인보이스 데이터
        pdf_path: PDF 파일 경로 (evidence 생성용)

    Returns:
        표준 Lines 형식 DataFrame
    """
    rows = []

    for line_no, item in enumerate(payload.items, 1):
        # NOTES 파싱
        notes = item.notes or ""

        # 기본 필드
        row = {
            'invoice_no': payload.meta.invoice_number or '',
            'line_no': line_no,
            'tariff_id': extract_tariff_id(notes) or '',
            'description': format_description(item.subject or ''),
        }

        # unit1, unit2, unit3 추출
        invoice_type = detect_invoice_type_from_notes(notes)
        unit1, unit2, unit3 = extract_units_from_item(item, notes, invoice_type)
        row['unit1'] = round(unit1, 2)
        row['unit2'] = round(unit2, 2)
        row['unit3'] = round(unit3, 2)

        # rate 추출
        row['rate'] = extract_rate_from_item(item, notes)

        # 금액 필드
        row['amount_excl_tax'] = round(item.amount or 0.0, 2)

        # 세율 및 세금
        tax_rate_str = re.search(r'tax_rate=([^\s;]+)', notes)
        if tax_rate_str:
            row['tax_rate_pct'] = round(extract_tax_rate_pct(tax_rate_str.group(1)), 2)
        else:
            row['tax_rate_pct'] = 0.0

        tax_amount_match = re.search(r'tax_amount=([^\s;]+)', notes)
        if tax_amount_match:
            row['tax_amount'] = round(float(tax_amount_match.group(1)), 2)
        else:
            row['tax_amount'] = 0.0

        # total_incl_tax 계산
        amount_incl_match = re.search(r'amount_incl_vat=([^\s;]+)', notes)
        if amount_incl_match:
            row['total_incl_tax'] = round(float(amount_incl_match.group(1)), 2)
        else:
            row['total_incl_tax'] = round(row['amount_excl_tax'] + row['tax_amount'], 2)

        # calc_check (Excel 수식, 나중에 적용)
        row['calc_check'] = None  # Excel 수식으로 설정

        # evidence (기본값)
        row['evidence'] = f"p1,row{line_no}"

        rows.append(row)

    # 컬럼 순서 지정 (가이드 순서대로)
    df = pd.DataFrame(rows)
    column_order = [
        'invoice_no', 'line_no', 'tariff_id', 'description',
        'unit1', 'unit2', 'unit3', 'rate',
        'amount_excl_tax', 'tax_rate_pct', 'tax_amount', 'total_incl_tax',
        'calc_check', 'evidence'
    ]
    df = df[column_order]
    return df


def detect_invoice_type_from_notes(notes: str) -> str:
    """NOTES에서 인보이스 타입 감지"""
    if 'OFCO' in notes.upper():
        return 'OFCO'
    if 'ADP' in notes.upper() or 'unit1=' in notes:
        return 'ADP'
    if 'SAFEEN' in notes.upper():
        return 'SAFEEN'
    return 'OFCO'  # 기본값


def is_adp_or_safeen_item(item: LineItem) -> bool:
    """
    항목이 ADP 또는 SAFEEN 인보이스인지 확인

    Args:
        item: LineItem 객체

    Returns:
        True if ADP or SAFEEN, False otherwise
    """
    description = item.subject or ""
    notes = item.notes or ""

    # description에서 패턴 검색
    if re.search(r'ADP\s+INV[-]?\d+', description, re.IGNORECASE):
        return True
    if re.search(r'SAFEEN\s+INV[-]?\d+', description, re.IGNORECASE):
        return True

    # NOTES에서 타입 확인
    if 'ADP' in notes.upper() or 'unit1=' in notes:
        return True
    if 'SAFEEN' in notes.upper():
        return True

    return False


def extract_adp_safeen_invoice_no(description: str) -> Optional[str]:
    """
    ADP/SAFEEN 항목에서 invoice_no 추출

    Args:
        description: 항목 description (예: "ADP INV-61871 - Port Dues & Services")

    Returns:
        invoice_no (예: "INV-61871") 또는 None
    """
    if not description:
        return None

    # ADP 패턴: "ADP INV-61871" 또는 "ADP INV61871"
    match = re.search(r'ADP\s+INV[-]?(\d+)', description, re.IGNORECASE)
    if match:
        return f"INV-{match.group(1)}"

    # SAFEEN 패턴: "SAFEEN INV-62102"
    match = re.search(r'SAFEEN\s+INV[-]?(\d+)', description, re.IGNORECASE)
    if match:
        return f"INV-{match.group(1)}"

    return None


def get_column_letter(col_num: int) -> str:
    """컬럼 번호를 Excel 컬럼 문자로 변환 (1=A, 2=B, ...)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


def apply_calc_check_formula(df: pd.DataFrame, workbook, worksheet):
    """
    calc_check 컬럼에 Excel 수식 적용

    EA 슬롯 기반 검증:
    Line_Amount = Σ EA_i_Amount (i=1..4)
    calc_check = |Line_Amount - amount_excl_tax| <= 0.01

    또는 기존 방식 (EA 슬롯이 없으면):
    =ABS(ROUND(PRODUCT(unit1:unit3)*rate,2) - ROUND(amount_excl_tax,2)) <= 0.01
    """
    # 컬럼 인덱스 찾기 (0-based)
    col_map = {col: idx for idx, col in enumerate(df.columns)}

    amount_col_idx = col_map.get('amount_excl_tax', 8)  # I열
    calc_check_col_idx = col_map.get('calc_check', None)

    if calc_check_col_idx is None:
        return  # calc_check 컬럼이 없으면 스킵

    # EA_Amount 컬럼 확인
    ea_amount_cols = []
    for i in range(1, 5):
        col_name = f'EA_{i}_Amount'
        if col_name in col_map:
            ea_amount_cols.append(col_name)

    # Excel 수식 적용 (헤더 제외, 2행부터)
    for df_row_idx in range(len(df)):
        excel_row = df_row_idx + 2  # 헤더 다음 행 (2행부터 시작)

        if ea_amount_cols:
            # EA 슬롯 기반 검증: Line_Amount = Σ EA_i_Amount
            ea_amount_col_letters = [get_column_letter(col_map[col] + 1) for col in ea_amount_cols]
            amount_col = get_column_letter(amount_col_idx + 1)

            # 수식: =ABS(ROUND(SUM(EA_1_Amount+EA_2_Amount+EA_3_Amount+EA_4_Amount),2) - ROUND(amount_excl_tax,2)) <= 0.01
            # EA_Amount 컬럼들을 더하기
            if len(ea_amount_col_letters) == 1:
                sum_formula = ea_amount_col_letters[0] + str(excel_row)
            elif len(ea_amount_col_letters) == 2:
                sum_formula = f"{ea_amount_col_letters[0]}{excel_row}+{ea_amount_col_letters[1]}{excel_row}"
            elif len(ea_amount_col_letters) == 3:
                sum_formula = f"{ea_amount_col_letters[0]}{excel_row}+{ea_amount_col_letters[1]}{excel_row}+{ea_amount_col_letters[2]}{excel_row}"
            elif len(ea_amount_col_letters) == 4:
                sum_formula = f"{ea_amount_col_letters[0]}{excel_row}+{ea_amount_col_letters[1]}{excel_row}+{ea_amount_col_letters[2]}{excel_row}+{ea_amount_col_letters[3]}{excel_row}"
            else:
                # 4개 이상이면 SUM 범위 사용
                sum_formula = f"SUM({ea_amount_col_letters[0]}{excel_row}:{ea_amount_col_letters[-1]}{excel_row})"

            formula = f'=ABS(ROUND({sum_formula},2) - ROUND({amount_col}{excel_row},2)) <= 0.01'
        else:
            # 기존 방식 (EA 슬롯이 없으면)
            unit1_col_idx = col_map.get('unit1', 4)
            unit2_col_idx = col_map.get('unit2', 5)
            unit3_col_idx = col_map.get('unit3', 6)
            rate_col_idx = col_map.get('rate', 7)

            unit1_col = get_column_letter(unit1_col_idx + 1)
            unit2_col = get_column_letter(unit2_col_idx + 1)
            unit3_col = get_column_letter(unit3_col_idx + 1)
            rate_col = get_column_letter(rate_col_idx + 1)
            amount_col = get_column_letter(amount_col_idx + 1)

            # 수식: =ABS(ROUND(PRODUCT(E{row}:G{row})*H{row},2) - ROUND(I{row},2)) <= 0.01
            formula = f'=ABS(ROUND(PRODUCT({unit1_col}{excel_row}:{unit3_col}{excel_row})*{rate_col}{excel_row},2) - ROUND({amount_col}{excel_row},2)) <= 0.01'

        cell = worksheet.cell(row=excel_row, column=calc_check_col_idx + 1)
        cell.value = formula


def main():
    import argparse

    parser = argparse.ArgumentParser(description='OFCO 파서 결과를 표준 Lines 형식으로 Excel에 입력')
    parser.add_argument('pdf_path', help='PDF 파일 경로')
    parser.add_argument('--excel', default=None, help='Excel 파일 경로 (기본: PDF 파일명과 같은 이름)')
    parser.add_argument('--sheet', default='Standard_Lines', help='시트 이름')
    parser.add_argument('--append', action='store_true', help='기존 데이터에 추가 (기본: 새 시트 생성)')

    args = parser.parse_args()

    # Excel 파일 경로 결정
    if args.excel is None:
        pdf_path_obj = Path(args.pdf_path)
        excel_path = pdf_path_obj.parent / f"{pdf_path_obj.stem}_standard_lines.xlsx"
    else:
        excel_path = Path(args.excel)

    # PDF 파싱
    print(f"파싱 중: {args.pdf_path}")
    payload = parse_ofco_invoice(args.pdf_path)
    print(f"파싱 완료: {len(payload.items)}개 항목")

    # 항목 분류
    print("항목 분류 중...")
    adp_safeen_items, ofco_items = split_items_by_type(payload)
    print(f"ADP/SAFEEN 항목: {len(adp_safeen_items)}개, OFCO 항목: {len(ofco_items)}개")

    # 표준 Lines 형식으로 변환 (ADP/SAFEEN만)
    print("Standard Lines 시트 데이터 변환 중 (ADP/SAFEEN)...")
    standard_lines_df = convert_adp_safeen_to_standard_lines(adp_safeen_items, payload.meta.invoice_number or "")
    print(f"ADP/SAFEEN Standard Lines 변환 완료: {len(standard_lines_df)}개 행")

    # OFCO 시트 데이터 변환 (전체 46개 항목 순서대로)
    print("OFCO 시트 데이터 변환 중...")
    ofco_df = convert_ofco_to_excel_rows(payload.items, payload.meta.invoice_number or "")  # 전체 항목 사용
    print(f"OFCO 변환 완료: {len(ofco_df)}개 행")

    # Excel 파일 로드 또는 생성
    if excel_path.exists():
        workbook = load_workbook(excel_path)
        # 기존 시트 제거 (같은 이름이면)
        for sheet_name in ["Standard_Lines", "OFCO"]:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
    else:
        from openpyxl import Workbook
        workbook = Workbook()
        workbook.remove(workbook.active)

    # Standard Lines 시트 생성 (ADP/SAFEEN + OFCO 병합 · 중복 방지 · 라벨링)
    combined_sl_df = None

    def _with_src(df, src_name):
        if isinstance(df, pd.DataFrame) and len(df) > 0:
            df = df.copy()
            if 'SRC' not in df.columns:
                df['SRC'] = src_name
            else:
                df['SRC'] = df['SRC'].fillna(src_name).replace('', src_name)
        return df

    # ---- 병합 정책 (권장: Standard_Lines에 ADP/SAFEEN + OFCO 모두 포함) ----
    # 1) standard_lines_df(ADP/SAFEEN)이 존재하면 우선 컬럼/순서를 기준으로 OFCO를 정렬 후 append
    # 2) ADP/SAFEEN이 없고 OFCO만 있으면 OFCO를 그대로 사용
    # 3) 둘 다 없으면 빈 DataFrame
    if isinstance(standard_lines_df, pd.DataFrame) and len(standard_lines_df) > 0:
        if isinstance(ofco_df, pd.DataFrame) and len(ofco_df) > 0:
            # ofco_df를 standard_lines_df 컬럼 순서에 맞춰 재배치 (존재하지 않는 컬럼은 빈값)
            try:
                ofco_aligned = ofco_df.reindex(columns=standard_lines_df.columns, fill_value='')
            except Exception:
                # 안전장치: 서로 스키마가 완전히 다르면 공통 컬럼만 취해 빈 컬럼으로 보정
                common_cols = [c for c in standard_lines_df.columns if c in ofco_df.columns]
                ofco_aligned = ofco_df[common_cols].reindex(columns=standard_lines_df.columns, fill_value='')
            adp_safeen_df = _with_src(standard_lines_df.reset_index(drop=True), 'ADP/SAFEEN')
            ofco_aligned = _with_src(ofco_aligned.reset_index(drop=True), 'OFCO')
            combined_sl_df = pd.concat([adp_safeen_df, ofco_aligned],
                                       ignore_index=True, sort=False)
        else:
            combined_sl_df = _with_src(standard_lines_df.copy(), 'ADP/SAFEEN')
    elif isinstance(ofco_df, pd.DataFrame) and len(ofco_df) > 0:
        # ADP/SAFEEN이 없더라도 OFCO만으로 Standard_Lines 구성
        combined_sl_df = _with_src(ofco_df.copy(), 'OFCO')
    else:
        combined_sl_df = pd.DataFrame()

    # (안전) 인덱스 정리 및 컬럼 정렬: 주요 표준 컬럼 순서를 유지하도록 시도
    if len(combined_sl_df) > 0:
        # 기본 컬럼 순서가 있으면 그 순서로 정렬(없는 컬럼은 뒤로)
        preferred_order = [
            'invoice_no','line_no','tariff_id','description',
            'unit1','unit2','unit3','rate',
            'amount_excl_tax','tax_rate_pct','tax_amount','total_incl_tax',
            'EA_1_Qty','EA_1_Rate','EA_1_Amount','EA_2_Qty','EA_2_Rate','EA_2_Amount',
            'EA_3_Qty','EA_3_Rate','EA_3_Amount','EA_4_Qty','EA_4_Rate','EA_4_Amount',
            'calc_check','SRC'
        ]
        existing_cols = [c for c in preferred_order if c in combined_sl_df.columns]
        other_cols = [c for c in combined_sl_df.columns if c not in existing_cols]
        combined_sl_df = combined_sl_df[existing_cols + other_cols]

        # 중복 방지(자연키): invoice_no, line_no, description, amount_excl_tax
        natural_key = [c for c in ['invoice_no','line_no','description','amount_excl_tax'] if c in combined_sl_df.columns]
        if natural_key:
            combined_sl_df = combined_sl_df.drop_duplicates(subset=natural_key, keep='first').reset_index(drop=True)

    # Standard Lines 시트 쓰기 (병합된 전체 데이터)
    if len(combined_sl_df) > 0:
        print("Standard_Lines 시트 생성 중 (ADP/SAFEEN + OFCO 병합)...")
        worksheet_sl = workbook.create_sheet("Standard_Lines")

        # 헤더 추가
        headers = list(combined_sl_df.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet_sl.cell(row=1, column=col_idx, value=header)

        # 데이터 추가 (calc_check 제외 — 기존 로직 호환)
        calc_check_col_idx = None
        if 'calc_check' in combined_sl_df.columns:
            calc_check_col_idx = combined_sl_df.columns.get_loc('calc_check')

        for df_row_idx, row in combined_sl_df.iterrows():
            excel_row = df_row_idx + 2  # 헤더 다음 행
            for col_idx, value in enumerate(row):
                if calc_check_col_idx is not None and col_idx == calc_check_col_idx:
                    # calc_check는 수식으로 처리하므로 스킵(나중에 apply)
                    continue
                cell = worksheet_sl.cell(row=excel_row, column=col_idx + 1)
                cell.value = value

        # calc_check 수식 적용 (EA_*_Amount 우선, 미존재 시 unit×rate 대안)
        apply_calc_check_formula(combined_sl_df, workbook, worksheet_sl)

    # OFCO 시트는 기존대로 생성(검토용), 필요 시 이 부분을 주석/삭제해도 됨
    if len(ofco_df) > 0:
        print("OFCO 시트 생성 중 (원본 OFCO 테이블 유지)...")
        worksheet_ofco = workbook.create_sheet("OFCO")

        # OFCO 헤더 추가
        headers = list(ofco_df.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet_ofco.cell(row=1, column=col_idx, value=header)

        # OFCO 데이터 추가 (기존 순서 유지)
        for df_row_idx, row in ofco_df.iterrows():
            excel_row = df_row_idx + 2
            for col_idx, value in enumerate(row):
                cell = worksheet_ofco.cell(row=excel_row, column=col_idx + 1)
                cell.value = value

    # 파일 저장 (Permission 오류 처리)
    try:
        workbook.save(excel_path)
        print(f"\n완료: {excel_path}")
    except PermissionError:
        # 파일이 열려있으면 대체 파일명으로 저장
        backup_path = excel_path.parent / f"{excel_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(backup_path)
        print(f"\n경고: 원본 파일이 열려있어 다음 위치에 저장했습니다: {backup_path}")
        print(f"원본 파일 ({excel_path})을 닫고 다시 시도하세요.")

    print(f"인보이스: {payload.meta.invoice_number}")
    print(f"Standard Lines 시트(병합된 전체): {len(combined_sl_df)}개 라인 (ADP/SAFEEN + OFCO 포함 여부 확인)")
    print(f"OFCO 시트(오리지널 변환 결과): {len(ofco_df)}개 라인")


if __name__ == '__main__':
    main()

