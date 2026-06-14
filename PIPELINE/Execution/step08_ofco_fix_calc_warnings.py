#!/usr/bin/env python3
"""
OFCO calc_check 경고 자동 수정 스크립트

calc_check 실패 원인:
- 단가(Unit Price) * 수량(Quantity) ≠ 금액(Amount)
- 주요 원인: PDF에서 금액 필드 OCR 추출 실패

해결 방법:
- 단가와 수량이 있으면 금액을 자동 계산
- 계산된 금액에 'Calculated' 출처 표시

Usage:
    python step08_ofco_fix_calc_warnings.py --input ofco_pipeline_pdf_XXX.xlsx --sheet EA_Slots
    python step08_ofco_fix_calc_warnings.py --input "OFCO INVOICE.xlsx" --sheet Sheet1 --backup

Options:
    --input: 입력 파일 경로
    --sheet: 대상 시트명 (기본값: EA_Slots)
    --backup: 백업 파일 생성 (기본값: True)
    --tolerance: 금액 오차 허용 범위 (기본값: 0.01 AED)
"""

import pandas as pd
import sys
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Tuple, List
import argparse


def fix_missing_amounts(df: pd.DataFrame, tolerance: float = 0.01) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    누락된 금액을 단가×수량으로 자동 계산
    
    Args:
        df: 대상 DataFrame
        tolerance: 금액 오차 허용 범위 (AED)
    
    Returns:
        (수정된 DataFrame, 통계 정보)
    """
    stats = {
        'total_rows': len(df),
        'calc_check_failures': 0,
        'amount_missing': 0,
        'amount_calculated': 0,
        'amount_corrected': 0,
        'calc_check_fixed': 0
    }
    
    # EA 슬롯별 처리 (EA_1 ~ EA_4)
    for i in range(1, 5):
        qty_col = f'EA_{i}_Qty'
        rate_col = f'EA_{i}_Rate'
        amount_col = f'EA_{i}_Amount'
        
        if qty_col not in df.columns or rate_col not in df.columns or amount_col not in df.columns:
            continue
        
        # 1. 금액이 누락되었지만 단가와 수량이 있는 경우
        mask_missing = (
            df[amount_col].isna() & 
            df[qty_col].notna() & 
            df[rate_col].notna()
        )
        
        if mask_missing.any():
            df.loc[mask_missing, amount_col] = df.loc[mask_missing, qty_col] * df.loc[mask_missing, rate_col]
            df.loc[mask_missing, 'Data_Source'] = f'{amount_col}_Calculated'
            count = mask_missing.sum()
            stats['amount_calculated'] += count
            print(f"[INFO] {amount_col}: {count}개 누락 금액 자동 계산")
        
        # 2. 금액이 있지만 단가×수량과 다른 경우 (오차 범위 초과)
        mask_incorrect = (
            df[amount_col].notna() & 
            df[qty_col].notna() & 
            df[rate_col].notna()
        )
        
        if mask_incorrect.any():
            calculated = df.loc[mask_incorrect, qty_col] * df.loc[mask_incorrect, rate_col]
            actual = df.loc[mask_incorrect, amount_col]
            diff = abs(calculated - actual)
            
            mask_needs_correction = mask_incorrect & (diff > tolerance)
            
            if mask_needs_correction.any():
                df.loc[mask_needs_correction, amount_col] = calculated[mask_needs_correction]
                df.loc[mask_needs_correction, 'Data_Source'] = f'{amount_col}_Corrected'
                count = mask_needs_correction.sum()
                stats['amount_corrected'] += count
                print(f"[INFO] {amount_col}: {count}개 금액 오차 수정 (허용 범위: ±{tolerance} AED)")
    
    # 3. EA_Total_Amount 재계산
    if all(f'EA_{i}_Amount' in df.columns for i in range(1, 5)) and 'EA_Total_Amount' in df.columns:
        ea_amounts = [df[f'EA_{i}_Amount'].fillna(0) for i in range(1, 5)]
        df['EA_Total_Amount'] = sum(ea_amounts)
        print(f"[INFO] EA_Total_Amount 재계산 완료")
    
    # 4. calc_check 재검증
    if 'calc_check' in df.columns:
        stats['calc_check_failures'] = (~df['calc_check']).sum()
        print(f"[INFO] 수정 전 calc_check 실패: {stats['calc_check_failures']}행")
        
        # calc_check 재계산 (간단한 검증)
        # 실제 calc_check 로직이 복잡하면 해당 함수를 import하여 사용
        for i in range(1, 5):
            qty_col = f'EA_{i}_Qty'
            rate_col = f'EA_{i}_Rate'
            amount_col = f'EA_{i}_Amount'
            
            if all(col in df.columns for col in [qty_col, rate_col, amount_col]):
                mask = df[qty_col].notna() & df[rate_col].notna() & df[amount_col].notna()
                if mask.any():
                    calculated = df.loc[mask, qty_col] * df.loc[mask, rate_col]
                    actual = df.loc[mask, amount_col]
                    df.loc[mask, 'calc_check'] = abs(calculated - actual) <= tolerance
        
        stats['calc_check_fixed'] = (~df['calc_check']).sum()
        improvement = stats['calc_check_failures'] - stats['calc_check_fixed']
        print(f"[INFO] 수정 후 calc_check 실패: {stats['calc_check_fixed']}행 (개선: {improvement}행)")
    
    return df, stats


def fix_standard_lines_amounts(df: pd.DataFrame, tolerance: float = 0.01) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    Standard_Lines 시트의 금액 필드 수정
    
    Args:
        df: Standard_Lines DataFrame
        tolerance: 금액 오차 허용 범위
    
    Returns:
        (수정된 DataFrame, 통계 정보)
    """
    stats = {
        'total_rows': len(df),
        'amount_missing': 0,
        'amount_calculated': 0,
        'amount_corrected': 0
    }
    
    # unit_price, quantity, amount_excl_tax 컬럼 처리
    if all(col in df.columns for col in ['unit_price', 'quantity', 'amount_excl_tax']):
        # 1. 금액 누락 시 자동 계산
        mask_missing = (
            df['amount_excl_tax'].isna() & 
            df['unit_price'].notna() & 
            df['quantity'].notna()
        )
        
        if mask_missing.any():
            df.loc[mask_missing, 'amount_excl_tax'] = df.loc[mask_missing, 'unit_price'] * df.loc[mask_missing, 'quantity']
            df.loc[mask_missing, 'data_source'] = 'amount_excl_tax_Calculated'
            stats['amount_calculated'] = mask_missing.sum()
            print(f"[INFO] amount_excl_tax: {stats['amount_calculated']}개 누락 금액 자동 계산")
        
        # 2. 금액 오차 수정
        mask_incorrect = (
            df['amount_excl_tax'].notna() & 
            df['unit_price'].notna() & 
            df['quantity'].notna()
        )
        
        if mask_incorrect.any():
            calculated = df.loc[mask_incorrect, 'unit_price'] * df.loc[mask_incorrect, 'quantity']
            actual = df.loc[mask_incorrect, 'amount_excl_tax']
            diff = abs(calculated - actual)
            
            mask_needs_correction = mask_incorrect & (diff > tolerance)
            
            if mask_needs_correction.any():
                df.loc[mask_needs_correction, 'amount_excl_tax'] = calculated[mask_needs_correction]
                df.loc[mask_needs_correction, 'data_source'] = 'amount_excl_tax_Corrected'
                stats['amount_corrected'] = mask_needs_correction.sum()
                print(f"[INFO] amount_excl_tax: {stats['amount_corrected']}개 금액 오차 수정")
    
    # 3. calc_check 재검증
    if 'calc_check' in df.columns:
        calc_failures_before = (~df['calc_check']).sum()
        print(f"[INFO] 수정 전 calc_check 실패: {calc_failures_before}행")
        
        # calc_check 재계산
        if all(col in df.columns for col in ['unit_price', 'quantity', 'amount_excl_tax']):
            mask = df['unit_price'].notna() & df['quantity'].notna() & df['amount_excl_tax'].notna()
            if mask.any():
                calculated = df.loc[mask, 'unit_price'] * df.loc[mask, 'quantity']
                actual = df.loc[mask, 'amount_excl_tax']
                df.loc[mask, 'calc_check'] = abs(calculated - actual) <= tolerance
        
        calc_failures_after = (~df['calc_check']).sum()
        improvement = calc_failures_before - calc_failures_after
        print(f"[INFO] 수정 후 calc_check 실패: {calc_failures_after}행 (개선: {improvement}행)")
    
    return df, stats


def process_file(
    input_file: str,
    sheet_name: str = 'EA_Slots',
    backup: bool = True,
    tolerance: float = 0.01
) -> Tuple[bool, Dict[str, Any]]:
    """
    파일 처리 메인 함수
    
    Args:
        input_file: 입력 파일 경로
        sheet_name: 대상 시트명
        backup: 백업 생성 여부
        tolerance: 금액 오차 허용 범위
    
    Returns:
        (성공 여부, 통계 정보)
    """
    overall_stats = {
        'input_file': input_file,
        'sheet_name': sheet_name,
        'success': False,
        'backup_created': False,
        'errors': []
    }
    
    try:
        input_path = Path(input_file)
        
        if not input_path.exists():
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_file}")
        
        print(f"[INFO] 입력 파일: {input_path}")
        print(f"[INFO] 대상 시트: {sheet_name}")
        print(f"[INFO] 금액 오차 허용 범위: ±{tolerance} AED")
        
        # 백업 생성
        if backup:
            backup_path = input_path.with_suffix(f'.xlsx.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            shutil.copy2(input_path, backup_path)
            overall_stats['backup_created'] = True
            overall_stats['backup_path'] = str(backup_path)
            print(f"[OK] 백업 파일 생성: {backup_path}")
        
        # 파일 읽기
        print(f"\n[Step 1/3] 파일 읽기...")
        df = pd.read_excel(input_path, sheet_name=sheet_name)
        print(f"[OK] {len(df)}행, {len(df.columns)}컬럼 로드")
        
        # 데이터 수정
        print(f"\n[Step 2/3] 금액 데이터 수정...")
        
        if sheet_name in ['EA_Slots', 'Complete_Pipeline']:
            df, stats = fix_missing_amounts(df, tolerance=tolerance)
        elif sheet_name == 'Standard_Lines':
            df, stats = fix_standard_lines_amounts(df, tolerance=tolerance)
        else:
            # LIST 시트 등 다른 시트는 EA 컬럼 기반으로 처리
            df, stats = fix_missing_amounts(df, tolerance=tolerance)
        
        overall_stats.update(stats)
        
        # 파일 저장
        print(f"\n[Step 3/3] 파일 저장...")
        
        # 기존 파일의 다른 시트 유지
        from openpyxl import load_workbook
        wb = load_workbook(input_path)
        
        # 대상 시트 업데이트
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
        
        # DataFrame을 시트에 쓰기
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 임시 파일로 저장 후 이동
        temp_path = input_path.with_suffix('.xlsx.tmp')
        wb.save(temp_path)
        wb.close()
        
        if temp_path.exists():
            shutil.move(str(temp_path), str(input_path))
            print(f"[OK] 파일 저장 완료: {input_path}")
        else:
            raise Exception("임시 파일 저장 실패")
        
        overall_stats['success'] = True
        
        # 통계 출력
        print(f"\n{'='*80}")
        print(f"calc_check 경고 수정 완료!")
        print(f"{'='*80}")
        print(f"전체 행 수: {stats.get('total_rows', len(df))}")
        print(f"금액 자동 계산: {stats.get('amount_calculated', 0)}개")
        print(f"금액 오차 수정: {stats.get('amount_corrected', 0)}개")
        
        if 'calc_check_failures' in stats:
            print(f"calc_check 실패 (수정 전): {stats['calc_check_failures']}행")
            print(f"calc_check 실패 (수정 후): {stats.get('calc_check_fixed', 0)}행")
            improvement = stats['calc_check_failures'] - stats.get('calc_check_fixed', 0)
            if improvement > 0:
                improvement_pct = (improvement / stats['calc_check_failures'] * 100) if stats['calc_check_failures'] > 0 else 0
                print(f"개선율: {improvement}행 ({improvement_pct:.1f}%)")
        
        return True, overall_stats
        
    except FileNotFoundError as e:
        error_msg = f"파일을 찾을 수 없습니다: {e}"
        print(f"\n[ERROR] {error_msg}")
        overall_stats['errors'].append(error_msg)
        return False, overall_stats
    
    except Exception as e:
        error_msg = f"처리 중 오류 발생: {type(e).__name__} - {e}"
        print(f"\n[ERROR] {error_msg}")
        overall_stats['errors'].append(error_msg)
        
        # 롤백
        if backup and overall_stats.get('backup_created') and 'backup_path' in overall_stats:
            backup_path = Path(overall_stats['backup_path'])
            if backup_path.exists():
                print(f"\n[INFO] 롤백 시도...")
                try:
                    shutil.copy2(backup_path, input_path)
                    print(f"[OK] 롤백 완료")
                except Exception as rollback_error:
                    print(f"[ERROR] 롤백 실패: {rollback_error}")
        
        return False, overall_stats


def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(
        description='OFCO calc_check 경고 자동 수정 스크립트 v1.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  # EA_Slots 시트 처리
  python step08_ofco_fix_calc_warnings.py --input ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx
  
  # Standard_Lines 시트 처리
  python step08_ofco_fix_calc_warnings.py --input ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx --sheet Standard_Lines
  
  # 백업 생성 + 오차 허용 범위 설정
  python step08_ofco_fix_calc_warnings.py --input "OFCO INVOICE.xlsx" --sheet Sheet1 --backup --tolerance 0.05
        """
    )
    
    parser.add_argument('--input', type=str, required=True,
                        help='입력 파일 경로')
    parser.add_argument('--sheet', type=str, default='EA_Slots',
                        help='대상 시트명 (기본값: EA_Slots)')
    parser.add_argument('--backup', action='store_true', default=True,
                        help='백업 파일 생성 (기본값: True)')
    parser.add_argument('--no-backup', action='store_false', dest='backup',
                        help='백업 파일 생성 안 함')
    parser.add_argument('--tolerance', type=float, default=0.01,
                        help='금액 오차 허용 범위 (기본값: 0.01 AED)')
    
    args = parser.parse_args()
    
    print("="*80)
    print("OFCO calc_check 경고 자동 수정 스크립트")
    print("="*80)
    print()
    
    success, stats = process_file(
        input_file=args.input,
        sheet_name=args.sheet,
        backup=args.backup,
        tolerance=args.tolerance
    )
    
    if success:
        print(f"\n[SUCCESS] 처리가 완료되었습니다!")
        sys.exit(0)
    else:
        print(f"\n[FAILURE] 처리 중 오류가 발생했습니다.")
        if stats.get('errors'):
            print(f"\n오류 목록:")
            for error in stats['errors']:
                print(f"  - {error}")
        sys.exit(1)


if __name__ == '__main__':
    main()
