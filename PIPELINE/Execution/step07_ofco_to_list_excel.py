#!/usr/bin/env python3
"""
OFCO 파서 결과 → LIST 시트 입력 스크립트

파싱된 인보이스 데이터를 LIST 시트에 순서대로 추가
"""

import pandas as pd
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook

# 같은 폴더의 파서 모듈 import
from ofco_parse_pdf_patched import parse_ofco_invoice, OFCOPayload, LineItem

# EA 슬롯 알고리즘 v1.3 import
from ofco_ea_slot_v13 import process_row_to_ea_slots, process_dataframe, validate_invoice


def normalize_price_center_for_column(pc: str) -> str:
    """
    PRICE CENTER 이름을 LIST 시트 컬럼 형식으로 변환

    예: "AGENCY FEE FOR CARGO CLEARANCE" → "AGENCY_FEE_FOR_CARGO_CLEARANCE"
    """
    if not pc:
        return ""
    # 대문자로 변환하고 공백/특수문자를 언더스코어로 변경
    normalized = re.sub(r'[^\w\s]', '', pc.upper())
    normalized = re.sub(r'\s+', '_', normalized.strip())
    return normalized


def find_price_center_columns(df: pd.DataFrame, price_center: str) -> Optional[Dict[str, str]]:
    """
    PRICE CENTER에 해당하는 QTY/AMOUNT 컬럼 찾기

    Returns:
        {'QTY': 'AGENCY_FEE_FOR_CARGO_CLEARANCE_QTY', 'AMOUNT': 'AGENCY_FEE_FOR_CARGO_CLEARANCE_AMOUNT'} 또는 None
    """
    if not price_center:
        return None

    normalized_pc = normalize_price_center_for_column(price_center)
    qty_col = f"{normalized_pc}_QTY"
    amount_col = f"{normalized_pc}_AMOUNT"

    # 정확히 일치하는 컬럼 찾기
    if qty_col in df.columns and amount_col in df.columns:
        return {'QTY': qty_col, 'AMOUNT': amount_col}

    # 부분 일치로 찾기 (유연한 매칭)
    for col in df.columns:
        col_str = str(col).upper()
        if col_str.endswith('_QTY') or col_str.endswith('_AMOUNT'):
            base_name = col_str.rsplit('_', 1)[0]
            if normalized_pc in base_name or base_name in normalized_pc:
                # 매칭되는 컬럼 찾았으면 다른 쌍도 찾기
                pair_col = col_str.replace('_QTY', '_AMOUNT') if col_str.endswith('_QTY') else col_str.replace('_AMOUNT', '_QTY')
                for pair in df.columns:
                    if str(pair).upper() == pair_col:
                        qty_name = str(col) if col_str.endswith('_QTY') else str(pair)
                        amount_name = str(col) if col_str.endswith('_AMOUNT') else str(pair)
                        return {'QTY': qty_name, 'AMOUNT': amount_name}

    return None


def extract_ea_pairs_from_notes(notes: str, ea_rate_pairs: List[List[float]]) -> List[Tuple[float, float]]:
    """
    NOTES와 ea_rate_pairs에서 EA 쌍 추출 (최대 4개)

    Returns:
        [(qty, rate), ...] 리스트 (최대 4개)
    """
    ea_pairs = []

    # 1. ea_rate_pairs 우선 사용
    if ea_rate_pairs:
        for pair in ea_rate_pairs[:4]:
            if len(pair) >= 2:
                qty, rate = float(pair[0]), float(pair[1])
                if qty > 0 and rate > 0:
                    ea_pairs.append((qty, rate))

    # 2. NOTES에서 추가 추출
    if notes and len(ea_pairs) < 4:
        # qty와 unit_price 추출
        qty_match = re.search(r'qty=([^\s;]+)', notes)
        rate_match = re.search(r'unit_price=([^\s;]+)', notes)
        if qty_match and rate_match:
            try:
                qty = float(qty_match.group(1))
                rate = float(rate_match.group(1))
                if qty > 0 and rate > 0:
                    # 중복 체크
                    if not any(abs(p[0] - qty) < 0.01 and abs(p[1] - rate) < 0.01 for p in ea_pairs):
                        ea_pairs.append((qty, rate))
            except (ValueError, AttributeError):
                pass

    return ea_pairs[:4]  # 최대 4개


def convert_item_to_list_row(
    item: Dict[str, Any],
    invoice_meta: Dict[str, Any],
    row_num: int,
    voyage_no: Optional[str] = None
) -> Dict[str, Any]:
    """
    파서 출력 항목을 LIST 시트 행 형식으로 변환

    Args:
        item: 파서 출력 항목 (SUBJECT, EA_RATE_PAIRS, Amount (AED), COST MAIN, etc.)
        invoice_meta: 인보이스 메타데이터
        row_num: 행 번호
        voyage_no: Voyage No (선택)

    Returns:
        LIST 시트 행 데이터 (dict)
    """
    row = {}

    # 기본 정보
    row['NO'] = row_num
    row['SUBJECT'] = item.get('SUBJECT', '')
    row['INVOICE NUMBER'] = invoice_meta.get('invoice_number', '')
    invoice_date = invoice_meta.get('invoice_date')
    if invoice_date:
        if isinstance(invoice_date, str):
            try:
                dt = datetime.strptime(invoice_date.split()[0], '%Y-%m-%d')
                row['INVOICE DATE'] = dt
                row['INVOICE DATE_YEAR_MONTH'] = dt.strftime('%Y-%m-%d')
            except:
                row['INVOICE DATE'] = invoice_date
                row['INVOICE DATE_YEAR_MONTH'] = invoice_date
        else:
            row['INVOICE DATE'] = invoice_date
            row['INVOICE DATE_YEAR_MONTH'] = invoice_date

    row['Voyage No'] = voyage_no or invoice_meta.get('voyage_no', '')

    # Cost Center 정보
    row['COST MAIN'] = item.get('COST MAIN', '')
    row['COST CENTER A'] = item.get('COST CENTER A', '')
    row['COST CENTER B'] = item.get('COST CENTER B', '')
    row['PRICE CENTER'] = item.get('PRICE CENTER', '')

    # EA 슬롯 (EA_1~EA_4)
    ea_rate_pairs = item.get('EA_RATE_PAIRS', [])
    notes = item.get('NOTES', '')
    ea_pairs = extract_ea_pairs_from_notes(notes, ea_rate_pairs)

    for i in range(4):
        if i < len(ea_pairs):
            qty, rate = ea_pairs[i]
            row[f'EA_{i+1}_Qty'] = qty
            row[f'EA_{i+1}_Rate'] = rate
        else:
            row[f'EA_{i+1}_Qty'] = None
            row[f'EA_{i+1}_Rate'] = None

    # PRICE CENTER별 QTY/AMOUNT 컬럼 찾기 및 입력
    price_center = item.get('PRICE CENTER', '')
    amount = item.get('Amount (AED)', 0.0)

    # PRICE CENTER 컬럼 매핑 (나중에 df가 로드되면 찾을 것)
    row['_price_center_normalized'] = normalize_price_center_for_column(price_center)
    row['_amount'] = amount

    # EA 쌍에서 QTY 추출 (첫 번째 쌍의 qty 사용)
    if ea_pairs:
        row['_qty'] = ea_pairs[0][0]
    else:
        # NOTES에서 qty 추출 시도
        qty_match = re.search(r'qty=([^\s;]+)', notes)
        if qty_match:
            try:
                row['_qty'] = float(qty_match.group(1))
            except:
                row['_qty'] = 1.0  # 기본값
        else:
            row['_qty'] = 1.0

    # TOTAL AMOUNT 관련 정보 저장
    row['_amount_excl_vat'] = amount
    # NOTES에서 tax_amount와 amount_incl_vat 추출
    tax_amount_match = re.search(r'tax_amount=([^\s;]+)', notes)
    if tax_amount_match:
        try:
            row['_tax_amount'] = float(tax_amount_match.group(1))
        except:
            row['_tax_amount'] = 0.0
    else:
        row['_tax_amount'] = 0.0

    amount_incl_match = re.search(r'amount_incl_vat=([^\s;]+)', notes)
    if amount_incl_match:
        try:
            row['_amount_incl_vat'] = float(amount_incl_match.group(1))
        except:
            row['_amount_incl_vat'] = amount + row['_tax_amount']
    else:
        row['_amount_incl_vat'] = amount + row['_tax_amount']

    return row


def add_parsed_data_to_list_sheet(
    excel_path: Path,
    pdf_path: str,
    voyage_no: Optional[str] = None
) -> None:
    """
    파서 결과를 LIST 시트에 추가

    Args:
        excel_path: Excel 파일 경로
        pdf_path: PDF 파일 경로
        voyage_no: Voyage No (선택)
    """
    # PDF 파싱
    print(f"파싱 중: {pdf_path}")
    payload = parse_ofco_invoice(pdf_path)
    print(f"파싱 완료: {len(payload.items)}개 항목")

    # Excel 파일 로드
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 파일이 없습니다: {excel_path}")

    workbook = load_workbook(excel_path)
    if 'LIST' not in workbook.sheetnames:
        raise ValueError("LIST 시트가 없습니다.")

    # LIST 시트 읽기
    df_list = pd.read_excel(excel_path, sheet_name='LIST')
    print(f"기존 LIST 시트: {len(df_list)}개 행")

    # 파서 결과를 LIST 행 형식으로 변환
    invoice_meta_dict = {
        'invoice_number': payload.meta.invoice_number or '',
        'invoice_date': payload.meta.invoice_date,
        'voyage_no': payload.meta.voyage_no or voyage_no,
    }

    new_rows = []
    for idx, item in enumerate(payload.items, 1):
        # LineItem을 dict로 변환
        item_dict = {
            'SUBJECT': item.subject or '',
            'EA_RATE_PAIRS': item.ea_rates or [],
            'Amount (AED)': item.amount or 0.0,
            'COST MAIN': item.cost_main or '',
            'COST CENTER A': item.cost_center_a or '',
            'COST CENTER B': item.cost_center_b or '',
            'PRICE CENTER': item.price_center or '',
            'NOTES': item.notes or '',
        }
        row = convert_item_to_list_row(item_dict, invoice_meta_dict, idx, voyage_no)
        new_rows.append(row)

    # 새로운 행들을 DataFrame으로 변환
    new_df = pd.DataFrame(new_rows)

    # EA 슬롯 알고리즘 v1.3 적용
    print("\nEA 슬롯 알고리즘 v1.3 적용 중...")

    # PRICE CENTER별 QTY/AMOUNT 컬럼 매핑 및 입력
    for idx, row in new_df.iterrows():
        price_center_norm = row.get('_price_center_normalized', '')
        qty = row.get('_qty', 1.0)
        amount = row.get('_amount', 0.0)

        if price_center_norm:
            # 정확히 일치하는 컬럼 찾기
            qty_col = f"{price_center_norm}_QTY"
            amount_col = f"{price_center_norm}_AMOUNT"

            if qty_col in df_list.columns and amount_col in df_list.columns:
                new_df.at[idx, qty_col] = qty
                new_df.at[idx, amount_col] = amount

    # EA 슬롯 알고리즘으로 재처리 (PRICE CENTER별 컬럼 → EA 슬롯)
    # 기존 PRICE CENTER별 QTY/AMOUNT 컬럼을 EA 슬롯으로 변환
    ea_results = []
    for idx, row in new_df.iterrows():
        ea_result = process_row_to_ea_slots(row)
        ea_results.append(ea_result)

    # EA 슬롯 결과를 DataFrame에 추가
    for idx, ea_result in enumerate(ea_results):
        for key, value in ea_result.items():
            if key != '_overflow_items':
                if key not in new_df.columns:
                    new_df[key] = None
                new_df.at[idx, key] = value

    # TOTAL AMOUNT 컬럼 입력
    # 전체 인보이스 총액 (payload.meta에서)
    total_excl_vat = payload.meta.total_excl_vat or 0.0
    total_vat = payload.meta.total_vat or 0.0
    total_incl_vat = payload.meta.total_incl_vat or (total_excl_vat + total_vat)

    # USD 변환 정보
    exchange_rate = payload.meta.exchange_rate or 1.0
    amount_usd = payload.meta.alt_excl_vat or (total_excl_vat / exchange_rate if exchange_rate > 0 else 0.0)
    vat_usd = payload.meta.alt_vat or (total_vat / exchange_rate if exchange_rate > 0 else 0.0)
    total_amount_usd = payload.meta.alt_incl_vat or (total_incl_vat / exchange_rate if exchange_rate > 0 else 0.0)

    # 각 행에 개별 금액 및 전체 총액 입력
    for idx, row in new_df.iterrows():
        amount_excl_vat = row.get('_amount_excl_vat', 0.0)
        amount_incl_vat = row.get('_amount_incl_vat', 0.0)

        # 각 행의 금액 (Amount_A_AED는 개별 행의 amount_excl_vat)
        if 'Amount_A_AED' in df_list.columns:
            new_df.at[idx, 'Amount_A_AED'] = amount_excl_vat

        # 각 행의 Total_Amount_AED (개별 행의 amount_incl_vat)
        if 'Total_Amount_AED' in df_list.columns:
            new_df.at[idx, 'Total_Amount_AED'] = amount_incl_vat

        # USD 변환 (exchange_rate 사용)
        if exchange_rate and exchange_rate > 0:
            amount_usd_row = amount_excl_vat / exchange_rate
            total_amount_usd_row = amount_incl_vat / exchange_rate

            if 'Amount_USD' in df_list.columns:
                new_df.at[idx, 'Amount_USD'] = round(amount_usd_row, 2)
            if 'Total_Amount_USD' in df_list.columns:
                new_df.at[idx, 'Total_Amount_USD'] = round(total_amount_usd_row, 2)

        # VAT_USD 컬럼도 있으면 계산
        if 'VAT_USD' in df_list.columns and exchange_rate and exchange_rate > 0:
            tax_amount = row.get('_tax_amount', 0.0)
            vat_usd_row = tax_amount / exchange_rate
            new_df.at[idx, 'VAT_USD'] = round(vat_usd_row, 2)

    # 전체 인보이스 총액은 첫 번째 행에만 입력 (또는 모든 행에)
    # 사용 예시: 각 행의 Total_Amount_AED는 해당 행의 금액이고,
    # 전체 인보이스 총액은 별도로 관리하거나 계산 컬럼으로 처리

    # 임시 컬럼 제거
    new_df = new_df.drop(columns=['_price_center_normalized', '_amount', '_qty', '_amount_excl_vat', '_tax_amount', '_amount_incl_vat'], errors='ignore')

    # 전체 순번, 청구 회차 설정
    if '전체 순번' in df_list.columns:
        max_seq = df_list['전체 순번'].max() if len(df_list) > 0 else 0
        new_df['전체 순번'] = range(max_seq + 1, max_seq + len(new_df) + 1)
    if '청구 회차' in df_list.columns:
        new_df['청구 회차'] = 1  # 기본값

    # 기존 컬럼과 일치시키기 (없는 컬럼은 None으로)
    missing_cols = {col: None for col in df_list.columns if col not in new_df.columns}
    if missing_cols:
        missing_df = pd.DataFrame([missing_cols] * len(new_df))
        new_df = pd.concat([new_df, missing_df], axis=1)

    # 컬럼 순서 맞추기
    new_df = new_df[df_list.columns]

    # 라인 분할 처리 (overflow가 있는 경우)
    split_rows_list = []
    for idx, ea_result in enumerate(ea_results):
        overflow = ea_result.get('_overflow_items', [])
        if len(overflow) > 0:
            row = new_df.iloc[idx]
            line_no = row.get('line_no') or row.get('NO') or (idx + 1)
            from ofco_ea_slot_v13 import split_and_reconcile
            split_rows = split_and_reconcile(row, overflow, line_no)

            # 분할된 행들을 DataFrame 형식으로 변환
            for split_row_dict in split_rows:
                # 기존 컬럼과 일치시키기
                split_row = pd.Series(split_row_dict)
                # 컬럼 순서 맞추기
                for col in df_list.columns:
                    if col not in split_row.index:
                        split_row[col] = None
                split_rows_list.append(split_row[df_list.columns])

    if split_rows_list:
        print(f"라인 분할: {len(split_rows_list)}개 새 행 추가")
        split_df = pd.DataFrame(split_rows_list)
        new_df = pd.concat([new_df, split_df], ignore_index=True)

    # 기존 데이터에 추가
    df_combined = pd.concat([df_list, new_df], ignore_index=True)

    # Excel에 저장
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_combined.to_excel(writer, sheet_name='LIST', index=False)

    print(f"\n완료: {excel_path}")
    print(f"LIST 시트: {len(df_combined)}개 행 (기존 {len(df_list)}개 + 신규 {len(new_df)}개)")


def main():
    import argparse

    parser = argparse.ArgumentParser(description='OFCO 파서 결과를 LIST 시트에 추가')
    parser.add_argument('pdf_path', help='PDF 파일 경로')
    parser.add_argument('--excel', default=None, help='Excel 파일 경로 (기본: OFCO INVOICE.xlsx)')
    parser.add_argument('--voyage', default=None, help='Voyage No')

    args = parser.parse_args()

    # Excel 파일 경로 결정
    if args.excel is None:
        excel_path = Path(__file__).parent.parent.parent / "OFCO INVOICE.xlsx"
    else:
        excel_path = Path(args.excel)

    add_parsed_data_to_list_sheet(excel_path, args.pdf_path, args.voyage)


if __name__ == '__main__':
    main()

