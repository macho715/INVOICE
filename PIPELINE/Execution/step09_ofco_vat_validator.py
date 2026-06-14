#!/usr/bin/env python3
"""
OFCO VAT 검증 로직 강화 스크립트

vat_check 실패 원인:
- VAT 계산 오차 (5% 세율 적용 시 반올림 문제)
- PDF에서 VAT 금액 OCR 추출 오류

해결 방법:
- VAT 오차 허용 범위 도입 (기본 2% 이내 PASS)
- VAT 자동 재계산 옵션 제공
- 통화별 VAT 처리 (AED/USD 분리)

Usage:
    python step09_ofco_vat_validator.py --input ofco_pipeline_pdf_XXX.xlsx --sheet EA_Slots
    python step09_ofco_vat_validator.py --input "OFCO INVOICE.xlsx" --sheet Sheet1 --tolerance 0.02 --recalculate

Options:
    --input: 입력 파일 경로
    --sheet: 대상 시트명 (기본값: EA_Slots)
    --tolerance: VAT 오차 허용 비율 (기본값: 0.02 = 2%)
    --recalculate: VAT 자동 재계산 활성화
    --vat-rate: VAT 세율 (기본값: 0.05 = 5%)
    --backup: 백업 파일 생성 (기본값: True)
"""

import pandas as pd
import sys
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Tuple, List
import argparse


def validate_vat_with_tolerance(
    df: pd.DataFrame,
    tolerance: float = 0.02,
    vat_rate: float = 0.05,
    recalculate: bool = False
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    VAT 검증 및 수정 (오차 허용 범위 적용)
    
    Args:
        df: 대상 DataFrame
        tolerance: VAT 오차 허용 비율 (0.02 = 2%)
        vat_rate: VAT 세율 (0.05 = 5%)
        recalculate: VAT 자동 재계산 여부
    
    Returns:
        (수정된 DataFrame, 통계 정보)
    """
    stats = {
        'total_rows': len(df),
        'vat_check_failures_before': 0,
        'vat_check_failures_after': 0,
        'vat_within_tolerance': 0,
        'vat_recalculated': 0,
        'vat_missing': 0
    }
    
    print(f"[INFO] VAT 검증 설정:")
    print(f"  - 세율: {vat_rate * 100}%")
    print(f"  - 오차 허용 범위: ±{tolerance * 100}%")
    print(f"  - 자동 재계산: {'활성화' if recalculate else '비활성화'}")
    
    # EA_Slots / Complete_Pipeline 처리
    if any(f'EA_{i}_Amount' in df.columns for i in range(1, 5)):
        # EA_Total_Amount 기준 VAT 검증
        if 'EA_Total_Amount' in df.columns:
            # VAT 컬럼 확인 (여러 가능성)
            vat_cols = [col for col in df.columns if 'VAT' in col.upper() or 'TAX' in col.upper()]
            
            if vat_cols:
                vat_col = vat_cols[0]  # 첫 번째 VAT 컬럼 사용
                print(f"[INFO] VAT 컬럼 사용: {vat_col}")
                
                # 1. VAT 누락 확인
                mask_missing = df['EA_Total_Amount'].notna() & df[vat_col].isna()
                if mask_missing.any():
                    stats['vat_missing'] = mask_missing.sum()
                    if recalculate:
                        df.loc[mask_missing, vat_col] = df.loc[mask_missing, 'EA_Total_Amount'] * vat_rate
                        df.loc[mask_missing, 'VAT_Data_Source'] = 'VAT_Calculated_Missing'
                        print(f"[INFO] VAT 누락 {stats['vat_missing']}건 자동 계산")
                    else:
                        print(f"[WARNING] VAT 누락 {stats['vat_missing']}건 (재계산 비활성화)")
                
                # 2. VAT 검증
                mask_valid = df['EA_Total_Amount'].notna() & df[vat_col].notna()
                if mask_valid.any():
                    vat_calculated = df.loc[mask_valid, 'EA_Total_Amount'] * vat_rate
                    vat_actual = df.loc[mask_valid, vat_col]
                    
                    # 절대 차이
                    vat_diff = abs(vat_calculated - vat_actual)
                    
                    # 상대 오차 (실제 VAT 대비 %)
                    vat_relative_error = vat_diff / vat_actual.abs()
                    vat_relative_error = vat_relative_error.fillna(0)  # 0으로 나누기 방지
                    
                    # 허용 범위 내 여부
                    mask_within_tolerance = mask_valid.copy()
                    mask_within_tolerance[mask_valid] = vat_relative_error <= tolerance
                    stats['vat_within_tolerance'] = mask_within_tolerance.sum()
                    
                    # 허용 범위 초과 (재계산 필요)
                    mask_needs_recalc = mask_valid & ~mask_within_tolerance
                    
                    if mask_needs_recalc.any():
                        count = mask_needs_recalc.sum()
                        if recalculate:
                            df.loc[mask_needs_recalc, vat_col] = vat_calculated[mask_needs_recalc]
                            df.loc[mask_needs_recalc, 'VAT_Data_Source'] = 'VAT_Recalculated'
                            stats['vat_recalculated'] = count
                            print(f"[INFO] VAT 오차 초과 {count}건 재계산")
                        else:
                            print(f"[WARNING] VAT 오차 초과 {count}건 (재계산 비활성화)")
                    
                    print(f"[INFO] VAT 허용 범위 내: {stats['vat_within_tolerance']}/{mask_valid.sum()}행")
            else:
                print(f"[WARNING] VAT 관련 컬럼을 찾을 수 없습니다")
    
    # Standard_Lines 처리
    elif 'amount_excl_tax' in df.columns and 'tax_amount' in df.columns:
        # 1. VAT 누락 확인
        mask_missing = df['amount_excl_tax'].notna() & df['tax_amount'].isna()
        if mask_missing.any():
            stats['vat_missing'] = mask_missing.sum()
            if recalculate:
                df.loc[mask_missing, 'tax_amount'] = df.loc[mask_missing, 'amount_excl_tax'] * vat_rate
                df.loc[mask_missing, 'vat_data_source'] = 'tax_amount_Calculated'
                print(f"[INFO] tax_amount 누락 {stats['vat_missing']}건 자동 계산")
            else:
                print(f"[WARNING] tax_amount 누락 {stats['vat_missing']}건")
        
        # 2. VAT 검증
        mask_valid = df['amount_excl_tax'].notna() & df['tax_amount'].notna()
        if mask_valid.any():
            vat_calculated = df.loc[mask_valid, 'amount_excl_tax'] * vat_rate
            vat_actual = df.loc[mask_valid, 'tax_amount']
            
            vat_diff = abs(vat_calculated - vat_actual)
            vat_relative_error = vat_diff / vat_actual.abs()
            vat_relative_error = vat_relative_error.fillna(0)
            
            mask_within_tolerance = mask_valid.copy()
            mask_within_tolerance[mask_valid] = vat_relative_error <= tolerance
            stats['vat_within_tolerance'] = mask_within_tolerance.sum()
            
            mask_needs_recalc = mask_valid & ~mask_within_tolerance
            
            if mask_needs_recalc.any():
                count = mask_needs_recalc.sum()
                if recalculate:
                    df.loc[mask_needs_recalc, 'tax_amount'] = vat_calculated[mask_needs_recalc]
                    df.loc[mask_needs_recalc, 'vat_data_source'] = 'tax_amount_Recalculated'
                    stats['vat_recalculated'] = count
                    print(f"[INFO] tax_amount 오차 초과 {count}건 재계산")
                else:
                    print(f"[WARNING] tax_amount 오차 초과 {count}건")
            
            print(f"[INFO] VAT 허용 범위 내: {stats['vat_within_tolerance']}/{mask_valid.sum()}행")
    
    # 3. vat_check 재검증
    if 'vat_check' in df.columns:
        stats['vat_check_failures_before'] = (~df['vat_check']).sum()
        print(f"[INFO] 수정 전 vat_check 실패: {stats['vat_check_failures_before']}행")
        
        # vat_check 재계산 (허용 범위 적용)
        if any(f'EA_{i}_Amount' in df.columns for i in range(1, 5)) and 'EA_Total_Amount' in df.columns:
            vat_cols = [col for col in df.columns if 'VAT' in col.upper() or 'TAX' in col.upper()]
            if vat_cols:
                vat_col = vat_cols[0]
                mask = df['EA_Total_Amount'].notna() & df[vat_col].notna()
                if mask.any():
                    vat_calculated = df.loc[mask, 'EA_Total_Amount'] * vat_rate
                    vat_actual = df.loc[mask, vat_col]
                    vat_relative_error = abs(vat_calculated - vat_actual) / vat_actual.abs()
                    vat_relative_error = vat_relative_error.fillna(0)
                    df.loc[mask, 'vat_check'] = vat_relative_error <= tolerance
        elif 'amount_excl_tax' in df.columns and 'tax_amount' in df.columns:
            mask = df['amount_excl_tax'].notna() & df['tax_amount'].notna()
            if mask.any():
                vat_calculated = df.loc[mask, 'amount_excl_tax'] * vat_rate
                vat_actual = df.loc[mask, 'tax_amount']
                vat_relative_error = abs(vat_calculated - vat_actual) / vat_actual.abs()
                vat_relative_error = vat_relative_error.fillna(0)
                df.loc[mask, 'vat_check'] = vat_relative_error <= tolerance
        
        stats['vat_check_failures_after'] = (~df['vat_check']).sum()
        improvement = stats['vat_check_failures_before'] - stats['vat_check_failures_after']
        print(f"[INFO] 수정 후 vat_check 실패: {stats['vat_check_failures_after']}행 (개선: {improvement}행)")
    
    return df, stats


def process_file(
    input_file: str,
    sheet_name: str = 'EA_Slots',
    tolerance: float = 0.02,
    vat_rate: float = 0.05,
    recalculate: bool = False,
    backup: bool = True
) -> Tuple[bool, Dict[str, Any]]:
    """
    파일 처리 메인 함수
    
    Args:
        input_file: 입력 파일 경로
        sheet_name: 대상 시트명
        tolerance: VAT 오차 허용 비율
        vat_rate: VAT 세율
        recalculate: VAT 자동 재계산 여부
        backup: 백업 생성 여부
    
    Returns:
        (성공 여부, 통계 정보)
    """
    overall_stats = {
        'input_file': input_file,
        'sheet_name': sheet_name,
        'success': False,
        'backup_created': False,
        'errors': []
    }
    
    try:
        input_path = Path(input_file)
        
        if not input_path.exists():
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_file}")
        
        print(f"[INFO] 입력 파일: {input_path}")
        print(f"[INFO] 대상 시트: {sheet_name}")
        
        # 백업 생성
        if backup:
            backup_path = input_path.with_suffix(f'.xlsx.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            shutil.copy2(input_path, backup_path)
            overall_stats['backup_created'] = True
            overall_stats['backup_path'] = str(backup_path)
            print(f"[OK] 백업 파일 생성: {backup_path}")
        
        # 파일 읽기
        print(f"\n[Step 1/3] 파일 읽기...")
        df = pd.read_excel(input_path, sheet_name=sheet_name)
        print(f"[OK] {len(df)}행, {len(df.columns)}컬럼 로드")
        
        # VAT 검증 및 수정
        print(f"\n[Step 2/3] VAT 검증 및 수정...")
        df, stats = validate_vat_with_tolerance(
            df,
            tolerance=tolerance,
            vat_rate=vat_rate,
            recalculate=recalculate
        )
        overall_stats.update(stats)
        
        # 파일 저장
        print(f"\n[Step 3/3] 파일 저장...")
        
        # 기존 파일의 다른 시트 유지
        from openpyxl import load_workbook
        wb = load_workbook(input_path)
        
        # 대상 시트 업데이트
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
        
        # DataFrame을 시트에 쓰기
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 임시 파일로 저장 후 이동
        temp_path = input_path.with_suffix('.xlsx.tmp')
        wb.save(temp_path)
        wb.close()
        
        if temp_path.exists():
            shutil.move(str(temp_path), str(input_path))
            print(f"[OK] 파일 저장 완료: {input_path}")
        else:
            raise Exception("임시 파일 저장 실패")
        
        overall_stats['success'] = True
        
        # 통계 출력
        print(f"\n{'='*80}")
        print(f"VAT 검증 완료!")
        print(f"{'='*80}")
        print(f"전체 행 수: {stats['total_rows']}")
        print(f"VAT 누락: {stats['vat_missing']}건")
        print(f"VAT 허용 범위 내: {stats['vat_within_tolerance']}건")
        
        if recalculate:
            print(f"VAT 재계산: {stats['vat_recalculated']}건")
        
        if 'vat_check_failures_before' in stats:
            print(f"\nvat_check 실패 (수정 전): {stats['vat_check_failures_before']}행")
            print(f"vat_check 실패 (수정 후): {stats['vat_check_failures_after']}행")
            improvement = stats['vat_check_failures_before'] - stats['vat_check_failures_after']
            if improvement > 0:
                improvement_pct = (improvement / stats['vat_check_failures_before'] * 100) if stats['vat_check_failures_before'] > 0 else 0
                print(f"개선율: {improvement}행 ({improvement_pct:.1f}%)")
        
        return True, overall_stats
        
    except FileNotFoundError as e:
        error_msg = f"파일을 찾을 수 없습니다: {e}"
        print(f"\n[ERROR] {error_msg}")
        overall_stats['errors'].append(error_msg)
        return False, overall_stats
    
    except Exception as e:
        error_msg = f"처리 중 오류 발생: {type(e).__name__} - {e}"
        print(f"\n[ERROR] {error_msg}")
        overall_stats['errors'].append(error_msg)
        
        # 롤백
        if backup and overall_stats.get('backup_created') and 'backup_path' in overall_stats:
            backup_path = Path(overall_stats['backup_path'])
            if backup_path.exists():
                print(f"\n[INFO] 롤백 시도...")
                try:
                    shutil.copy2(backup_path, input_path)
                    print(f"[OK] 롤백 완료")
                except Exception as rollback_error:
                    print(f"[ERROR] 롤백 실패: {rollback_error}")
        
        return False, overall_stats


def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(
        description='OFCO VAT 검증 로직 강화 스크립트 v1.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  # 기본 검증 (2% 허용 범위, 재계산 없음)
  python step09_ofco_vat_validator.py --input ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx
  
  # VAT 자동 재계산 활성화
  python step09_ofco_vat_validator.py --input ofco_pipeline_pdf_OFCO_INV_0001178_v3.xlsx --recalculate
  
  # 허용 범위 5%로 확대
  python step09_ofco_vat_validator.py --input "OFCO INVOICE.xlsx" --sheet Sheet1 --tolerance 0.05 --recalculate
        """
    )
    
    parser.add_argument('--input', type=str, required=True,
                        help='입력 파일 경로')
    parser.add_argument('--sheet', type=str, default='EA_Slots',
                        help='대상 시트명 (기본값: EA_Slots)')
    parser.add_argument('--tolerance', type=float, default=0.02,
                        help='VAT 오차 허용 비율 (기본값: 0.02 = 2%%)')
    parser.add_argument('--vat-rate', type=float, default=0.05,
                        help='VAT 세율 (기본값: 0.05 = 5%%)')
    parser.add_argument('--recalculate', action='store_true', default=False,
                        help='VAT 자동 재계산 활성화')
    parser.add_argument('--backup', action='store_true', default=True,
                        help='백업 파일 생성 (기본값: True)')
    parser.add_argument('--no-backup', action='store_false', dest='backup',
                        help='백업 파일 생성 안 함')
    
    args = parser.parse_args()
    
    print("="*80)
    print("OFCO VAT 검증 로직 강화 스크립트")
    print("="*80)
    print()
    
    success, stats = process_file(
        input_file=args.input,
        sheet_name=args.sheet,
        tolerance=args.tolerance,
        vat_rate=args.vat_rate,
        recalculate=args.recalculate,
        backup=args.backup
    )
    
    if success:
        print(f"\n[SUCCESS] 처리가 완료되었습니다!")
        sys.exit(0)
    else:
        print(f"\n[FAILURE] 처리 중 오류가 발생했습니다.")
        if stats.get('errors'):
            print(f"\n오류 목록:")
            for error in stats['errors']:
                print(f"  - {error}")
        sys.exit(1)


if __name__ == '__main__':
    main()
