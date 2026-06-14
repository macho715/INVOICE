# ==============================
# core/duplicate_handler.py
# ==============================
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, Iterable, List, Tuple, Literal, Callable, Optional
import re
import unicodedata

import numpy as np
import pandas as pd

# Optional semantic matcher wiring (graceful fallback)
try:
    from semantic_matcher import SemanticMatcher  # project-local
except Exception:  # pragma: no cover
    SemanticMatcher = None  # type: ignore


# -------- Normalization Profile ---------
@dataclass(frozen=True)
class NormalizationProfile:
    squeeze_spaces: bool = True
    uppercase: bool = True
    keep_chars: str = r"0-9A-Za-z\-_/"  # allowed charset before sep unification
    unify_sep: bool = True  # unify -, _, / into '-'
    zero_pad_digits: Optional[int] = None  # e.g., 6 -> '000123'


def _normalize_case_no_scalar(x: object, prof: NormalizationProfile) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    s = str(x)
    # NFKC (fullwidth -> halfwidth) + strip
    s = unicodedata.normalize("NFKC", s).strip()
    if prof.squeeze_spaces:
        s = re.sub(r"\s+", "", s)
    # keep only whitelisted chars first
    s = re.sub(fr"[^{prof.keep_chars}]", "", s)
    if prof.unify_sep:
        s = re.sub(r"[_/]", "-", s)
        s = re.sub(r"-+", "-", s)
    if prof.uppercase:
        s = s.upper()
    # optional zero-pad for trailing digits
    if prof.zero_pad_digits is not None and prof.zero_pad_digits > 0:
        m = re.match(r"^(.*?)(\d+)$", s)
        if m:
            head, digits = m.groups()
            s = f"{head}{int(digits):0{prof.zero_pad_digits}d}"
    return s


def normalize_case_series(series: pd.Series, profile: NormalizationProfile | None = None) -> pd.Series:
    prof = profile or NormalizationProfile()
    return series.astype("object").map(lambda v: _normalize_case_no_scalar(v, prof))


# -------- Column Resolution ---------
SemanticFinder = Callable[[pd.DataFrame, str], Optional[str]]


def default_semantic_finder(df: pd.DataFrame, tag: str, required: bool = False) -> Optional[str]:
    """Try to use project SemanticMatcher if available. Fallback to simple heuristics."""
    # 1) try SemanticMatcher
    if SemanticMatcher is not None:
        try:
            m = SemanticMatcher()
            return m.find_column(df, tag, required=required)
        except Exception:
            pass
    # 2) heuristic fallback
    candidates = [c for c in df.columns if re.search(r"(case|package)", str(c), re.I)]
    if candidates:
        return candidates[0]
    if required:
        raise KeyError(f"Column for tag '{tag}' not found")
    return None


# -------- Core Duplicate Logic ---------
_CASE_KEY = "__case_key__"
_DUP_LABEL = "__dup_label__"


def detect_case_column(df: pd.DataFrame, semantic_finder: Optional[Callable[..., Optional[str]]] = None,
                       required: bool = True) -> Optional[str]:
    finder = semantic_finder or default_semantic_finder
    return finder(df, "case_number", required=required)  # type: ignore[arg-type]


def mark_duplicates(
    df: pd.DataFrame,
    semantic_finder: Optional[Callable[..., Optional[str]]] = None,
    keep: Literal["first", "last", "False"] = "first",
    profile: NormalizationProfile | None = None,
) -> Tuple[pd.DataFrame, pd.Series]:
    """
    Create normalized key and within-sheet duplicate mask.
    Returns (df_with_key_and_label, dup_mask)
    """
    out = df.copy()
    col = detect_case_column(out, semantic_finder=semantic_finder, required=True)
    key = normalize_case_series(out[col], profile)  # type: ignore[index]
    out[_CASE_KEY] = key
    dup_mask = out.duplicated(subset=[_CASE_KEY], keep=(None if keep == "False" else keep))
    out[_DUP_LABEL] = np.where(dup_mask, "중복", "고유")
    return out, dup_mask


def compute_global_dup_flags(
    sheet_dfs: Dict[str, pd.DataFrame],
    semantic_finder: Optional[Callable[..., Optional[str]]] = None,
    profile: NormalizationProfile | None = None,
) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Make a global frequency across all sheets and assign 고유/중복 per row.
    Returns (updated_sheet_dfs, summary_df[File Name, Sheet Name, 고유, 중복, 총합계])
    """
    prof = profile or NormalizationProfile()
    # 1) build keys list
    keys: List[pd.Series] = []
    resolved: Dict[str, Tuple[pd.DataFrame, Optional[str]]] = {}
    for name, df in sheet_dfs.items():
        tdf = df.copy()
        try:
            col = detect_case_column(tdf, semantic_finder=semantic_finder, required=True)
        except Exception:
            col = None
        if col:
            tdf[_CASE_KEY] = normalize_case_series(tdf[col], prof)
            keys.append(tdf[_CASE_KEY])
        else:
            tdf[_CASE_KEY] = ""
            tdf[_DUP_LABEL] = ""  # header unresolved
        resolved[name] = (tdf, col)

    # 2) global frequency
    freq = pd.concat(keys).value_counts(dropna=False) if keys else pd.Series(dtype=int)

    # 3) assign labels + build summary
    summary_rows: List[Dict[str, object]] = []
    updated: Dict[str, pd.DataFrame] = {}
    for name, (tdf, col) in resolved.items():
        if col:
            cnt = tdf[_CASE_KEY].map(freq).fillna(0).astype(int)
            dup = cnt.gt(1)
            tdf[_DUP_LABEL] = np.where(dup, "중복", "고유")
            unique_cnt = int((~dup).sum())
            dup_cnt = int(dup.sum())
        else:
            unique_cnt = int(len(tdf))
            dup_cnt = 0
        summary_rows.append({
            "File Name": "",
            "Sheet Name": name,
            "고유": unique_cnt,
            "중복": dup_cnt,
            "총합계": int(len(tdf)),
        })
        updated[name] = tdf

    summary = pd.DataFrame(summary_rows, columns=["File Name", "Sheet Name", "고유", "중복", "총합계"])
    return updated, summary


def drop_duplicates_by_case(
    df: pd.DataFrame,
    semantic_finder: Optional[Callable[..., Optional[str]]] = None,
    keep: Literal["first", "last", "False"] = "first",
    profile: NormalizationProfile | None = None,
) -> pd.DataFrame:
    col = detect_case_column(df, semantic_finder=semantic_finder, required=False)
    if not col:
        return df
    key = normalize_case_series(df[col], profile)  # type: ignore[index]
    mask = ~key.duplicated(keep=(None if keep == "False" else keep))
    return df.loc[mask].copy()


def drop_duplicates_global(
    dfs: Iterable[pd.DataFrame],
    semantic_finder: Optional[Callable[..., Optional[str]]] = None,
    keep: Literal["first", "last", "False"] = "first",
    profile: NormalizationProfile | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    diag_rows: List[Dict[str, object]] = []
    tagged: List[pd.DataFrame] = []
    for idx, df in enumerate(dfs, start=1):
        tdf, dup = mark_duplicates(df, semantic_finder=semantic_finder, keep="first", profile=profile)
        diag_rows.append({
            "part": f"sheet_{idx}",
            "rows": int(len(tdf)),
            "unique": int((~dup).sum()),
            "dups": int(dup.sum()),
        })
        tagged.append(tdf)
    pre_diag = pd.DataFrame(diag_rows, columns=["part", "rows", "unique", "dups"])  # before-global
    merged = pd.concat(tagged, ignore_index=True) if tagged else pd.DataFrame()
    if not merged.empty:
        merged = merged.drop_duplicates(subset=[_CASE_KEY], keep=(None if keep == "False" else keep))
    return pre_diag, merged


# -------- Excel Adapter (values injection; formula is optional) ---------
from openpyxl import load_workbook  # heavy import left here intentionally
from openpyxl.utils import get_column_letter


def inject_excel_dup_column(
    xlsx_path: str,
    sheet_name: str,
    dup_values: List[str],
    header_row: int,
    data_start_row: Optional[int] = None,
) -> None:
    """
    Write computed dup labels into column A. Avoids style breakage from column insert.
    If you must use formulas instead of values, adapt here.
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    hdr_cell = ws.cell(row=header_row, column=1)
    hdr_cell.value = "고유 중복"

    start = data_start_row or (header_row + 1)
    for i, v in enumerate(dup_values, start=start):
        ws.cell(row=i, column=1, value=v)
    wb.save(xlsx_path)


# -------- Validation Summary ---------

def validate_duplicate_removal(diag_df: pd.DataFrame) -> Dict[str, float]:
    total_rows = float(diag_df.get("rows", pd.Series(dtype=float)).sum()) if not diag_df.empty else 0.0
    total_dups = float(diag_df.get("dups", pd.Series(dtype=float)).sum()) if not diag_df.empty else 0.0
    rate = (total_dups / total_rows) if total_rows else 0.0
    return {
        "total_parts": float(diag_df.shape[0]),
        "total_rows": total_rows,
        "total_dups": total_dups,
        "dup_rate": rate,
    }


# ==============================
# data_synchronizer_v30.py — Unified Diff (context-aware, git-apply --3way 권장)
# ==============================
# 아래 패치는 import, Stage-1 per-sheet 저장 직전(A), 병합 저장 직전(B)에 훅을 추가합니다.

*** Begin Patch
*** Update File: data_synchronizer_v30.py
@@
-import pandas as pd
+import pandas as pd
+import numpy as np
@@
-from typing import Any, Dict, List
+from typing import Any, Dict, List
+
+# DUP-GUARD imports (new)
+from core.duplicate_handler import (
+    compute_global_dup_flags,
+    drop_duplicates_by_case,
+    NormalizationProfile,
+    mark_duplicates,
+    detect_case_column,
+)
+try:
+    from semantic_matcher import SemanticMatcher
+    _SEM = SemanticMatcher()
+    def _find_col(df, tag, required=False):
+        return _SEM.find_column(df, tag, required=required)
+except Exception:
+    def _find_col(df, tag, required=False):
+        from core.duplicate_handler import default_semantic_finder
+        return default_semantic_finder(df, tag, required=required)
@@  # Stage 1: per-sheet processing loop (right before writing each sheet)
-    # existing: df is ready to be saved to Excel
+    # DUP-GUARD (A): within-sheet duplicate flags + A-column injection
+    try:
+        df_with_flag, dup_mask = mark_duplicates(df, semantic_finder=_find_col, keep="first")
+        df = df_with_flag
+        # place visible label as first column for Excel output
+        if "고유 중복" not in df.columns:
+            df.insert(0, "고유 중복", np.where(dup_mask, "중복", "고유"))
+    except Exception as e:
+        logger.warning(f"[DUP-GUARD][{sheet_name}] skip marking duplicates: {e}")
@@  # Stage 1: after concatenating sheets into merged dataframe
-merged_df = pd.concat(sheet_frames, ignore_index=True)
+merged_df = pd.concat(sheet_frames, ignore_index=True)
+# DUP-GUARD (B): workbook-level dedup by Case key
+try:
+    merged_df = drop_duplicates_by_case(merged_df, semantic_finder=_find_col, keep="first")
+except Exception as e:
+    logger.warning(f"[DUP-GUARD][merge] skip global dedup: {e}")
*** End Patch
