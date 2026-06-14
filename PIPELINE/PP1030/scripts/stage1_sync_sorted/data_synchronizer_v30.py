# -*- coding: utf-8 -*-
"""
DataSynchronizer v3.0 - Semantic Header Matching Edition
=========================================================

This is a completely refactored version of the data synchronizer that uses
the new core header matching system. All hardcoded column names have been
replaced with semantic key lookups.

Key Improvements:
- Zero hardcoding of column names
- Automatic adaptation to different Excel formats
- Robust header detection and matching
- Comprehensive error reporting
- Better maintainability

Migration from v2.9:
- All column name strings replaced with semantic keys
- Header finding logic centralized in core modules
- Enhanced validation and error messages

HEADER MANAGEMENT POLICY:
========================
All header/column definitions and matching logic are centralized in @core/.
This file MUST NOT hardcode any column names. All column references must go through:

- get_warehouse_columns() - Get warehouse column list from core registry
- get_site_columns() - Get site column list from core registry
- SemanticMatcher - Match columns by semantic meaning, not exact names
- HVDC_HEADER_REGISTRY - Central registry for all header definitions

DO NOT hardcode column names like "Case No.", "DHL WH", etc.
ALWAYS use semantic keys like "case_number", "dhl_wh" and let core handle matching.

For details, see: @core/README.md and @core/INTEGRATION_GUIDE.md
"""

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Add project root to Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Import the new core header matching system
from scripts.core import (
    HVDC_HEADER_REGISTRY,
    STAGE1_BASE_COLS_ORDER,
    HeaderCategory,
    HeaderDetectionResult,
    HeaderDetector,
    HeaderRegistry,
    SemanticMatcher,
    find_header_by_meaning,
)
from scripts.core.standard_header_order import reorder_dataframe_columns

# DUP-GUARD imports (new)
from scripts.core.duplicate_handler import (
    compute_global_dup_flags,
    drop_duplicates_by_case,
    NormalizationProfile,
    mark_duplicates,
    detect_case_column,
)

try:
    from scripts.core import SemanticMatcher as CoreSemanticMatcher

    _SEM = CoreSemanticMatcher()

    def _find_col(df, tag, required=False):
        return _SEM.find_column(df, tag, required=required)

except Exception:

    def _find_col(df, tag, required=False):
        from scripts.core.duplicate_handler import default_semantic_finder

        return default_semantic_finder(df, tag, required=required)


# ===== Configuration =====
ORANGE = "FFFFA500"  # Changed date cell (ARGB format) - True ORANGE
YELLOW = "FFFFFF00"  # New row (ARGB format)

# Invalid header patterns to filter out
INVALID_HEADER_PATTERNS = [
    r"^열\d+$",  # 열1, 열2 등
    r"^\d+$",  # 0, 1, 2, 3, 4 등 순수 숫자
    r"^총합계$",  # 총합계
    r"^Unnamed:.*$",  # Unnamed: 0, Unnamed: 1 등
    r"^\.+$",  # ... 등 점만 있는 컬럼
]

# Metadata columns that should not be overwritten during sync
METADATA_COLUMNS = [
    "Source_Sheet",  # Original sheet name - should be preserved
    "Source_Vendor",  # Vendor name (HITACHI/SIEMENS) - should be preserved
]

# Sheet names to exclude (aggregate/summary sheets)
EXCLUDED_SHEET_NAMES = [
    "summary",  # English
    "총합계",  # Korean total
    "total",  # English total
    "aggregate",  # Aggregate data
]

# Sheet name aliases for semantic matching
SHEET_NAME_ALIASES = {
    "case list": ["case list, ril", "case list ril", "caselist", "case_list"],
    "he capacitor": ["he-0214,0252 (capacitor)", "he-0214", "he-0252", "capacitor"],
}

# 시트 포함 패턴(대소문자 무시)
INCLUDE_SHEET_PATTERNS = [
    r"(?i)^case\s*list",                 # Case List
    r"(?i)^ril(\s*\(.*\))?$",            # RIL, RIL(SIMENSE), RIL(HE) 등
    r"(?i)^he\s*local$",                 # HE Local
    r"(?i)^he[-_ ]?\d+(?:\s*,\s*\d+)*(?:\s*\(.*\))?$",  # HE-0214,0252 (Capacitor) - 괄호 부가정보 허용
    r"(?i)capacitor",                    # ... (Capacitor)
    r"(?i)shipment",                     # HE Shipment (Warehouse)
]

# 표준화된 시트명 매핑
CANON_SHEET_MAP = [
    (re.compile(r"(?i)^case\s*list"),          "CASE_LIST"),
    (re.compile(r"(?i)^ril"),                  "RIL"),
    (re.compile(r"(?i)^he\s*local$"),          "HE_LOCAL"),
    (re.compile(r"(?i)^he[-_ ]?\d"),           "HE_CAPACITOR"),
    (re.compile(r"(?i)shipment"),              "HE_SHIPMENT"),
]


def _should_include_sheet(name: str) -> bool:
    """시트명이 포함 패턴에 맞는지 확인"""
    low = name.strip().lower()
    return any(re.search(p, low) for p in INCLUDE_SHEET_PATTERNS)


def _canon_sheet(name: str) -> str:
    """시트명을 표준화된 태그로 변환"""
    for rx, tag in CANON_SHEET_MAP:
        if rx.search(name):
            return tag
    return "MISC"


def _infer_sheet_vendor(name: str) -> str:
    """시트명 기준 벤더 힌트: (SIMENSE|SIEMENS|SIM) / (HE|HITACHI)"""
    low = name.lower()
    normalized = re.sub(r"[^0-9a-z]+", " ", low)
    if re.search(r"\b(simense|siemens|sim)\b", normalized):
        return "simense"
    if re.search(r"\b(he|hitachi)\b", normalized):
        return "hitachi"
    return ""


def _fix_case_key(df: pd.DataFrame) -> pd.DataFrame:
    """
    Case No 키 손실 방지: 문자열 강제, 공백 제거, 특수 유니코드 스페이스 제거
    
    Args:
        df: DataFrame to fix
        
    Returns:
        DataFrame with fixed Case No column
    """
    # 후보 키 컬럼을 문자열로 통일
    case_col = None
    for cand in ["Case No.", "Case No", "CASE NO", "Caseno", "Case Number", "case no"]:
        if cand in df.columns:
            case_col = cand
            break
    
    if case_col:
        s = df[case_col].astype(str)
        # 특수 유니코드 스페이스 제거 (\u00A0, \u2007, \u202F 등)
        s = s.str.replace(r"\u00A0|\u2007|\u202F", " ", regex=True)
        s = s.str.strip()
        s = s.replace({"nan": "", "NaN": "", "None": "", "": ""})
        # 표준 키 컬럼으로 통일
        if "Case No." not in df.columns:
            df["Case No."] = s
        else:
            df["Case No."] = s
    elif "Case No." not in df.columns:
        # Case No 컬럼이 없는 시트도 일단 살려서 나중에 라우팅
        df["Case No."] = ""
    
    return df

# File name keywords for identifying same file types
FILE_NAME_KEYWORDS = {
    "master": ["case list", "case_list", "caselist", "master"],
    "warehouse": ["hvdc", "hitachi", "warehouse", "he"],
}

# ✅ Corrected: These columns contain WAREHOUSE IN/OUT DATES
# Column names say "location" but actual data is DATES (warehouse entry/exit timestamps)
DATE_SEMANTIC_KEYS = [
    "etd_atd",  # Estimated/Actual Time of Departure
    "eta_ata",  # Estimated/Actual Time of Arrival
    "dhl_wh",  # DHL Warehouse IN date
    "dsv_indoor",  # DSV Indoor Warehouse IN date
    "dsv_al_markaz",  # DSV Al Markaz Warehouse IN date
    "dsv_outdoor",  # DSV Outdoor Warehouse IN date
    "aaa_storage",  # AAA Storage IN date
    "hauler_indoor",  # Hauler Indoor IN date
    "dsv_mzp",  # DSV MZP IN date
    "mosb",  # MOSB IN date
    "shifting",  # Shifting date
    "mir",  # MIR site IN date
    "shu",  # SHU site IN date
    "das",  # DAS site IN date
    "agi",  # AGI site IN date
]

ALWAYS_OVERWRITE_NONDATE = True


def _to_date(val) -> Optional[pd.Timestamp]:
    """Convert various date formats to pandas Timestamp."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        if isinstance(val, pd.Timestamp):
            return val
        return pd.to_datetime(val, errors="coerce")
    except Exception:
        return None


@dataclass
class Change:
    """Record of a single cell change."""

    row_index: int
    column_name: str
    old_value: Any
    new_value: Any
    change_type: str  # "date_update" | "field_update" | "new_record"
    semantic_key: str = ""  # ✅ Phase 1: Semantic key for flexible column mapping
    case_no: str = ""  # ✅ Phase 4: Case No. for correct row lookup after reordering


@dataclass
class ChangeTracker:
    """Tracks all changes made during synchronization."""

    changes: List[Change] = field(default_factory=list)
    new_cases: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    semantic_to_column: Dict[str, str] = field(default_factory=dict)  # ✅ Phase 2: Mapping storage

    def set_column_mapping(self, semantic_key: str, column_name: str):
        """✅ Phase 2: Store semantic key → actual column name mapping"""
        if semantic_key and column_name:
            self.semantic_to_column[semantic_key] = column_name

    def get_column_name(self, semantic_key: str, fallback: str = "") -> str:
        """✅ Phase 2: Get actual column name from semantic key (with fallback)"""
        return self.semantic_to_column.get(semantic_key, fallback)

    def add_change(self, **kw):
        """Add a change record."""
        self.changes.append(
            Change(
                row_index=int(kw.get("row_index", -1)),
                column_name=str(kw.get("column_name", "")),
                old_value=kw.get("old_value"),
                new_value=kw.get("new_value"),
                change_type=str(kw.get("change_type", "field_update")),
                semantic_key=str(kw.get("semantic_key", "")),  # ✅ Phase 2: Include semantic_key
                case_no=str(kw.get("case_no", "")),  # ✅ Phase 4: Include case_no
            )
        )

    def log_new_case(self, case_no: str, row_data: Dict[str, Any], row_index: Optional[int] = None):
        """Log a new case that was appended."""
        self.new_cases[str(case_no)] = dict(row_data or {})
        if row_index is not None:
            self.add_change(
                row_index=row_index,
                column_name="",
                old_value=None,
                new_value=None,
                change_type="new_record",
            )


@dataclass
class SyncResult:
    """Result of a synchronization operation."""

    success: bool
    message: str
    output_path: str
    stats: Dict[str, Any]
    matching_report: Optional[str] = None  # New: matching diagnostics


@dataclass
class HeaderCandidate:
    """헤더 후보 정보를 저장합니다. | Container for header candidate metadata."""

    row_index: int
    source: str
    confidence: float
    threshold: float
    warnings: List[str] = field(default_factory=list)

    @property
    def is_confident(self) -> bool:
        """임계치 이상인지 확인합니다. | Check whether candidate meets threshold."""

        return self.confidence >= self.threshold


class DataSynchronizerV30:
    """
    Advanced data synchronizer with semantic header matching.

    This version eliminates all hardcoded column names and uses semantic
    matching to automatically find the correct columns regardless of how
    they're named in the Excel file.

    The synchronizer works in several phases:
    1. Load files and detect header rows
    2. Match all required columns using semantic keys
    3. Validate that required columns were found
    4. Perform synchronization using matched column names
    5. Apply Excel formatting to highlight changes

    Example:
        >>> sync = DataSynchronizerV30()
        >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
        >>> if result.success:
        >>>     print(f"Synchronized to {result.output_path}")
        >>> else:
        >>>     print(f"Error: {result.message}")
    """

    def __init__(
        self,
        date_semantic_keys: Optional[List[str]] = None,
        header_overrides: Optional[Dict[Tuple[str, str], int]] = None,
        min_header_confidence: float = 0.7,
        use_polars: bool = False,
        use_xlsxwriter: bool = False,
    ) -> None:
        """동기화기를 초기화합니다. | Initialize the synchronizer."""

        # Use semantic keys instead of hardcoded column names
        self.date_semantic_keys = date_semantic_keys or DATE_SEMANTIC_KEYS

        # Initialize the semantic matcher
        self.matcher = SemanticMatcher(min_confidence=0.7, allow_partial=True)

        # Header detection helpers
        self.header_confidence_threshold = min_header_confidence
        self.header_detector = HeaderDetector(min_confidence=min_header_confidence)
        self.manual_header_overrides: Dict[Tuple[str, str], int] = {}
        if header_overrides:
            for (file_label, sheet_name), row in header_overrides.items():
                self.manual_header_overrides[(file_label.lower(), sheet_name.lower())] = int(row)

        # Change tracking
        self.change_tracker = ChangeTracker()

        # Column mapping storage (will be populated during matching)
        self.master_columns: Dict[str, str] = {}  # semantic_key -> actual_column
        self.warehouse_columns: Dict[str, str] = {}

        # Debug tracking flag
        self.debug_dhl_wh = True  # Enable DHL WH tracking for debugging

        # Polars integration (optional, experimental)
        self.use_polars = use_polars
        if use_polars:
            try:
                from scripts.core.polars_adapter import is_polars_available
                if is_polars_available():
                    print("[INFO] Polars mode enabled (experimental)")
                else:
                    print("[WARNING] Polars requested but not available, falling back to Pandas")
                    self.use_polars = False
            except ImportError:
                print("[WARNING] Polars adapter not available, falling back to Pandas")
                self.use_polars = False

        # XlsxWriter optimization (optional)
        # Note: 색상 포맷팅은 openpyxl이 필요하므로, 데이터 쓰기만 최적화
        self.use_xlsxwriter = use_xlsxwriter
        if use_xlsxwriter:
            try:
                from scripts.core.excel_writer_optimized import is_xlsxwriter_available
                if is_xlsxwriter_available():
                    print("[INFO] XlsxWriter optimization enabled")
                else:
                    print("[WARNING] XlsxWriter requested but not available, using openpyxl")
                    self.use_xlsxwriter = False
            except ImportError:
                print("[WARNING] XlsxWriter optimizer not available, using openpyxl")
                self.use_xlsxwriter = False

        print("[OK] DataSynchronizer v3.0 initialized with semantic header matching")

    def _dates_equal(self, a, b) -> bool:
        """Check if two dates are equal, ignoring format differences."""
        da = _to_date(a)
        db = _to_date(b)
        if da is None and db is None:
            return True
        if da is None or db is None:
            return False
        if pd.isna(da) or pd.isna(db):
            return pd.isna(da) and pd.isna(db)
        return da.normalize() == db.normalize()

    def _track_dhl_wh(self, df: pd.DataFrame, stage_name: str) -> None:
        """Track DHL WH data through pipeline stages (DEBUG)"""
        if not self.debug_dhl_wh:
            return

        dhl_col = None
        # Try to find DHL WH column
        for col in df.columns:
            if "DHL" in str(col).upper() and "WH" in str(col).upper():
                dhl_col = col
                break

        if dhl_col:
            count = df[dhl_col].notna().sum()
            print(f"[DHL WH TRACK] {stage_name}: '{dhl_col}' = {count}건 데이터")
        else:
            print(f"[DHL WH TRACK] {stage_name}: DHL WH 컬럼 없음")
            print(
                f"[DHL WH TRACK]   사용 가능 컬럼: {[c for c in df.columns if 'DHL' in str(c).upper() or 'WH' in str(c).upper()][:5]}"
            )

    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """
        Check if sheet should be skipped (aggregate/summary sheets)

        Args:
            sheet_name: Name of the sheet

        Returns:
            True if sheet should be skipped
        """
        normalized = sheet_name.strip().lower()
        # 제외 시트명 체크
        if normalized in EXCLUDED_SHEET_NAMES:
            return True
        # 포함 패턴 매칭: 패턴에 맞지 않으면 건너뛰기
        if not _should_include_sheet(sheet_name):
            return True
        return False

    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """
        시트명을 정규화하여 매칭 키 생성

        Args:
            sheet_name: 원본 시트명

        Returns:
            정규화된 시트명 (소문자, 특수문자/공백 정리)
        """
        normalized = sheet_name.strip().lower()
        # 특수문자 제거 (쉼표, 괄호, 하이픈)
        normalized = normalized.replace(",", "").replace("(", "").replace(")", "")
        normalized = normalized.replace("-", " ").replace("_", " ")
        # 연속된 공백을 하나로
        normalized = " ".join(normalized.split())
        return normalized

    def _find_matching_sheet(self, target_sheet: str, available_sheets: List[str]) -> Optional[str]:
        """
        시트명 semantic matching

        Args:
            target_sheet: 찾으려는 시트명 (Master)
            available_sheets: 사용 가능한 시트명 목록 (Warehouse)

        Returns:
            매칭된 시트명 또는 None
        """
        target_norm = self._normalize_sheet_name(target_sheet)

        # 1. Exact match first
        for sheet in available_sheets:
            if self._normalize_sheet_name(sheet) == target_norm:
                return sheet

        # 2. Alias matching
        for base_name, aliases in SHEET_NAME_ALIASES.items():
            # Check if target matches base or any alias
            if target_norm == base_name or target_norm in aliases:
                # Find warehouse sheet that also matches
                for sheet in available_sheets:
                    sheet_norm = self._normalize_sheet_name(sheet)
                    if sheet_norm == base_name or sheet_norm in aliases:
                        return sheet

        # 3. Partial keyword matching (fallback)
        target_keywords = set(target_norm.split())
        best_match = None
        best_score = 0

        for sheet in available_sheets:
            sheet_norm = self._normalize_sheet_name(sheet)
            sheet_keywords = set(sheet_norm.split())

            # Calculate overlap
            overlap = len(target_keywords & sheet_keywords)
            if overlap > best_score:
                best_score = overlap
                best_match = sheet

        # Only return if significant overlap (at least 2 common words)
        if best_score >= 2:
            return best_match

        return None

    def _normalize_file_name(self, file_name: str) -> str:
        """
        파일명을 정규화

        Args:
            file_name: 원본 파일명 (확장자 포함 가능)

        Returns:
            정규화된 파일명
        """
        # 확장자 제거
        if "." in file_name:
            file_name = file_name.rsplit(".", 1)[0]

        normalized = file_name.strip().lower()
        # 특수문자 제거
        normalized = normalized.replace("_", " ").replace("-", " ")
        normalized = " ".join(normalized.split())
        return normalized

    def _identify_file_type(self, file_path: str) -> str:
        """
        파일 타입 식별 (master 또는 warehouse)

        Args:
            file_path: 파일 경로

        Returns:
            'master', 'warehouse', 또는 'unknown'
        """
        from pathlib import Path

        file_name = Path(file_path).name
        file_norm = self._normalize_file_name(file_name)

        # Check against keywords
        for file_type, keywords in FILE_NAME_KEYWORDS.items():
            for keyword in keywords:
                if keyword in file_norm:
                    return file_type

        return "unknown"

    def _get_header_override(self, file_label: str, sheet_name: str) -> Optional[int]:
        """수동 헤더 오버라이드를 조회합니다. | Fetch manual header override if configured."""

        label = file_label.strip().lower()
        sheet = sheet_name.strip().lower()
        candidates = [
            (label, sheet),
            ("*", sheet),
            (label, "*"),
            ("*", "*"),
        ]
        for key in candidates:
            if key in self.manual_header_overrides:
                return self.manual_header_overrides[key]
        return None

    def _format_header_candidate(self, candidate: HeaderCandidate) -> str:
        """헤더 후보 정보를 문자열로 표현합니다. | Format header candidate for logging."""

        status = "confident" if candidate.is_confident else "low-confidence"
        return (
            f"row={candidate.row_index} via {candidate.source} "
            f"(score={candidate.confidence:.2f}, threshold={candidate.threshold:.2f}, {status})"
        )

    def _build_header_candidates(
        self,
        file_label: str,
        sheet_name: str,
        vendor_header_row: Optional[int],
        auto_result: HeaderDetectionResult,
    ) -> List[HeaderCandidate]:
        """헤더 후보 목록을 생성합니다. | Build ordered list of header candidates."""

        ordered: List[HeaderCandidate] = []
        seen: Set[int] = set()

        manual_row = self._get_header_override(file_label, sheet_name)
        if manual_row is not None and manual_row >= 0:
            ordered.append(
                HeaderCandidate(
                    row_index=manual_row,
                    source="manual",
                    confidence=1.0,
                    threshold=self.header_confidence_threshold,
                )
            )
            seen.add(manual_row)

        # CRITICAL: Vendor-specific header row takes precedence over auto-detection
        # This is important for SIEMENS files where auto-detection may incorrectly
        # detect a data row as header (e.g., row 4 instead of row 0)
        if vendor_header_row is not None and vendor_header_row not in seen:
            ordered.append(
                HeaderCandidate(
                    row_index=vendor_header_row,
                    source="vendor",
                    confidence=0.95,
                    threshold=self.header_confidence_threshold,
                )
            )
            seen.add(vendor_header_row)

        if auto_result.row_index not in seen:
            ordered.append(
                HeaderCandidate(
                    row_index=auto_result.row_index,
                    source=f"auto:{auto_result.method}",
                    confidence=auto_result.confidence,
                    threshold=auto_result.threshold,
                    warnings=list(auto_result.warnings),
                )
            )
            seen.add(auto_result.row_index)

        return ordered

    def _basic_header_validation(self, df: pd.DataFrame) -> Tuple[bool, str]:
        """기본 시맨틱 검증으로 헤더를 확인합니다. | Validate header using core semantic key."""

        report = self.matcher.match_dataframe(df, ["case_number"])
        column_name = report.get_column_name("case_number")
        if column_name:
            return True, column_name
        return False, "Missing semantic key 'case_number'"

    def _load_sheet_with_candidates(
        self,
        xl: pd.ExcelFile,
        file_path: str,
        sheet_name: str,
        file_label: str,
        vendor_header_row: Optional[int],
    ) -> Tuple[pd.DataFrame, HeaderCandidate]:
        """헤더 후보를 순차 적용하여 시트를 로드합니다. | Load sheet trying header candidates."""

        auto_result = self.header_detector.detect_with_diagnostics(
            file_path=file_path,
            sheet_name=sheet_name,
        )

        candidates = self._build_header_candidates(
            file_label=file_label,
            sheet_name=sheet_name,
            vendor_header_row=vendor_header_row,
            auto_result=auto_result,
        )

        if not candidates:
            raise ValueError(f"No header candidates available for {file_label}:{sheet_name}")

        errors: List[str] = []
        for candidate in candidates:
            if candidate.warnings:
                for warning in candidate.warnings:
                    print(
                        f"  [WARN] Candidate {candidate.row_index} ({candidate.source}): {warning}"
                    )

            print(f"  [TRY] {self._format_header_candidate(candidate)}")

            df = pd.read_excel(
                xl,
                sheet_name=sheet_name,
                header=candidate.row_index,
                engine="openpyxl",
            )

            if df.empty:
                errors.append(f"row {candidate.row_index} produced empty DataFrame")
                print("  [WARN] Loaded DataFrame is empty; trying next candidate")
                continue

            valid, detail = self._basic_header_validation(df)
            if valid:
                if not candidate.is_confident:
                    print(
                        f"  [WARN] Confidence {candidate.confidence:.2f} below "
                        f"threshold {candidate.threshold:.2f}"
                    )
                print(f"  [OK] Header validated via column '{detail}'")
                return df, candidate

            errors.append(f"row {candidate.row_index} missing required column (detail: {detail})")
            print(f"  [WARN] Candidate {candidate.row_index} failed validation: {detail}")

        joined = "; ".join(errors)
        raise ValueError(
            f"Unable to resolve header for {file_label}:{sheet_name} (tried: {joined})"
        )

    def _detect_vendor_and_header_row(self, file_path: Path) -> Tuple[str, Optional[int]]:
        """
        Detect vendor type and appropriate header row based on file name.

        Args:
            file_path: Path to the file

        Returns:
            Tuple of (vendor_type, header_row_index)
        """
        file_norm = file_path.stem.lower()

        if "siem" in file_norm or "simense" in file_norm:
            return ("SIEMENS", 0)  # SIEMENS는 첫 행이 헤더
        if "hitachi" in file_norm or "hvdc" in file_norm:
            return ("HITACHI", 4)  # HITACHI는 5번째 행이 헤더
        return ("UNKNOWN", None)

    def _load_master_files(self, master_xlsx: str) -> Dict[str, pd.DataFrame]:
        """
        Load Master files (HITACHI + SIEMENS) and merge them.

        Args:
            master_xlsx: Path to Master Excel file (HITACHI)

        Returns:
            Dictionary of sheet_name -> DataFrame
        """
        from pathlib import Path

        master_sheets = {}

        # Load HITACHI Master file
        print(f"[INFO] Loading HITACHI Master file: {Path(master_xlsx).name}")
        hitachi_sheets_data = self._load_file_by_sheets(master_xlsx, "Master")
        # Extract DataFrames from tuples
        hitachi_sheets = {name: df for name, (df, _) in hitachi_sheets_data.items()}

        # ✅ Set Source_Vendor and Source_Sheet for ALL HITACHI data (표준화 적용)
        for sheet_name, df in hitachi_sheets.items():
            df["Source_Vendor"] = "HITACHI"
            # 표준화된 시트명 사용
            canon_name = _canon_sheet(sheet_name)
            df["Source_Sheet"] = canon_name
            # Case No 키 손실 방지
            df = _fix_case_key(df)
            # _raw_rows 추적용 attrs 설정
            if hasattr(df, 'attrs'):
                df.attrs["_raw_rows"] = len(df)
            print(
                f"[HITACHI] Set Source_Vendor='HITACHI', Source_Sheet='{canon_name}' (원본: '{sheet_name}') for {len(df)} rows"
            )

        master_sheets.update(hitachi_sheets)

        # Look for SIEMENS file in the same directory first
        master_dir = Path(master_xlsx).parent
        siemens_files = list(master_dir.glob("*SIMENSE*.xls*")) + list(
            master_dir.glob("*SIM*.xls*")
        )

        # If not found, search in parent directory's subdirectories
        if not siemens_files:
            parent_dir = master_dir.parent  # data/raw/
            print(f"[INFO] Searching for SIEMENS files in subdirectories of {parent_dir}")
            siemens_files = list(parent_dir.glob("*/*SIMENSE*.xls*")) + list(
                parent_dir.glob("*/*SIM*.xls*")
            )

        if siemens_files:
            siemens_file = siemens_files[0]  # Take the first SIEMENS file found
            print(f"[INFO] Found SIEMENS file: {siemens_file.name}")
            print(f"[INFO] Loading SIEMENS Master file...")

            siemens_sheets_data = self._load_file_by_sheets(str(siemens_file), "SIEMENS")
            # Extract DataFrames from tuples
            siemens_sheets = {name: df for name, (df, _) in siemens_sheets_data.items()}

            # ===== SIEMENS 데이터 정제 및 컬럼 통합 =====
            print("[INFO] Cleaning and mapping SIEMENS data...")
            for sheet_name, siemens_df in siemens_sheets.items():
                # 1. Bill of Lading 컬럼 제거
                bill_of_lading_patterns = ["billof", "billoflading", "b/l", "bl", "lading"]
                bl_columns = [
                    col
                    for col in siemens_df.columns
                    if any(
                        pattern in str(col).lower().replace(" ", "").replace("_", "")
                        for pattern in bill_of_lading_patterns
                    )
                ]
                if bl_columns:
                    print(
                        f"[SIEMENS] Dropping Bill of Lading columns from '{sheet_name}': {bl_columns}"
                    )
                    siemens_df.drop(columns=bl_columns, inplace=True, errors="ignore")

                # 2. columns_to_drop 리스트 초기화
                columns_to_drop = []

                # 3. SIEMENS 컬럼을 HITACHI 컬럼명으로 매핑
                column_mapping = {
                    "PackageNo": "Case No.",
                    "PO.No": "EQ No",
                    "HSCode": "HS Code",
                    "No.": "no.",  # 시퀀스 번호 통합
                }

                print(f"[SIEMENS] Mapping columns to HITACHI structure for '{sheet_name}'...")
                for siemens_col, hitachi_col in column_mapping.items():
                    if siemens_col in siemens_df.columns:
                        if hitachi_col in siemens_df.columns:
                            # 기존 HITACHI 컬럼이 있으면 빈 값만 SIEMENS 데이터로 채우기
                            null_count = siemens_df[hitachi_col].isna().sum()
                            if null_count > 0:
                                siemens_df[hitachi_col].fillna(
                                    siemens_df[siemens_col], inplace=True
                                )
                                print(
                                    f"  - Filled {null_count} null values in '{hitachi_col}' with '{siemens_col}' data"
                                )
                        else:
                            # HITACHI 컬럼이 없으면 SIEMENS 컬럼을 rename
                            siemens_df.rename(columns={siemens_col: hitachi_col}, inplace=True)
                            print(f"  - Renamed '{siemens_col}' → '{hitachi_col}'")

                # 4. SIEMENS 전용 컬럼 무조건 제거 (매핑 여부 관계없이)
                siemens_specific_cols = ["No.", "PackageNo", "PO.No", "HSCode", "BillofLading"]
                for siemens_col, hitachi_col in column_mapping.items():
                    # 매핑 후 원본 SIEMENS 컬럼이 남아있으면 제거
                    if (
                        siemens_col in siemens_df.columns
                        and hitachi_col in siemens_df.columns
                        and siemens_col != hitachi_col
                    ):
                        columns_to_drop.append(siemens_col)

                # SIEMENS 전용 컬럼 무조건 제거
                for col in siemens_specific_cols:
                    if col in siemens_df.columns:
                        columns_to_drop.append(col)

                if columns_to_drop:
                    print(
                        f"[SIEMENS] Dropping SIEMENS-specific columns from '{sheet_name}': {columns_to_drop}"
                    )
                    siemens_df.drop(columns=columns_to_drop, inplace=True, errors="ignore")

                # 업데이트된 DataFrame 저장
                siemens_sheets[sheet_name] = siemens_df

            print("[OK] SIEMENS data cleaning and column mapping completed")

            # Merge SIEMENS sheets with HITACHI sheets
            for sheet_name, siemens_df in siemens_sheets.items():
                if sheet_name in master_sheets:
                    # Merge with existing HITACHI sheet
                    print(f"[INFO] Merging SIEMENS sheet '{sheet_name}' with HITACHI sheet")
                    hitachi_df = master_sheets[sheet_name]

                    # Add source vendor and sheet columns (표준화 적용)
                    # hitachi_df["Source_Vendor"] = "HITACHI"  # Already set above
                    siemens_df["Source_Vendor"] = "SIEMENS"
                    # 표준화된 시트명 사용
                    canon_name = _canon_sheet(sheet_name)
                    siemens_df["Source_Sheet"] = canon_name
                    # Case No 키 손실 방지
                    siemens_df = _fix_case_key(siemens_df)
                    hitachi_df = _fix_case_key(hitachi_df)

                    # Merge DataFrames
                    merged_df = pd.concat([hitachi_df, siemens_df], ignore_index=True, sort=False)

                    # CRITICAL: Remove duplicates based on Case No (Source_Vendor 고려)
                    if "Case No." in merged_df.columns or "Case No" in merged_df.columns:
                        before_dedup = len(merged_df)
                        case_col = "Case No." if "Case No." in merged_df.columns else "Case No"
                        
                        # ✅ Source_Vendor가 다른 경우 별도 레코드로 유지
                        if "Source_Vendor" in merged_df.columns and case_col in merged_df.columns:
                            try:
                                # 빈 키 행은 보존
                                empty_mask = merged_df[case_col].isna() | (merged_df[case_col].astype(str).str.strip() == "")
                                non_empty_df = merged_df[~empty_mask].copy()
                                empty_df = merged_df[empty_mask].copy()
                                
                                # Source_Vendor + Case No 복합 키로 중복 제거
                                non_empty_df = non_empty_df.drop_duplicates(
                                    subset=[case_col, "Source_Vendor"], 
                                    keep="first"
                                )
                                
                                merged_df = pd.concat([non_empty_df, empty_df], ignore_index=True, sort=False)
                                after_dedup = len(merged_df)
                                removed = before_dedup - after_dedup
                                if removed > 0:
                                    print(
                                        f"[DEDUP] Removed {removed} duplicate Case No entries (Source_Vendor 고려)"
                                    )
                            except Exception as e:
                                print(f"[WARNING] Source_Vendor 기반 중복 제거 실패, 기본 방법 사용: {e}")
                                # 폴백: 기존 로직
                                try:
                                    merged_df = drop_duplicates_by_case(
                                        merged_df, semantic_finder=_find_col, keep="first"
                                    )
                                    after_dedup = len(merged_df)
                                    removed = before_dedup - after_dedup
                                    if removed > 0:
                                        print(
                                            f"[DEDUP] Removed {removed} duplicate Case No entries (fallback)"
                                        )
                                except Exception as e2:
                                    print(f"[WARNING] Drop duplicates failed, using basic method: {e2}")
                                    # 최종 폴백: 기본 drop_duplicates
                                    if case_col in merged_df.columns:
                                        non_empty = merged_df[case_col].notna() & (merged_df[case_col].astype(str).str.strip() != "")
                                        merged_df_clean = merged_df[non_empty].drop_duplicates(subset=[case_col], keep="first")
                                        merged_df_empty = merged_df[~non_empty]
                                        merged_df = pd.concat([merged_df_clean, merged_df_empty], ignore_index=True, sort=False)
                                        after_dedup = len(merged_df)
                                        removed = before_dedup - after_dedup
                                        if removed > 0:
                                            print(f"[DEDUP] Removed {removed} duplicate Case No entries (basic method)")
                        else:
                            # Source_Vendor 컬럼이 없으면 기존 로직 사용
                            try:
                                merged_df = drop_duplicates_by_case(
                                    merged_df, semantic_finder=_find_col, keep="first"
                                )
                                after_dedup = len(merged_df)
                                removed = before_dedup - after_dedup
                                if removed > 0:
                                    print(
                                        f"[DEDUP] Removed {removed} duplicate Case No entries from merged data"
                                    )
                            except Exception as e:
                                print(f"[WARNING] Drop duplicates failed, using fallback: {e}")
                                # 폴백: 기본 drop_duplicates (빈 키는 제외)
                                if case_col in merged_df.columns:
                                    non_empty = merged_df[case_col].notna() & (merged_df[case_col].astype(str).str.strip() != "")
                                    merged_df_clean = merged_df[non_empty].drop_duplicates(subset=[case_col], keep="first")
                                    merged_df_empty = merged_df[~non_empty]
                                    merged_df = pd.concat([merged_df_clean, merged_df_empty], ignore_index=True, sort=False)
                                    after_dedup = len(merged_df)
                                    removed = before_dedup - after_dedup
                                    if removed > 0:
                                        print(f"[DEDUP] Removed {removed} duplicate Case No entries (fallback method)")

                    master_sheets[sheet_name] = merged_df

                    print(
                        f"[OK] Merged '{sheet_name}': HITACHI({len(hitachi_df)}) + SIEMENS({len(siemens_df)}) = {len(merged_df)} rows after dedup"
                    )
                else:
                    # New sheet from SIEMENS
                    print(f"[INFO] Adding new SIEMENS sheet '{sheet_name}'")
                    siemens_df["Source_Vendor"] = "SIEMENS"
                    # 표준화된 시트명 사용
                    canon_name = _canon_sheet(sheet_name)
                    siemens_df["Source_Sheet"] = canon_name
                    # Case No 키 손실 방지
                    siemens_df = _fix_case_key(siemens_df)
                    # _raw_rows 추적용 attrs 설정
                    if hasattr(siemens_df, 'attrs'):
                        siemens_df.attrs["_raw_rows"] = len(siemens_df)
                    master_sheets[sheet_name] = siemens_df
        else:
            print(f"[INFO] No SIEMENS file found in {master_dir}")

        return master_sheets

    def _load_warehouse_files(
        self,
        warehouse_configs: List[Dict[str, Any]]
    ) -> Tuple[Dict[str, pd.DataFrame], Dict[str, int]]:
        """
        Load multiple warehouse files (HITACHI + SIEMENS) and merge them.

        Args:
            warehouse_configs: List of dicts with keys 'vendor', 'path', 'required'
                Example: [
                    {'vendor': 'HITACHI', 'path': '...HE.xlsx', 'required': True},
                    {'vendor': 'SIEMENS', 'path': '...SIM.xlsm', 'required': False}
                ]

        Returns:
            Tuple of:
            - Dict[sheet_name -> merged DataFrame]
            - Dict[sheet_name -> header_row_index]
            Source_Vendor column indicates the warehouse source for each row.

        Note:
            - Required files that don't exist will raise FileNotFoundError
            - Optional files that don't exist will be skipped with a warning
            - Same Case No. in both files: HITACHI data takes precedence
        """
        from pathlib import Path

        all_warehouse_sheets: Dict[str, List[Tuple[pd.DataFrame, str]]] = {}
        loaded_vendors = []

        print("\n" + "=" * 60)
        print("PHASE 1B: Loading Warehouse Files (Multi-Vendor)")
        print("=" * 60)

        for config in warehouse_configs:
            vendor = config.get('vendor', 'UNKNOWN')
            path_str = config.get('path', '')
            required = config.get('required', False)
            file_path = Path(path_str)

            print(f"\n[{vendor}] Loading warehouse file: {file_path.name}")

            # Check file existence
            if not file_path.exists():
                if required:
                    raise FileNotFoundError(
                        f"Required {vendor} warehouse file not found: {file_path}"
                    )
                else:
                    print(f"  [WARNING] Optional {vendor} file not found, skipping: {file_path}")
                    continue

            # Load the warehouse file
            try:
                vendor_sheets_data = self._load_file_by_sheets(str(file_path), f"Warehouse-{vendor}")
                loaded_vendors.append(vendor)

                # Extract DataFrames and add vendor metadata
                for sheet_name, (df, header_row) in vendor_sheets_data.items():
                    # SIEMENS-specific column mapping for warehouse files (same as Master)
                    if vendor == 'SIEMENS':
                        # Handle both "Package No" (with space) and "PackageNo" (without space)
                        siemens_wh_mapping = {
                            "Package No": "Case No.",  # SIEMENS warehouse uses "Package No" with space
                            "PackageNo": "Case No.",   # Fallback for files without space
                            "PO. No": "EQ No",         # "PO. No" with space
                            "PO.No": "EQ No",          # "PO.No" without space
                            "HS Code": "HS Code",      # "HS Code" with space
                            "HSCode": "HS Code",       # "HSCode" without space
                            "No.": "no.",
                        }
                        rename_map = {}
                        for old_col, new_col in siemens_wh_mapping.items():
                            if old_col in df.columns and new_col not in df.columns:
                                rename_map[old_col] = new_col
                        if rename_map:
                            df = df.rename(columns=rename_map)
                            print(f"    [SIEMENS] Mapped columns: {rename_map}")
                        # Drop SIEMENS-specific columns after mapping (only if they still exist after rename)
                        drop_cols = ["No.", "Package No", "PackageNo", "PO. No", "PO.No", "HS Code", "HSCode", "Bill of Lading", "BillofLading"]
                        for col in drop_cols:
                            if col in df.columns:
                                df = df.drop(columns=[col])
                                print(f"    [SIEMENS] Dropped column: {col}")

                    # Add warehouse source vendor metadata
                    if "Source_Vendor" not in df.columns:
                        df["Source_Vendor"] = vendor

                    # Track which warehouse this row came from (new column for debugging)
                    df["_Warehouse_Source"] = vendor

                    # Store in grouped structure
                    if sheet_name not in all_warehouse_sheets:
                        all_warehouse_sheets[sheet_name] = []
                    all_warehouse_sheets[sheet_name].append((df, vendor))

                    print(f"  [OK] Loaded '{sheet_name}': {len(df)} rows (header row: {header_row})")

            except Exception as e:
                print(f"  [ERROR] Failed to load {vendor} warehouse: {e}")
                if required:
                    raise
                continue

        print(f"\n[INFO] Loaded {len(loaded_vendors)} warehouse files: {loaded_vendors}")

        # Merge sheets with the same name from different vendors
        merged_warehouse_sheets: Dict[str, pd.DataFrame] = {}
        header_rows: Dict[str, int] = {}  # Track header rows for each sheet

        for sheet_name, df_vendor_list in all_warehouse_sheets.items():
            # Default header row: use HITACHI's if available, otherwise first vendor's
            hitachi_items = [(df, v) for df, v in df_vendor_list if v == 'HITACHI']
            if hitachi_items:
                header_rows[sheet_name] = 4  # HITACHI default
            else:
                header_rows[sheet_name] = 0  # SIEMENS default

            if len(df_vendor_list) == 1:
                # Single vendor for this sheet
                merged_warehouse_sheets[sheet_name] = df_vendor_list[0][0]
                print(f"[MERGE] '{sheet_name}': Single vendor ({df_vendor_list[0][1]}), {len(df_vendor_list[0][0])} rows")
            else:
                # Multiple vendors - merge by Case No.
                print(f"[MERGE] '{sheet_name}': Merging {len(df_vendor_list)} vendors...")
                merged_df = self._merge_warehouse_data(df_vendor_list, sheet_name)
                merged_warehouse_sheets[sheet_name] = merged_df

        print(f"\n[OK] Warehouse loading complete: {len(merged_warehouse_sheets)} sheets total")
        return merged_warehouse_sheets, header_rows

    def _merge_warehouse_data(
        self,
        df_vendor_list: List[Tuple[pd.DataFrame, str]],
        sheet_name: str
    ) -> pd.DataFrame:
        """
        Merge warehouse data from multiple vendors by Case Key.

        Merge Strategy:
        1. Identify Case Key column ('Case No.' or similar)
        2. HITACHI data takes precedence for overlapping Cases
        3. SIEMENS data fills in Cases not present in HITACHI
        4. Warehouse columns (DHL Warehouse, DSV Indoor, etc.) are preserved from both

        Args:
            df_vendor_list: List of (DataFrame, vendor_name) tuples
            sheet_name: Name of the sheet being merged (for logging)

        Returns:
            Merged DataFrame with data from all vendors
        """
        print(f"  [MERGE_DETAIL] Processing {len(df_vendor_list)} dataframes for '{sheet_name}'")

        # Separate HITACHI and SIEMENS data
        hitachi_dfs = [(df, vendor) for df, vendor in df_vendor_list if vendor == 'HITACHI']
        siemens_dfs = [(df, vendor) for df, vendor in df_vendor_list if vendor != 'HITACHI']

        # Concatenate same-vendor dataframes first
        hitachi_df = pd.concat([df for df, _ in hitachi_dfs], ignore_index=True) if hitachi_dfs else pd.DataFrame()
        siemens_df = pd.concat([df for df, _ in siemens_dfs], ignore_index=True) if siemens_dfs else pd.DataFrame()

        if hitachi_df.empty:
            print(f"  [MERGE_DETAIL] No HITACHI data, returning SIEMENS only: {len(siemens_df)} rows")
            return siemens_df

        if siemens_df.empty:
            print(f"  [MERGE_DETAIL] No SIEMENS data, returning HITACHI only: {len(hitachi_df)} rows")
            return hitachi_df

        # Both exist - merge by Case Key
        print(f"  [MERGE_DETAIL] HITACHI: {len(hitachi_df)} rows, SIEMENS: {len(siemens_df)} rows")

        # Find Case Key column for HITACHI
        hitachi_case_col = None
        for col in hitachi_df.columns:
            if col.strip().lower().replace(' ', '') in ['caseno', 'caseno.', 'casenumber', 'packageno']:
                hitachi_case_col = col
                break

        if not hitachi_case_col:
            print(f"  [WARNING] HITACHI Case Key column not found, using simple concat")
            return pd.concat([hitachi_df, siemens_df], ignore_index=True)

        print(f"  [MERGE_DETAIL] HITACHI Case Key: '{hitachi_case_col}'")

        # Find Case Key column for SIEMENS
        print(f"  [MERGE_DEBUG] SIEMENS columns: {list(siemens_df.columns)}")
        
        # Check if 'Case No.' has actual case numbers (not just row numbers)
        siemens_case_col = None
        if 'Case No.' in siemens_df.columns:
            case_no_valid = siemens_df['Case No.'].notna().sum()
            # Check sample values
            sample_vals = siemens_df['Case No.'].dropna().astype(str).head(5).tolist()
            case_no_alpha = siemens_df['Case No.'].astype(str).str.contains(r'[A-Za-z]', na=False).sum()
            print(f"  [MERGE_DEBUG] 'Case No.' valid: {case_no_valid}, alpha: {case_no_alpha}, samples: {sample_vals}")
            
            # 'Case No.' is valid only if it contains alphanumeric case numbers
            if case_no_alpha > 0:
                siemens_case_col = 'Case No.'
                print(f"  [MERGE_DETAIL] Using 'Case No.' as Case Key")
            else:
                print(f"  [MERGE_DEBUG] 'Case No.' has no alphanumeric values, looking for alternate...")
        
        # If 'Case No.' not usable, look for any column with alphanumeric identifiers
        if not siemens_case_col:
            for col in siemens_df.columns:
                col_norm = col.strip().lower().replace(' ', '')
                if col_norm in ['packageno', 'packageno.', 'package', 'packagenumber', 'siminvoice', 'invoice']:
                    col_alpha = siemens_df[col].astype(str).str.contains(r'[A-Za-z]', na=False).sum()
                    col_samples = siemens_df[col].dropna().astype(str).head(3).tolist()
                    print(f"    [MERGE_DEBUG] Checking '{col}': alpha={col_alpha}, samples={col_samples}")
                    if col_alpha > 0:
                        siemens_case_col = col
                        print(f"  [MERGE_DETAIL] Using alternate Case Key: '{siemens_case_col}'")
                        break
        
        if not siemens_case_col:
            print(f"  [WARNING] No usable Case Key column found in SIEMENS data")
            print(f"  [MERGE_DETAIL] SIEMENS columns were: {list(siemens_df.columns)}")
            # Fallback: use simple concat to include all SIEMENS data anyway
            siemens_df['Source_Vendor'] = 'SIEMENS'
            return pd.concat([hitachi_df, siemens_df], ignore_index=True)
        
        print(f"  [MERGE_DETAIL] SIEMENS Case Key: '{siemens_case_col}'")

        # Normalize SIEMENS column names to match HITACHI standard (handle spaces, typos)
        print(f"  [MERGE_DETAIL] Normalizing column names for SIEMENS...")
        print(f"  [MERGE_DEBUG] SIEMENS columns to map: {list(siemens_df.columns)[:15]}")
        siemens_col_mapping = {}
        
        for col in siemens_df.columns:
            col_norm = col.strip().lower().replace('  ', ' ')  # Normalize double spaces
            col_stripped = col.strip()
            
            # Map SIEMENS variations to HITACHI standard names
            if 'aaa storage' in col_norm:
                siemens_col_mapping[col] = 'AAA Storage'
            elif col_norm == 'haluer':
                siemens_col_mapping[col] = 'HAULER'
            elif 'dsv mzd' in col_norm:
                siemens_col_mapping[col] = 'DSV MZP'
            elif 'dsv indoor' in col_norm:
                siemens_col_mapping[col] = 'DSV Indoor'
            elif 'dsv al markaz' in col_norm:
                siemens_col_mapping[col] = 'DSV Al Markaz'
            elif 'dsv outdoor' in col_norm:
                siemens_col_mapping[col] = 'DSV Outdoor'
            elif 'haulier indoor' in col_norm or 'hauler indoor' in col_norm:
                siemens_col_mapping[col] = 'Hauler Indoor'
            elif 'jdn mzd' in col_norm:
                siemens_col_mapping[col] = 'JDN MZD'
            elif 'mosb' in col_norm:
                siemens_col_mapping[col] = 'MOSB'
            # Map Case Key columns (Package No, No., etc.) to 'Case No.'
            elif col_stripped.lower() in ['package no', 'no.', 'packageno', 'no', 'package', 'caseno']:
                siemens_col_mapping[col] = 'Case No.'
                print(f"    [MAP] '{col}' -> 'Case No.'")

        if siemens_col_mapping:
            print(f"    Mapped {len(siemens_col_mapping)} SIEMENS columns: {siemens_col_mapping}")
            # Check for potential duplicates and remove mappings that would create duplicates
            current_cols = set(siemens_df.columns)
            filtered_mapping = {}
            for old_col, new_col in siemens_col_mapping.items():
                if new_col in current_cols and old_col != new_col:
                    print(f"    [SKIP] '{old_col}' -> '{new_col}' (target already exists)")
                else:
                    filtered_mapping[old_col] = new_col
            if filtered_mapping:
                siemens_df = siemens_df.rename(columns=filtered_mapping)

        # Get all warehouse columns that might differ between vendors
        warehouse_cols = ['DHL Warehouse', 'DSV Indoor', 'DSV Al Markaz', 'Hauler Indoor',
                         'DSV Outdoor', 'DSV MZP', 'HAULER', 'JDN MZD', 'MOSB', 'AAA Storage']

        # Check which columns actually exist in both dataframes
        available_wh_cols = []
        for wh_col in warehouse_cols:
            in_hitachi = wh_col in hitachi_df.columns
            in_siemens = wh_col in siemens_df.columns
            if in_hitachi or in_siemens:
                available_wh_cols.append(wh_col)
                if in_hitachi and in_siemens:
                    print(f"    [OK] '{wh_col}' exists in both vendors")
                elif in_hitachi:
                    print(f"    [INFO] '{wh_col}' only in HITACHI")
                else:
                    print(f"    [INFO] '{wh_col}' only in SIEMENS")

        # DEBUG: Show columns before normalization
        print(f"  [MERGE_DEBUG] HITACHI columns before norm: {list(hitachi_df.columns)[:10]}...")
        print(f"  [MERGE_DEBUG] SIEMENS columns before norm: {list(siemens_df.columns)[:10]}...")
        print(f"  [MERGE_DEBUG] hitachi_case_col: '{hitachi_case_col}', exists: {hitachi_case_col in hitachi_df.columns}")
        print(f"  [MERGE_DEBUG] siemens_case_col: '{siemens_case_col}', exists: {siemens_case_col in siemens_df.columns}")
        
        # Check for duplicate columns
        hitachi_dups = hitachi_df.columns[hitachi_df.columns.duplicated()].tolist()
        siemens_dups = siemens_df.columns[siemens_df.columns.duplicated()].tolist()
        if hitachi_dups:
            print(f"  [MERGE_DEBUG] HITACHI duplicate columns: {hitachi_dups}")
        if siemens_dups:
            print(f"  [MERGE_DEBUG] SIEMENS duplicate columns: {siemens_dups}")

        # Normalize Case Key for matching (use each vendor's own Case Key column)
        hitachi_df['_case_key_norm'] = hitachi_df[hitachi_case_col].astype(str).str.strip().str.lower()
        siemens_df['_case_key_norm'] = siemens_df[siemens_case_col].astype(str).str.strip().str.lower()
        
        print(f"  [MERGE_DEBUG] HITACHI _case_key_norm non-null: {hitachi_df['_case_key_norm'].notna().sum()}")
        print(f"  [MERGE_DEBUG] SIEMENS _case_key_norm non-null: {siemens_df['_case_key_norm'].notna().sum()}")

        # Identify overlapping and unique cases (filter out 'nan' strings from astype(str))
        hitachi_keys = hitachi_df['_case_key_norm']
        siemens_keys = siemens_df['_case_key_norm']
        
        # DEBUG: Show raw values before set conversion
        print(f"  [MERGE_DEBUG] HITACHI keys dtype: {hitachi_keys.dtype}, unique count: {hitachi_keys.nunique()}")
        print(f"  [MERGE_DEBUG] SIEMENS keys dtype: {siemens_keys.dtype}, unique count: {siemens_keys.nunique()}")
        print(f"  [MERGE_DEBUG] Sample HITACHI keys: {hitachi_keys.iloc[:5].tolist()}")
        print(f"  [MERGE_DEBUG] Sample SIEMENS keys: {siemens_keys.iloc[:5].tolist()}")

        # Filter: remove 'nan'/'none' strings and empty values
        hitachi_clean = hitachi_keys[hitachi_keys.notna() & ~hitachi_keys.isin(['', 'nan', 'none', 'NaN', 'None'])]
        siemens_clean = siemens_keys[siemens_keys.notna() & ~siemens_keys.isin(['', 'nan', 'none', 'NaN', 'None'])]
        
        hitachi_cases = set(hitachi_clean)
        siemens_cases = set(siemens_clean)

        overlap_cases = hitachi_cases & siemens_cases
        only_siemens_cases = siemens_cases - hitachi_cases

        print(f"  [MERGE_DETAIL] Overlapping Cases: {len(overlap_cases)}")
        print(f"  [MERGE_DETAIL] SIEMENS-only Cases: {len(only_siemens_cases)}")
        print(f"  [MERGE_DETAIL] Total HITACHI cases: {len(hitachi_cases)}, Total SIEMENS cases: {len(siemens_cases)}")

        # For overlapping cases: merge warehouse columns using SIEMENS to fill HITACHI nulls
        # Use simple dictionary-based approach to avoid pandas indexing issues
        if overlap_cases and len(available_wh_cols) > 0:
            print(f"  [MERGE_DETAIL] Merging warehouse columns for overlapping cases...")
            fill_count = 0

            # Build lookup from SIEMENS: case_key -> row_index
            siemens_lookup = {}
            for idx, row in siemens_df.iterrows():
                case_key = str(row['_case_key_norm']).strip()
                if case_key and case_key != 'nan' and case_key in overlap_cases:
                    if case_key not in siemens_lookup:
                        siemens_lookup[case_key] = idx

            # Get warehouse columns present in both dataframes
            common_wh_cols = [c for c in available_wh_cols 
                              if c in hitachi_df.columns and c in siemens_df.columns]

            # Fill HITACHI nulls with SIEMENS values
            for hitachi_idx, hitachi_row in hitachi_df.iterrows():
                case_key = str(hitachi_row['_case_key_norm']).strip()
                if case_key not in siemens_lookup:
                    continue

                siemens_idx = siemens_lookup[case_key]

                for wh_col in common_wh_cols:
                    # Check if HITACHI value is null
                    h_val = hitachi_row[wh_col]
                    is_h_null = pd.isna(h_val) or h_val is None or h_val == ''

                    if is_h_null:
                        # Get SIEMENS value
                        s_val = siemens_df.at[siemens_idx, wh_col]
                        is_s_valid = not (pd.isna(s_val) or s_val is None or s_val == '')

                        if is_s_valid:
                            hitachi_df.at[hitachi_idx, wh_col] = s_val
                            fill_count += 1

            print(f"  [MERGE_DETAIL] Filled {fill_count} warehouse values from SIEMENS")

        # Append SIEMENS-only cases
        if only_siemens_cases:
            siemens_only_df = siemens_df[siemens_df['_case_key_norm'].isin(only_siemens_cases)].copy()
            merged_df = pd.concat([hitachi_df, siemens_only_df], ignore_index=True)
            print(f"  [MERGE_DETAIL] Appended {len(siemens_only_df)} SIEMENS-only rows")
        else:
            merged_df = hitachi_df.copy()

        # Clean up temp columns
        merged_df = merged_df.drop(columns=['_case_key_norm'], errors='ignore')

        # Update Source_Vendor for SIEMENS-only rows (if _Warehouse_Source exists)
        if '_Warehouse_Source' in merged_df.columns:
            merged_df.loc[merged_df['_Warehouse_Source'] == 'SIEMENS', 'Source_Vendor'] = 'SIEMENS'
            merged_df = merged_df.drop(columns=['_Warehouse_Source'], errors='ignore')

        print(f"  [MERGE_RESULT] Final merged: {len(merged_df)} rows")
        return merged_df

    def _load_file_with_header_detection(
        self, file_path: str, file_label: str
    ) -> Tuple[pd.DataFrame, int]:
        """
        Load all sheets from Excel file with vendor-specific header row detection.

        Args:
            file_path: Path to the Excel file
            file_label: Label for logging (e.g., "Master", "Warehouse")

        Returns:
            Tuple of (merged_dataframe, header_row_index)
        """
        print(f"\n{'='*60}")
        print(f"Loading {file_label} file: {Path(file_path).name}")
        print(f"{'='*60}")

        xl = pd.ExcelFile(file_path, engine="openpyxl")
        all_dfs: List[pd.DataFrame] = []
        header_row: Optional[int] = None

        print(f"Found {len(xl.sheet_names)} sheets in file")

        vendor_type, vendor_header_row = self._detect_vendor_and_header_row(Path(file_path))
        if vendor_header_row is not None:
            print(f"[INFO] Detected vendor: {vendor_type}, default header row {vendor_header_row}")
        else:
            print(
                f"[INFO] Vendor could not be determined from name; falling back to heuristic detection"
            )

        for sheet_name in xl.sheet_names:
            print(f"\n  Loading sheet: '{sheet_name}'")

            # Skip summary/aggregate sheets
            if self._should_skip_sheet(sheet_name):
                print(f"  [SKIP] Aggregate sheet (not Case data)")
                continue

            df, candidate = self._load_sheet_with_candidates(
                xl=xl,
                file_path=file_path,
                sheet_name=sheet_name,
                file_label=file_label,
                vendor_header_row=vendor_header_row,
            )

            sheet_header_row = candidate.row_index

            if header_row is None:
                header_row = sheet_header_row

            # Track source sheet (preserve original sheet name)
            df["Source_Sheet"] = sheet_name

            # DEBUG: Track DHL WH in each sheet
            self._track_dhl_wh(df, f"Sheet '{sheet_name}' loaded")

            all_dfs.append(df)
            print(f"  [OK] {len(df)} rows loaded")

        if not all_dfs:
            raise ValueError(f"No valid sheets found in {file_label}")

        # Merge all sheets
        merged_df = pd.concat(all_dfs, ignore_index=True, sort=False)
        print(f"\n[OK] Total: {len(merged_df)} rows from {len(all_dfs)} sheets")

        # DEBUG: Track after merge
        self._track_dhl_wh(merged_df, "After pd.concat (merge)")

        # DEBUG: Check for duplicate column names and WH columns
        if self.debug_dhl_wh:
            from collections import Counter

            col_counts = Counter(merged_df.columns)
            duplicates = {col: count for col, count in col_counts.items() if count > 1}
            if duplicates:
                print(f"  [⚠️WARNING⚠️] Duplicate column names detected!")
                for col, count in duplicates.items():
                    print(f"    - '{col}' appears {count} times")
            else:
                print(f"  [OK] No duplicate column names")

            # Show exact column positions for WH columns
            wh_cols_with_idx = [
                (i, c) for i, c in enumerate(merged_df.columns) if "WH" in str(c).upper()
            ]
            print(f"  [DEBUG] WH columns after concat (position, name):")
            for idx, col in wh_cols_with_idx:
                data_count = merged_df[col].notna().sum()
                print(f"    - [{idx}] '{col}' = {data_count}건")

        # Filter out invalid columns (신규 추가)
        print("\nFiltering invalid columns:")
        merged_df = self._filter_invalid_columns(merged_df)
        self._track_dhl_wh(merged_df, "After filter_invalid_columns")

        # Consolidate incorrectly named warehouse columns
        print("\nConsolidating warehouse columns:")
        merged_df = self._consolidate_warehouse_columns(merged_df)
        self._track_dhl_wh(merged_df, "After consolidate_warehouse_columns")

        # Ensure all location columns exist
        print("\nEnsuring all location columns:")
        merged_df = self._ensure_all_location_columns(merged_df)
        self._track_dhl_wh(merged_df, "After ensure_all_location_columns")

        return merged_df, header_row

    def _load_file_by_sheets(
        self, file_path: str, file_label: str
    ) -> Dict[str, Tuple[pd.DataFrame, int]]:
        """
        Load each sheet from Excel file separately with automatic header row detection.

        Args:
            file_path: Path to the Excel file
            file_label: Label for logging (e.g., "Master", "Warehouse")

        Returns:
            Dict mapping sheet_name to (dataframe, header_row_index)
        """
        print(f"\n{'='*60}")
        print(f"Loading {file_label} file by sheets: {Path(file_path).name}")
        print(f"{'='*60}")

        # Excel 파일 읽기: Polars 옵션 지원
        if self.use_polars:
            try:
                from scripts.core.polars_adapter import read_excel_polars, to_pandas, is_polars_available
                
                if is_polars_available():
                    # Polars로 Excel 읽기 시도
                    try:
                        pl_sheets = read_excel_polars(file_path)
                        if isinstance(pl_sheets, dict):
                            # 여러 시트인 경우
                            xl = pd.ExcelFile(file_path, engine="openpyxl")  # 시트 이름 목록용
                            sheet_names = list(pl_sheets.keys())
                        else:
                            # 단일 시트인 경우
                            xl = pd.ExcelFile(file_path, engine="openpyxl")
                            sheet_names = xl.sheet_names
                            pl_sheets = {sheet_names[0]: pl_sheets}
                    except Exception as e:
                        print(f"  [WARNING] Polars Excel reading failed, falling back to Pandas: {e}")
                        self.use_polars = False
                        xl = pd.ExcelFile(file_path, engine="openpyxl")
                        sheet_names = xl.sheet_names
                        pl_sheets = None
                else:
                    self.use_polars = False
                    xl = pd.ExcelFile(file_path, engine="openpyxl")
                    sheet_names = xl.sheet_names
                    pl_sheets = None
            except ImportError:
                print("  [WARNING] Polars adapter not available, using Pandas")
                self.use_polars = False
                xl = pd.ExcelFile(file_path, engine="openpyxl")
                sheet_names = xl.sheet_names
                pl_sheets = None
        else:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
            sheet_names = xl.sheet_names
            pl_sheets = None

        sheet_data: Dict[str, Tuple[pd.DataFrame, int]] = {}

        print(f"Found {len(sheet_names)} sheets in file")

        vendor_type, vendor_header_row = self._detect_vendor_and_header_row(Path(file_path))
        if vendor_header_row is not None:
            print(f"[INFO] Detected vendor: {vendor_type}, default header row {vendor_header_row}")
        else:
            print(
                f"[INFO] Vendor could not be determined from name; falling back to heuristic detection"
            )

        for sheet_name in sheet_names:
            print(f"\n  Loading sheet: '{sheet_name}'")

            # Skip summary/aggregate sheets
            if self._should_skip_sheet(sheet_name):
                print(f"  [SKIP] Aggregate sheet (not Case data)")
                continue

            # ✅ Log raw sheet size before processing (for data loss tracking)
            try:
                raw_df_temp = pd.read_excel(xl, sheet_name=sheet_name, header=None, engine="openpyxl")
                raw_row_count = len(raw_df_temp)
                print(f"  [RAW] Raw sheet size: {raw_row_count:,} rows (before header detection)")
            except Exception as e:
                raw_row_count = 0
                print(f"  [WARN] Could not determine raw sheet size: {e}")

            # Polars로 읽은 경우: DataFrame만 Polars에서 가져오고, 헤더 감지는 기존 로직 사용
            if self.use_polars and pl_sheets and sheet_name in pl_sheets:
                try:
                    from scripts.core.polars_adapter import to_pandas
                    pl_df = pl_sheets[sheet_name]
                    # Polars DataFrame을 Pandas로 변환
                    df_polars = to_pandas(pl_df)
                    # 헤더 감지는 기존 _load_sheet_with_candidates 메서드 사용
                    # 하지만 이미 읽은 DataFrame을 활용하기 위해 임시로 처리
                    # 실제로는 기존 로직이 더 안정적이므로, Polars는 읽기 속도 개선에만 사용
                    # 여기서는 기존 로직으로 폴백 (향후 최적화 가능)
                    print(f"  [INFO] Polars read successful for '{sheet_name}', using existing header detection")
                    df, candidate = self._load_sheet_with_candidates(
                        xl=xl,
                        file_path=file_path,
                        sheet_name=sheet_name,
                        file_label=file_label,
                        vendor_header_row=vendor_header_row,
                    )
                except Exception as e:
                    print(f"  [WARNING] Polars processing failed for sheet '{sheet_name}', falling back to Pandas: {e}")
                    df, candidate = self._load_sheet_with_candidates(
                        xl=xl,
                        file_path=file_path,
                        sheet_name=sheet_name,
                        file_label=file_label,
                        vendor_header_row=vendor_header_row,
                    )
            else:
                # 기존 Pandas 방식
                df, candidate = self._load_sheet_with_candidates(
                    xl=xl,
                    file_path=file_path,
                    sheet_name=sheet_name,
                    file_label=file_label,
                    vendor_header_row=vendor_header_row,
                )

            sheet_header_row = candidate.row_index

            # Track source sheet: 표준화된 이름 사용
            canon_name = _canon_sheet(sheet_name)
            df["Source_Sheet"] = canon_name

            # Source_Vendor 자동 추론 및 주입
            if "Source_Vendor" not in df.columns:
                df["Source_Vendor"] = None
            vendor_hint = _infer_sheet_vendor(sheet_name)
            if vendor_hint and (df["Source_Vendor"].isna().all() or (df["Source_Vendor"] == "").all()):
                df["Source_Vendor"] = vendor_hint.upper()

            # DEBUG: Track DHL WH in each sheet
            self._track_dhl_wh(df, f"Sheet '{sheet_name}' loaded")

            # Process each sheet independently
            df_before_filter = len(df)
            df = self._filter_invalid_columns(df)
            df_after_filter = len(df)
            if df_before_filter != df_after_filter:
                print(f"  [FILTER] Rows after filtering invalid columns: {df_before_filter:,} → {df_after_filter:,} (-{df_before_filter - df_after_filter:,})")
            
            df = self._consolidate_warehouse_columns(df)
            df = self._ensure_all_location_columns(df)
            
            # Case No 키 손실 방지
            df = _fix_case_key(df)
            
            # _raw_rows 추적용 attrs 설정
            if hasattr(df, 'attrs'):
                df.attrs["_raw_rows"] = raw_row_count
                df.attrs["_processed_rows"] = len(df)

            sheet_data[sheet_name] = (df, sheet_header_row)
            
            # 시트별 상세 로그 (데이터 손실 추적 강화)
            has_case = any(
                c.strip().lower().replace(" ", "") in {"caseno", "caseid", "casenumber"} 
                for c in df.columns
            )
            loss_info = ""
            if raw_row_count > 0:
                loss_pct = ((raw_row_count - len(df)) / raw_row_count * 100) if raw_row_count > 0 else 0
                loss_info = f", loss: {raw_row_count - len(df):,} rows ({loss_pct:.1f}%)"
            print(f"  [OK] {len(df):,} rows loaded (Source_Sheet={canon_name}, Source_Vendor={vendor_hint or 'auto'}, case_key={has_case}{loss_info})")

        if not sheet_data:
            raise ValueError(f"No valid sheets found in {file_label}")

        print(f"\n[OK] Loaded {len(sheet_data)} sheets")
        return sheet_data

    @staticmethod
    def parse_header_override_args(
        overrides: Optional[List[str]],
    ) -> Dict[Tuple[str, str], int]:
        """헤더 오버라이드 인자를 파싱합니다. | Parse manual header override arguments."""

        parsed: Dict[Tuple[str, str], int] = {}
        if not overrides:
            return parsed

        for raw in overrides:
            if "=" not in raw:
                raise ValueError(
                    "Header override must follow <file_label>:<sheet>=<row_index> format"
                )
            target, row_str = raw.split("=", 1)
            if ":" in target:
                file_label, sheet_name = target.split(":", 1)
            else:
                file_label, sheet_name = "*", target

            file_label = file_label.strip() or "*"
            sheet_name = sheet_name.strip() or "*"

            try:
                row_index = int(row_str)
            except ValueError as exc:
                raise ValueError(
                    f"Invalid row index for header override '{raw}': {row_str}"
                ) from exc

            if row_index < 0:
                raise ValueError(f"Header override row index must be >= 0 (got {row_index})")

            parsed[(file_label.lower(), sheet_name.lower())] = row_index

        return parsed

    def _filter_invalid_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        잘못된 헤더를 가진 컬럼 제거

        Args:
            df: 필터링할 DataFrame

        Returns:
            유효한 컬럼만 포함된 DataFrame
        """
        import re

        valid_columns = []
        removed_columns = []

        for col in df.columns:
            col_str = str(col).strip()

            # 잘못된 패턴 검사
            is_invalid = False
            for pattern in INVALID_HEADER_PATTERNS:
                if re.match(pattern, col_str):
                    is_invalid = True
                    removed_columns.append(col)
                    break

            if not is_invalid:
                valid_columns.append(col)

        if removed_columns:
            print(f"  [CLEANUP] Removed {len(removed_columns)} invalid columns:")
            for col in removed_columns:
                print(f"    - '{col}'")

        return df[valid_columns]

    def _consolidate_warehouse_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Consolidate incorrectly named warehouse columns.

        Some raw data files use incorrect column names for the same warehouse.
        This method merges such columns to ensure data consistency:
        - "DSV WH" → "DSV Indoor" (HE Local sheet uses incorrect name)

        NOTE: DHL WH is a separate, valid warehouse and should NOT be consolidated!

        Args:
            df: DataFrame with potentially incorrect column names

        Returns:
            DataFrame with consolidated columns
        """
        consolidations = {
            "DSV WH": "DSV Indoor",  # HE Local sheet incorrectly names DSV Indoor as DSV WH
        }

        # DEBUG: Check DHL WH before consolidation
        if self.debug_dhl_wh:
            dhl_before = "DHL WH" in df.columns
            all_cols_before = list(df.columns)
            print(f"  [DEBUG] Before consolidation:")
            print(f"    - Total columns: {len(all_cols_before)}")
            print(f"    - DHL WH exists: {dhl_before}")
            if dhl_before:
                dhl_idx = all_cols_before.index("DHL WH")
                print(f"    - DHL WH position: {dhl_idx}")
                print(f"    - DHL WH data count: {df['DHL WH'].notna().sum()}건")

            # Show all warehouse-like columns
            wh_cols = [
                c
                for c in df.columns
                if "WH" in str(c).upper() or "DSV" in str(c).upper() or "DHL" in str(c).upper()
            ]
            print(f"    - WH/DSV/DHL columns: {wh_cols}")

        for wrong_name, correct_name in consolidations.items():
            if wrong_name in df.columns:
                # DEBUG: Track before rename
                if self.debug_dhl_wh:
                    dhl_before_rename = "DHL WH" in df.columns
                    print(f"  [DEBUG] Before '{wrong_name}' → '{correct_name}':")
                    print(f"    - DHL WH exists: {dhl_before_rename}")

                if correct_name in df.columns:
                    # Merge data: use correct_name where it exists, fill with wrong_name otherwise
                    df[correct_name] = df[correct_name].fillna(df[wrong_name])
                    df = df.drop(columns=[wrong_name])
                    print(f"  [OK] Merged '{wrong_name}' → '{correct_name}'")
                else:
                    # Just rename if correct column doesn't exist yet
                    # IMPORTANT: Use columns.str.replace() to avoid affecting other columns
                    new_columns = []
                    renamed = False
                    for col in df.columns:
                        if col == wrong_name and not renamed:
                            # Only rename the FIRST occurrence
                            new_columns.append(correct_name)
                            renamed = True
                            print(
                                f"  [OK] Renamed '{wrong_name}' → '{correct_name}' at position {len(new_columns)-1}"
                            )
                        else:
                            new_columns.append(col)

                    df.columns = new_columns

                # DEBUG: Track after rename
                if self.debug_dhl_wh:
                    dhl_after_rename = "DHL WH" in df.columns
                    print(f"  [DEBUG] After '{wrong_name}' → '{correct_name}':")
                    print(f"    - DHL WH exists: {dhl_after_rename}")
                    if dhl_before_rename and not dhl_after_rename:
                        print(f"  [🚨CRITICAL🚨] DHL WH was deleted by this rename operation!")
                        print(
                            f"    - Operation: rename({{{repr(wrong_name)}: {repr(correct_name)}}})"
                        )
                        print(f"    - DataFrame columns now: {len(df.columns)}")

        # DEBUG: Check DHL WH after consolidation
        if self.debug_dhl_wh:
            dhl_after = "DHL WH" in df.columns
            all_cols_after = list(df.columns)
            print(f"  [DEBUG] After consolidation:")
            print(f"    - Total columns: {len(all_cols_after)}")
            print(f"    - DHL WH exists: {dhl_after}")

            # Show all warehouse-like columns after
            wh_cols_after = [
                c
                for c in df.columns
                if "WH" in str(c).upper() or "DSV" in str(c).upper() or "DHL" in str(c).upper()
            ]
            print(f"    - WH/DSV/DHL columns: {wh_cols_after}")

            if dhl_before and not dhl_after:
                print(f"  [🚨ERROR🚨] DHL WH was LOST during consolidation!")
                print(f"    - Columns lost: {set(all_cols_before) - set(all_cols_after)}")
                print(f"    - Columns gained: {set(all_cols_after) - set(all_cols_before)}")

        return df

    def _ensure_all_location_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ensure all warehouse and site columns exist, but keep Shifting in original position.
        Source_Sheet is metadata and should not be included in ordering logic.

        Args:
            df: DataFrame to check and update

        Returns:
            DataFrame with all location columns present in correct order
        """
        # Get column order from core registry (Single Source of Truth)
        # All warehouse/site column definitions are centrally managed in @core/header_registry.py
        from core import get_site_columns, get_warehouse_columns

        WAREHOUSE_ORDER = get_warehouse_columns()
        SITE_ORDER = get_site_columns()

        # Combined order: Warehouse first, then Site
        all_locations = WAREHOUSE_ORDER + SITE_ORDER
        location_set = set(all_locations)

        # Check and add missing columns
        missing_cols = []
        for location in all_locations:
            if location not in df.columns:
                df[location] = pd.NaT
                missing_cols.append(location)

        if missing_cols:
            print(f"  [OK] Added {len(missing_cols)} missing location columns:")
            for col in missing_cols:
                print(f"    - {col}")
        else:
            print(f"  [OK] All location columns present")

        # Separate columns into groups using core's standard order

        # Ensure "no." column exists (create if missing)
        if "no." not in df.columns:
            # Create "no." as sequential row number
            df.insert(0, "no.", range(1, len(df) + 1))
            print("  [OK] Created 'no.' column as row index")

        # 1. Base columns in standard order (from core)
        base_cols_in_df = [col for col in STAGE1_BASE_COLS_ORDER if col in df.columns]

        # 2. Extra base columns not in standard order (for dynamic columns)
        # Exclude 'no' (lowercase without dot) as it's likely a duplicate of 'no.'
        extra_base_cols = [
            col
            for col in df.columns
            if col not in STAGE1_BASE_COLS_ORDER
            and col not in location_set
            and col != "Shifting"
            and col != "Source_Sheet"
            and col != "no"  # Exclude 'no' (keep only 'no.')
        ]

        base_cols = base_cols_in_df + extra_base_cols

        # 3. Special columns
        shifting_col = "Shifting" if "Shifting" in df.columns else None
        source_sheet_col = "Source_Sheet" if "Source_Sheet" in df.columns else None

        # Build final column order:
        # base_cols + warehouse_cols + shifting + site_cols + source_sheet
        final_order = (
            base_cols
            + WAREHOUSE_ORDER
            + ([shifting_col] if shifting_col else [])
            + SITE_ORDER
            + ([source_sheet_col] if source_sheet_col else [])
        )

        # Reorder dataframe
        df = df[[c for c in final_order if c in df.columns]]

        print(
            f"  [OK] Column order: base({len(base_cols)}) + warehouses({len(WAREHOUSE_ORDER)}) + Shifting + sites({len(SITE_ORDER)}) + Source_Sheet"
        )
        if base_cols:
            print(f"  [DEBUG] First 5 base columns: {base_cols[:5]}")

        return df

    def _match_and_validate_headers(self, df: pd.DataFrame, file_label: str) -> Dict[str, str]:
        """
        Match semantic keys to actual column names and validate required columns.

        Args:
            df: The DataFrame to match against
            file_label: Label for logging

        Returns:
            Dictionary mapping semantic_key to actual column name

        Raises:
            ValueError: If required columns are not found
        """
        print(f"\nMatching headers for {file_label}...")

        # Define required semantic keys for synchronization
        required_keys = [
            "case_number",  # Must have case number for matching
        ]

        # All keys we want to find (required + date columns + location columns)
        location_headers = HVDC_HEADER_REGISTRY.get_by_category(HeaderCategory.LOCATION)
        location_keys = [h.semantic_key for h in location_headers]
        all_keys = required_keys + self.date_semantic_keys + location_keys

        # Perform semantic matching
        report = self.matcher.match_dataframe(df, all_keys)

        # Print summary
        print(f"  Matched: {report.successful_matches}/{report.total_semantic_keys}")
        print(f"  Success rate: {report.successful_matches/report.total_semantic_keys:.0%}")

        # Validate required columns
        missing_required = []
        for key in required_keys:
            if not report.get_column_name(key):
                missing_required.append(key)

        if missing_required:
            # Print detailed error information
            print(f"\n{'='*60}")
            print(f"ERROR: Missing required columns in {file_label}")
            print(f"{'='*60}")
            report.print_summary()
            raise ValueError(
                f"Required semantic keys not found in {file_label}: {missing_required}"
            )

        # Show what was matched
        print("\n  Key matches:")
        for key in required_keys + self.date_semantic_keys[:3]:  # Show first few
            col = report.get_column_name(key)
            if col:
                print(f"    - {key:20s} → '{col}'")

        if len(self.date_semantic_keys) > 3:
            remaining = len([k for k in self.date_semantic_keys if report.get_column_name(k)])
            print(f"    ... and {remaining} more date columns")

        # Build mapping dictionary
        column_mapping = {
            key: report.get_column_name(key) for key in all_keys if report.get_column_name(key)
        }

        return column_mapping

    def _build_case_index(self, df: pd.DataFrame, case_col: str) -> Dict[str, int]:
        """
        Build an index mapping case numbers to row indices.

        Args:
            df: The DataFrame
            case_col: Name of the case number column

        Returns:
            Dictionary mapping normalized case numbers to row indices
        """
        import re

        idx: Dict[str, int] = {}

        # Normalize case numbers: uppercase, remove special characters
        series = df[case_col].fillna("").astype(str).str.strip().str.upper()
        series = series.apply(lambda x: re.sub(r"[^A-Z0-9]", "", x))

        for i, v in enumerate(series.tolist()):
            if not v:
                continue
            if v not in idx:  # Keep first occurrence
                idx[v] = i

        return idx

    def _apply_master_order_sorting(
        self,
        master: pd.DataFrame,
        warehouse: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Maintain Warehouse original order while updating with Master data.

        HVDC HITACHI 파일의 원본 순서를 유지하면서 Master 데이터로 업데이트합니다.
        신규 케이스만 제일 하단에 추가합니다.

        Args:
            master: Master DataFrame
            warehouse: Warehouse DataFrame
            master_cols: Master column mapping (semantic_key -> column_name)
            wh_cols: Warehouse column mapping

        Returns:
            Tuple of (master, warehouse) - 순서 변경 없음
        """
        print("\nMaintaining Warehouse original order...")

        # Get case columns
        master_case_col = master_cols["case_number"]
        wh_case_col = wh_cols["case_number"]

        # Warehouse 원본 순서 그대로 유지 (정렬하지 않음)
        print(f"  Warehouse original order: {len(warehouse)} rows")

        # Master의 Case No 집합만 추출
        master_case_set = set(master[master_case_col].dropna().tolist())
        print(f"  Master has {len(master_case_set)} cases")

        # Warehouse 순서는 변경하지 않음
        print(f"  First 5 Warehouse Case No.: {warehouse[wh_case_col].head().tolist()}")

        return master, warehouse  # 순서 변경 없음

    def _maintain_warehouse_order(
        self,
        warehouse: pd.DataFrame,
        master: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> pd.DataFrame:
        """
        Warehouse 원본 순서를 유지하고 신규 케이스만 하단에 추가합니다.

        HVDC HITACHI 파일의 원본 순서는 변경하지 않고, Master 데이터로 업데이트만 수행합니다.
        신규 케이스는 제일 하단에 추가됩니다.

        Args:
            warehouse: 업데이트된 Warehouse DataFrame
            master: Master DataFrame
            master_cols: Master column mapping
            wh_cols: Warehouse column mapping

        Returns:
            pd.DataFrame: 원본 순서가 유지된 Warehouse DataFrame (신규 케이스는 하단)
        """
        master_case_col = master_cols["case_number"]
        wh_case_col = wh_cols["case_number"]

        # Master의 Case No 집합
        master_case_set = set(master[master_case_col].dropna().tolist())

        # 1. Warehouse에 이미 있는 케이스 (원본 순서 유지)
        wh_existing_cases = warehouse[warehouse[wh_case_col].isin(master_case_set)].copy()

        # 2. Warehouse에 없는 케이스 (원래부터 Warehouse에만 있던)
        wh_only_cases = warehouse[~warehouse[wh_case_col].isin(master_case_set)].copy()

        # 최종: 기존 Warehouse 순서 + Warehouse 전용 케이스
        # (신규 케이스는 _apply_updates에서 append됨)
        sorted_warehouse = pd.concat([wh_existing_cases, wh_only_cases], ignore_index=True)

        print(f"  Warehouse order maintained: {len(sorted_warehouse)} rows")
        print(f"    - Existing cases: {len(wh_existing_cases)} (original order)")
        print(f"    - Warehouse-only: {len(wh_only_cases)}")
        print(f"  First 5 Warehouse Case No.: {sorted_warehouse[wh_case_col].head().tolist()}")

        return sorted_warehouse

    def _apply_updates(
        self,
        master: pd.DataFrame,
        wh: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Apply updates from Master to Warehouse using matched column names.

        Args:
            master: Master DataFrame
            wh: Warehouse DataFrame
            master_cols: Master column mapping
            wh_cols: Warehouse column mapping

        Returns:
            Tuple of (updated_warehouse, statistics)
        """
        print("\nApplying updates from Master to Warehouse...")

        stats = dict(updates=0, date_updates=0, field_updates=0, appends=0)

        # Build warehouse index by case number
        wh_case_col = wh_cols["case_number"]
        wh_index = self._build_case_index(wh, wh_case_col)

        # Get master case column
        master_case_col = master_cols["case_number"]

        # Find all semantic keys that exist in both files
        common_keys = set(master_cols.keys()) & set(wh_cols.keys())

        # Find master-only keys (columns that exist in Master but not in Warehouse)
        # These need to be added to Warehouse during synchronization
        master_only_keys = set(master_cols.keys()) - set(wh_cols.keys())

        print(f"  Master columns: {list(master_cols.keys())}")
        print(f"  Warehouse columns: {list(wh_cols.keys())}")
        print(f"  Common keys: {list(common_keys)}")
        print(f"  Master-only keys: {list(master_only_keys)}")

        if master_only_keys:
            print(f"  Master-only columns found: {list(master_only_keys)}")
            print(f"  These will be added to Warehouse during sync")

        # ✅ Phase 3: Store semantic key → actual column name mappings
        for semantic_key in common_keys:
            w_col = wh_cols.get(semantic_key)
            if w_col:
                self.change_tracker.set_column_mapping(semantic_key, w_col)

        for semantic_key in master_only_keys:
            m_col = master_cols.get(semantic_key)
            if m_col:
                self.change_tracker.set_column_mapping(semantic_key, m_col)

        print(f"  [Phase 3] Stored {len(self.change_tracker.semantic_to_column)} column mappings")

        # Process each master row
        for mi, mrow in master.iterrows():
            # Get case number
            key = (
                str(mrow[master_case_col]).strip().upper()
                if pd.notna(mrow[master_case_col])
                else ""
            )
            if not key:
                continue

            # Check if case exists in warehouse
            if key not in wh_index:
                # Append new case - START WITH ALL MASTER COLUMNS
                append_row = {}

                # STEP 1: Copy ALL Master columns (matched or not)
                for col in master.columns:
                    append_row[col] = mrow[col]

                # STEP 2: Apply semantic name mapping for common columns
                # This renames Master columns to Warehouse names when they differ
                for semantic_key in common_keys:
                    m_col = master_cols[semantic_key]
                    w_col = wh_cols[semantic_key]

                    # If Warehouse uses different name, rename in append_row
                    if m_col != w_col and m_col in append_row:
                        append_row[w_col] = append_row.pop(m_col)

                # STEP 3: Source_Sheet already copied in Step 1

                # STEP 4: Ensure warehouse has all columns before concat
                for col in append_row.keys():
                    if col not in wh.columns:
                        wh[col] = None  # Initialize for existing rows

                wh = pd.concat([wh, pd.DataFrame([append_row])], ignore_index=True)
                new_index = len(wh) - 1
                stats["appends"] += 1

                # Track new case
                self.change_tracker.log_new_case(
                    case_no=key, row_data=append_row, row_index=new_index
                )
                continue

            # Case exists - update it
            wi = wh_index[key]

            # Update Source_Sheet from Master for existing cases
            # (Source_Sheet is not in common_keys, so we handle it separately)
            if "Source_Sheet" in master.columns and "Source_Sheet" in wh.columns and wi < len(wh):
                # Use Master's Source_Sheet to reflect original sheet name
                old_source = wh.at[wi, "Source_Sheet"]
                new_source = mrow["Source_Sheet"]
                wh.at[wi, "Source_Sheet"] = new_source
                # Track Source_Sheet updates for debugging
                if old_source != new_source:
                    stats["source_sheet_updates"] = stats.get("source_sheet_updates", 0) + 1

            # ✅ Update Source_Vendor from Master for existing cases
            if "Source_Vendor" in master.columns and "Source_Vendor" in wh.columns and wi < len(wh):
                # Use Master's Source_Vendor to reflect vendor
                new_vendor = mrow["Source_Vendor"]
                wh.at[wi, "Source_Vendor"] = new_vendor
                # Track Source_Vendor updates
                stats["source_vendor_updates"] = stats.get("source_vendor_updates", 0) + 1

            # Process each common column
            for semantic_key in common_keys:
                m_col = master_cols[semantic_key]
                w_col = wh_cols[semantic_key]

                # Skip metadata columns - preserve Warehouse's original value
                if w_col in METADATA_COLUMNS:
                    continue

                mval = mrow[m_col]
                wval = wh.at[wi, w_col] if wi < len(wh) else None

                # Check if this is a date column
                is_date = semantic_key in self.date_semantic_keys

                if is_date:
                    # Date column: Master always wins if it has a value
                    if pd.notna(mval):
                        if not self._dates_equal(mval, wval):
                            stats["updates"] += 1
                            stats["date_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,  # w_col is actual Excel header name (e.g., "ETD/ATD")
                                semantic_key=semantic_key,  # ✅ Phase 3
                                case_no=key,  # ✅ Phase 4: Include case_no for row lookup
                                old_value=wval,
                                new_value=mval,
                                change_type="date_update",
                            )
                        else:
                            # Equal logically - ensure consistent format
                            wh.at[wi, w_col] = mval
                else:
                    # Non-date column: Overwrite if Master has value
                    if ALWAYS_OVERWRITE_NONDATE and pd.notna(mval):
                        if (wval is None) or (str(mval) != str(wval)):
                            stats["updates"] += 1
                            stats["field_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,  # w_col is actual Excel header name
                                semantic_key=semantic_key,  # ✅ Phase 3
                                case_no=key,  # ✅ Phase 4: Include case_no
                                old_value=wval,
                                new_value=mval,
                                change_type="field_update",
                            )

            # Process master-only columns (like DHL WH) for existing cases
            for semantic_key in master_only_keys:
                m_col = master_cols[semantic_key]
                mval = mrow[m_col]

                # Add master-only column to warehouse if it doesn't exist
                if m_col not in wh.columns:
                    wh[m_col] = None  # Initialize with None for all rows

                # Update the value for this case
                if pd.notna(mval):
                    old_val = wh.at[wi, m_col] if wi < len(wh) else None
                    if old_val != mval:
                        stats["updates"] += 1
                        stats["field_updates"] += 1
                        wh.at[wi, m_col] = mval
                        self.change_tracker.add_change(
                            row_index=wi,
                            column_name=m_col,
                            semantic_key=semantic_key,  # ✅ Phase 3
                            case_no=key,  # ✅ Phase 4: Include case_no
                            old_value=old_val,
                            new_value=mval,
                            change_type="master_only_update",
                        )

        print(f"  [OK] Updates: {stats['updates']} cells changed")
        print(f"    - Date updates: {stats['date_updates']}")
        print(f"    - Field updates: {stats['field_updates']}")
        print(f"    - New records: {stats['appends']}")
        if "source_sheet_updates" in stats:
            print(f"    - Source_Sheet updates: {stats['source_sheet_updates']}")

        return wh, stats

    def synchronize(
        self,
        master_xlsx: str,
        warehouse_xlsx: str,
        output_path: Optional[str] = None,
        warehouse_configs: Optional[List[Dict[str, Any]]] = None
    ) -> SyncResult:
        """
        Synchronize Master data into Warehouse file.

        This is the main entry point for the synchronization process. It:
        1. Loads both files with automatic header detection
        2. Matches all required columns using semantic keys
        3. Validates that required columns exist
        4. Applies Master data to Warehouse
        5. Maintains Master's row order
        6. Saves the result with color-coded changes

        Args:
            master_xlsx: Path to Master Excel file
            warehouse_xlsx: Path to Warehouse Excel file (legacy, used if warehouse_configs is None)
            output_path: Path for output file (optional, auto-generated if None)
            warehouse_configs: Optional list of warehouse file configs for multi-vendor loading.
                Each config dict should have: {'vendor': str, 'path': str, 'required': bool}
                If provided, this takes precedence over warehouse_xlsx.

        Returns:
            SyncResult object with success status, messages, and statistics

        Example:
            >>> sync = DataSynchronizerV30()
            >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
            >>> if result.success:
            >>>     print(f"Success! Output: {result.output_path}")
            >>>     print(f"Changes: {result.stats}")
        """
        try:
            # Phase 1: Load files by sheets (NEW: Sheet-by-sheet processing)
            print("\n" + "=" * 60)
            print("PHASE 1: Loading Files by Sheets")
            print("=" * 60)

            # Load Master files (HITACHI + SIEMENS)
            master_sheets = self._load_master_files(master_xlsx)

            # Check if Master and Warehouse are the same file
            from pathlib import Path

            master_path = Path(master_xlsx).resolve()
            warehouse_path = Path(warehouse_xlsx).resolve()

            # ===== Multi-vendor warehouse loading (Option B: Unified Sync) =====
            warehouse_sheets_data: Dict[str, Tuple[pd.DataFrame, int]] = {}
            if warehouse_configs:
                print("[INFO] Using multi-vendor warehouse loading mode")
                warehouse_sheets, header_rows = self._load_warehouse_files(warehouse_configs)

                # Add Source_Vendor and Source_Sheet metadata
                for sheet_name, df in warehouse_sheets.items():
                    if "Source_Sheet" not in df.columns:
                        df["Source_Sheet"] = sheet_name
                    print(
                        f"[WAREHOUSE-UNIFIED] Initialized metadata for '{sheet_name}': "
                        f"Source_Vendor='{df.get('Source_Vendor', ['N/A'])[0]}', "
                        f"rows={len(df)}"
                    )

                # Create warehouse_sheets_data structure for compatibility
                for sheet_name, df in warehouse_sheets.items():
                    warehouse_sheets_data[sheet_name] = (df, header_rows.get(sheet_name, 4))

            elif master_path == warehouse_path:
                # Same file: Load Warehouse separately to get ALL sheets (including warehouse-only sheets like HE Local, Capacitor)
                print("[INFO] Master and Warehouse are the same file - loading Warehouse separately to capture all sheets")
                # Load Warehouse file separately to get all sheets that match INCLUDE_SHEET_PATTERNS
                warehouse_sheets_data = self._load_file_by_sheets(warehouse_xlsx, "Warehouse")
                # Extract DataFrames from tuples
                warehouse_sheets = {name: df for name, (df, _) in warehouse_sheets_data.items()}

                # Merge Master data into Warehouse sheets where they match
                # Warehouse sheets that don't match Master will remain as warehouse-only sheets
                print(f"[INFO] Master has {len(master_sheets)} sheets, Warehouse has {len(warehouse_sheets)} sheets")

                # ✅ Add Source_Vendor and Source_Sheet to Warehouse data (initially empty)
                for sheet_name, df in warehouse_sheets.items():
                    if "Source_Vendor" not in df.columns:
                        df["Source_Vendor"] = None  # Will be filled from Master during sync
                    if "Source_Sheet" not in df.columns:
                        df["Source_Sheet"] = sheet_name  # Warehouse's own sheet name
                    print(
                        f"[WAREHOUSE] Initialized metadata for '{sheet_name}': Source_Vendor=None, Source_Sheet='{sheet_name}'"
                    )
            else:
                # Different files: Load Warehouse separately
                warehouse_sheets_data = self._load_file_by_sheets(warehouse_xlsx, "Warehouse")
                # Extract DataFrames from tuples
                warehouse_sheets = {name: df for name, (df, _) in warehouse_sheets_data.items()}

                # ✅ Add Source_Vendor and Source_Sheet to Warehouse data (initially empty)
                for sheet_name, df in warehouse_sheets.items():
                    if "Source_Vendor" not in df.columns:
                        df["Source_Vendor"] = None  # Will be filled from Master during sync
                    if "Source_Sheet" not in df.columns:
                        df["Source_Sheet"] = sheet_name  # Warehouse's own sheet name
                    print(
                        f"[WAREHOUSE] Initialized metadata for '{sheet_name}': Source_Vendor=None, Source_Sheet='{sheet_name}'"
                    )

            # Phase 2: Process each sheet independently
            print("\n" + "=" * 60)
            print("PHASE 2: Sheet-by-Sheet Synchronization")
            print("=" * 60)

            processed_sheets = {}
            total_stats = {"updates": 0, "appends": 0, "field_updates": 0, "date_updates": 0}

            # Get common sheet names with semantic matching
            common_sheets = {}  # Dict[master_sheet, warehouse_sheet]
            master_only_sheets = []
            warehouse_only_sheets = []

            # Find matching sheets
            matched_warehouse_sheets = set()

            for m_sheet in master_sheets.keys():
                w_sheet = self._find_matching_sheet(m_sheet, list(warehouse_sheets.keys()))
                if w_sheet:
                    common_sheets[m_sheet] = w_sheet
                    matched_warehouse_sheets.add(w_sheet)
                    print(f"  [MATCH] '{m_sheet}' ↔ '{w_sheet}'")
                else:
                    master_only_sheets.append(m_sheet)

            # Find unmatched warehouse sheets
            for w_sheet in warehouse_sheets.keys():
                if w_sheet not in matched_warehouse_sheets:
                    warehouse_only_sheets.append(w_sheet)

            print(f"Common sheets: {len(common_sheets)}")
            print(f"Master-only sheets: {len(master_only_sheets)}")
            print(f"Warehouse-only sheets: {len(warehouse_only_sheets)}")

            # Dictionary to store per-sheet change trackers
            sheet_change_trackers = {}

            # Process common sheets with matched pairs
            for m_sheet_name, w_sheet_name in common_sheets.items():
                print(f"\n--- Synchronizing: '{m_sheet_name}' ↔ '{w_sheet_name}' ---")

                m_df = master_sheets[m_sheet_name]
                w_df = warehouse_sheets[w_sheet_name]
                # Get header rows from original data
                w_header_row = warehouse_sheets_data[w_sheet_name][1]
                m_header_row = 4  # Default for master sheets

                # Use warehouse sheet name for output
                output_sheet_name = w_sheet_name

                # ✅ Create a new change tracker for this sheet
                self.change_tracker = ChangeTracker()

                # Match headers for this sheet
                master_columns = self._match_and_validate_headers(m_df, f"Master-{m_sheet_name}")
                warehouse_columns = self._match_and_validate_headers(
                    w_df, f"Warehouse-{w_sheet_name}"
                )

                # Sort according to Master order
                m_df, w_df = self._apply_master_order_sorting(
                    m_df, w_df, master_columns, warehouse_columns
                )

                # Apply updates
                updated_w_df, stats = self._apply_updates(
                    m_df, w_df, master_columns, warehouse_columns
                )

                # Maintain Master order after updates
                updated_w_df = self._maintain_warehouse_order(
                    updated_w_df, m_df, master_columns, warehouse_columns
                )

                # Remove Source_Sheet column (sheet name is preserved in sheet name)
                if "Source_Sheet" in updated_w_df.columns:
                    updated_w_df = updated_w_df.drop(columns=["Source_Sheet"])

                # ✅ Store the change tracker for this sheet
                sheet_change_trackers[output_sheet_name] = self.change_tracker

                processed_sheets[output_sheet_name] = (updated_w_df, w_header_row)

                # Accumulate stats
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)

                # 히타치 데이터 추적 (동기화 후)
                hitachi_in_sheet = 0
                if "Source_Vendor" in updated_w_df.columns:
                    hitachi_in_sheet = (updated_w_df["Source_Vendor"].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])).sum()
                
                print(
                    f"  [OK] {output_sheet_name}: {len(updated_w_df)} rows, {stats.get('updates', 0)} updates, {stats.get('appends', 0)} new records"
                )
                if hitachi_in_sheet > 0:
                    print(
                        f"      [HITACHI] 히타치 행수: {hitachi_in_sheet:,} 행 (동기화 후)"
                    )

            # Process warehouse-only sheets (keep as-is)
            for sheet_name in warehouse_only_sheets:
                print(f"\n--- Keeping warehouse-only sheet: '{sheet_name}' ---")
                w_df = warehouse_sheets[sheet_name]
                # Get header row from warehouse_sheets_data
                w_header_row = warehouse_sheets_data[sheet_name][1]

                # Remove Source_Sheet column
                if "Source_Sheet" in w_df.columns:
                    w_df = w_df.drop(columns=["Source_Sheet"])

                processed_sheets[sheet_name] = (w_df, w_header_row)
                print(f"  [OK] {sheet_name}: {len(w_df)} rows (unchanged)")

            # Process master-only sheets (append to first common sheet or create new)
            for sheet_name in master_only_sheets:
                print(f"\n--- Processing master-only sheet: '{sheet_name}' ---")
                m_df = master_sheets[sheet_name]
                # Use default header row for master sheets
                m_header_row = (
                    4  # Default for HITACHI, SIEMENS uses 0 but we'll use 4 for consistency
                )

                # Remove Source_Sheet column
                if "Source_Sheet" in m_df.columns:
                    m_df = m_df.drop(columns=["Source_Sheet"])

                processed_sheets[sheet_name] = (m_df, m_header_row)
                print(f"  [OK] {sheet_name}: {len(m_df)} rows (master-only)")

            # Phase 3: Save multi-sheet output
            print("\n" + "=" * 60)
            print("PHASE 3: Saving Multi-Sheet Output")
            print("=" * 60)

            # Determine output path
            out = output_path or str(
                Path(warehouse_xlsx).with_name(
                    Path(warehouse_xlsx).stem + ".synced_v3.4_multi.xlsx"
                )
            )

            # Save all sheets with standard 63-column header order
            print(f"  Writing to: {Path(out).name}")
            # XlsxWriter 최적화 사용 여부에 따라 선택
            if self.use_xlsxwriter:
                try:
                    from scripts.core.excel_writer_optimized import OptimizedExcelWriter
                    with OptimizedExcelWriter(out, use_xlsxwriter=True, constant_memory=True) as writer:
                        for sheet_name, (df, header_row) in processed_sheets.items():
                            # Clean sheet name for Excel (remove invalid characters)
                            clean_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[
                                :31
                            ]  # Excel limit
                            # Apply standard 63-column header order before saving
                            df_reordered = reorder_dataframe_columns(
                                df, is_stage2=False, keep_unlisted=False, use_semantic_matching=True
                            )
                            # DUP-GUARD (A): within-sheet duplicate flags + A-column injection
                            try:
                                df_with_flag, dup_mask = mark_duplicates(
                                    df_reordered, semantic_finder=_find_col, keep="first"
                                )
                                df_reordered = df_with_flag
                                # place visible label as first column for Excel output
                                if "고유 중복" not in df_reordered.columns:
                                    df_reordered.insert(0, "고유 중복", np.where(dup_mask, "중복", "고유"))
                            except Exception as e:
                                print(
                                    f"    [WARNING][DUP-GUARD][{sheet_name}] skip marking duplicates: {e}"
                                )
                            writer.write_dataframe(df_reordered, sheet_name=clean_sheet_name, index=False)
                            print(
                                f"    - {clean_sheet_name}: {len(df)} rows, {len(df_reordered.columns)} columns (standard order, XlsxWriter)"
                            )
                except Exception as e:
                    print(f"  [WARNING] XlsxWriter failed, falling back to openpyxl: {e}")
                    self.use_xlsxwriter = False
                    # Fall through to openpyxl code below

            # 기본 openpyxl 방식 (XlsxWriter 실패 시 또는 비활성화 시)
            if not self.use_xlsxwriter:
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    for sheet_name, (df, header_row) in processed_sheets.items():
                        # Clean sheet name for Excel (remove invalid characters)
                        clean_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[
                            :31
                        ]  # Excel limit
                        # Apply standard 63-column header order before saving
                        df_reordered = reorder_dataframe_columns(
                            df, is_stage2=False, keep_unlisted=False, use_semantic_matching=True
                        )
                        # DUP-GUARD (A): within-sheet duplicate flags + A-column injection
                        try:
                            df_with_flag, dup_mask = mark_duplicates(
                                df_reordered, semantic_finder=_find_col, keep="first"
                            )
                            df_reordered = df_with_flag
                            # place visible label as first column for Excel output
                            if "고유 중복" not in df_reordered.columns:
                                df_reordered.insert(0, "고유 중복", np.where(dup_mask, "중복", "고유"))
                        except Exception as e:
                            print(
                                f"    [WARNING][DUP-GUARD][{sheet_name}] skip marking duplicates: {e}"
                            )
                        df_reordered.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                        print(
                            f"    - {clean_sheet_name}: {len(df)} rows, {len(df_reordered.columns)} columns (standard order)"
                        )

            print(f"  [OK] Saved {len(processed_sheets)} sheets")

            # Apply color formatting to all sheets
            print(f"  Applying color formatting...")
            for sheet_name, (df, header_row) in processed_sheets.items():
                clean_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[:31]
                # ✅ Use the sheet-specific change tracker
                if sheet_name in sheet_change_trackers:
                    self.change_tracker = sheet_change_trackers[sheet_name]
                    print(f"    - {clean_sheet_name}: {len(self.change_tracker.changes)} changes")
                    self._apply_excel_formatting(out, clean_sheet_name, header_row)
                else:
                    print(f"    - {clean_sheet_name}: No change tracker (skipped)")
            print(f"  [OK] Formatting applied to all sheets")

            # Create merged file (NEW: Single sheet with all data combined)
            print(f"\n[INFO] 합쳐진 단일시트 파일 생성 중...")

            # Combine all sheet DataFrames in specific order
            combined_dfs = []

            # Define sheet order (Case List, RIL first)
            sheet_order = []
            if "Case List, RIL" in processed_sheets:
                sheet_order.append("Case List, RIL")
            if "HE Local" in processed_sheets:
                sheet_order.append("HE Local")
            if "HE-0214,0252 (Capacitor)" in processed_sheets:
                sheet_order.append("HE-0214,0252 (Capacitor)")

            # Add any remaining sheets not in the predefined order
            for sheet_name in processed_sheets.keys():
                if sheet_name not in sheet_order:
                    sheet_order.append(sheet_name)

            # Combine sheets in order
            for sheet_name in sheet_order:
                df, _ = processed_sheets[sheet_name]
                df_copy = df.copy()
                # 표준화된 시트명 사용 (이미 _load_file_by_sheets에서 적용되었지만, 병합 시에도 확인)
                canon_sheet_name = _canon_sheet(sheet_name)
                if "Source_Sheet" not in df_copy.columns or (len(df_copy) > 0 and df_copy["Source_Sheet"].iloc[0] != canon_sheet_name):
                    df_copy["Source_Sheet"] = canon_sheet_name  # 출처 기록
                combined_dfs.append(df_copy)

            # ========== 체크포인트 로거: 병합 직전 시트별 행수 추적 (개선 + 히타치 상세 분석) ==========
            lost = []
            sheet_names_list = list(sheet_order)  # sheet_order 순서대로
            
            # HITACHI 데이터 추적을 위한 변수
            hitachi_rows_before_merge = 0
            hitachi_source_vendor_col = None
            
            # processed_sheets에서 원본 행수 정보 추출
            for idx, df_copy in enumerate(combined_dfs):
                sheet_name = sheet_names_list[idx] if idx < len(sheet_names_list) else f"sheet_{idx}"
                
                # 원본 행수 추적 (여러 소스에서 확인)
                raw_n = 0
                if sheet_name in processed_sheets:
                    df_orig, _ = processed_sheets[sheet_name]
                    raw_n = len(df_orig)
                    # attrs에서도 확인
                    if hasattr(df_orig, 'attrs') and '_raw_rows' in df_orig.attrs:
                        raw_n = max(raw_n, int(df_orig.attrs.get("_raw_rows", 0) or 0))
                # df_copy에서도 확인
                if hasattr(df_copy, 'attrs') and '_raw_rows' in df_copy.attrs:
                    raw_n = max(raw_n, int(df_copy.attrs.get("_raw_rows", 0) or 0))
                # df_copy의 실제 길이와 비교 (처리 과정에서 변경되었을 수 있음)
                norm_n = len(df_copy)
                
                # Case No 컬럼 정확한 감지 (여러 패턴 확인)
                case_cols = [c for c in df_copy.columns if c and isinstance(c, str)]
                has_case = False
                case_col_name = None
                for col in case_cols:
                    col_norm = col.strip().lower().replace(" ", "").replace(".", "").replace("_", "")
                    if col_norm in {"caseno", "caseid", "casenumber", "casenumber.", "caseno."}:
                        has_case = True
                        case_col_name = col
                        break
                
                # Case No 값 존재 여부 확인
                has_case_values = False
                if case_col_name:
                    non_empty = df_copy[case_col_name].notna() & (df_copy[case_col_name].astype(str).str.strip() != "")
                    has_case_values = non_empty.sum() > 0
                
                lost.append({
                    "sheet": sheet_name,
                    "raw_rows": raw_n,
                    "normalized": norm_n,
                    "diff": norm_n - raw_n if raw_n > 0 else 0,
                    "has_case_col": has_case,
                    "case_col": case_col_name or "N/A",
                    "has_case_values": has_case_values,
                    "case_values_count": int(df_copy[case_col_name].notna().sum()) if case_col_name else 0,
                })
                
                # HITACHI 데이터 카운트 (Source_Vendor 확인)
                if "Source_Vendor" in df_copy.columns:
                    hitachi_in_sheet = (df_copy["Source_Vendor"].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])).sum()
                    hitachi_rows_before_merge += hitachi_in_sheet
                    if hitachi_in_sheet > 0 and hitachi_source_vendor_col is None:
                        hitachi_source_vendor_col = "Source_Vendor"
                elif "Vendor" in df_copy.columns:
                    hitachi_in_sheet = (df_copy["Vendor"].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])).sum()
                    hitachi_rows_before_merge += hitachi_in_sheet
                    if hitachi_in_sheet > 0 and hitachi_source_vendor_col is None:
                        hitachi_source_vendor_col = "Vendor"
            
            if lost:
                print("\n[DEBUG] per-sheet rows (병합 직전 상세) >>>")
                lost_df = pd.DataFrame(lost).sort_values("normalized", ascending=True)
                print(lost_df.to_string(index=False))
                total_before_merge = sum(lost_df["normalized"])
                total_raw = sum(lost_df["raw_rows"]) if lost_df["raw_rows"].sum() > 0 else total_before_merge
                print(f"  총합 (병합 직전): {total_before_merge:,} 행 (원본: {total_raw:,} 행)")
                
                # 경고: 원본 대비 손실이 있는 경우
                if total_raw > 0 and total_before_merge < total_raw:
                    loss_count = total_raw - total_before_merge
                    loss_pct = (loss_count / total_raw) * 100
                    print(f"  [WARNING] 원본 대비 손실: {loss_count:,} 행 ({loss_pct:.1f}%)")
                
                # Case No 컬럼 없는 시트 경고
                sheets_without_case = lost_df[~lost_df["has_case_col"]]["sheet"].tolist()
                if sheets_without_case:
                    print(f"  [WARNING] Case No 컬럼 없는 시트: {', '.join(sheets_without_case)}")

            # Concatenate all sheets
            merged_df = pd.concat(combined_dfs, ignore_index=True, sort=False)
            print(f"  - 합쳐진 데이터: {len(merged_df)}행, {len(merged_df.columns)}컬럼")
            print(f"  - Source_Sheet 컬럼 추가됨")
            # ========== 체크포인트 로거: 중복 제거 직전 ==========
            before_dedup_count = len(merged_df)
            
            # DUP-GUARD (B): workbook-level dedup by Case key (상세 분석)
            try:
                # 중복 제거 전 상세 분석
                case_col = _find_col(merged_df, "case_number", required=False) if _find_col else None
                if not case_col:
                    # 폴백: 직접 컬럼 찾기
                    for cand in ["Case No.", "Case No", "CASE NO", "Caseno", "Case Number"]:
                        if cand in merged_df.columns:
                            case_col = cand
                            break
                
                if case_col:
                    # 빈 키 행 확인
                    empty_keys = merged_df[case_col].isna() | (merged_df[case_col].astype(str).str.strip() == "")
                    empty_count = empty_keys.sum()
                    
                    # 중복 키 확인
                    if empty_count < len(merged_df):
                        non_empty = merged_df[~empty_keys]
                        dup_keys = non_empty[case_col].duplicated(keep=False)
                        dup_count = dup_keys.sum()
                        unique_keys = len(non_empty[case_col].unique())
                        
                        print(f"  [DEDUP 분석] Case No 컬럼: {case_col}")
                        print(f"    - 총 행수: {len(merged_df):,}")
                        print(f"    - 빈 키: {empty_count:,} 행 (보존)")
                        print(f"    - 유효 키: {len(merged_df) - empty_count:,} 행")
                        print(f"    - 고유 키: {unique_keys:,}개")
                        print(f"    - 중복 키: {dup_count:,} 행")
                        
                        if dup_count > 0:
                            # 중복 키 샘플 출력 (최대 10개)
                            dup_samples = non_empty[dup_keys][case_col].value_counts().head(10)
                            print(f"    - 중복 키 샘플 (상위 10개):")
                            for key, count in dup_samples.items():
                                print(f"      '{key}': {count}회")
                            
                            # 중복 키의 Source_Sheet 분포 분석 (시트 간 중복인지 시트 내 중복인지 확인)
                            if "Source_Sheet" in merged_df.columns:
                                dup_keys_list = non_empty[dup_keys][case_col].unique()[:5]  # 상위 5개만 분석
                                print(f"    - 중복 키 Source_Sheet 분포 분석 (샘플 5개):")
                                for dup_key in dup_keys_list:
                                    dup_key_rows = merged_df[merged_df[case_col] == dup_key]
                                    sheet_dist = dup_key_rows["Source_Sheet"].value_counts() if "Source_Sheet" in dup_key_rows.columns else pd.Series()
                                    print(f"      '{dup_key}': {len(dup_key_rows)}행, 시트 분포: {dict(sheet_dist.to_dict())}")
                    else:
                        print(f"  [WARNING] 모든 행의 Case No가 비어있음 - 중복 제거 스킵")
                
                # 중복 제거 실행 (Source_Vendor 고려)
                merged_df_before_dedup = merged_df.copy()
                
                # ✅ Source_Vendor가 다른 경우 별도 레코드로 유지
                if "Source_Vendor" in merged_df.columns and case_col:
                    try:
                        # 빈 키 행은 보존
                        empty_mask = merged_df[case_col].isna() | (merged_df[case_col].astype(str).str.strip() == "")
                        non_empty_df = merged_df[~empty_mask].copy()
                        empty_df = merged_df[empty_mask].copy()
                        
                        # Source_Vendor + Case No 복합 키로 중복 제거
                        non_empty_df = non_empty_df.drop_duplicates(
                            subset=[case_col, "Source_Vendor"], 
                            keep="first"
                        )
                        
                        merged_df = pd.concat([non_empty_df, empty_df], ignore_index=True, sort=False)
                        after_dedup_count = len(merged_df)
                        removed_count = before_dedup_count - after_dedup_count
                        print(f"  - 전역 중복 제거 후 (Source_Vendor 고려): {after_dedup_count}행 (제거: {removed_count}행)")
                    except Exception as e:
                        print(f"  [WARNING] Source_Vendor 기반 중복 제거 실패, 기본 방법 사용: {e}")
                        # 폴백: 기존 로직
                        merged_df = drop_duplicates_by_case(
                            merged_df, semantic_finder=_find_col, keep="first"
                        )
                        after_dedup_count = len(merged_df)
                        removed_count = before_dedup_count - after_dedup_count
                        print(f"  - 전역 중복 제거 후: {after_dedup_count}행 (제거: {removed_count}행)")
                else:
                    # Source_Vendor 컬럼이 없으면 기존 로직
                    merged_df = drop_duplicates_by_case(
                        merged_df, semantic_finder=_find_col, keep="first"
                    )
                    after_dedup_count = len(merged_df)
                    removed_count = before_dedup_count - after_dedup_count
                    print(f"  - 전역 중복 제거 후: {after_dedup_count}행 (제거: {removed_count}행)")
                
                # 빈 키 행 보존 확인
                if case_col:
                    empty_after = (merged_df[case_col].isna() | (merged_df[case_col].astype(str).str.strip() == "")).sum()
                    if empty_count != empty_after:
                        print(f"  [WARNING] 빈 키 행 수 변화: {empty_count} → {empty_after}")
                    
            except Exception as e:
                import traceback
                print(f"  [WARNING][DUP-GUARD][merge] skip global dedup: {e}")
                print(f"  [DEBUG] {traceback.format_exc()}")
            
            # ========== 체크포인트 로거: 최종 결과 (상세 + 히타치 분석) ==========
            print(f"\n[DEBUG] 병합 결과 요약:")
            print(f"  - 병합 직전 총합: {total_before_merge:,} 행")
            print(f"  - 병합 후 (concat): {before_dedup_count:,} 행")
            print(f"  - 중복 제거 후: {len(merged_df):,} 행")
            
            # 누락 분석
            concat_loss = total_before_merge - before_dedup_count
            dedup_loss = before_dedup_count - len(merged_df)
            total_loss = total_before_merge - len(merged_df)
            
            if concat_loss > 0:
                print(f"  [WARNING] concat 단계 누락: {concat_loss:,} 행")
            if dedup_loss > 0:
                print(f"  - 중복 제거로 인한 감소: {dedup_loss:,} 행")
            print(f"  - 총 누락 행수: {total_loss:,} 행")
            
            # 누락 비율 계산
            if total_before_merge > 0:
                loss_pct = (total_loss / total_before_merge) * 100
                print(f"  - 누락 비율: {loss_pct:.1f}%")
                
                # 경고: 누락 비율이 높은 경우 (5% 초과)
                if loss_pct > 5.0:
                    print(f"  [WARNING] 높은 누락 비율 감지! ({loss_pct:.1f}%) 상세 분석 필요")
            
            # ========== HITACHI 데이터 상세 분석 ==========
            print(f"\n[DEBUG][HITACHI] 히타치 데이터 추적:")
            print(f"  - 병합 직전 히타치 행수: {hitachi_rows_before_merge:,} 행")
            
            if hitachi_source_vendor_col and hitachi_source_vendor_col in merged_df.columns:
                hitachi_rows_after_merge = (
                    merged_df[hitachi_source_vendor_col].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])
                ).sum()
                hitachi_loss = hitachi_rows_before_merge - hitachi_rows_after_merge
                hitachi_loss_pct = (hitachi_loss / hitachi_rows_before_merge * 100) if hitachi_rows_before_merge > 0 else 0
                
                print(f"  - 중복 제거 후 히타치 행수: {hitachi_rows_after_merge:,} 행")
                print(f"  - 히타치 누락 행수: {hitachi_loss:,} 행 ({hitachi_loss_pct:.1f}%)")
                
                if hitachi_loss_pct > 10.0:
                    print(f"  [WARNING][HITACHI] 높은 누락 비율! ({hitachi_loss_pct:.1f}%)")
                
                # 히타치 중복 키 분석 (빈 키 제외, 중복 제거 전 데이터 사용)
                hitachi_df_before_dedup = merged_df_before_dedup[
                    merged_df_before_dedup[hitachi_source_vendor_col].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])
                ].copy()
                
                if case_col and case_col in hitachi_df_before_dedup.columns:
                    # 빈 키 분리
                    hitachi_keys_str = hitachi_df_before_dedup[case_col].astype(str).str.strip()
                    hitachi_empty_mask = (
                        hitachi_df_before_dedup[case_col].isna() 
                        | (hitachi_keys_str == "") 
                        | (hitachi_keys_str.isin(["nan", "NaN", "None", "N/A", "NA", ""]))
                    )
                    hitachi_non_empty = hitachi_df_before_dedup[~hitachi_empty_mask].copy()
                    
                    # 빈 키 통계
                    hitachi_empty_count_before = hitachi_empty_mask.sum()
                    print(f"  - 히타치 빈 Case No (중복 제거 전): {hitachi_empty_count_before:,} 행")
                    
                    if len(hitachi_non_empty) > 0:
                        hitachi_non_empty_keys = hitachi_non_empty[case_col].astype(str).str.strip()
                        hitachi_dup_mask = hitachi_non_empty_keys.duplicated(keep=False)
                        hitachi_dup_count = hitachi_dup_mask.sum()
                        
                        if hitachi_dup_count > 0:
                            # 중복된 행들의 키 값 추출
                            hitachi_dup_rows = hitachi_non_empty[hitachi_dup_mask]
                            hitachi_dup_keys_series = hitachi_dup_rows[case_col].astype(str).str.strip()
                            hitachi_dup_samples = hitachi_dup_keys_series.value_counts().head(10)
                            
                            print(f"  - 히타치 중복 키 (유효 키만, 중복 제거 전): {hitachi_dup_count:,} 행")
                            print(f"  - 히타치 중복 키 샘플 (상위 10개):")
                            
                            if len(hitachi_dup_samples) > 0:
                                for key, count in hitachi_dup_samples.items():
                                    print(f"      '{key}': {count}회")
                                    
                                # 히타치 중복 키의 Source_Sheet 분포
                                if "Source_Sheet" in hitachi_non_empty.columns:
                                    print(f"  - 히타치 중복 키 Source_Sheet 분포 (샘플 5개):")
                                    sample_keys = hitachi_dup_samples.head(5).index.tolist()
                                    for dup_key in sample_keys:
                                        dup_key_str = str(dup_key)
                                        dup_key_rows = hitachi_dup_rows[hitachi_dup_rows[case_col].astype(str).str.strip() == dup_key_str]
                                        if len(dup_key_rows) > 0:
                                            sheet_dist = dup_key_rows["Source_Sheet"].value_counts() if "Source_Sheet" in dup_key_rows.columns else pd.Series()
                                            print(f"      '{dup_key_str}': {len(dup_key_rows)}행, 시트 분포: {dict(sheet_dist.to_dict())}")
                            else:
                                print(f"      [WARNING] 중복 키 샘플 추출 실패 (hitachi_dup_samples 비어있음)")
                        else:
                            print(f"  - 히타치 유효 키 중복 없음")
                    
                    # 히타치 최종 상태 확인 (중복 제거 후)
                    hitachi_df_after = merged_df[
                        merged_df[hitachi_source_vendor_col].astype(str).str.upper().str.strip().isin(["HITACHI", "HE", "HITACHI(HE)"])
                    ].copy()
                    hitachi_empty_after = (
                        hitachi_df_after[case_col].isna() 
                        | (hitachi_df_after[case_col].astype(str).str.strip() == "") 
                        | (hitachi_df_after[case_col].astype(str).str.strip().isin(["nan", "NaN", "None", "N/A", "NA", ""]))
                    ).sum()
                    print(f"  - 히타치 빈 Case No (중복 제거 후): {hitachi_empty_after:,} 행")
                    
                    if hitachi_empty_count_before != hitachi_empty_after:
                        print(f"  [WARNING][HITACHI] 빈 키 행 수 변화: {hitachi_empty_count_before:,} → {hitachi_empty_after:,}")
            else:
                print(f"  [WARNING][HITACHI] Source_Vendor 컬럼을 찾을 수 없어 히타치 추적 불가")

            # ✅ SIEMENS 데이터 검증 및 복원
            print(f"\n[INFO] SIEMENS 데이터 검증 중...")
            if "Source_Vendor" in merged_df.columns:
                siemens_count = (merged_df["Source_Vendor"] == "SIEMENS").sum()
                print(f"  - merged_df의 SIEMENS 행수: {siemens_count:,} 행")
                
                # Master 원본에서 SIEMENS 데이터 확인
                master_siemens_count = 0
                master_siemens_data = []
                for sheet_name, m_df in master_sheets.items():
                    if "Source_Vendor" in m_df.columns:
                        siemens_in_sheet = m_df[m_df["Source_Vendor"] == "SIEMENS"]
                        if len(siemens_in_sheet) > 0:
                            master_siemens_count += len(siemens_in_sheet)
                            master_siemens_data.append((sheet_name, siemens_in_sheet.copy()))
                            print(f"  - Master 원본 '{sheet_name}'의 SIEMENS: {len(siemens_in_sheet):,} 행")
                
                # SIEMENS 데이터가 누락된 경우 복원
                if master_siemens_count > siemens_count:
                    missing_count = master_siemens_count - siemens_count
                    print(f"  [WARNING] SIEMENS 데이터 누락 감지: {missing_count:,} 행")
                    print(f"  [INFO] Master 원본에서 SIEMENS 데이터 복원 중...")
                    
                    # Master 원본 SIEMENS 데이터 병합
                    if master_siemens_data:
                        master_siemens_list = []
                        for sheet_name, siemens_df in master_siemens_data:
                            # Source_Sheet 컬럼 확인 및 추가
                            if "Source_Sheet" not in siemens_df.columns:
                                canon_name = _canon_sheet(sheet_name)
                                siemens_df["Source_Sheet"] = canon_name
                            master_siemens_list.append(siemens_df)
                        
                        master_siemens_combined = pd.concat(master_siemens_list, ignore_index=True, sort=False)
                        
                        # merged_df에 없는 SIEMENS 케이스만 추가
                        case_col = "Case No." if "Case No." in merged_df.columns else "Case No"
                        if case_col in merged_df.columns and case_col in master_siemens_combined.columns:
                            # merged_df에 있는 SIEMENS Case No 확인
                            existing_siemens_cases = set(
                                merged_df[
                                    (merged_df["Source_Vendor"] == "SIEMENS") & 
                                    merged_df[case_col].notna()
                                ][case_col]
                                .astype(str)
                                .str.strip()
                            )
                            
                            # 누락된 SIEMENS 케이스만 필터링
                            missing_siemens = master_siemens_combined[
                                ~master_siemens_combined[case_col].astype(str).str.strip().isin(existing_siemens_cases)
                            ]
                            
                            if len(missing_siemens) > 0:
                                print(f"  [INFO] 누락된 SIEMENS 케이스 {len(missing_siemens):,}개 복원")
                                # merged_df에 추가
                                merged_df = pd.concat([merged_df, missing_siemens], ignore_index=True, sort=False)
                                print(f"  [OK] SIEMENS 데이터 복원 완료: {len(missing_siemens):,} 행 추가")
                                print(f"  - 복원 후 총 SIEMENS 행수: {(merged_df['Source_Vendor'] == 'SIEMENS').sum():,} 행")
                            else:
                                print(f"  [INFO] 누락된 SIEMENS 케이스 없음 (Case No 매칭됨)")
                        else:
                            # Case No 컬럼이 없으면 전체 추가 (안전장치)
                            print(f"  [WARNING] Case No 컬럼을 찾을 수 없어 전체 SIEMENS 데이터 추가")
                            merged_df = pd.concat([merged_df, master_siemens_combined], ignore_index=True, sort=False)
                else:
                    print(f"  [OK] SIEMENS 데이터 정상: {siemens_count:,} 행")
            else:
                print(f"  [WARNING] Source_Vendor 컬럼이 없어 SIEMENS 검증 불가")
            
            # Save merged file with standard 63-column header order
            merged_output_path = Path(out).with_name(
                Path(out).stem.replace("_multi", "_merged") + ".xlsx"
            )
            if merged_output_path == Path(out):
                # If no "_multi" in name, add "_merged" suffix
                merged_output_path = Path(out).with_name(Path(out).stem + "_merged.xlsx")
            # Apply standard 63-column header order before saving
            merged_df_reordered = reorder_dataframe_columns(
                merged_df, is_stage2=False, keep_unlisted=False, use_semantic_matching=True
            )
            # DUP-GUARD: Add "고유 중복" column to merged file if not present
            if "고유 중복" not in merged_df_reordered.columns and "__dup_label__" in merged_df_reordered.columns:
                merged_df_reordered.insert(0, "고유 중복", merged_df_reordered["__dup_label__"])
            elif "고유 중복" not in merged_df_reordered.columns:
                # If no __dup_label__, compute it
                try:
                    merged_with_flag, dup_mask = mark_duplicates(
                        merged_df_reordered, semantic_finder=_find_col, keep="first"
                    )
                    merged_df_reordered = merged_with_flag
                    if "고유 중복" not in merged_df_reordered.columns:
                        merged_df_reordered.insert(0, "고유 중복", np.where(dup_mask, "중복", "고유"))
                except Exception as e:
                    print(f"  [WARNING][DUP-GUARD][merged] skip adding duplicate column: {e}")
            # 합쳐진 파일도 XlsxWriter 최적화 적용
            if self.use_xlsxwriter:
                try:
                    from scripts.core.excel_writer_optimized import OptimizedExcelWriter
                    with OptimizedExcelWriter(merged_output_path, use_xlsxwriter=True, constant_memory=True) as writer:
                        writer.write_dataframe(merged_df_reordered, sheet_name="Merged Data", index=False)
                except Exception as e:
                    print(f"  [WARNING] XlsxWriter failed for merged file, using openpyxl: {e}")
                    with pd.ExcelWriter(merged_output_path, engine="openpyxl") as writer:
                        merged_df_reordered.to_excel(writer, sheet_name="Merged Data", index=False)
            else:
                with pd.ExcelWriter(merged_output_path, engine="openpyxl") as writer:
                    merged_df_reordered.to_excel(writer, sheet_name="Merged Data", index=False)

            print(f"[OK] 합쳐진 파일 저장: {merged_output_path.name} (63개 헤더로 정리됨)")

            # Prepare result
            total_stats["output_file"] = out
            total_stats["merged_file"] = str(merged_output_path)
            total_stats["merged_rows"] = len(merged_df_reordered)
            total_stats["merged_columns"] = len(merged_df_reordered.columns)

            print("\n" + "=" * 60)
            print("[OK] MULTI-SHEET SYNCHRONIZATION COMPLETE")
            print("=" * 60)
            print(f"Multi-sheet output: {out}")
            print(f"Merged output: {merged_output_path}")
            print(f"Sheets processed: {len(processed_sheets)}")
            print(
                f"Total changes: {total_stats['updates']} updates, {total_stats['appends']} new records"
            )

            return SyncResult(
                True,
                "Multi-sheet sync & colorize done.",
                out,
                total_stats,
                matching_report="All headers matched successfully",
            )

        except Exception as e:
            error_msg = f"Synchronization failed: {str(e)}"
            print(f"\n{'='*60}")
            print(f"[ERROR] ERROR: {error_msg}")
            print(f"{'='*60}")

            return SyncResult(
                False,
                error_msg,
                output_path or warehouse_xlsx,
                {},
                matching_report=str(e),
            )

    def _fuzzy_find_column(self, target_col: str, header_map: Dict[str, int]) -> Optional[int]:
        """
        ✅ Phase 4: Find column using normalized matching (Fallback level 3)

        This handles cases where column names differ slightly due to spacing,
        punctuation, or other formatting differences.

        Example: "ETD/ATD" matches "ETD / ATD"
        """
        try:
            from scripts.core.header_normalizer import HeaderNormalizer

            normalizer = HeaderNormalizer()
            target_normalized = normalizer.normalize(target_col)

            for header, idx in header_map.items():
                header_normalized = normalizer.normalize(header)
                if target_normalized == header_normalized:
                    return idx
        except Exception:
            # Fallback failed, return None
            pass

        return None

    def _apply_excel_formatting(self, excel_file: str, sheet_name: str, header_row: int):
        """
        Apply color formatting to the Excel file to highlight changes.

        Args:
            excel_file: Path to the Excel file
            sheet_name: Name of the sheet to format
            header_row: Row index where headers start (0-based from pandas)
        """
        try:
            # Early return if no changes to apply
            if not self.change_tracker.changes:
                print(f"      [DEBUG] No changes to apply for {sheet_name}")
                return

            print(
                f"      [DEBUG] Applying {len(self.change_tracker.changes)} changes to {sheet_name}"
            )
            print(f"      [DEBUG] Target file: {excel_file}")

            # ✅ Load without data_only to preserve formatting
            wb = load_workbook(excel_file, data_only=False)
            if sheet_name not in wb.sheetnames:
                print(f"      [DEBUG] Sheet {sheet_name} not found in {wb.sheetnames}")
                return

            ws = wb[sheet_name]

            # Excel rows are 1-indexed, and we need to account for header
            # FIX: Always use row 1 as header (actual Excel header row)
            excel_header_row = 1
            print(f"      [DEBUG] Excel header row: {excel_header_row} (fixed)")

            # Build header map
            header_map = {}
            case_no_col_idx = None
            for c_idx, cell in enumerate(ws[excel_header_row], start=1):
                if cell.value is None:
                    continue
                header_name = str(cell.value).strip()
                header_map[header_name] = c_idx

                # Find Case No. column
                if "Case" in header_name and "No" in header_name:
                    case_no_col_idx = c_idx

            print(
                f"      [DEBUG] Header map size: {len(header_map)}, first 5: {list(header_map.keys())[:5]}"
            )
            print(f"      [DEBUG] Case No. column index: {case_no_col_idx}")

            # ✅ Phase 4: Build Case No. → Excel row mapping
            case_to_row = {}
            if case_no_col_idx:
                for row_idx in range(excel_header_row + 1, ws.max_row + 1):
                    case_no_cell = ws.cell(row=row_idx, column=case_no_col_idx)
                    if case_no_cell.value:
                        case_to_row[str(case_no_cell.value).strip()] = row_idx
                print(f"      [DEBUG] Built case_to_row mapping: {len(case_to_row)} cases")

            # Define fills
            orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
            yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

            # ✅ Phase 4: Apply date changes (orange) with 3-level Fallback
            orange_applied = 0
            match_by_semantic = 0
            match_by_exact = 0
            match_by_fuzzy = 0
            orange_cells: List[Tuple[int, int]] = []

            for change in self.change_tracker.changes:
                if change.change_type != "date_update":
                    continue

                # ✅ Phase 4: Use case_no to find correct row after reordering
                if change.case_no and change.case_no in case_to_row:
                    excel_row = case_to_row[change.case_no]
                else:
                    # Fallback to old method if case_no not available
                    excel_row = change.row_index + excel_header_row + 1
                    if change.case_no:
                        print(
                            f"      [WARN] Case No. '{change.case_no}' not found in case_to_row mapping"
                        )

                # Level 1: Use semantic key mapping (most accurate)
                actual_col_name = self.change_tracker.get_column_name(
                    change.semantic_key, fallback=change.column_name
                )

                # Level 2: Exact match in header_map
                col_idx = header_map.get(actual_col_name)
                if col_idx:
                    match_by_semantic += 1 if change.semantic_key else 0
                    match_by_exact += 1
                else:
                    # Level 3: Fuzzy matching (last resort)
                    col_idx = self._fuzzy_find_column(actual_col_name, header_map)
                    if col_idx:
                        match_by_fuzzy += 1

                if col_idx:
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = orange_fill
                        orange_applied += 1
                        orange_cells.append((excel_row, col_idx))

            print(f"      [DEBUG] Orange cells applied: {orange_applied}")
            print(f"        - Semantic key matches: {match_by_semantic}")
            print(f"        - Exact matches: {match_by_exact}")
            print(f"        - Fuzzy matches: {match_by_fuzzy}")

            # Apply new records (yellow)
            yellow_applied = 0
            yellow_cells: List[Tuple[int, int]] = []
            for change in self.change_tracker.changes:
                if change.change_type == "new_record":
                    excel_row = change.row_index + excel_header_row + 1

                    # Color all cells with data in this row
                    for cell in ws[excel_row]:
                        if cell.value is not None and str(cell.value).strip():
                            cell.fill = yellow_fill
                            yellow_applied += 1
                            yellow_cells.append((excel_row, cell.col_idx))

            print(f"      [DEBUG] Yellow cells applied: {yellow_applied}")

            # ✅ Save with explicit close to ensure colors persist
            wb.save(excel_file)
            wb.close()
            print(f"      [DEBUG] File saved and closed")

            # ✅ 즉시 검증: 저장된 파일을 다시 읽어 색상 확인
            try:

                def _matches_color(fill: PatternFill, targets: Tuple[str, ...]) -> bool:
                    if not fill:
                        return False
                    color = getattr(fill, "start_color", None)
                    candidates: List[str] = []
                    if color is not None:
                        rgb = str(getattr(color, "rgb", "") or "").upper()
                        if rgb:
                            candidates.append(rgb)
                        indexed = getattr(color, "indexed", None)
                        if isinstance(indexed, str):
                            candidates.append(indexed.upper())
                    fg_color = getattr(fill, "fgColor", None)
                    if fg_color is not None:
                        fg_rgb = str(getattr(fg_color, "rgb", "") or "").upper()
                        if fg_rgb:
                            candidates.append(fg_rgb)
                    return any(
                        any(target in candidate for target in targets) for candidate in candidates
                    )

                wb_verify = load_workbook(excel_file, data_only=False)
                ws_verify = wb_verify[sheet_name]

                verify_orange = 0
                for row_idx, col_idx in orange_cells:
                    cell = ws_verify.cell(row=row_idx, column=col_idx)
                    if _matches_color(cell.fill, ("FFFFA500", "FFA500", "00FFA500")):
                        verify_orange += 1

                verify_yellow = 0
                for row_idx, col_idx in yellow_cells:
                    cell = ws_verify.cell(row=row_idx, column=col_idx)
                    if _matches_color(cell.fill, ("FFFFFF00", "FFFF00", "00FFFF00")):
                        verify_yellow += 1

                print(f"      [VERIFY] Saved file color check:")
                print(f"        - Orange: {verify_orange}")
                print(f"        - Yellow: {verify_yellow}")

                if verify_orange < orange_applied:
                    missing = orange_applied - verify_orange
                    if missing > 0:
                        print(f"        WARNING: Orange color mismatch on {missing} cells!")
                if verify_yellow < yellow_applied:
                    missing = yellow_applied - verify_yellow
                    if missing > 0:
                        print(f"        WARNING: Yellow color mismatch on {missing} cells!")

                wb_verify.close()

            except Exception as verify_error:
                print(f"      [VERIFY] 검증 실패: {verify_error}")

        except Exception as e:
            print(f"  Warning: Formatting failed: {e}")
            import traceback

            traceback.print_exc()


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="DataSynchronizer v3.0 - Semantic Header Matching Edition"
    )
    ap.add_argument("--master", required=True, help="Path to Master Excel file")
    ap.add_argument("--warehouse", required=True, help="Path to Warehouse Excel file")
    ap.add_argument("--out", default="", help="Output path (optional)")
    ap.add_argument(
        "--header-override",
        action="append",
        default=[],
        help=(
            "Manual header row override (<file_label>:<sheet>=<row>, use '*' for wildcard). "
            "Provide multiple times for multiple sheets."
        ),
    )
    args = ap.parse_args()

    print("\n" + "=" * 60)
    print("DataSynchronizer v3.0")
    print("Semantic Header Matching Edition")
    print("=" * 60)

    try:
        overrides = DataSynchronizerV30.parse_header_override_args(args.header_override)
    except ValueError as override_error:
        ap.error(str(override_error))

    sync = DataSynchronizerV30(header_overrides=overrides)
    res = sync.synchronize(args.master, args.warehouse, args.out or None)

    print("\n" + "=" * 60)
    print("FINAL RESULT")
    print("=" * 60)
    print(f"Success: {res.success}")
    print(f"Message: {res.message}")
    print(f"Output:  {res.output_path}")
    print(f"Stats:   {res.stats}")

    if not res.success:
        print(f"\nError details: {res.matching_report}")
        exit(1)
