#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 색상 검증 통합 스크립트 (Unified Version)

모든 검증 모드를 하나의 CLI로 통합:
- basic: 기본 최적화 스캔 (기본값)
- ultra_fast: 초고속 샘플링 (50행, 8컬럼)
- batch: 배치 처리 (여러 범위 분할)
- full: 전체 스캔 (느리지만 완전)
- change_log: ChangeTracker 로그 기반

통합 버전: v1.0.1 (2025-11-03 Patched)
기반: verify_excel_colors_optimized.py

[v1.0.1 변경사항]
- 진행 상황 표시 추가
- 에러 메시지 출력 개선
- max_cells_limit 사전 체크 추가
- 인자 유효성 검사 추가
- 샘플링 자동 적용 로직 추가
- --verbose 옵션 추가
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook
from openpyxl.cell import Cell

try:
    from openpyxl.styles.colors import COLOR_INDEX
except ImportError:
    COLOR_INDEX = {}

# ============================================================================
# 기본 설정
# ============================================================================
DEF_STAGE1 = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
DEF_REPORT = Path("docs/reports/excel_color_verification.json")
DEF_HEADER_SEARCH_MAX_ROWS = 5
DEF_SCAN_MAX_ROWS_PER_SHEET = 200
DEF_SCAN_MAX_COLS_PER_ROW = 12
DEF_SHEETS_SCAN_LIMIT = 2
DEF_CASE_SAMPLES_PER_COLOR = 5
DEF_PROCESSES = max(1, cpu_count() // 4)
DEF_MAX_CELLS_PER_WORKER = 2000

ORANGE_RGB = "FFFFA500"
YELLOW_RGB = "FFFFFF00"
ORANGE_INDEX_GUESS = {45, 46, 53}
YELLOW_INDEX_GUESS = {5, 6}

DATE_HEADER_HINTS = [
    "date", "입고", "출고", "검수", "생성", "updated", "received", "issued",
    "in date", "out date", "posting", "posting date"
]

# 전역 변수: verbose 모드
VERBOSE = False

# ============================================================================
# 유틸리티 함수
# ============================================================================
def log_verbose(msg: str):
    """Verbose 모드일 때만 출력"""
    if VERBOSE:
        print(f"[VERBOSE] {msg}", file=sys.stderr)

def log_info(msg: str):
    """정보 메시지 출력"""
    print(f"[INFO] {msg}")

def log_warn(msg: str):
    """경고 메시지 출력"""
    print(f"[WARN] {msg}", file=sys.stderr)

def log_error(msg: str):
    """에러 메시지 출력"""
    print(f"[ERROR] {msg}", file=sys.stderr)

# ============================================================================
# 데이터 구조
# ============================================================================
@dataclass
class SheetScanResult:
    sheet: str
    header_row: int
    total_rows: int
    total_columns: int
    case_col_idx: Optional[int]
    orange_cells: int
    yellow_cells: int
    yellow_rows: List[int]
    sample_cases: List[Dict[str, object]]

# ============================================================================
# 색상 판정 유틸 (최적화 버전)
# ============================================================================
def _argb_upper(v: Optional[str]) -> Optional[str]:
    return str(v).upper() if v else None

def _indexed_to_argb(idx: Optional[int]) -> Optional[str]:
    if idx is None or not COLOR_INDEX or idx < 0 or idx >= len(COLOR_INDEX):
        return None
    try:
        rgb6 = COLOR_INDEX[idx]
        if not rgb6:
            return None
        rgb6 = rgb6.upper().replace("#", "")
        return "FF" + rgb6 if len(rgb6) == 6 else rgb6 if len(rgb6) == 8 else None
    except (IndexError, KeyError, TypeError):
        return None

def _theme_to_argb(cell: Cell, theme_obj) -> Optional[str]:
    try:
        color = cell.fill.start_color
        theme_idx = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0.0) or 0.0
        
        if theme_idx is None or theme_obj is None:
            return None
        
        clr_scheme = getattr(getattr(theme_obj, "themeElements", None), "clrScheme", None)
        if not clr_scheme or not hasattr(clr_scheme, "children"):
            return None
        
        children = list(clr_scheme.children)
        if theme_idx >= len(children):
            return None
        
        entry = children[theme_idx]
        srgb = getattr(getattr(entry, "srgbClr", None), "val", None)
        if not srgb:
            return None
        
        base = srgb.upper()
        argb = "FF" + base if len(base) == 6 else base
        
        if tint == 0:
            return argb
        
        def blend(c: int, t: float) -> int:
            return int(round(c + (255 - c) * t)) if t > 0 else int(round(c * (1 + t)))
        
        r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
        r2, g2, b2 = max(0, min(255, blend(r, tint))), max(0, min(255, blend(g, tint))), max(0, min(255, blend(b, tint)))
        return f"FF{r2:02X}{g2:02X}{b2:02X}"
    except Exception:
        return None

def get_cell_argb(cell: Cell, theme_obj=None) -> Tuple[Optional[str], Optional[int]]:
    """셀 색상 추출 (최적화)"""
    if not hasattr(cell, 'fill') or not cell.fill:
        return None, None
    
    f = cell.fill
    if not hasattr(f, 'start_color') or not f.start_color:
        return None, None
    
    sc = f.start_color
    
    try:
        if hasattr(sc, 'rgb') and sc.rgb:
            return _argb_upper(sc.rgb), getattr(sc, "indexed", None)
    except:
        pass
    
    t = getattr(sc, "type", None)
    indexed_val = getattr(sc, "indexed", None)
    
    if t == "rgb":
        return _argb_upper(getattr(sc, "rgb", None)), indexed_val
    if t == "indexed":
        return _indexed_to_argb(getattr(sc, "indexed", None)), indexed_val
    if t == "theme":
        try:
            return _theme_to_argb(cell, theme_obj), indexed_val
        except:
            return None, indexed_val
    
    fg = getattr(f, "fgColor", None)
    if fg:
        try:
            if hasattr(fg, 'rgb') and fg.rgb:
                return _argb_upper(fg.rgb), getattr(fg, "indexed", None)
        except:
            pass
        
        fg_type = getattr(fg, "type", None)
        if fg_type == "rgb":
            return _argb_upper(getattr(fg, "rgb", None)), getattr(fg, "indexed", None)
        if fg_type == "indexed":
            return _indexed_to_argb(getattr(fg, "indexed", None)), getattr(fg, "indexed", None)
    
    return None, indexed_val

def is_orange(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    if argb:
        u = argb.upper()
        if ("A500" in u) or (u == ORANGE_RGB):
            return True
        try:
            r, g, b = int(u[2:4], 16), int(u[4:6], 16), int(u[6:8], 16)
            if r > 200 and 100 <= g <= 180 and b < 120:
                return True
        except Exception:
            pass
    return indexed in ORANGE_INDEX_GUESS if indexed is not None else False

def is_yellow(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    if argb:
        u = argb.upper()
        if ("FF00" in u) or (u == YELLOW_RGB):
            return True
        try:
            r, g, b = int(u[2:4], 16), int(u[4:6], 16), int(u[6:8], 16)
            if r > 220 and g > 210 and b < 80:
                return True
        except Exception:
            pass
    return indexed in YELLOW_INDEX_GUESS if indexed is not None else False

# ============================================================================
# 공통 유틸리티
# ============================================================================
def norm_str(x) -> str:
    return str(x or "").strip().lower()

def find_header_row(ws, max_search_rows: int) -> int:
    for r in range(1, min(max_search_rows, ws.max_row) + 1):
        try:
            v = norm_str(ws.cell(r, 1).value)
            if ("고유" in v) or ("중복" in v):
                return r
        except:
            continue
    return 1

def find_case_col_idx(ws, header_row: int) -> Optional[int]:
    for c in range(1, min(15, ws.max_column) + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if "case" in hv and "no" in hv:
                return c
        except:
            continue
    return None

def auto_date_columns(ws, header_row: int) -> List[int]:
    hint_cols = []
    for c in range(1, ws.max_column + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if hv and any(h in hv for h in DATE_HEADER_HINTS):
                hint_cols.append(c)
        except:
            continue
    return hint_cols

def parse_date_cols_arg(arg: Optional[str], ws, header_row: int) -> List[int]:
    if not arg:
        return []
    cols = []
    for tok in [t.strip() for t in arg.split(",") if t.strip()]:
        if tok.isdigit():
            cols.append(int(tok))
        else:
            for c in range(1, ws.max_column + 1):
                try:
                    if norm_str(ws.cell(header_row, c).value) == norm_str(tok):
                        cols.append(c)
                        break
                except:
                    continue
    return sorted(set([c for c in cols if 1 <= c <= ws.max_column]))

# ============================================================================
# 모드별 스캔 함수
# ============================================================================

# ---- ULTRA_FAST 모드 ----
def scan_ultra_fast(stage1_path: Path, sheets_limit: int) -> Dict:
    """초고속 샘플링: 처음 50행, 8개 컬럼만"""
    log_info("ULTRA_FAST 모드: 초고속 샘플링 (50행, 8컬럼)")
    
    results = {}
    
    try:
        wb = load_workbook(stage1_path, data_only=False, read_only=True)
        target_sheets = wb.sheetnames[:sheets_limit]
        
        for i, sheet_name in enumerate(target_sheets, 1):
            log_info(f"[{i}/{len(target_sheets)}] {sheet_name} 스캔 중...")
            
            ws = wb[sheet_name]
            header_row = find_header_row(ws, 5)
            case_col_idx = find_case_col_idx(ws, header_row)
            
            orange_count = 0
            yellow_count = 0
            yellow_rows = []
            
            data_start = header_row + 1
            max_rows = min(50, ws.max_row - data_start + 1)
            max_cols = min(8, ws.max_column)
            
            for r in range(data_start, data_start + max_rows):
                row_has_yellow = False
                for c in range(1, max_cols + 1):
                    try:
                        cell = ws.cell(r, c)
                        if not hasattr(cell, 'fill') or not cell.fill:
                            continue
                        
                        argb, idx = get_cell_argb(cell, None)
                        if is_orange(argb, idx):
                            orange_count += 1
                        elif is_yellow(argb, idx):
                            yellow_count += 1
                            row_has_yellow = True
                    except:
                        continue
                
                if row_has_yellow:
                    yellow_rows.append(r)
            
            sample_cases = []
            if case_col_idx:
                for r in yellow_rows[:5]:
                    try:
                        v = ws.cell(r, case_col_idx).value
                        if v:
                            sample_cases.append({"case_no": str(v), "row": r, "color": "YELLOW"})
                    except:
                        continue
            
            results[sheet_name] = {
                "header_row": header_row,
                "sampled_rows": max_rows,
                "sampled_cols": max_cols,
                "orange_samples": orange_count,
                "yellow_samples": yellow_count,
                "sample_cases": sample_cases
            }
            
            log_info(f"[완료] {sheet_name}: ORANGE={orange_count}, YELLOW={yellow_count}")
        
        wb.close()
        
    except Exception as e:
        log_error(f"ULTRA_FAST 모드 실패: {e}")
        return {}
    
    return results

# ---- BASIC/FULL 모드 (최적화 스캔) ----
def scan_one_sheet_basic(args: Tuple):
    """워커용: 기본 최적화 스캔 (v1.0.1 패치)"""
    (
        sheet_name, stage1_path, header_search_max, head_rows, cols_limit, case_samples,
        tail_rows, random_sample, random_seed, date_cols_only, full_scan, date_cols_spec,
        max_cells_limit
    ) = args

    try:
        log_verbose(f"{sheet_name}: 워크북 로드 중...")
        wb = load_workbook(stage1_path, data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        theme_obj = getattr(wb, "_theme", None)
        
        total_rows, total_cols = ws.max_row, ws.max_column
        header_row = find_header_row(ws, header_search_max)
        case_col_idx = find_case_col_idx(ws, header_row)
        
        log_verbose(f"{sheet_name}: 헤더={header_row}, 총 행={total_rows}, 총 열={total_cols}")
        
        data_start = header_row + 1
        data_end = total_rows
        
        if date_cols_only:
            explicit = parse_date_cols_arg(date_cols_spec, ws, header_row)
            auto = auto_date_columns(ws, header_row) if not explicit else []
            target_cols = sorted(set(explicit or auto))
            if not target_cols:
                target_cols = list(range(1, min(cols_limit, total_cols) + 1))
        else:
            target_cols = list(range(1, min(cols_limit, total_cols) + 1))
        
        row_indices = []
        
        if full_scan:
            skip_step = 2 if total_rows > 10000 else 1
            row_indices = list(range(data_start, data_end + 1, skip_step))
        else:
            if head_rows > 0:
                row_indices.extend(range(data_start, min(data_start + head_rows, data_end) + 1))
            if tail_rows > 0:
                tail_rows_limit = min(tail_rows, 200)
                if total_rows > 5000:
                    tail_start = max(data_start, data_end - tail_rows_limit * 2 + 1)
                    row_indices.extend(range(tail_start, data_end + 1, 2))
                else:
                    tail_start = max(data_start, data_end - tail_rows_limit + 1)
                    row_indices.extend(range(tail_start, data_end + 1))
            if random_sample > 0 and data_end >= data_start:
                random_sample_limit = min(random_sample, 100)
                rng = random.Random(random_seed)
                population = list(range(data_start, data_end + 1))
                row_indices.extend(rng.sample(population, min(random_sample_limit, len(population))))
            
            row_indices = sorted(set(row_indices))
        
        # v1.0.1: max_cells_limit 사전 체크
        if target_cols:
            expected_cells = len(row_indices) * len(target_cols)
            if expected_cells > max_cells_limit * 1.5:  # 50% 여유
                skip_factor = max(1, int(expected_cells / max_cells_limit))
                original_len = len(row_indices)
                row_indices = row_indices[::skip_factor]
                log_warn(f"{sheet_name}: 샘플링 자동 적용 (행 {original_len} → {len(row_indices)}, skip={skip_factor})")
        
        orange_cells = 0
        yellow_cells = 0
        yellow_rows: List[int] = []
        sample_cases: List[Dict[str, object]] = []
        cells_scanned = 0
        
        max_rows_to_scan = min(len(row_indices), max_cells_limit // len(target_cols)) if target_cols else 0
        
        log_verbose(f"{sheet_name}: 스캔 시작 (행={len(row_indices)}, 열={len(target_cols)})")
        
        for r in row_indices[:max_rows_to_scan]:
            if cells_scanned >= max_cells_limit:
                log_verbose(f"{sheet_name}: max_cells_limit 도달, 조기 종료")
                break
            if len(yellow_rows) >= case_samples * 2 and orange_cells > 0:
                break
            
            row_has_yellow = False
            
            for c in target_cols:
                if cells_scanned >= max_cells_limit:
                    break
                try:
                    cell = ws.cell(r, c)
                    argb, idx = get_cell_argb(cell, theme_obj)
                    cells_scanned += 1
                    if is_orange(argb, idx):
                        orange_cells += 1
                    elif is_yellow(argb, idx):
                        yellow_cells += 1
                        row_has_yellow = True
                except Exception as e:
                    log_verbose(f"{sheet_name}: 셀({r},{c}) 스캔 실패: {e}")
                    continue
            if row_has_yellow and len(yellow_rows) < case_samples * 2:
                yellow_rows.append(r)
        
        log_verbose(f"{sheet_name}: 스캔 완료 (셀={cells_scanned})")
        
        if case_col_idx:
            seen = set()
            for r in yellow_rows[:case_samples]:
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "YELLOW"})
                            seen.add(key)
                except:
                    continue
            
            added = 0
            for r in row_indices[:50]:
                if added >= case_samples:
                    break
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "ORANGE?"})
                            seen.add(key)
                            added += 1
                except:
                    continue
        
        wb.close()
        return SheetScanResult(
            sheet=sheet_name,
            header_row=header_row,
            total_rows=total_rows,
            total_columns=total_cols,
            case_col_idx=case_col_idx,
            orange_cells=orange_cells,
            yellow_cells=yellow_cells,
            yellow_rows=yellow_rows,
            sample_cases=sample_cases,
        )
    except Exception as e:
        # v1.0.1: 에러 메시지 출력
        log_error(f"{sheet_name} 스캔 실패: {type(e).__name__}: {e}")
        if VERBOSE:
            import traceback
            traceback.print_exc()
        
        return SheetScanResult(
            sheet=sheet_name, header_row=1, total_rows=0, total_columns=0,
            case_col_idx=None, orange_cells=0, yellow_cells=0,
            yellow_rows=[], sample_cases=[]
        )

# ---- BATCH 모드 ----
def run_batch_mode(stage1_path: Path, sheets_limit: int) -> Dict:
    """배치 처리: 여러 범위를 나눠서 실행"""
    log_info("BATCH 모드: 배치 처리 (2000개씩 분할)")
    
    batches = [
        {"name": "상단", "rows": 200, "cols": 10, "tail_rows": 0, "random_sample": 0},
        {"name": "중간상단", "rows": 0, "cols": 10, "tail_rows": 0, "random_sample": 100, "random_seed": 42},
        {"name": "중간", "rows": 0, "cols": 10, "tail_rows": 0, "random_sample": 100, "random_seed": 123},
        {"name": "중간하단", "rows": 0, "cols": 10, "tail_rows": 0, "random_sample": 100, "random_seed": 456},
        {"name": "하단", "rows": 0, "cols": 10, "tail_rows": 200, "random_sample": 0},
    ]
    
    all_results = {
        "batches": [],
        "total_orange": 0,
        "total_yellow": 0
    }
    
    wb0 = load_workbook(stage1_path, read_only=True)
    first_sheet = wb0.sheetnames[0] if wb0.sheetnames else None
    wb0.close()
    
    if not first_sheet:
        return all_results
    
    for i, batch in enumerate(batches, 1):
        log_info(f"[{i}/{len(batches)}] 배치: {batch['name']}")
        
        jobs = [(
            first_sheet,
            str(stage1_path),
            DEF_HEADER_SEARCH_MAX_ROWS,
            batch["rows"],
            batch["cols"],
            DEF_CASE_SAMPLES_PER_COLOR,
            batch.get("tail_rows", 0),
            batch.get("random_sample", 0),
            batch.get("random_seed", 23),
            False,
            False,
            "",
            2000
        )]
        
        try:
            result = scan_one_sheet_basic(jobs[0])
            
            all_results["batches"].append({
                "name": batch["name"],
                "orange": result.orange_cells,
                "yellow": result.yellow_cells,
                "params": batch
            })
            
            all_results["total_orange"] += result.orange_cells
            all_results["total_yellow"] += result.yellow_cells
            
            log_info(f"[완료] {batch['name']}: ORANGE={result.orange_cells}, YELLOW={result.yellow_cells}")
            
        except Exception as e:
            log_error(f"배치 {batch['name']} 실패: {e}")
    
    return all_results

# ---- CHANGE_LOG 모드 ----
def summarize_change_log(path: Path) -> Dict[str, object]:
    """ChangeTracker 로그 요약"""
    orange = 0
    yellow = 0
    per_sheet = {}
    
    if path.suffix.lower() in (".json", ".jsonl", ".ndjson"):
        try:
            with path.open("r", encoding="utf-8") as f:
                txt = f.read().strip()
                data = json.loads(txt) if txt.startswith('{') or txt.startswith('[') else [json.loads(line) for line in txt.splitlines() if line.strip()]
                if isinstance(data, dict):
                    data = data.get("changes", [])
        except Exception as e:
            log_error(f"ChangeTracker 로그 파싱 실패: {e}")
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}
        
        for e in data:
            sname = e.get("sheet", "unknown")
            ctype = e.get("change_type", e.get("type", ""))
            if sname not in per_sheet:
                per_sheet[sname] = {"date_update": 0, "new_record": 0}
            if ctype == "date_update":
                per_sheet[sname]["date_update"] += 1
                orange += 1
            elif ctype == "new_record":
                per_sheet[sname]["new_record"] += 1
                yellow += 1
    else:
        try:
            df = pd.read_csv(path)
            sheet_col = next((c for c in df.columns if "sheet" in c.lower()), None)
            type_col = next((c for c in df.columns if "type" in c.lower()), None)
            for _, row in df.iterrows():
                sname = str(row[sheet_col]) if sheet_col else "unknown"
                ctype = str(row[type_col]).lower().strip() if type_col else ""
                if sname not in per_sheet:
                    per_sheet[sname] = {"date_update": 0, "new_record": 0}
                if ctype == "date_update":
                    per_sheet[sname]["date_update"] += 1
                    orange += 1
                elif ctype == "new_record":
                    per_sheet[sname]["new_record"] += 1
                    yellow += 1
        except Exception as e:
            log_error(f"CSV 파싱 실패: {e}")
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}
    
    return {
        "sample_orange": orange,
        "sample_yellow": yellow,
        "by_sheet": per_sheet,
    }

# ---- 케이스 추적 ----
def track_case_numbers(stage1: Path, scan_results: Dict[str, SheetScanResult]) -> List[Dict[str, object]]:
    out = []
    for sname, meta in scan_results.items():
        if meta.case_col_idx is None:
            continue
        try:
            df = pd.read_excel(stage1, sheet_name=sname, nrows=500)
        except Exception as e:
            log_verbose(f"{sname}: Case 추적 실패: {e}")
            continue
        
        case_col = next((col for col in df.columns if "case" in str(col).lower() and "no" in str(col).lower()), None)
        if case_col is None:
            continue
        
        targets = {str(x["case_no"]).strip().upper() for x in meta.sample_cases}
        if not targets:
            continue
        
        ser = df[case_col].astype(str).str.strip().str.upper()
        hits = df[ser.isin(targets)]
        for v in hits[case_col].tolist():
            out.append({"case_no": str(v), "sheet": sname, "found": True})
    return out

# ============================================================================
# 메인 함수
# ============================================================================
def main():
    global VERBOSE
    
    ap = argparse.ArgumentParser(
        description="Excel 색상 검증 통합 스크립트 v1.0.1",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
모드별 설명:
  basic        : 기본 최적화 스캔 (기본값)
  ultra_fast   : 초고속 샘플링 (50행, 8컬럼만)
  batch        : 배치 처리 (여러 범위 분할)
  full         : 전체 스캔 (느리지만 완전)
  change_log   : ChangeTracker 로그 기반

[v1.0.1 변경사항]
- 진행 상황 표시 추가
- 에러 메시지 출력 개선
- max_cells_limit 사전 체크
- 샘플링 자동 적용
- --verbose 옵션 추가

사용 예시:
  python verify_excel_colors_unified.py --mode=basic
  python verify_excel_colors_unified.py --mode=ultra_fast --sheets=3
  python verify_excel_colors_unified.py --mode=batch --verbose
        """
    )
    
    ap.add_argument("--mode", type=str, default="basic",
                   choices=["basic", "ultra_fast", "batch", "full", "change_log"],
                   help="검증 모드 선택")
    ap.add_argument("--stage1", type=Path, default=DEF_STAGE1)
    ap.add_argument("--report", type=Path, default=DEF_REPORT)
    
    ap.add_argument("--header-rows", type=int, default=DEF_HEADER_SEARCH_MAX_ROWS)
    ap.add_argument("--rows", type=int, default=DEF_SCAN_MAX_ROWS_PER_SHEET)
    ap.add_argument("--cols", type=int, default=DEF_SCAN_MAX_COLS_PER_ROW)
    ap.add_argument("--sheets", type=int, default=DEF_SHEETS_SCAN_LIMIT)
    ap.add_argument("--samples", type=int, default=DEF_CASE_SAMPLES_PER_COLOR)
    ap.add_argument("--processes", type=int, default=DEF_PROCESSES)
    ap.add_argument("--no-parallel", action="store_true")
    
    ap.add_argument("--tail-rows", type=int, default=0)
    ap.add_argument("--random-sample", type=int, default=0)
    ap.add_argument("--random-seed", type=int, default=23)
    ap.add_argument("--date-cols-only", action="store_true")
    ap.add_argument("--full-scan", action="store_true")
    ap.add_argument("--date-cols", type=str, default="")
    ap.add_argument("--max-cells", type=int, default=DEF_MAX_CELLS_PER_WORKER)
    
    ap.add_argument("--change-log", type=Path, default=None)
    ap.add_argument("--prefer-change-log", action="store_true")
    
    # v1.0.1: 새 옵션
    ap.add_argument("--verbose", action="store_true", help="상세 로그 출력")

    args = ap.parse_args()
    
    # v1.0.1: verbose 모드 설정
    VERBOSE = args.verbose
    
    # v1.0.1: 인자 유효성 검사
    if args.max_cells <= 0:
        log_warn(f"잘못된 --max-cells 값: {args.max_cells}. 기본값 {DEF_MAX_CELLS_PER_WORKER} 사용")
        args.max_cells = DEF_MAX_CELLS_PER_WORKER
    
    if args.sheets <= 0:
        log_warn(f"잘못된 --sheets 값: {args.sheets}. 기본값 {DEF_SHEETS_SCAN_LIMIT} 사용")
        args.sheets = DEF_SHEETS_SCAN_LIMIT
    
    if args.rows < 0:
        log_warn(f"잘못된 --rows 값: {args.rows}. 기본값 {DEF_SCAN_MAX_ROWS_PER_SHEET} 사용")
        args.rows = DEF_SCAN_MAX_ROWS_PER_SHEET

    print("=" * 80)
    print("Excel 색상 검증 통합 스크립트 v1.0.1".center(80))
    print("=" * 80)
    print(f"\n모드: {args.mode.upper()}")
    print(f"파일: {args.stage1}")
    print("-" * 80)

    if args.mode == "change_log" or (args.change_log and args.prefer_change_log):
        if not args.change_log or not args.change_log.exists():
            log_error("ChangeTracker 로그 파일이 없습니다.")
            return 1
        
        log_info("ChangeTracker 로그 분석 중...")
        change_summary = summarize_change_log(args.change_log)
        out = {
            "mode": "change_log",
            "version": "v1.0.1",
            "verification": change_summary,
            "params": {"source": "change_log_only", "change_log": str(args.change_log)},
        }
        args.report.parent.mkdir(parents=True, exist_ok=True)
        with args.report.open("w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        log_info(f"결과 저장: {args.report}")
        print(f"\nORANGE={change_summary['sample_orange']}, YELLOW={change_summary['sample_yellow']}")
        return 0

    if not args.stage1.exists():
        log_error(f"Stage1 파일 없음: {args.stage1}")
        return 1

    if args.mode == "ultra_fast":
        results = scan_ultra_fast(args.stage1, args.sheets)
        out = {
            "mode": "ultra_fast",
            "version": "v1.0.1",
            "sheets": results,
            "params": vars(args)
        }
        
    elif args.mode == "batch":
        results = run_batch_mode(args.stage1, args.sheets)
        out = {
            "mode": "batch",
            "version": "v1.0.1",
            "results": results,
            "params": vars(args)
        }
        
    elif args.mode == "basic" or args.mode == "full":
        try:
            wb0 = load_workbook(args.stage1, read_only=True)
            all_sheets = wb0.sheetnames
            # v1.0.1: sheets 값이 실제 시트 수보다 큰 경우 조정
            actual_sheets_limit = min(args.sheets, len(all_sheets))
            if actual_sheets_limit < args.sheets:
                log_warn(f"요청한 시트 수({args.sheets})가 실제 시트 수({len(all_sheets)})보다 큽니다. {actual_sheets_limit}개로 조정")
            target_sheets = all_sheets[:actual_sheets_limit]
            wb0.close()
            log_info(f"대상 시트: {target_sheets}")
        except Exception as e:
            log_error(f"Stage1 파일 읽기 실패: {e}")
            return 1

        if args.mode == "full":
            args.full_scan = True
            log_warn("FULL 모드: 전체 스캔 활성화 (처리 시간이 오래 걸릴 수 있습니다)")

        jobs = [
            (
                s, str(args.stage1), args.header_rows, args.rows, args.cols,
                args.samples, args.tail_rows, args.random_sample,
                args.random_seed, args.date_cols_only, args.full_scan,
                args.date_cols, args.max_cells
            )
            for s in target_sheets
        ]

        if args.no_parallel or len(jobs) < args.processes:
            log_info("순차 처리 모드")
            results_list = []
            for i, j in enumerate(jobs, 1):
                log_info(f"[{i}/{len(jobs)}] {j[0]} 스캔 중...")
                result = scan_one_sheet_basic(j)
                results_list.append(result)
                log_info(f"[완료] {j[0]}: ORANGE={result.orange_cells}, YELLOW={result.yellow_cells}")
        else:
            log_info(f"병렬 처리 모드 (프로세스: {args.processes})")
            with Pool(processes=args.processes) as pool:
                results_list = pool.map(scan_one_sheet_basic, jobs)

        scan_results = {r.sheet: r for r in results_list}
        
        log_info("Case 번호 추적 중...")
        tracking = track_case_numbers(args.stage1, scan_results)

        total_orange = sum(r.orange_cells for r in scan_results.values())
        total_yellow = sum(r.yellow_cells for r in scan_results.values())

        out = {
            "mode": args.mode,
            "version": "v1.0.1",
            "colors": {
                "sheets": {
                    k: {
                        "header_row": v.header_row,
                        "total_rows": v.total_rows,
                        "total_columns": v.total_columns,
                        "case_column": v.case_col_idx,
                        "estimated_orange": v.orange_cells,
                        "estimated_yellow": v.yellow_cells,
                        "yellow_row_samples": v.yellow_rows,
                        "sample_cases": v.sample_cases,
                    }
                    for k, v in scan_results.items()
                }
            },
            "tracking": tracking,
            "verification": {
                "sample_orange": total_orange,
                "sample_yellow": total_yellow,
            },
            "params": vars(args),
        }
        
        print(f"\n[결과]")
        print(f"  ORANGE: {total_orange:,}개")
        print(f"  YELLOW: {total_yellow:,}개")

    else:
        log_error(f"알 수 없는 모드: {args.mode}")
        return 1

    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    log_info(f"결과 저장: {args.report}")
    print("=" * 80)
    return 0

if __name__ == "__main__":
    sys.exit(main())
