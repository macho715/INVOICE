#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
컬러 작업 로직 엑셀 파일 빠른 확인 스크립트 (최적화 버전)
Raw 파일부터 케이스 번호 전체 과정 추적 및 색상 확인
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional
import sys
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))

def print_section(title: str, char="="):
    """섹션 제목 출력"""
    print(f"\n{char * 80}")
    print(f"{title:^80}")
    print(f"{char * 80}\n")

# 색상 정의
ORANGE = "FFFFA500"
YELLOW = "FFFFFF00"

def get_cell_color_fast(cell) -> Optional[str]:
    """셀의 색상 정보 빠르게 추출 (최적화)"""
    if not cell.fill:
        return None
    
    try:
        # start_color 우선 확인
        if hasattr(cell.fill, 'start_color'):
            color = cell.fill.start_color
            rgb = getattr(color, 'rgb', None)
            if rgb:
                return str(rgb).upper()
        
        # fgColor 확인
        if hasattr(cell.fill, 'fgColor'):
            fg_color = cell.fill.fgColor
            fg_rgb = getattr(fg_color, 'rgb', None)
            if fg_rgb:
                return str(fg_rgb).upper()
    except:
        pass
    
    return None

def is_orange_fast(color: Optional[str]) -> bool:
    """ORANGE 색상인지 빠르게 확인"""
    if not color:
        return False
    return "A500" in color or "FFFFA500" in color

def is_yellow_fast(color: Optional[str]) -> bool:
    """YELLOW 색상인지 빠르게 확인"""
    if not color:
        return False
    return "FF00" in color or "FFFF00" in color

def analyze_raw_files_fast():
    """Raw 파일 시트 구조 빠른 분석 (최적화)"""
    print_section("1. Raw 파일 분석 (빠른 스캔)")
    
    results = {
        'master_file': {},
        'warehouse_file': {}
    }
    
    # Master 파일
    master_file = Path("data/raw/Case List(HE).xlsx")
    if master_file.exists():
        print(f"[1] Master 파일: {master_file.name}")
        print("-" * 80)
        
        master_xl = pd.ExcelFile(master_file)
        print(f"  시트 목록: {master_xl.sheet_names}")
        
        for sheet in master_xl.sheet_names[:2]:  # 최대 2개 시트만
            try:
                df = pd.read_excel(master_file, sheet_name=sheet, header=4, nrows=100)
                
                case_col = None
                for col in df.columns[:10]:  # 처음 10개 컬럼만 체크
                    if 'case' in str(col).lower() and 'no' in str(col).lower():
                        case_col = col
                        break
                
                results['master_file'][sheet] = {
                    'case_column': case_col,
                    'sample_rows': len(df)
                }
                
                print(f"\n  시트: '{sheet}'")
                print(f"    - Case No 컬럼: {case_col or 'N/A'}")
            except Exception as e:
                print(f"    [ERROR] {e}")
    
    # Warehouse 파일
    warehouse_file = Path("data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx")
    if warehouse_file.exists():
        print(f"\n[2] Warehouse 파일: {warehouse_file.name}")
        print("-" * 80)
        
        warehouse_xl = pd.ExcelFile(warehouse_file)
        print(f"  시트 목록: {warehouse_xl.sheet_names}")
        
        for sheet in warehouse_xl.sheet_names[:3]:  # 최대 3개 시트만
            try:
                # 원본 행수만 빠르게 확인
                raw_df = pd.read_excel(warehouse_file, sheet_name=sheet, header=None, nrows=1)
                
                # 전체 행수는 ExcelFile로 확인
                full_xl = pd.ExcelFile(warehouse_file)
                sheet_obj = full_xl.sheet_names.index(sheet)
                # 실제로는 파일을 다시 열어야 하지만, 빠른 확인을 위해 건너뜀
                
                results['warehouse_file'][sheet] = {
                    'has_data': True
                }
                
                print(f"\n  시트: '{sheet}'")
                print(f"    - 데이터 있음")
            except Exception as e:
                print(f"    [ERROR] {e}")
    
    return results

def read_excel_colors_fast():
    """Stage 1 출력 엑셀 파일에서 실제 색상 빠르게 읽기 (최적화)"""
    print_section("2. Stage 1 출력 파일 색상 확인 (빠른 스캔)")
    
    results = {
        'sheets': {},
        'color_stats': defaultdict(int)
    }
    
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if not stage1_file.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {stage1_file}")
        return None
    
    print(f"[1] 파일 로드: {stage1_file.name}")
    print("-" * 80)
    
    wb = load_workbook(stage1_file, data_only=False, read_only=True)  # read_only로 빠르게
    
    print(f"  시트 목록: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames[:2]:  # 최대 2개 시트만
        print(f"\n[2] 시트 분석: '{sheet_name}' (빠른 샘플링)")
        print("-" * 80)
        
        ws = wb[sheet_name]
        
        # 헤더 행 빠르게 찾기
        header_row = 1
        for row_idx in range(1, min(5, ws.max_row + 1)):
            try:
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value and ('고유' in str(cell_value) or '중복' in str(cell_value)):
                    header_row = row_idx
                    break
            except:
                continue
        
        print(f"  헤더 행: {header_row}")
        print(f"  총 행수: {ws.max_row}")
        print(f"  총 컬럼: {ws.max_column}")
        
        # Case No 컬럼 빠르게 찾기
        case_col_idx = None
        case_col_name = None
        for col_idx in range(1, min(15, ws.max_column + 1)):  # 처음 15개 컬럼만
            try:
                header_cell = ws.cell(row=header_row, column=col_idx)
                header_value = str(header_cell.value or '').lower().strip()
                if 'case' in header_value and 'no' in header_value:
                    case_col_idx = col_idx
                    case_col_name = header_cell.value
                    break
            except:
                continue
        
        print(f"  Case No 컬럼: {case_col_name} (컬럼 {get_column_letter(case_col_idx) if case_col_idx else 'N/A'})")
        
        # 빠른 색상 스캔: 처음 200행만 + 샘플링
        orange_samples = []
        yellow_samples = []
        yellow_row_samples = []
        
        data_start_row = header_row + 1
        max_scan_rows = min(200, ws.max_row - data_start_row + 1)  # 최대 200행만
        
        print(f"  빠른 스캔: 처음 {max_scan_rows}행만 확인")
        
        orange_count = 0
        yellow_count = 0
        
        # 제한된 컬럼만 스캔 (처음 10개 컬럼)
        max_cols_to_scan = min(10, ws.max_column)
        
        for row_idx in range(data_start_row, min(data_start_row + max_scan_rows, ws.max_row + 1)):
            row_has_yellow = False
            row_orange_cols = []
            
            for col_idx in range(1, max_cols_to_scan + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    color = get_cell_color_fast(cell)
                    
                    if is_orange_fast(color):
                        orange_count += 1
                        if len(orange_samples) < 10:
                            orange_samples.append((row_idx, col_idx))
                    elif is_yellow_fast(color):
                        yellow_count += 1
                        row_has_yellow = True
                        if len(yellow_samples) < 10:
                            yellow_samples.append((row_idx, col_idx))
                except:
                    continue
            
            if row_has_yellow and len(yellow_row_samples) < 10:
                yellow_row_samples.append(row_idx)
        
        # Case No 추출 (샘플만)
        sample_cases = []
        if case_col_idx:
            for row_idx in yellow_row_samples[:5]:
                try:
                    case_cell = ws.cell(row=row_idx, column=case_col_idx)
                    if case_cell.value:
                        sample_cases.append({
                            'case_no': str(case_cell.value),
                            'row': row_idx,
                            'color': 'YELLOW'
                        })
                except:
                    continue
            
            for row_idx, _ in orange_samples[:5]:
                try:
                    case_cell = ws.cell(row=row_idx, column=case_col_idx)
                    if case_cell.value:
                        sample_cases.append({
                            'case_no': str(case_cell.value),
                            'row': row_idx,
                            'color': 'ORANGE'
                        })
                except:
                    continue
        
        results['sheets'][sheet_name] = {
            'header_row': header_row,
            'total_rows': ws.max_row,
            'total_columns': ws.max_column,
            'case_column': case_col_idx,
            'orange_samples': len(orange_samples),
            'yellow_samples': len(yellow_row_samples),
            'sample_cases': sample_cases[:10],
            'estimated_orange': orange_count,
            'estimated_yellow': yellow_count
        }
        
        print(f"\n  색상 샘플링 결과:")
        print(f"    - ORANGE 셀 샘플: {len(orange_samples)}개 (추정: {orange_count}개)")
        print(f"    - YELLOW 행 샘플: {len(yellow_row_samples)}개 (추정: {yellow_count}개)")
        
        if sample_cases:
            print(f"\n  샘플 Case No:")
            for item in sample_cases[:5]:
                print(f"    - Case No: {item['case_no']}, 행: {item['row']}, 색상: {item['color']}")
    
    wb.close()
    
    return results

def track_case_numbers_fast(raw_results, color_results):
    """케이스 번호 빠른 추적"""
    print_section("3. 케이스 번호 추적 (빠른 확인)")
    
    sample_cases = []
    
    for sheet_name, sheet_data in color_results['sheets'].items():
        for case_info in sheet_data.get('sample_cases', [])[:3]:  # 최대 3개만
            sample_cases.append({
                'case_no': case_info['case_no'],
                'sheet': sheet_name,
                'color': case_info['color']
            })
    
    if not sample_cases:
        print("[INFO] 추적할 샘플 케이스가 없습니다.")
        return None
    
    print(f"[1] 추적 대상: {len(sample_cases)}개 케이스")
    print("-" * 80)
    
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    tracking_results = []
    
    for case_info in sample_cases:
        case_no = case_info['case_no']
        sheet_name = case_info['sheet']
        
        print(f"\n  Case No: {case_no} ({case_info['color']})")
        
        try:
            # pandas로 빠르게 읽기
            stage1_df = pd.read_excel(stage1_file, sheet_name=sheet_name, nrows=1000)
            
            case_col = None
            for col in stage1_df.columns:
                if 'case' in str(col).lower() and 'no' in str(col).lower():
                    case_col = col
                    break
            
            if case_col:
                matches = stage1_df[
                    stage1_df[case_col].astype(str).str.strip().str.upper() == case_no.strip().upper()
                ]
                
                if len(matches) > 0:
                    row = matches.iloc[0]
                    print(f"    - 발견: 시트 '{sheet_name}', 행 {matches.index[0] + 1}")
                    
                    tracking_results.append({
                        'case_no': case_no,
                        'sheet': sheet_name,
                        'color': case_info['color'],
                        'found': True
                    })
        except Exception as e:
            print(f"    [ERROR] {e}")
    
    return tracking_results

def verify_color_logic_fast():
    """컬러 작업 로직 빠른 검증"""
    print_section("4. 컬러 작업 로직 검증")
    
    print("[1] 색상 적용 규칙 (코드 기준)")
    print("-" * 80)
    print("""
    data_synchronizer_v30.py 기준:
    
    1. ORANGE (FFFFA500): 날짜 컬럼 변경
       - change_type == "date_update"
       - Line 2528-2573
       - 해당 셀에만 색상 적용
    
    2. YELLOW (FFFFFF00): 신규 레코드
       - change_type == "new_record"
       - Line 2578-2590
       - 해당 행의 모든 데이터 셀에 색상 적용
    """)
    
    # 실제 파일에서 빠른 샘플 확인
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if stage1_file.exists():
        wb = load_workbook(stage1_file, data_only=False, read_only=True)
        
        total_orange = 0
        total_yellow = 0
        
        for sheet_name in wb.sheetnames[:1]:  # 첫 번째 시트만
            ws = wb[sheet_name]
            
            # 헤더 찾기
            header_row = 1
            for r in range(1, min(5, ws.max_row + 1)):
                if '고유' in str(ws.cell(r, 1).value or ''):
                    header_row = r
                    break
            
            # 빠른 샘플링: 처음 100행만
            data_start = header_row + 1
            max_rows = min(100, ws.max_row - data_start + 1)
            
            for row_idx in range(data_start, data_start + max_rows):
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    try:
                        cell = ws.cell(row_idx, col_idx)
                        color = get_cell_color_fast(cell)
                        
                        if is_orange_fast(color):
                            total_orange += 1
                        elif is_yellow_fast(color):
                            total_yellow += 1
                    except:
                        continue
        
        wb.close()
        
        print(f"\n[2] 빠른 샘플 검증 (처음 100행 기준)")
        print("-" * 80)
        print(f"  샘플 ORANGE 셀: {total_orange}개")
        print(f"  샘플 YELLOW 셀: {total_yellow}개")
        print(f"\n  [참고] 전체 통계는 실제 파일을 열어서 확인하세요.")
        
        return {
            'sample_orange': total_orange,
            'sample_yellow': total_yellow
        }
    
    return None

def main():
    """메인 실행 함수 (최적화)"""
    print("=" * 80)
    print("컬러 작업 로직 빠른 확인 (최적화 버전)".center(80))
    print("=" * 80)
    
    all_results = {}
    
    try:
        # 1. Raw 파일 빠른 분석
        all_results['raw_files'] = analyze_raw_files_fast()
        
        # 2. 색상 빠른 읽기
        all_results['colors'] = read_excel_colors_fast()
        
        # 3. 케이스 번호 빠른 추적
        if all_results.get('colors'):
            all_results['tracking'] = track_case_numbers_fast(
                all_results['raw_files'],
                all_results['colors']
            )
        
        # 4. 컬러 로직 빠른 검증
        all_results['verification'] = verify_color_logic_fast()
        
        # 결과 저장
        import json
        results_file = Path("docs/reports/excel_color_verification_fast.json")
        results_file.parent.mkdir(parents=True, exist_ok=True)
        
        json_results = {}
        for key, value in all_results.items():
            if value is None:
                continue
            if isinstance(value, dict):
                json_results[key] = {
                    k: v for k, v in value.items()
                    if not isinstance(v, (pd.DataFrame, defaultdict))
                }
        
        with open(results_file, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n[OK] 검증 결과 저장: {results_file}")
        
        print("\n" + "=" * 80)
        print("빠른 검증 완료".center(80))
        print("=" * 80)
        
        return all_results
        
    except Exception as e:
        import traceback
        print(f"\n[ERROR] 검증 중 오류 발생: {e}")
        print(traceback.format_exc())
        return None

if __name__ == "__main__":
    main()

