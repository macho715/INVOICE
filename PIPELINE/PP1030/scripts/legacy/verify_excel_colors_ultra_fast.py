#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
컬러 작업 로직 초고속 확인 스크립트
문제점 해결: ws.max_row 호출과 전체 셀 스캔 제거
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
import sys

sys.path.insert(0, str(Path(__file__).parent))

def print_section(title: str, char="="):
    """섹션 제목 출력"""
    print(f"\n{char * 80}")
    print(f"{title:^80}")
    print(f"{char * 80}\n")

# 색상 정의
ORANGE_TARGETS = ["A500", "FFFFA500", "FFA500"]
YELLOW_TARGETS = ["FF00", "FFFF00", "FFFFFF00"]

def get_cell_color_ultra_fast(cell) -> Optional[str]:
    """최적화된 색상 추출 - 예외 처리 최소화"""
    try:
        fill = cell.fill
        if not fill:
            return None
        
        # start_color만 빠르게 확인
        if hasattr(fill, 'start_color'):
            rgb = getattr(fill.start_color, 'rgb', None)
            if rgb:
                return str(rgb).upper()
        return None
    except:
        return None

def is_colored_fast(color: Optional[str], targets: List[str]) -> bool:
    """색상 매칭 (최적화)"""
    if not color:
        return False
    return any(t in color for t in targets)

def analyze_from_pandas():
    """Pandas로 먼저 구조 확인 (매우 빠름)"""
    print_section("1. Raw 파일 구조 확인 (Pandas)")
    
    results = {}
    
    # Master 파일
    master_file = Path("data/raw/Case List(HE).xlsx")
    if master_file.exists():
        print(f"[1] Master 파일: {master_file.name}")
        try:
            df = pd.read_excel(master_file, sheet_name="Case List", header=4, nrows=10)
            case_col = next((c for c in df.columns if 'case' in str(c).lower() and 'no' in str(c).lower()), None)
            print(f"  - Case No 컬럼: {case_col}")
            results['master'] = {'case_column': case_col}
        except Exception as e:
            print(f"  [ERROR] {e}")
    
    # Warehouse 파일
    warehouse_file = Path("data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx")
    if warehouse_file.exists():
        print(f"\n[2] Warehouse 파일: {warehouse_file.name}")
        try:
            xl = pd.ExcelFile(warehouse_file)
            print(f"  - 시트 목록: {xl.sheet_names}")
            results['warehouse'] = {'sheets': xl.sheet_names}
        except Exception as e:
            print(f"  [ERROR] {e}")
    
    return results

def check_colors_minimal():
    """최소한의 색상 확인만 (초고속)"""
    print_section("2. Stage 1 출력 색상 확인 (초고속 샘플링)")
    
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if not stage1_file.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {stage1_file}")
        return None
    
    print(f"[1] 파일: {stage1_file.name}")
    print("-" * 80)
    
    # read_only + 첫 번째 시트만 + 매우 작은 샘플
    try:
        wb = load_workbook(stage1_file, data_only=False, read_only=True)
        sheet_name = wb.sheetnames[0]  # 첫 번째 시트만
        print(f"  분석 시트: '{sheet_name}'")
        
        ws = wb[sheet_name]
        
        # 헤더 행 찾기 (최대 5행만)
        header_row = 1
        for r in [1, 2, 3, 4, 5]:
            try:
                val = str(ws.cell(r, 1).value or '')
                if '고유' in val or '중복' in val:
                    header_row = r
                    break
            except:
                continue
        
        print(f"  헤더 행: {header_row}")
        
        # Case No 컬럼 찾기 (최대 10개 컬럼만)
        case_col_idx = None
        for c in range(1, 11):
            try:
                val = str(ws.cell(header_row, c).value or '').lower()
                if 'case' in val and 'no' in val:
                    case_col_idx = c
                    print(f"  Case No 컬럼: 컬럼 {get_column_letter(c)}")
                    break
            except:
                continue
        
        # 초고속 샘플링: 처음 50행만, 처음 8개 컬럼만
        print(f"\n[2] 초고속 샘플링 (처음 50행, 8개 컬럼만)")
        print("-" * 80)
        
        orange_found = []
        yellow_found = []
        
        data_start = header_row + 1
        max_rows = 50  # 고정
        
        for row in range(data_start, min(data_start + max_rows, data_start + 50)):
            row_has_yellow = False
            
            # 최대 8개 컬럼만 체크
            for col in range(1, min(9, 9)):
                try:
                    cell = ws.cell(row, col)
                    
                    # fill 속성만 빠르게 체크
                    if not hasattr(cell, 'fill') or not cell.fill:
                        continue
                    
                    color = get_cell_color_ultra_fast(cell)
                    
                    if is_colored_fast(color, ORANGE_TARGETS):
                        if len(orange_found) < 5:
                            orange_found.append((row, col))
                    elif is_colored_fast(color, YELLOW_TARGETS):
                        if not row_has_yellow:
                            yellow_found.append(row)
                            row_has_yellow = True
                except:
                    continue
        
        # Case No 추출
        sample_cases = []
        if case_col_idx:
            for row in yellow_found[:3]:
                try:
                    case_val = ws.cell(row, case_col_idx).value
                    if case_val:
                        sample_cases.append({
                            'case_no': str(case_val),
                            'row': row,
                            'color': 'YELLOW'
                        })
                except:
                    pass
        
        wb.close()
        
        print(f"  샘플 결과:")
        print(f"    - ORANGE 셀 샘플: {len(orange_found)}개")
        print(f"    - YELLOW 행 샘플: {len(yellow_found)}개")
        
        if sample_cases:
            print(f"\n  샘플 Case No:")
            for item in sample_cases:
                print(f"    - {item['case_no']} (행 {item['row']}, {item['color']})")
        
        return {
            'sheet': sheet_name,
            'header_row': header_row,
            'case_column': case_col_idx,
            'orange_samples': len(orange_found),
            'yellow_samples': len(yellow_found),
            'sample_cases': sample_cases
        }
        
    except Exception as e:
        import traceback
        print(f"[ERROR] {e}")
        print(traceback.format_exc())
        return None

def track_case_quick():
    """케이스 번호 빠른 추적"""
    print_section("3. 케이스 번호 추적 (빠른 확인)")
    
    # pandas로 먼저 읽어서 샘플 찾기
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if not stage1_file.exists():
        return None
    
    try:
        # 첫 번째 시트만, 처음 1000행만
        df = pd.read_excel(stage1_file, sheet_name=0, nrows=1000)
        
        case_col = next((c for c in df.columns if 'case' in str(c).lower() and 'no' in str(c).lower()), None)
        
        if case_col:
            print(f"[1] Case No 컬럼: {case_col}")
            print(f"    샘플 Case No (처음 10개):")
            sample_cases = df[case_col].dropna().head(10).tolist()
            for case in sample_cases:
                print(f"      - {case}")
            
            return {'case_column': case_col, 'samples': sample_cases}
    except Exception as e:
        print(f"[ERROR] {e}")
    
    return None

def verify_color_logic_quick():
    """컬러 로직 빠른 검증"""
    print_section("4. 컬러 작업 로직 검증")
    
    print("[1] 색상 적용 규칙 (코드 기준)")
    print("-" * 80)
    print("""
    코드 위치: scripts/stage1_sync_sorted/data_synchronizer_v30.py
    
    1. ORANGE (FFFFA500):
       - Line 2528-2573: _apply_excel_formatting()
       - change_type == "date_update"
       - 해당 셀에만 색상 적용
    
    2. YELLOW (FFFFFF00):
       - Line 2578-2590: _apply_excel_formatting()
       - change_type == "new_record"
       - 해당 행의 모든 데이터 셀에 색상 적용
    
    3. 적용 순서:
       - Line 2056-2067: 각 시트별로 _apply_excel_formatting() 호출
       - 시트별 ChangeTracker 사용 (sheet_change_trackers)
    """)
    
    # 실제 파일 빠른 확인
    color_results = check_colors_minimal()
    
    if color_results:
        print(f"\n[2] 실제 파일 검증 결과")
        print("-" * 80)
        print(f"  시트: {color_results['sheet']}")
        print(f"  ORANGE 셀 샘플: {color_results['orange_samples']}개")
        print(f"  YELLOW 행 샘플: {color_results['yellow_samples']}개")
        print(f"  샘플 Case No: {len(color_results.get('sample_cases', []))}개")
    
    return color_results

def main():
    """메인 실행 함수 (초고속)"""
    print("=" * 80)
    print("컬러 작업 로직 초고속 확인".center(80))
    print("=" * 80)
    
    all_results = {}
    
    try:
        # 1. Pandas로 빠른 구조 확인
        all_results['structure'] = analyze_from_pandas()
        
        # 2. 최소한의 색상 확인 (초고속)
        all_results['colors'] = check_colors_minimal()
        
        # 3. 케이스 번호 빠른 추적
        all_results['tracking'] = track_case_quick()
        
        # 4. 컬러 로직 검증
        all_results['verification'] = verify_color_logic_quick()
        
        print("\n" + "=" * 80)
        print("초고속 검증 완료".center(80))
        print("=" * 80)
        
        return all_results
        
    except Exception as e:
        import traceback
        print(f"\n[ERROR] {e}")
        print(traceback.format_exc())
        return None

if __name__ == "__main__":
    main()

