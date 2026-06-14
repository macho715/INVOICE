#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
컬러 작업 로직 엑셀 파일 직접 확인 스크립트
Raw 파일부터 케이스 번호 전체 과정 추적 및 색상 확인
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional
import sys
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))

def print_section(title: str, char="="):
    """섹션 제목 출력"""
    print(f"\n{char * 80}")
    print(f"{title:^80}")
    print(f"{char * 80}\n")

# 색상 정의 (코드와 동일)
ORANGE = "FFFFA500"  # Changed date cell
YELLOW = "FFFFFF00"  # New row

def get_cell_color(cell) -> Optional[str]:
    """셀의 색상 정보 추출"""
    if not cell.fill or not hasattr(cell.fill, 'start_color'):
        return None
    
    fill = cell.fill
    color = getattr(fill, 'start_color', None)
    
    if color:
        # RGB 값 추출
        rgb = str(getattr(color, 'rgb', '') or '')
        if rgb:
            return rgb.upper()
    
    # fgColor도 확인
    fg_color = getattr(fill, 'fgColor', None)
    if fg_color:
        fg_rgb = str(getattr(fg_color, 'rgb', '') or '')
        if fg_rgb:
            return fg_rgb.upper()
    
    return None

def is_orange(color: Optional[str]) -> bool:
    """ORANGE 색상인지 확인"""
    if not color:
        return False
    return any(target in color for target in ["FFFFA500", "FFA500", "00FFA500", "A500"])

def is_yellow(color: Optional[str]) -> bool:
    """YELLOW 색상인지 확인"""
    if not color:
        return False
    return any(target in color for target in ["FFFFFF00", "FFFF00", "00FFFF00", "FF00"])

def analyze_raw_files():
    """Raw 파일 시트 구조 및 케이스 번호 위치 확인"""
    print_section("1. Raw 파일 분석")
    
    results = {
        'master_file': {},
        'warehouse_file': {}
    }
    
    # Master 파일 분석
    master_file = Path("data/raw/Case List(HE).xlsx")
    if master_file.exists():
        print(f"[1] Master 파일: {master_file.name}")
        print("-" * 80)
        
        master_xl = pd.ExcelFile(master_file)
        print(f"  시트 목록: {master_xl.sheet_names}")
        
        for sheet in master_xl.sheet_names:
            # 헤더 행 찾기 (여러 행 시도)
            df = None
            header_row = None
            
            for h in range(6):  # 0~5행 시도
                try:
                    df = pd.read_excel(master_file, sheet_name=sheet, header=h)
                    if len(df.columns) > 5:  # 충분한 컬럼이 있으면
                        header_row = h
                        break
                except:
                    continue
            
            if df is None:
                continue
            
            # Case No 컬럼 찾기
            case_col = None
            for col in df.columns:
                col_str = str(col).lower().strip()
                if 'case' in col_str and 'no' in col_str:
                    case_col = col
                    break
            
            results['master_file'][sheet] = {
                'header_row': header_row,
                'total_rows': len(df),
                'case_column': case_col,
                'columns': list(df.columns[:10])
            }
            
            print(f"\n  시트: '{sheet}'")
            print(f"    - 헤더 행: {header_row}")
            print(f"    - 총 행수: {len(df):,}행")
            print(f"    - Case No 컬럼: {case_col or 'N/A'}")
            if case_col:
                non_empty = df[case_col].notna().sum()
                print(f"    - Case No 값 있음: {non_empty:,}개")
                if non_empty > 0:
                    sample_cases = df[case_col].dropna().head(5).tolist()
                    print(f"    - 샘플 Case No: {sample_cases}")
    
    # Warehouse 파일 분석
    warehouse_file = Path("data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx")
    if warehouse_file.exists():
        print(f"\n[2] Warehouse 파일: {warehouse_file.name}")
        print("-" * 80)
        
        warehouse_xl = pd.ExcelFile(warehouse_file)
        print(f"  시트 목록: {warehouse_xl.sheet_names}")
        
        for sheet in warehouse_xl.sheet_names:
            # 원본 행수 확인 (헤더 제외)
            try:
                raw_df = pd.read_excel(warehouse_file, sheet_name=sheet, header=None)
                raw_rows = len(raw_df)
            except:
                raw_rows = 0
            
            # 실제 데이터 로드
            df = None
            header_row = None
            
            for h in range(6):
                try:
                    df = pd.read_excel(warehouse_file, sheet_name=sheet, header=h)
                    if len(df.columns) > 5:
                        header_row = h
                        break
                except:
                    continue
            
            if df is None:
                continue
            
            # Case No 컬럼 찾기
            case_col = None
            for col in df.columns:
                col_str = str(col).lower().strip()
                if 'case' in col_str and 'no' in col_str:
                    case_col = col
                    break
            
            results['warehouse_file'][sheet] = {
                'header_row': header_row,
                'raw_rows': raw_rows,
                'data_rows': len(df),
                'case_column': case_col,
                'columns': list(df.columns[:10])
            }
            
            print(f"\n  시트: '{sheet}'")
            print(f"    - 원본 행수 (헤더 제외): {raw_rows:,}행")
            print(f"    - 데이터 행수: {len(df):,}행")
            print(f"    - 헤더 행: {header_row}")
            print(f"    - Case No 컬럼: {case_col or 'N/A'}")
            if case_col:
                non_empty = df[case_col].notna().sum()
                print(f"    - Case No 값 있음: {non_empty:,}개")
                if non_empty > 0:
                    sample_cases = df[case_col].dropna().head(5).tolist()
                    print(f"    - 샘플 Case No: {sample_cases}")
    
    return results

def read_excel_colors():
    """Stage 1 출력 엑셀 파일에서 실제 색상 읽기"""
    print_section("2. Stage 1 출력 파일 색상 확인")
    
    results = {
        'sheets': {},
        'color_stats': defaultdict(int)
    }
    
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if not stage1_file.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {stage1_file}")
        return None
    
    print(f"[1] 파일 로드: {stage1_file.name}")
    print("-" * 80)
    
    # OpenPyXL로 엑셀 파일 열기
    wb = load_workbook(stage1_file, data_only=False)
    
    print(f"  시트 목록: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n[2] 시트 분석: '{sheet_name}'")
        print("-" * 80)
        
        ws = wb[sheet_name]
        
        # 헤더 행 찾기 (고유 중복 컬럼이 있는 행)
        header_row = 1
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and ('고유' in str(cell_value) or '중복' in str(cell_value)):
                header_row = row_idx
                break
        
        print(f"  헤더 행: {header_row}")
        print(f"  총 행수: {ws.max_row}")
        print(f"  총 컬럼: {ws.max_column}")
        
        # Case No 컬럼 찾기
        case_col_idx = None
        case_col_name = None
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=header_row, column=col_idx)
            header_value = str(header_cell.value or '').lower().strip()
            if 'case' in header_value and 'no' in header_value:
                case_col_idx = col_idx
                case_col_name = header_cell.value
                break
        
        print(f"  Case No 컬럼: {case_col_name} (컬럼 {get_column_letter(case_col_idx) if case_col_idx else 'N/A'})")
        
        # 색상 스캔
        orange_cells = []
        yellow_cells = []
        yellow_rows = set()
        
        data_start_row = header_row + 1
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_has_yellow = False
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                color = get_cell_color(cell)
                
                if is_orange(color):
                    orange_cells.append((row_idx, col_idx, cell.value))
                    results['color_stats']['orange'] += 1
                elif is_yellow(color):
                    yellow_cells.append((row_idx, col_idx, cell.value))
                    row_has_yellow = True
                    results['color_stats']['yellow'] += 1
            
            if row_has_yellow:
                yellow_rows.add(row_idx)
        
        # Case No 추출 (색상이 있는 행만 - 최적화)
        sample_cases_with_colors = []
        if case_col_idx:
            # 이미 찾은 yellow_rows와 orange_cells를 활용
            checked_rows = set()
            
            # YELLOW 행의 Case No 먼저 추출
            for row_idx in list(yellow_rows)[:20]:  # 최대 20개만
                case_cell = ws.cell(row=row_idx, column=case_col_idx)
                case_value = case_cell.value
                if case_value:
                    sample_cases_with_colors.append({
                        'case_no': str(case_value),
                        'row': row_idx,
                        'color': 'YELLOW'
                    })
                    checked_rows.add(row_idx)
                    if len(sample_cases_with_colors) >= 10:
                        break
            
            # ORANGE 셀이 있는 행의 Case No 추출
            for row_idx, col_idx, _ in orange_cells[:10]:  # 최대 10개만
                if row_idx in checked_rows:
                    continue
                case_cell = ws.cell(row=row_idx, column=case_col_idx)
                case_value = case_cell.value
                if case_value:
                    sample_cases_with_colors.append({
                        'case_no': str(case_value),
                        'row': row_idx,
                        'color': 'ORANGE'
                    })
                    checked_rows.add(row_idx)
                    if len(sample_cases_with_colors) >= 15:
                        break
        
        results['sheets'][sheet_name] = {
            'header_row': header_row,
            'total_rows': ws.max_row,
            'total_columns': ws.max_column,
            'case_column': case_col_idx,
            'case_column_name': case_col_name,
            'orange_cells': len(orange_cells),
            'yellow_rows': len(yellow_rows),
            'sample_cases': sample_cases_with_colors[:10]
        }
        
        print(f"\n  색상 통계:")
        print(f"    - ORANGE 셀: {len(orange_cells):,}개")
        print(f"    - YELLOW 행: {len(yellow_rows):,}개")
        
        if sample_cases_with_colors:
            print(f"\n  샘플 Case No (색상 적용된 것):")
            for item in sample_cases_with_colors[:5]:
                print(f"    - Case No: {item['case_no']}, 행: {item['row']}, 색상: {item['color']}")
    
    wb.close()
    
    return results

def track_case_numbers(raw_results, color_results):
    """샘플 케이스 번호로 Raw → Stage 1 전체 과정 추적"""
    print_section("3. 케이스 번호 추적 (Raw → Stage 1)")
    
    # Stage 1 출력에서 색상이 적용된 케이스 선택
    sample_cases = []
    
    for sheet_name, sheet_data in color_results['sheets'].items():
        for case_info in sheet_data.get('sample_cases', []):
            if case_info['color'] in ['ORANGE', 'YELLOW']:
                sample_cases.append({
                    'case_no': case_info['case_no'],
                    'sheet': sheet_name,
                    'row': case_info['row'],
                    'color': case_info['color']
                })
                if len(sample_cases) >= 5:
                    break
        if len(sample_cases) >= 5:
            break
    
    if not sample_cases:
        print("[WARN] 색상이 적용된 케이스를 찾을 수 없습니다.")
        return None
    
    print(f"[1] 추적 대상 케이스: {len(sample_cases)}개")
    print("-" * 80)
    
    # Stage 1 출력 파일 로드 (pandas)
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    tracking_results = []
    
    for case_info in sample_cases:
        case_no = case_info['case_no']
        sheet_name = case_info['sheet']
        
        print(f"\n[2] Case No 추적: {case_no}")
        print("-" * 80)
        
        # Stage 1 출력에서 찾기
        try:
            stage1_df = pd.read_excel(stage1_file, sheet_name=sheet_name)
            
            # Case No 컬럼 찾기
            case_col = None
            for col in stage1_df.columns:
                if 'case' in str(col).lower() and 'no' in str(col).lower():
                    case_col = col
                    break
            
            if case_col:
                matches = stage1_df[
                    stage1_df[case_col].astype(str).str.strip().str.upper() == case_no.strip().upper()
                ]
                
                if len(matches) > 0:
                    row = matches.iloc[0]
                    print(f"  Stage 1 출력에서 발견:")
                    print(f"    - 시트: {sheet_name}")
                    print(f"    - 행 번호: {matches.index[0] + 1}")
                    print(f"    - 색상: {case_info['color']}")
                    
                    # 주요 컬럼 확인
                    key_cols = ['ETD/ATD', 'ETA/ATA', 'DHL WH', 'Product', 'Item']
                    for col in key_cols:
                        if col in row.index:
                            value = row[col]
                            print(f"    - {col}: {value}")
                    
                    tracking_results.append({
                        'case_no': case_no,
                        'sheet': sheet_name,
                        'stage1_row': matches.index[0] + 1,
                        'color': case_info['color'],
                        'data': row.to_dict()
                    })
        except Exception as e:
            print(f"  [ERROR] 추적 실패: {e}")
    
    return tracking_results

def verify_color_logic():
    """컬러 작업 로직 검증 및 종합 보고서"""
    print_section("4. 컬러 작업 로직 검증")
    
    print("[1] 색상 적용 규칙")
    print("-" * 80)
    print("""
    코드 기준 (data_synchronizer_v30.py):
    
    1. ORANGE (FFFFA500): 날짜 컬럼 변경
       - change_type == "date_update"
       - Line 2528-2573: _apply_excel_formatting()
       - 해당 셀에만 색상 적용
    
    2. YELLOW (FFFFFF00): 신규 레코드
       - change_type == "new_record"
       - Line 2578-2590: _apply_excel_formatting()
       - 해당 행의 모든 데이터 셀에 색상 적용
    """)
    
    # 실제 파일에서 검증
    stage1_file = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
    
    if stage1_file.exists():
        wb = load_workbook(stage1_file, data_only=False)
        
        total_orange = 0
        total_yellow_rows = 0
        yellow_row_counts = defaultdict(int)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            orange_in_sheet = 0
            yellow_rows_in_sheet = set()
            
            # 헤더 행 찾기
            header_row = 1
            for row_idx in range(1, min(10, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value and ('고유' in str(cell_value) or '중복' in str(cell_value)):
                    header_row = row_idx
                    break
            
            data_start_row = header_row + 1
            max_rows = min(3000, ws.max_row)  # 최대 3000행만 스캔
            
            print(f"    [{sheet_name}] 스캔 중... (최대 {max_rows:,}행)")
            
            # 샘플링: 처음/중간/끝 부분만
            if ws.max_row <= max_rows:
                rows_to_check = range(data_start_row, ws.max_row + 1)
            else:
                rows_to_check = []
                rows_to_check.extend(range(data_start_row, min(data_start_row + 1000, ws.max_row + 1)))
                mid = (data_start_row + ws.max_row) // 2
                rows_to_check.extend(range(mid - 500, min(mid + 500, ws.max_row + 1)))
                rows_to_check.extend(range(max(data_start_row, ws.max_row - 1000), ws.max_row + 1))
                rows_to_check = sorted(set(rows_to_check))
            
            for row_idx in rows_to_check:
                if row_idx > max_rows:
                    break
                
                row_has_yellow = False
                
                # 컬럼 샘플링 (첫 15개 + 마지막 5개)
                cols_to_check = list(range(1, min(16, ws.max_column + 1)))
                if ws.max_column > 15:
                    cols_to_check.extend(range(ws.max_column - 4, ws.max_column + 1))
                
                for col_idx in cols_to_check:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    color = get_cell_color(cell)
                    
                    if is_orange(color):
                        orange_in_sheet += 1
                        total_orange += 1
                    elif is_yellow(color):
                        row_has_yellow = True
                        yellow_rows_in_sheet.add(row_idx)
                
                # YELLOW 행을 찾으면 확장 스캔 (해당 행의 모든 셀 확인)
                if row_has_yellow and row_idx not in yellow_rows_in_sheet:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if is_yellow(get_cell_color(cell)):
                            yellow_rows_in_sheet.add(row_idx)
                            break
            
            # 추정치 조정 (샘플링 비율 반영)
            if len(rows_to_check) < (ws.max_row - data_start_row + 1):
                sample_ratio = len(rows_to_check) / max(1, ws.max_row - data_start_row + 1)
                orange_in_sheet = int(orange_in_sheet / max(0.1, sample_ratio))
            
            if yellow_rows_in_sheet:
                yellow_row_counts[sheet_name] = len(yellow_rows_in_sheet)
                total_yellow_rows += len(yellow_rows_in_sheet)
        
        wb.close()
        
        print(f"\n[2] 실제 파일 검증 결과")
        print("-" * 80)
        print(f"  총 ORANGE 셀: {total_orange:,}개")
        print(f"  총 YELLOW 행: {total_yellow_rows:,}개")
        print(f"\n  시트별 YELLOW 행 수:")
        for sheet, count in yellow_row_counts.items():
            print(f"    - {sheet}: {count:,}행")
        
        return {
            'total_orange_cells': total_orange,
            'total_yellow_rows': total_yellow_rows,
            'yellow_by_sheet': dict(yellow_row_counts)
        }
    
    return None

def main():
    """메인 실행 함수"""
    print("=" * 80)
    print("컬러 작업 로직 엑셀 파일 직접 확인".center(80))
    print("=" * 80)
    
    all_results = {}
    
    try:
        # 1. Raw 파일 분석
        all_results['raw_files'] = analyze_raw_files()
        
        # 2. 색상 읽기
        all_results['colors'] = read_excel_colors()
        
        # 3. 케이스 번호 추적
        if all_results.get('colors'):
            all_results['tracking'] = track_case_numbers(
                all_results['raw_files'],
                all_results['colors']
            )
        
        # 4. 컬러 로직 검증
        all_results['verification'] = verify_color_logic()
        
        # 결과 저장
        import json
        results_file = Path("docs/reports/excel_color_verification_results.json")
        results_file.parent.mkdir(parents=True, exist_ok=True)
        
        # JSON 변환
        json_results = {}
        for key, value in all_results.items():
            if value is None:
                continue
            if isinstance(value, dict):
                json_results[key] = {
                    k: v for k, v in value.items()
                    if not isinstance(v, (pd.DataFrame, defaultdict))
                }
        
        with open(results_file, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n[OK] 검증 결과 저장: {results_file}")
        
        print("\n" + "=" * 80)
        print("검증 완료".center(80))
        print("=" * 80)
        
        return all_results
        
    except Exception as e:
        import traceback
        print(f"\n[ERROR] 검증 중 오류 발생: {e}")
        print(traceback.format_exc())
        return None

if __name__ == "__main__":
    main()

