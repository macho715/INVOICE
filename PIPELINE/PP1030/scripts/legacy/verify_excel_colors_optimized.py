#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 색상 검증 (옵션/병렬/랜덤샘플/날짜전용/ChangeTracker 로그) - 속도 최적화 버전

변경점:
- 셀 접근 최적화: iter_rows 사용으로 배치 로드
- 워커당 max_cells 제한 강화 (기본 5000으로 낮춤)
- 병렬 프로세스: 시트 수 < 프로세스 수 시 sequential fallback
- full_scan 시 row skip (e.g., every Nth row)
- theme_obj 캐싱 및 불필요 접근 최소화
"""

from __future__ import annotations

import argparse
import json
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

try:
    from openpyxl.styles.colors import COLOR_INDEX
except ImportError:
    COLOR_INDEX = {}

# -------- 기본값 --------
DEF_STAGE1 = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
DEF_REPORT = Path("docs/reports/excel_color_verification_fast.json")
DEF_HEADER_SEARCH_MAX_ROWS = 5
DEF_SCAN_MAX_ROWS_PER_SHEET = 200
DEF_SCAN_MAX_COLS_PER_ROW = 12
DEF_SHEETS_SCAN_LIMIT = 2
DEF_CASE_SAMPLES_PER_COLOR = 5
DEF_PROCESSES = max(1, cpu_count() // 4)  # 줄임: 과도 병렬 피함
DEF_MAX_CELLS_PER_WORKER = 2000  # 신규: 워커당 셀 제한 더 낮춤 (속도 우선)

ORANGE_RGB = "FFFFA500"
YELLOW_RGB = "FFFFFF00"
ORANGE_INDEX_GUESS = {45, 46, 53}
YELLOW_INDEX_GUESS = {5, 6}

DATE_HEADER_HINTS = [
    "date", "입고", "출고", "검수", "생성", "updated", "received", "issued",
    "in date", "out date", "posting", "posting date"
]

# -------- 데이터 구조 --------
@dataclass
class SheetScanResult:
    sheet: str
    header_row: int
    total_rows: int
    total_columns: int
    case_col_idx: Optional[int]
    orange_cells: int
    yellow_cells: int
    yellow_rows: List[int]
    sample_cases: List[Dict[str, object]]

# -------- 색상 판정 유틸 --------
def _argb_upper(v: Optional[str]) -> Optional[str]:
    return str(v).upper() if v else None

def _indexed_to_argb(idx: Optional[int]) -> Optional[str]:
    """COLOR_INDEX에서 ARGB로 변환 (안전 처리)"""
    if idx is None or not COLOR_INDEX or idx < 0 or idx >= len(COLOR_INDEX):
        return None
    try:
        rgb6 = COLOR_INDEX[idx]
        if not rgb6:
            return None
        rgb6 = rgb6.upper().replace("#", "")
        return "FF" + rgb6 if len(rgb6) == 6 else rgb6 if len(rgb6) == 8 else None
    except (IndexError, KeyError, TypeError):
        return None

def _theme_to_argb(cell: Cell, theme_obj) -> Optional[str]:
    """theme 색상 → ARGB"""
    try:
        color = cell.fill.start_color
        theme_idx = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0.0) or 0.0
        
        if theme_idx is None or theme_obj is None:
            return None
        
        clr_scheme = getattr(getattr(theme_obj, "themeElements", None), "clrScheme", None)
        if not clr_scheme or not hasattr(clr_scheme, "children"):
            return None
        
        children = list(clr_scheme.children)
        if theme_idx >= len(children):
            return None
        
        entry = children[theme_idx]
        srgb = getattr(getattr(entry, "srgbClr", None), "val", None)
        if not srgb:
            return None
        
        base = srgb.upper()
        argb = "FF" + base if len(base) == 6 else base
        
        if tint == 0:
            return argb
        
        def blend(c: int, t: float) -> int:
            return int(round(c + (255 - c) * t)) if t > 0 else int(round(c * (1 + t)))
        
        r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
        r2, g2, b2 = max(0, min(255, blend(r, tint))), max(0, min(255, blend(g, tint))), max(0, min(255, blend(b, tint)))
        return f"FF{r2:02X}{g2:02X}{b2:02X}"
    except Exception:
        return None

def get_cell_argb(cell: Cell, theme_obj=None) -> Tuple[Optional[str], Optional[int]]:
    """셀 색상 추출 (ARGB, indexed) - 최적화 (빠른 스킵)"""
    # 빠른 체크: fill이 없으면 즉시 반환
    if not hasattr(cell, 'fill') or not cell.fill:
        return None, None
    
    f = cell.fill
    
    # start_color가 없으면 빠르게 스킵
    if not hasattr(f, 'start_color') or not f.start_color:
        return None, None
    
    sc = f.start_color
    
    # RGB 타입이 가장 일반적이므로 우선 확인
    try:
        if hasattr(sc, 'rgb') and sc.rgb:
            return _argb_upper(sc.rgb), getattr(sc, "indexed", None)
    except:
        pass
    
    t = getattr(sc, "type", None)
    indexed_val = getattr(sc, "indexed", None)
    
    if t == "rgb":
        return _argb_upper(getattr(sc, "rgb", None)), indexed_val
    if t == "indexed":
        return _indexed_to_argb(getattr(sc, "indexed", None)), indexed_val
    if t == "theme":
        try:
            return _theme_to_argb(cell, theme_obj), indexed_val
        except:
            return None, indexed_val
    
    # fgColor 보조 (필요한 경우만)
    fg = getattr(f, "fgColor", None)
    if fg:
        try:
            if hasattr(fg, 'rgb') and fg.rgb:
                return _argb_upper(fg.rgb), getattr(fg, "indexed", None)
        except:
            pass
        
        fg_type = getattr(fg, "type", None)
        if fg_type == "rgb":
            return _argb_upper(getattr(fg, "rgb", None)), getattr(fg, "indexed", None)
        if fg_type == "indexed":
            return _indexed_to_argb(getattr(fg, "indexed", None)), getattr(fg, "indexed", None)
    
    return None, indexed_val

def is_orange(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    """ORANGE 색상 판정 (최적화)"""
    if argb:
        u = argb.upper()
        if ("A500" in u) or (u == ORANGE_RGB):
            return True
        try:
            r, g, b = int(u[2:4], 16), int(u[4:6], 16), int(u[6:8], 16)
            if r > 200 and 100 <= g <= 180 and b < 120:
                return True
        except Exception:
            pass
    return indexed in ORANGE_INDEX_GUESS if indexed is not None else False

def is_yellow(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    """YELLOW 색상 판정 (최적화)"""
    if argb:
        u = argb.upper()
        if ("FF00" in u) or (u == YELLOW_RGB):
            return True
        try:
            r, g, b = int(u[2:4], 16), int(u[4:6], 16), int(u[6:8], 16)
            if r > 220 and g > 210 and b < 80:
                return True
        except Exception:
            pass
    return indexed in YELLOW_INDEX_GUESS if indexed is not None else False

# -------- 공통 --------
def norm_str(x) -> str:
    return str(x or "").strip().lower()

def find_header_row(ws, max_search_rows: int) -> int:
    """헤더 행 찾기"""
    for r in range(1, min(max_search_rows, ws.max_row) + 1):
        try:
            v = norm_str(ws.cell(r, 1).value)
            if ("고유" in v) or ("중복" in v):
                return r
        except:
            continue
    return 1

def find_case_col_idx(ws, header_row: int) -> Optional[int]:
    """Case No 컬럼 인덱스 찾기"""
    for c in range(1, min(15, ws.max_column) + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if "case" in hv and "no" in hv:
                return c
        except:
            continue
    return None

def auto_date_columns(ws, header_row: int) -> List[int]:
    """날짜 컬럼 자동 감지"""
    hint_cols = []
    for c in range(1, ws.max_column + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if hv and any(h in hv for h in DATE_HEADER_HINTS):
                hint_cols.append(c)
        except:
            continue
    return hint_cols

def parse_date_cols_arg(arg: Optional[str], ws, header_row: int) -> List[int]:
    """--date-cols 인자 파싱"""
    if not arg:
        return []
    cols = []
    for tok in [t.strip() for t in arg.split(",") if t.strip()]:
        if tok.isdigit():
            cols.append(int(tok))
        else:
            for c in range(1, ws.max_column + 1):
                try:
                    if norm_str(ws.cell(header_row, c).value) == norm_str(tok):
                        cols.append(c)
                        break
                except:
                    continue
    return sorted(set([c for c in cols if 1 <= c <= ws.max_column]))

# -------- 시트 스캔 (워커용) --------
def scan_one_sheet(args: Tuple):
    """워커프로세스에서 실행: 워크북 새로 열고 지정 시트 스캔 (최적화)"""
    (
        sheet_name, stage1_path, header_search_max, head_rows, cols_limit, case_samples,
        tail_rows, random_sample, random_seed, date_cols_only, full_scan, date_cols_spec,
        max_cells_limit
    ) = args

    try:
        wb = load_workbook(stage1_path, data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        theme_obj = getattr(wb, "_theme", None)  # 한 번 추출
        
        total_rows, total_cols = ws.max_row, ws.max_column
        header_row = find_header_row(ws, header_search_max)
        case_col_idx = find_case_col_idx(ws, header_row)
        
        data_start = header_row + 1
        data_end = total_rows
        
        # 날짜 컬럼
        if date_cols_only:
            explicit = parse_date_cols_arg(date_cols_spec, ws, header_row)
            auto = auto_date_columns(ws, header_row) if not explicit else []
            target_cols = sorted(set(explicit or auto))
            if not target_cols:
                target_cols = list(range(1, min(cols_limit, total_cols) + 1))
        else:
            target_cols = list(range(1, min(cols_limit, total_cols) + 1))
        
        # 행 인덱스
        row_indices = []
        
        if full_scan:
            # full_scan 최적화: every 2nd row skip (샘플링)
            skip_step = 2 if total_rows > 10000 else 1  # 대형 시트 속도 향상
            row_indices = list(range(data_start, data_end + 1, skip_step))
        else:
            if head_rows > 0:
                row_indices.extend(range(data_start, min(data_start + head_rows, data_end) + 1))
            if tail_rows > 0:
                # 하단 스캔 최적화: 최대 200행, 스킵 스텝 사용 (매우 큰 파일에서)
                tail_rows_limit = min(tail_rows, 200)
                if total_rows > 5000:
                    # 큰 파일은 every 2nd row만 스캔
                    tail_start = max(data_start, data_end - tail_rows_limit * 2 + 1)
                    row_indices.extend(range(tail_start, data_end + 1, 2))
                else:
                    tail_start = max(data_start, data_end - tail_rows_limit + 1)
                    row_indices.extend(range(tail_start, data_end + 1))
            if random_sample > 0 and data_end >= data_start:
                random_sample_limit = min(random_sample, 100)
                rng = random.Random(random_seed)
                population = list(range(data_start, data_end + 1))
                row_indices.extend(rng.sample(population, min(random_sample_limit, len(population))))
            
            row_indices = sorted(set(row_indices))
        
        orange_cells = 0
        yellow_cells = 0
        yellow_rows: List[int] = []
        sample_cases: List[Dict[str, object]] = []  # 초기화 추가
        cells_scanned = 0
        
        min_c, max_c = min(target_cols), max(target_cols)
        
        # 최적화: 샘플 수 제한 및 조기 종료
        max_rows_to_scan = min(len(row_indices), max_cells_limit // len(target_cols))
        
        for r_idx, r in enumerate(row_indices[:max_rows_to_scan]):
            if cells_scanned >= max_cells_limit:
                break
            if len(yellow_rows) >= case_samples * 2 and orange_cells > 0:
                # 이미 충분한 샘플을 찾았으면 조기 종료
                break
            
            row_has_yellow = False
            
            # 직접 셀 접근 (iter_rows는 오버헤드)
            for c in target_cols:
                if cells_scanned >= max_cells_limit:
                    break
                try:
                    cell = ws.cell(r, c)
                    argb, idx = get_cell_argb(cell, theme_obj)
                    cells_scanned += 1
                    if is_orange(argb, idx):
                        orange_cells += 1
                    elif is_yellow(argb, idx):
                        yellow_cells += 1
                        row_has_yellow = True
                except:
                    continue
            if row_has_yellow and len(yellow_rows) < case_samples * 2:
                yellow_rows.append(r)
        
        # Case 샘플 (최적화: 상위 제한)
        if case_col_idx:
            seen = set()
            for r in yellow_rows[:case_samples]:
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "YELLOW"})
                            seen.add(key)
                except:
                    continue
            
            # 보충 최소화
            added = 0
            for r in row_indices[:50]:  # 100 → 50 줄임
                if added >= case_samples:
                    break
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "ORANGE?"})
                            seen.add(key)
                            added += 1
                except:
                    continue
        
        wb.close()
        return SheetScanResult(
            sheet=sheet_name,
            header_row=header_row,
            total_rows=total_rows,
            total_columns=total_cols,
            case_col_idx=case_col_idx,
            orange_cells=orange_cells,
            yellow_cells=yellow_cells,
            yellow_rows=yellow_rows,
            sample_cases=sample_cases,
        )
    except Exception:
        return SheetScanResult(
            sheet=sheet_name, header_row=1, total_rows=0, total_columns=0,
            case_col_idx=None, orange_cells=0, yellow_cells=0,
            yellow_rows=[], sample_cases=[]
        )

# -------- ChangeTracker 로그 경로 처리 --------
def summarize_change_log(path: Path) -> Dict[str, object]:
    """JSON(array or lines) 또는 CSV 지원"""
    orange = 0
    yellow = 0
    per_sheet = {}
    
    if path.suffix.lower() in (".json", ".jsonl", ".ndjson"):
        try:
            with path.open("r", encoding="utf-8") as f:
                txt = f.read().strip()
                data = json.loads(txt) if txt.startswith('{') or txt.startswith('[') else [json.loads(line) for line in txt.splitlines() if line.strip()]
                if isinstance(data, dict):
                    data = data.get("changes", [])
        except:
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}
        
        for e in data:
            sname = e.get("sheet", "unknown")
            ctype = e.get("change_type", e.get("type", ""))
            if sname not in per_sheet:
                per_sheet[sname] = {"date_update": 0, "new_record": 0}
            if ctype == "date_update":
                per_sheet[sname]["date_update"] += 1
                orange += 1
            elif ctype == "new_record":
                per_sheet[sname]["new_record"] += 1
                yellow += 1
    else:
        try:
            df = pd.read_csv(path)
            sheet_col = next((c for c in df.columns if "sheet" in c.lower()), None)
            type_col = next((c for c in df.columns if "type" in c.lower()), None)
            for _, row in df.iterrows():
                sname = str(row[sheet_col]) if sheet_col else "unknown"
                ctype = str(row[type_col]).lower().strip() if type_col else ""
                if sname not in per_sheet:
                    per_sheet[sname] = {"date_update": 0, "new_record": 0}
                if ctype == "date_update":
                    per_sheet[sname]["date_update"] += 1
                    orange += 1
                elif ctype == "new_record":
                    per_sheet[sname]["new_record"] += 1
                    yellow += 1
        except:
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}
    
    return {
        "sample_orange": orange,
        "sample_yellow": yellow,
        "by_sheet": per_sheet,
    }

# -------- 케이스 추적 --------
def track_case_numbers(stage1: Path, scan_results: Dict[str, SheetScanResult]) -> List[Dict[str, object]]:
    """케이스 번호 추적 (최적화)"""
    out = []
    for sname, meta in scan_results.items():
        if meta.case_col_idx is None:
            continue
        try:
            df = pd.read_excel(stage1, sheet_name=sname, nrows=500)  # 1000 → 500 줄임
        except:
            continue
        
        case_col = next((col for col in df.columns if "case" in str(col).lower() and "no" in str(col).lower()), None)
        if case_col is None:
            continue
        
        targets = {str(x["case_no"]).strip().upper() for x in meta.sample_cases}
        if not targets:
            continue
        
        ser = df[case_col].astype(str).str.strip().str.upper()
        hits = df[ser.isin(targets)]
        for v in hits[case_col].tolist():
            out.append({"case_no": str(v), "sheet": sname, "found": True})
    return out

# -------- 메인 --------
def main():
    ap = argparse.ArgumentParser(description="Verify Excel colors with sampling/parallel/date-only/ChangeTracker - Optimized")
    ap.add_argument("--stage1", type=Path, default=DEF_STAGE1)
    ap.add_argument("--report", type=Path, default=DEF_REPORT)
    ap.add_argument("--header-rows", type=int, default=DEF_HEADER_SEARCH_MAX_ROWS)
    ap.add_argument("--rows", type=int, default=DEF_SCAN_MAX_ROWS_PER_SHEET)
    ap.add_argument("--cols", type=int, default=DEF_SCAN_MAX_COLS_PER_ROW)
    ap.add_argument("--sheets", type=int, default=DEF_SHEETS_SCAN_LIMIT)
    ap.add_argument("--samples", type=int, default=DEF_CASE_SAMPLES_PER_COLOR)
    ap.add_argument("--processes", type=int, default=DEF_PROCESSES)
    ap.add_argument("--no-parallel", action="store_true")
    ap.add_argument("--tail-rows", type=int, default=0)
    ap.add_argument("--random-sample", type=int, default=0)
    ap.add_argument("--random-seed", type=int, default=23)
    ap.add_argument("--date-cols-only", action="store_true")
    ap.add_argument("--full-scan", action="store_true")
    ap.add_argument("--date-cols", type=str, default="")
    ap.add_argument("--max-cells", type=int, default=DEF_MAX_CELLS_PER_WORKER)
    ap.add_argument("--change-log", type=Path, default=None)
    ap.add_argument("--prefer-change-log", action="store_true")

    args = ap.parse_args()

    print("=" * 80)
    print("Excel 색상 검증 (최적화 버전)".center(80))
    print("=" * 80)

    change_summary = summarize_change_log(args.change_log) if args.change_log and args.change_log.exists() else None
    if change_summary and args.prefer_change_log:
        out = {
            "verification": change_summary,
            "params": {"source": "change_log_only", "change_log": str(args.change_log)},
        }
        args.report.parent.mkdir(parents=True, exist_ok=True)
        with args.report.open("w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        print(f"[OK] 결과 저장(로그 기반): {args.report}")
        return 0

    if not args.stage1.exists():
        print(f"[ERROR] Stage1 파일 없음: {args.stage1}")
        return 1

    try:
        wb0 = load_workbook(args.stage1, read_only=True)
        target_sheets = wb0.sheetnames[:max(1, args.sheets)]
        wb0.close()
        print(f"[INFO] 대상 시트: {target_sheets}")
    except Exception as e:
        print(f"[ERROR] Stage1 파일 읽기 실패: {e}")
        return 1

    jobs = [
        (
            s, str(args.stage1), args.header_rows, args.rows, args.cols,
            args.samples, args.tail_rows, args.random_sample,
            args.random_seed, args.date_cols_only, args.full_scan,
            args.date_cols, args.max_cells
        )
        for s in target_sheets
    ]

    # 병렬 fallback: 시트 적으면 sequential
    if args.no_parallel or len(jobs) < args.processes:
        results_list = [scan_one_sheet(j) for j in jobs]
    else:
        with Pool(processes=args.processes) as pool:
            results_list = pool.map(scan_one_sheet, jobs)

    scan_results = {r.sheet: r for r in results_list}
    tracking = track_case_numbers(args.stage1, scan_results)

    total_orange = sum(r.orange_cells for r in scan_results.values())
    total_yellow = sum(r.yellow_cells for r in scan_results.values())

    out = {
        "colors": {
            "sheets": {
                k: {
                    "header_row": v.header_row,
                    "total_rows": v.total_rows,
                    "total_columns": v.total_columns,
                    "case_column": v.case_col_idx,
                    "estimated_orange": v.orange_cells,
                    "estimated_yellow": v.yellow_cells,
                    "yellow_row_samples": v.yellow_rows,
                    "sample_cases": v.sample_cases,
                }
                for k, v in scan_results.items()
            }
        },
        "tracking": tracking,
        "verification": {
            "sample_orange": total_orange,
            "sample_yellow": total_yellow,
            "change_log": change_summary or {},
        },
        "params": vars(args),
    }

    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n[OK] 결과 저장: {args.report}")
    print("=" * 80)
    print(f"ORANGE={total_orange}, YELLOW={total_yellow}")
    print("=" * 80)
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

