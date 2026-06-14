#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 색상 검증 (옵션/병렬/랜덤샘플/날짜전용/ChangeTracker 로그)
"""

from __future__ import annotations

import argparse
import json
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

try:
    from openpyxl.styles.colors import COLOR_INDEX
except ImportError:
    COLOR_INDEX = {}

# -------- 기본값 --------
DEF_STAGE1 = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
DEF_REPORT = Path("docs/reports/excel_color_verification_fast.json")
DEF_HEADER_SEARCH_MAX_ROWS = 5
DEF_SCAN_MAX_ROWS_PER_SHEET = 200
DEF_SCAN_MAX_COLS_PER_ROW = 12
DEF_SHEETS_SCAN_LIMIT = 2
DEF_CASE_SAMPLES_PER_COLOR = 5
DEF_PROCESSES = max(1, cpu_count() // 2)

ORANGE_RGB = "FFFFA500"
YELLOW_RGB = "FFFFFF00"
ORANGE_INDEX_GUESS = {45, 46, 53}
YELLOW_INDEX_GUESS = {5, 6}

DATE_HEADER_HINTS = [
    "date", "입고", "출고", "검수", "생성", "updated", "received", "issued",
    "in date", "out date", "posting", "posting date"
]

# -------- 데이터 구조 --------
@dataclass
class SheetScanResult:
    sheet: str
    header_row: int
    total_rows: int
    total_columns: int
    case_col_idx: Optional[int]
    orange_cells: int
    yellow_cells: int
    yellow_rows: List[int]
    sample_cases: List[Dict[str, object]]

# -------- 색상 판정 유틸 --------
def _argb_upper(v: Optional[str]) -> Optional[str]:
    return str(v).upper() if v else None

def _indexed_to_argb(idx: Optional[int]) -> Optional[str]:
    """COLOR_INDEX에서 ARGB로 변환 (안전 처리)"""
    if idx is None:
        return None
    try:
        # COLOR_INDEX가 비어있거나 범위를 벗어난 경우
        if not COLOR_INDEX or idx < 0 or idx >= len(COLOR_INDEX):
            return None
        rgb6 = COLOR_INDEX[idx]
        if not rgb6:
            return None
        rgb6 = rgb6.upper().replace("#", "")
        if len(rgb6) == 6:
            return "FF" + rgb6
        if len(rgb6) == 8:
            return rgb6
    except (IndexError, KeyError, TypeError, AttributeError):
        return None
    return None

def _theme_to_argb(cell: Cell, theme_obj) -> Optional[str]:
    """theme 색상 → ARGB (베스트에포트)"""
    try:
        color = cell.fill.start_color
        theme_idx = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0.0) or 0.0

        if theme_idx is None or theme_obj is None:
            return None

        clr_scheme = getattr(getattr(theme_obj, "themeElements", None), "clrScheme", None)
        if not clr_scheme or not hasattr(clr_scheme, "children"):
            return None

        children = list(clr_scheme.children)
        if theme_idx >= len(children):
            return None

        entry = children[theme_idx]
        srgb = getattr(getattr(entry, "srgbClr", None), "val", None)
        if not srgb:
            return None

        base = srgb.upper()
        argb = "FF" + base if len(base) == 6 else base

        if tint == 0:
            return argb

        def blend(c: int, t: float) -> int:
            if t > 0:
                return int(round(c + (255 - c) * t))
            else:
                return int(round(c * (1 + t)))

        r = int(argb[2:4], 16)
        g = int(argb[4:6], 16)
        b = int(argb[6:8], 16)
        r2 = max(0, min(255, blend(r, tint)))
        g2 = max(0, min(255, blend(g, tint)))
        b2 = max(0, min(255, blend(b, tint)))
        return f"FF{r2:02X}{g2:02X}{b2:02X}"
    except Exception:
        return None

def get_cell_argb(cell: Cell, theme_obj=None) -> Tuple[Optional[str], Optional[int]]:
    """셀 색상 추출 (ARGB, indexed)"""
    f = cell.fill
    if not f:
        return None, None
    sc = getattr(f, "start_color", None)
    if sc is None:
        return None, None

    t = getattr(sc, "type", None)
    indexed_val = getattr(sc, "indexed", None)
    
    if t == "rgb":
        return _argb_upper(getattr(sc, "rgb", None)), indexed_val
    if t == "indexed":
        return _indexed_to_argb(getattr(sc, "indexed", None)), indexed_val
    if t == "theme":
        return _theme_to_argb(cell, theme_obj), indexed_val

    # fgColor 보조
    fg = getattr(f, "fgColor", None)
    if fg is not None:
        fg_type = getattr(fg, "type", None)
        if fg_type == "rgb":
            return _argb_upper(getattr(fg, "rgb", None)), getattr(fg, "indexed", None)
        if fg_type == "indexed":
            return _indexed_to_argb(getattr(fg, "indexed", None)), getattr(fg, "indexed", None)
        if fg_type == "theme":
            return _theme_to_argb(cell, theme_obj), getattr(fg, "indexed", None)

    return None, indexed_val

def is_orange(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    """ORANGE 색상 판정"""
    if argb:
        u = argb.upper()
        if ("A500" in u) or (u == ORANGE_RGB):
            return True
        try:
            r = int(u[2:4], 16)
            g = int(u[4:6], 16)
            b = int(u[6:8], 16)
            if r > 200 and 100 <= g <= 180 and b < 120:
                return True
        except Exception:
            pass
    if indexed is not None and indexed in ORANGE_INDEX_GUESS:
        return True
    return False

def is_yellow(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    """YELLOW 색상 판정"""
    if argb:
        u = argb.upper()
        if ("FF00" in u) or (u == YELLOW_RGB):
            return True
        try:
            r = int(u[2:4], 16)
            g = int(u[4:6], 16)
            b = int(u[6:8], 16)
            if r > 220 and g > 210 and b < 80:
                return True
        except Exception:
            pass
    if indexed is not None and indexed in YELLOW_INDEX_GUESS:
        return True
    return False

# -------- 공통 --------
def norm_str(x) -> str:
    return str(x or "").strip().lower()

def find_header_row(ws, max_search_rows: int) -> int:
    """헤더 행 찾기"""
    total_rows = ws.max_row
    for r in range(1, min(max_search_rows, total_rows) + 1):
        try:
            v = norm_str(ws.cell(r, 1).value)
            if ("고유" in v) or ("중복" in v):
                return r
        except:
            continue
    return 1

def find_case_col_idx(ws, header_row: int) -> Optional[int]:
    """Case No 컬럼 인덱스 찾기"""
    total_cols = ws.max_column
    for c in range(1, min(15, total_cols) + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if "case" in hv and "no" in hv:
                return c
        except:
            continue
    return None

def auto_date_columns(ws, header_row: int) -> List[int]:
    """날짜 컬럼 자동 감지"""
    hint_cols = []
    for c in range(1, ws.max_column + 1):
        try:
            hv = norm_str(ws.cell(header_row, c).value)
            if not hv:
                continue
            if any(h in hv for h in DATE_HEADER_HINTS):
                hint_cols.append(c)
        except:
            continue
    return hint_cols

def parse_date_cols_arg(arg: Optional[str], ws, header_row: int) -> List[int]:
    """--date-cols 인자 파싱"""
    if not arg:
        return []
    cols = []
    for tok in [t.strip() for t in arg.split(",") if t.strip()]:
        if tok.isdigit():
            cols.append(int(tok))
        else:
            # 헤더명 매칭
            for c in range(1, ws.max_column + 1):
                try:
                    if norm_str(ws.cell(header_row, c).value) == norm_str(tok):
                        cols.append(c)
                        break
                except:
                    continue
    return sorted(set([c for c in cols if 1 <= c <= ws.max_column]))

# -------- 시트 스캔 (워커용) --------
def scan_one_sheet(args: Tuple):
    """워커프로세스에서 실행: 워크북 새로 열고 지정 시트 스캔"""
    (
        sheet_name, stage1_path, header_search_max, head_rows, cols_limit, case_samples,
        tail_rows, random_sample, random_seed, date_cols_only, full_scan, date_cols_spec,
        max_cells_limit
    ) = args

    try:
        wb = load_workbook(stage1_path, data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        
        # theme 객체 안전 접근
        theme_obj = None
        try:
            theme_obj = getattr(wb, "_theme", None)
        except:
            pass

        total_rows, total_cols = ws.max_row, ws.max_column
        header_row = find_header_row(ws, header_search_max)
        case_col_idx = find_case_col_idx(ws, header_row)

        # 스캔 대상 행 범위 계산
        data_start = header_row + 1
        data_end = total_rows

        # 날짜 컬럼 전용 여부 → 컬럼 집합 계산
        target_cols: List[int]
        if date_cols_only:
            explicit = parse_date_cols_arg(date_cols_spec, ws, header_row)
            auto = auto_date_columns(ws, header_row) if not explicit else []
            target_cols = sorted(set(explicit or auto))
            # 날짜 컬럼 못 찾으면 안전하게 좌측 cols_limit 범위로 축소
            if not target_cols:
                target_cols = list(range(1, min(cols_limit, total_cols) + 1))
        else:
            target_cols = list(range(1, min(cols_limit, total_cols) + 1))

        # 스캔 대상 행 인덱스 목록 생성
        row_indices: List[int] = []

        if full_scan:
            row_indices = list(range(data_start, data_end + 1))
        else:
            # 상단 head_rows
            if head_rows > 0:
                row_indices.extend(range(data_start, min(data_start + head_rows - 1, data_end) + 1))
            # 하단 tail_rows (최대 200으로 제한)
            if tail_rows > 0:
                tail_rows_limit = min(tail_rows, 200)  # 성능 보호
                tail_start = max(data_start, data_end - tail_rows_limit + 1)
                row_indices.extend(range(tail_start, data_end + 1))
            # 랜덤 샘플 (최대 100으로 제한)
            if random_sample > 0 and data_end >= data_start:
                random_sample_limit = min(random_sample, 100)  # 성능 보호
                rng = random.Random(random_seed)
                population = list(range(data_start, data_end + 1))
                sample_n = min(random_sample_limit, len(population))
                row_indices.extend(rng.sample(population, sample_n))

            # 중복 제거 & 정렬
            row_indices = sorted(set([r for r in row_indices if data_start <= r <= data_end]))

        orange_cells = 0
        yellow_cells = 0
        yellow_rows: List[int] = []
        sample_cases: List[Dict[str, object]] = []

        # 스캔 루프 (최적화: 샘플 수 제한 및 조기 종료)
        max_cells_to_scan = max_cells_limit
        cells_scanned = 0
        
        for r in row_indices:
            if cells_scanned >= max_cells_to_scan:
                break
            row_has_yellow = False
            for c in target_cols:
                if cells_scanned >= max_cells_to_scan:
                    break
                try:
                    cell = ws.cell(r, c)
                    argb, idx = get_cell_argb(cell, theme_obj=theme_obj)
                    cells_scanned += 1
                    if is_orange(argb, idx):
                        orange_cells += 1
                    elif is_yellow(argb, idx):
                        yellow_cells += 1
                        row_has_yellow = True
                except:
                    continue
            if row_has_yellow and len(yellow_rows) < case_samples * 2:
                yellow_rows.append(r)

        # Case 샘플 채집
        if case_col_idx:
            seen: set[str] = set()
            for r in yellow_rows[:case_samples]:
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "YELLOW"})
                            seen.add(key)
                except:
                    continue

            # 보충(상단 100행)
            added = 0
            for r in row_indices[:100]:
                if added >= case_samples:
                    break
                try:
                    v = ws.cell(r, case_col_idx).value
                    if v:
                        key = str(v).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "ORANGE?"})
                            seen.add(key)
                            added += 1
                except:
                    continue

        wb.close()
        return SheetScanResult(
            sheet=sheet_name,
            header_row=header_row,
            total_rows=total_rows,
            total_columns=total_cols,
            case_col_idx=case_col_idx,
            orange_cells=orange_cells,
            yellow_cells=yellow_cells,
            yellow_rows=yellow_rows,
            sample_cases=sample_cases,
        )
    except Exception as e:
        # 실패 시 빈 결과 반환
        return SheetScanResult(
            sheet=sheet_name,
            header_row=1,
            total_rows=0,
            total_columns=0,
            case_col_idx=None,
            orange_cells=0,
            yellow_cells=0,
            yellow_rows=[],
            sample_cases=[],
        )

# -------- ChangeTracker 로그 경로 처리 --------
def summarize_change_log(path: Path) -> Dict[str, object]:
    """JSON(array or lines) 또는 CSV 지원"""
    orange = 0
    yellow = 0
    per_sheet: Dict[str, Dict[str, int]] = {}

    def bump(sname: str, kind: str):
        nonlocal orange, yellow
        if sname not in per_sheet:
            per_sheet[sname] = {"date_update": 0, "new_record": 0}
        per_sheet[sname][kind] += 1
        if kind == "date_update":
            orange += 1
        elif kind == "new_record":
            yellow += 1

    if path.suffix.lower() in (".json", ".jsonl", ".ndjson"):
        try:
            with path.open("r", encoding="utf-8") as f:
                txt = f.read().strip()
                if not txt:
                    return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}
                # try array
                try:
                    data = json.loads(txt)
                    if isinstance(data, dict):
                        data = data.get("changes", [])
                except Exception:
                    # line-delimited
                    data = [json.loads(line) for line in txt.splitlines() if line.strip()]
        except Exception:
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}

        for e in data:
            sname = e.get("sheet", "unknown")
            ctype = e.get("change_type") or e.get("type") or ""
            if ctype == "date_update":
                bump(sname, "date_update")
            elif ctype == "new_record":
                bump(sname, "new_record")
    else:
        # CSV
        try:
            df = pd.read_csv(path)
            sheet_col = next((c for c in df.columns if c.lower() in ("sheet", "sheet_name")), None)
            type_col = next((c for c in df.columns if c.lower() in ("change_type", "type")), None)
            for _, row in df.iterrows():
                sname = str(row.get(sheet_col, "unknown")) if sheet_col else "unknown"
                ctype = str(row.get(type_col, "")).strip().lower() if type_col else ""
                if ctype == "date_update":
                    bump(sname, "date_update")
                elif ctype == "new_record":
                    bump(sname, "new_record")
        except Exception:
            return {"sample_orange": 0, "sample_yellow": 0, "by_sheet": {}}

    return {
        "sample_orange": orange,
        "sample_yellow": yellow,
        "by_sheet": per_sheet,
    }

# -------- 케이스 추적 --------
def track_case_numbers(stage1: Path, scan_results: Dict[str, SheetScanResult]) -> List[Dict[str, object]]:
    """케이스 번호 추적 (pandas로 로드)"""
    out: List[Dict[str, object]] = []
    for sname, meta in scan_results.items():
        if meta.case_col_idx is None:
            continue
        try:
            # 전체 로드 후 필요한 컬럼만 선택
            df = pd.read_excel(
                stage1, sheet_name=sname, nrows=1000, engine="openpyxl"
            )
        except Exception:
            continue

        case_col = None
        for col in df.columns:
            hv = str(col).lower()
            if "case" in hv and "no" in hv:
                case_col = col
                break
        if case_col is None:
            continue

        targets = {str(x["case_no"]).strip().upper() for x in meta.sample_cases}
        if not targets:
            continue

        ser = df[case_col].astype(str).str.strip().str.upper()
        hits = df[ser.isin(targets)]
        for v in hits[case_col].tolist():
            out.append({"case_no": str(v), "sheet": sname, "found": True})
    return out

# -------- 메인 --------
def main():
    ap = argparse.ArgumentParser(description="Verify Excel colors with sampling/parallel/date-only/ChangeTracker")
    ap.add_argument("--stage1", type=Path, default=DEF_STAGE1)
    ap.add_argument("--report", type=Path, default=DEF_REPORT)

    ap.add_argument("--header-rows", type=int, default=DEF_HEADER_SEARCH_MAX_ROWS)
    ap.add_argument("--rows", type=int, default=DEF_SCAN_MAX_ROWS_PER_SHEET)
    ap.add_argument("--cols", type=int, default=DEF_SCAN_MAX_COLS_PER_ROW)
    ap.add_argument("--sheets", type=int, default=DEF_SHEETS_SCAN_LIMIT)
    ap.add_argument("--samples", type=int, default=DEF_CASE_SAMPLES_PER_COLOR)
    ap.add_argument("--processes", type=int, default=DEF_PROCESSES)
    ap.add_argument("--no-parallel", action="store_true")

    # 확장: tail/랜덤/날짜전용/풀스캔
    ap.add_argument("--tail-rows", type=int, default=0, help="하단에서 추가 스캔할 행 수 (최대 200 권장)")
    ap.add_argument("--random-sample", type=int, default=0, help="무작위 샘플 행 수 (최대 100 권장)")
    ap.add_argument("--random-seed", type=int, default=23, help="랜덤 시드")
    ap.add_argument("--date-cols-only", action="store_true", help="날짜 컬럼만 스캔")
    ap.add_argument("--full-scan", action="store_true", help="지정 범위 무시하고 전체 행 스캔 (주의: 매우 느림)")
    ap.add_argument("--date-cols", type=str, default="", help='날짜 컬럼 지정: "IN Date,OUT Date" 또는 "5,7"')
    ap.add_argument("--max-cells", type=int, default=10000, help="최대 스캔 셀 수 제한 (기본: 10000)")

    # ChangeTracker
    ap.add_argument("--change-log", type=Path, default=None, help="ChangeTracker 로그(JSON/CSV)")
    ap.add_argument("--prefer-change-log", action="store_true", help="로그 있으면 파일 스캔 생략")

    args = ap.parse_args()

    print("=" * 80)
    print("Excel 색상 검증 (옵션/병렬/랜덤/날짜전용/ChangeTracker)".center(80))
    print("=" * 80)

    # 0) ChangeTracker 로그만으로 처리?
    change_summary = None
    if args.change_log and args.change_log.exists():
        change_summary = summarize_change_log(args.change_log)
        print(f"[INFO] ChangeTracker 요약: ORANGE={change_summary['sample_orange']} YELLOW={change_summary['sample_yellow']}")
        if args.prefer_change_log:
            # 결과 저장 후 종료
            out = {
                "verification": change_summary,
                "params": {
                    "source": "change_log_only",
                    "change_log": str(args.change_log),
                },
            }
            args.report.parent.mkdir(parents=True, exist_ok=True)
            with args.report.open("w", encoding="utf-8") as f:
                json.dump(out, f, indent=2, ensure_ascii=False)
            print(f"[OK] 결과 저장(로그 기반): {args.report}")
            return 0

    if not args.stage1.exists():
        print(f"[ERROR] Stage1 파일 없음: {args.stage1}")
        return 1

    # 1) 대상 시트
    try:
        wb0 = load_workbook(args.stage1, read_only=True)
        all_sheets = wb0.sheetnames
        target_sheets = all_sheets[:max(1, args.sheets)]
        wb0.close()
        print(f"[INFO] 대상 시트: {target_sheets}")
    except Exception as e:
        print(f"[ERROR] Stage1 파일 읽기 실패: {e}")
        return 1

    # 2) 병렬 스캔
    max_cells = getattr(args, 'max_cells', 10000)
    jobs = [
        (
            s, str(args.stage1), int(args.header_rows), int(args.rows), int(args.cols),
            int(args.samples), int(args.tail_rows), int(args.random_sample),
            int(args.random_seed), bool(args.date_cols_only), bool(args.full_scan),
            args.date_cols, int(max_cells)
        )
        for s in target_sheets
    ]

    if args.no_parallel or len(jobs) == 1:
        results_list = [scan_one_sheet(j) for j in jobs]
    else:
        with Pool(processes=max(1, int(args.processes))) as pool:
            results_list = pool.map(scan_one_sheet, jobs)

    scan_results: Dict[str, SheetScanResult] = {r.sheet: r for r in results_list}

    # 3) 케이스 추적
    tracking = track_case_numbers(args.stage1, scan_results)

    # 4) 합산
    total_orange = sum(r.orange_cells for r in scan_results.values())
    total_yellow = sum(r.yellow_cells for r in scan_results.values())

    # 5) 저장
    out = {
        "colors": {
            "sheets": {
                k: {
                    "header_row": v.header_row,
                    "total_rows": v.total_rows,
                    "total_columns": v.total_columns,
                    "case_column": v.case_col_idx,
                    "estimated_orange": v.orange_cells,
                    "estimated_yellow": v.yellow_cells,
                    "yellow_row_samples": v.yellow_rows,
                    "sample_cases": v.sample_cases,
                }
                for k, v in scan_results.items()
            }
        },
        "tracking": tracking,
        "verification": {
            "sample_orange": total_orange,
            "sample_yellow": total_yellow,
            "change_log": change_summary or {},
        },
        "params": {
            "rows": args.rows,
            "cols": args.cols,
            "tail_rows": args.tail_rows,
            "random_sample": args.random_sample,
            "random_seed": args.random_seed,
            "sheets": args.sheets,
            "samples": args.samples,
            "processes": (0 if args.no_parallel else args.processes),
            "date_cols_only": args.date_cols_only,
            "full_scan": args.full_scan,
            "date_cols": args.date_cols,
            "change_log": str(args.change_log) if args.change_log else None,
            "prefer_change_log": args.prefer_change_log,
        },
    }

    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n[OK] 결과 저장: {args.report}")
    print("=" * 80)
    print(f"ORANGE={total_orange}, YELLOW={total_yellow}")
    print("=" * 80)
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

