#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
컬러 작업 로직 엑셀 빠른 확인 스크립트 (CLI + 병렬 + 확장 색상 규칙)
"""

from __future__ import annotations

import argparse
import json
import math
import os
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
try:
    from openpyxl.styles.colors import COLOR_INDEX
except ImportError:
    # COLOR_INDEX가 없는 경우 대체
    COLOR_INDEX = {}
from multiprocessing import Pool, cpu_count

# --------------------------------------------------------------------------------------
# 기본값
# --------------------------------------------------------------------------------------
DEF_RAW_MASTER = Path("data/raw/Case List(HE).xlsx")
DEF_RAW_WH = Path("data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx")
DEF_STAGE1 = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
DEF_REPORT = Path("docs/reports/excel_color_verification_fast.json")

DEF_HEADER_SEARCH_MAX_ROWS = 5
DEF_SCAN_MAX_ROWS_PER_SHEET = 200
DEF_SCAN_MAX_COLS_PER_ROW = 12
DEF_SHEETS_SCAN_LIMIT = 2
DEF_CASE_SAMPLES_PER_COLOR = 5

# 색상 상수
ORANGE_RGB = "FFFFA500"
YELLOW_RGB = "FFFFFF00"
ORANGE_INDEX_GUESS = {45, 46, 53}
YELLOW_INDEX_GUESS = {5, 6}

# --------------------------------------------------------------------------------------
# 데이터 클래스
# --------------------------------------------------------------------------------------
@dataclass
class SheetScanResult:
    sheet: str
    header_row: int
    total_rows: int
    total_columns: int
    case_col_idx: Optional[int]
    orange_cells: int
    yellow_cells: int
    yellow_rows: List[int]
    sample_cases: List[Dict[str, object]]

# --------------------------------------------------------------------------------------
# 색상 도우미
# --------------------------------------------------------------------------------------
def _argb_upper(v: Optional[str]) -> Optional[str]:
    return str(v).upper() if v else None

def _indexed_to_argb(idx: Optional[int]) -> Optional[str]:
    """openpyxl COLOR_INDEX는 RGB 6자리(hex) 배열. ARGB로 보정."""
    if idx is None:
        return None
    try:
        if not COLOR_INDEX or idx >= len(COLOR_INDEX) or idx < 0:
            return None
        rgb6 = COLOR_INDEX[idx]
        if not rgb6:
            return None
        rgb6 = rgb6.upper().replace("#", "")
        if len(rgb6) == 6:
            return "FF" + rgb6
        if len(rgb6) == 8:
            return rgb6
    except (IndexError, KeyError, TypeError, AttributeError):
        return None
    return None

def _theme_to_argb(cell: Cell, theme_obj) -> Optional[str]:
    """theme 색상 → ARGB. 베스트에포트."""
    try:
        color = cell.fill.start_color
        theme_idx = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0.0) or 0.0

        if theme_idx is None or theme_obj is None:
            return None

        clr_scheme = getattr(getattr(theme_obj, "themeElements", None), "clrScheme", None)
        if not clr_scheme or not hasattr(clr_scheme, "children"):
            return None

        children = list(clr_scheme.children)
        if theme_idx >= len(children):
            return None

        entry = children[theme_idx]
        srgb = getattr(getattr(entry, "srgbClr", None), "val", None)
        if not srgb:
            return None

        base_rgb = srgb.upper()
        if len(base_rgb) == 6:
            argb = "FF" + base_rgb
        elif len(base_rgb) == 8:
            argb = base_rgb
        else:
            return None

        if abs(tint) < 1e-9:
            return argb

        def blend(c: int, t: float) -> int:
            if t > 0:
                return int(round(c + (255 - c) * t))
            else:
                return int(round(c * (1 + t)))

        r = int(argb[2:4], 16)
        g = int(argb[4:6], 16)
        b = int(argb[6:8], 16)
        r2 = max(0, min(255, blend(r, tint)))
        g2 = max(0, min(255, blend(g, tint)))
        b2 = max(0, min(255, blend(b, tint)))
        return f"FF{r2:02X}{g2:02X}{b2:02X}"
    except Exception:
        return None

def get_cell_argb(cell: Cell, theme_obj=None) -> Optional[str]:
    f = cell.fill
    if not f:
        return None
    sc = getattr(f, "start_color", None)
    if sc is None:
        return None

    t = getattr(sc, "type", None)
    if t == "rgb":
        return _argb_upper(getattr(sc, "rgb", None))
    if t == "indexed":
        return _indexed_to_argb(getattr(sc, "indexed", None))
    if t == "theme":
        return _theme_to_argb(cell, theme_obj)

    # fgColor 보조
    fg = getattr(f, "fgColor", None)
    if fg is not None:
        if getattr(fg, "type", None) == "rgb":
            return _argb_upper(getattr(fg, "rgb", None))
        if getattr(fg, "type", None) == "indexed":
            return _indexed_to_argb(getattr(fg, "indexed", None))
        if getattr(fg, "type", None) == "theme":
            return _theme_to_argb(cell, theme_obj)
    return None

def is_orange(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    if argb:
        u = argb.upper()
        if ("A500" in u) or (u == ORANGE_RGB):
            return True
        try:
            r = int(u[2:4], 16)
            g = int(u[4:6], 16)
            b = int(u[6:8], 16)
            if r > 200 and 100 <= g <= 180 and b < 120:
                return True
        except Exception:
            pass
    if indexed is not None and indexed in ORANGE_INDEX_GUESS:
        return True
    return False

def is_yellow(argb: Optional[str], indexed: Optional[int] = None) -> bool:
    if argb:
        u = argb.upper()
        if ("FF00" in u) or (u == YELLOW_RGB):
            return True
        try:
            r = int(u[2:4], 16)
            g = int(u[4:6], 16)
            b = int(u[6:8], 16)
            if r > 220 and g > 210 and b < 80:
                return True
        except Exception:
            pass
    if indexed is not None and indexed in YELLOW_INDEX_GUESS:
        return True
    return False

# --------------------------------------------------------------------------------------
# 스캔 로직
# --------------------------------------------------------------------------------------
def scan_one_sheet(args: Tuple[str, str, int, int, int, int]) -> SheetScanResult:
    """워커프로세스에서 실행: 워크북 새로 열고 지정 시트 스캔"""
    sheet_name, stage1_path, header_search_max, max_rows, max_cols, case_samples = args

    try:
        wb = load_workbook(stage1_path, data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        total_rows, total_cols = ws.max_row, ws.max_column
        # theme 객체 접근 시도 (없을 수 있음)
        theme_obj = None
        try:
            theme_obj = getattr(wb, "_theme", None)
        except:
            pass

        # 헤더 탐색
        header_row = 1
        for r in range(1, min(header_search_max, total_rows) + 1):
            try:
                v = str(ws.cell(r, 1).value or "").lower()
                if ("고유" in v) or ("중복" in v):
                    header_row = r
                    break
            except:
                continue

        # Case No 컬럼
        case_col_idx: Optional[int] = None
        for c in range(1, min(15, total_cols) + 1):
            try:
                hv = str(ws.cell(header_row, c).value or "").lower()
                if "case" in hv and "no" in hv:
                    case_col_idx = c
                    break
            except:
                continue

        data_start = header_row + 1
        last_row = min(total_rows, data_start + max_rows - 1)
        last_col = min(total_cols, max_cols)

        orange_cells = 0
        yellow_cells = 0
        yellow_rows: List[int] = []

        for r in range(data_start, last_row + 1):
            row_has_yellow = False
            for c in range(1, last_col + 1):
                try:
                    cell = ws.cell(r, c)
                    indexed_val = None
                    sc = getattr(getattr(cell, "fill", None), "start_color", None)
                    if sc is not None:
                        indexed_val = getattr(sc, "indexed", None)

                    argb = get_cell_argb(cell, theme_obj=theme_obj)
                    if is_orange(argb, indexed_val):
                        orange_cells += 1
                    elif is_yellow(argb, indexed_val):
                        yellow_cells += 1
                        row_has_yellow = True
                except:
                    continue

            if row_has_yellow and len(yellow_rows) < case_samples * 2:
                yellow_rows.append(r)

        # Case 샘플
        sample_cases: List[Dict[str, object]] = []
        if case_col_idx:
            seen: set[str] = set()
            # YELLOW 우선
            for r in yellow_rows[:case_samples]:
                try:
                    val = ws.cell(r, case_col_idx).value
                    if val:
                        key = str(val).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "YELLOW"})
                            seen.add(key)
                except:
                    continue

            # ORANGE 보충
            extra_end = min(last_row, data_start + 100)
            added = 0
            for r in range(data_start, extra_end + 1):
                if added >= case_samples:
                    break
                try:
                    val = ws.cell(r, case_col_idx).value
                    if val:
                        key = str(val).strip()
                        if key not in seen:
                            sample_cases.append({"case_no": key, "row": r, "color": "ORANGE?"})
                            seen.add(key)
                            added += 1
                except:
                    continue

        wb.close()
        return SheetScanResult(
            sheet=sheet_name,
            header_row=header_row,
            total_rows=total_rows,
            total_columns=total_cols,
            case_col_idx=case_col_idx,
            orange_cells=orange_cells,
            yellow_cells=yellow_cells,
            yellow_rows=yellow_rows,
            sample_cases=sample_cases,
        )
    except Exception as e:
        # 실패 시 빈 결과 반환
        return SheetScanResult(
            sheet=sheet_name,
            header_row=1,
            total_rows=0,
            total_columns=0,
            case_col_idx=None,
            orange_cells=0,
            yellow_cells=0,
            yellow_rows=[],
            sample_cases=[],
        )

# --------------------------------------------------------------------------------------
# 케이스 추적
# --------------------------------------------------------------------------------------
def track_case_numbers(stage1: Path, scan_results: Dict[str, SheetScanResult]) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    for sname, meta in scan_results.items():
        if meta.case_col_idx is None:
            continue
        try:
            # pandas의 usecols는 열 이름 또는 범위여야 하므로 수정
            df = pd.read_excel(
                stage1, sheet_name=sname, nrows=1000, engine="openpyxl"
            )
        except Exception:
            continue

        case_col = None
        for col in df.columns:
            hv = str(col).lower()
            if "case" in hv and "no" in hv:
                case_col = col
                break
        if case_col is None:
            continue

        targets = {str(x["case_no"]).strip().upper() for x in meta.sample_cases}
        if not targets:
            continue

        ser = df[case_col].astype(str).str.strip().str.upper()
        hits = df[ser.isin(targets)]
        for v in hits[case_col].tolist():
            out.append({"case_no": str(v), "sheet": sname, "found": True})
    return out

# --------------------------------------------------------------------------------------
# Raw 파일 메타
# --------------------------------------------------------------------------------------
def analyze_raw_meta(raw_master: Path, raw_wh: Path) -> Dict[str, Dict]:
    res: Dict[str, Dict] = {"master_file": {}, "warehouse_file": {}}
    if raw_master.exists():
        try:
            xl = pd.ExcelFile(raw_master)
            res["master_file"]["sheets"] = xl.sheet_names
        except Exception:
            pass
    if raw_wh.exists():
        try:
            xl = pd.ExcelFile(raw_wh)
            res["warehouse_file"]["sheets"] = xl.sheet_names
        except Exception:
            pass
    return res

# --------------------------------------------------------------------------------------
# 메인
# --------------------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description="Excel color scan (fast/parallel/theme/indexed)")
    p.add_argument("--stage1", type=Path, default=DEF_STAGE1, help="Stage1 XLSX 경로")
    p.add_argument("--raw-master", type=Path, default=DEF_RAW_MASTER, help="Raw Master 경로")
    p.add_argument("--raw-wh", type=Path, default=DEF_RAW_WH, help="Raw Warehouse 경로")
    p.add_argument("--report", type=Path, default=DEF_REPORT, help="결과 JSON 출력 경로")

    p.add_argument("--header-rows", type=int, default=DEF_HEADER_SEARCH_MAX_ROWS, help="헤더 탐색 최대 행")
    p.add_argument("--rows", type=int, default=DEF_SCAN_MAX_ROWS_PER_SHEET, help="시트당 스캔 행(상단)")
    p.add_argument("--cols", type=int, default=DEF_SCAN_MAX_COLS_PER_ROW, help="행당 스캔 열(좌측)")
    p.add_argument("--sheets", type=int, default=DEF_SHEETS_SCAN_LIMIT, help="스캔할 시트 수(좌측부터)")
    p.add_argument("--samples", type=int, default=DEF_CASE_SAMPLES_PER_COLOR, help="색상별 케이스 샘플 수")

    p.add_argument("--processes", type=int, default=max(1, cpu_count() // 2), help="병렬 프로세스 수")
    p.add_argument("--no-parallel", action="store_true", help="병렬 비활성화")

    args = p.parse_args()

    print("=" * 80)
    print("컬러 작업 로직 빠른 확인 (CLI+병렬+확장색)".center(80))
    print("=" * 80)

    if not args.stage1.exists():
        print(f"[ERROR] Stage1 파일 없음: {args.stage1}")
        return 1

    # 1) Raw 메타
    raw_meta = analyze_raw_meta(args.raw_master, args.raw_wh)

    # 2) 타겟 시트 목록
    try:
        wb0 = load_workbook(args.stage1, read_only=True)
        all_sheets = wb0.sheetnames
        target_sheets = all_sheets[:max(1, args.sheets)]
        wb0.close()
        print(f"[INFO] 대상 시트: {target_sheets}")
    except Exception as e:
        print(f"[ERROR] Stage1 파일 읽기 실패: {e}")
        return 1

    # 3) 시트 스캔 (병렬/단일)
    jobs = [
        (
            s,
            str(args.stage1),
            int(args.header_rows),
            int(args.rows),
            int(args.cols),
            int(args.samples),
        )
        for s in target_sheets
    ]

    if args.no_parallel or len(jobs) == 1:
        results_list = [scan_one_sheet(j) for j in jobs]
    else:
        procs = max(1, int(args.processes))
        with Pool(processes=procs) as pool:
            results_list = pool.map(scan_one_sheet, jobs)

    scan_results: Dict[str, SheetScanResult] = {r.sheet: r for r in results_list}

    # 4) 케이스 추적
    tracking = track_case_numbers(args.stage1, scan_results)

    # 5) 샘플 합산
    total_orange = sum(r.orange_cells for r in scan_results.values())
    total_yellow = sum(r.yellow_cells for r in scan_results.values())

    # 6) 결과 출력
    print(f"\n[결과]")
    print("-" * 80)
    for sname, result in scan_results.items():
        print(f"  시트: {sname}")
        print(f"    - ORANGE 셀: {result.orange_cells}개")
        print(f"    - YELLOW 셀: {result.yellow_cells}개")
        print(f"    - YELLOW 행 샘플: {len(result.yellow_rows)}개")
        if result.sample_cases:
            print(f"    - 샘플 Case No:")
            for case in result.sample_cases[:3]:
                print(f"      {case['case_no']} ({case['color']})")

    # 7) JSON 저장
    out = {
        "raw_files": raw_meta,
        "colors": {
            "sheets": {
                k: {
                    "header_row": v.header_row,
                    "total_rows": v.total_rows,
                    "total_columns": v.total_columns,
                    "case_column": v.case_col_idx,
                    "estimated_orange": v.orange_cells,
                    "estimated_yellow": v.yellow_cells,
                    "yellow_row_samples": v.yellow_rows,
                    "sample_cases": v.sample_cases,
                }
                for k, v in scan_results.items()
            }
        },
        "tracking": tracking,
        "verification": {"sample_orange": total_orange, "sample_yellow": total_yellow},
        "params": {
            "rows": args.rows,
            "cols": args.cols,
            "sheets": args.sheets,
            "samples": args.samples,
            "processes": (0 if args.no_parallel else args.processes),
        },
    }

    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n[OK] 결과 저장: {args.report}")
    print("=" * 80)
    print(f"샘플 ORANGE={total_orange}, YELLOW={total_yellow}")
    print("=" * 80)
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

