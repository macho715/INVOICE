"""
HVDC 파이프라인 Excel 런처 템플릿 생성 스크립트
- Excel 템플릿 파일(.xlsm) 생성
- CONTROL/LOG 시트 기본 구조 생성
- VBA 모듈은 수동 임포트 필요
"""
import openpyxl
from pathlib import Path
from datetime import datetime

def create_excel_template(output_path: str = "HVDC_Pipeline_Launcher.xlsm"):
    """Excel 템플릿 파일 생성"""
    wb = openpyxl.Workbook()
    
    # 기본 시트 삭제
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # CONTROL 시트 생성
    ws_control = wb.create_sheet("CONTROL")
    
    # 제목
    ws_control.merge_cells("A1:H1")
    ws_control["A1"] = "HVDC PIPELINE CONTROL (Excel Only Ops)"
    ws_control["A1"].font = openpyxl.styles.Font(size=16, bold=True)
    
    # 컬럼 너비
    ws_control.column_dimensions["A"].width = 26
    ws_control.column_dimensions["B"].width = 85
    ws_control.column_dimensions["C"].width = 12
    ws_control.column_dimensions["F"].width = 22
    
    # 섹션 헤더
    ws_control.merge_cells("A3:H3")
    ws_control["A3"] = "1) PATH SETTINGS"
    ws_control["A3"].font = openpyxl.styles.Font(bold=True)
    
    # PATH 라벨 (A4~A12)
    path_labels = [
        ("A4", "Project Root"),
        ("A5", "Config Dir"),
        ("A6", "pipeline_config.yaml"),
        ("A7", "stage2_derived_config.yaml"),
        ("A8", "stage4_anomaly.yaml"),
        ("A9", "Python EXE"),
        ("A10", "Scripts Dir"),
        ("A11", "Validator Py"),
        ("A12", "Log Dir"),
    ]
    
    for cell, label in path_labels:
        ws_control[cell] = label
    
    # 기본 수식 (B5~B12)
    ws_control["B5"] = '=IF(B4="","",B4 & "\\config")'
    ws_control["B6"] = '=IF(B5="","",B5 & "\\pipeline_config.yaml")'
    ws_control["B7"] = '=IF(B5="","",B5 & "\\stage2_derived_config.yaml")'
    ws_control["B8"] = '=IF(B5="","",B5 & "\\stage4_anomaly.yaml")'
    ws_control["B10"] = '=IF(B4="","",B4 & "\\scripts")'
    ws_control["B11"] = '=IF(B4="","",B4 & "\\tools\\config_validator.py")'
    ws_control["B12"] = '=IF(B4="","",B4 & "\\logs")'
    
    # Stage 섹션 헤더들
    stage_headers = [
        (14, "2) STAGE 1 (pipeline_config.yaml)"),
        (31, "3) STAGE 2 (stage2_derived_config.yaml)"),
        (37, "4) STAGE 3 (pipeline_config.yaml)"),
        (62, "5) STAGE 4 (stage4_anomaly.yaml)"),
    ]
    
    for row, header in stage_headers:
        ws_control.merge_cells(f"A{row}:H{row}")
        ws_control[f"A{row}"] = header
        ws_control[f"A{row}"].font = openpyxl.styles.Font(bold=True)
    
    # Stage1 기본값
    stage1_defaults = [
        ("A15", "Enabled", "B15", True),
        ("A16", "Master File", "B16", "auto"),
        ("A17", "Warehouse File", "B17", "data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx"),
        ("A18", "Output File (multi-sheet)", "B18", "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx"),
        ("A19", "Output File (merged)", "B19", '=IF(B18="","",LEFT(B18,LEN(B18)-5) & "_merged.xlsx")'),
        ("A25", "Sorting Enabled", "B25", True),
        ("A26", "Sort by Master No", "B26", True),
        ("A27", "Strict Case Key", "B27", True),
        ("A28", "Include All Sheets", "B28", True),
        ("A29", "Header Fallback", "B29", True),
    ]
    
    for label_cell, label, value_cell, value in stage1_defaults:
        ws_control[label_cell] = label
        ws_control[value_cell] = value
    
    # Include Patterns (B20 병합)
    ws_control["A20"] = "Include Patterns (1 line = 1 item)"
    ws_control.merge_cells("B20:H24")
    ws_control["B20"] = "Case List\nRIL\nHE Local\nHE-0214,0252\nSHU\nMIR\nDAS\nAGI"
    ws_control["B20"].alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    # Stage2 기본값
    stage2_defaults = [
        ("A32", "Synced(merged) Input", "B32", '=B19'),
        ("A33", "Derived Output", "B33", "data/processed/derived/HVDC WAREHOUSE_HITACHI(HE).xlsx"),
        ("A34", "Preserve Colors", "B34", True),
        ("A35", "Backup Enabled", "B35", True),
    ]
    
    for label_cell, label, value_cell, value in stage2_defaults:
        ws_control[label_cell] = label
        ws_control[value_cell] = value
    
    # Stage3 기본값
    stage3_defaults = [
        ("A38", "Enabled", "B38", True),
        ("A39", "Derived File", "B39", '=B33'),
        ("A40", "Report Pattern", "B40", "data/processed/reports/HVDC_입고로직_종합리포트_*.xlsx"),
    ]
    
    for label_cell, label, value_cell, value in stage3_defaults:
        ws_control[label_cell] = label
        ws_control[value_cell] = value
    
    # Stage3 Warehouse Table 헤더
    ws_control["A44"] = "Warehouse"
    ws_control["B44"] = "BaseCapacitySQM"
    ws_control["C44"] = "BillingMode"
    ws_control["D44"] = "SqmRate (blank=omit)"
    for cell in ["A44", "B44", "C44", "D44"]:
        ws_control[cell].font = openpyxl.styles.Font(bold=True)
    
    # Stage3 Warehouse 기본값 (A45~D54)
    warehouse_data = [
        ("DSV Al Markaz", 12000, "rate", 47),
        ("DSV Indoor", 8500, "rate", 47),
        ("DSV Outdoor", 15000, "rate", 18),
        ("DSV MZP", 1000, "rate", 33),
        ("DSV MZD", 1000, "rate", None),
        ("AAA Storage", 2000, "passthrough", None),
        ("Hauler Indoor", 1000, "passthrough", None),
        ("JDN MZD", 1000, "rate", 33),
        ("MOSB", 10000, "no-charge", None),
        ("DHL Warehouse", 1000, "passthrough", None),
    ]
    
    for i, (wh, cap, mode, rate) in enumerate(warehouse_data, start=45):
        ws_control[f"A{i}"] = wh
        ws_control[f"B{i}"] = cap
        ws_control[f"C{i}"] = mode
        if rate is not None:
            ws_control[f"D{i}"] = rate
    
    # Stage4 기본값
    stage4_defaults = [
        ("A63", "Input Report File", "B63", "reports/HVDC_입고로직_종합리포트_latest.xlsx"),
        ("A64", "Sheet Name", "B64", "통합_원본데이터_Fixed"),
        ("A65", "Excel Output", "B65", "reports/anomalies/anomaly_list.xlsx"),
        ("A66", "JSON Output", "B66", "reports/anomalies/anomaly_list.json"),
        ("A67", "Viz enable_by_default", "B67", False),
        ("A68", "Viz case_column", "B68", "Case No."),
        ("A69", "Backup Enabled", "B69", True),
        ("A70", "contamination", "B70", 0.02),
        ("A71", "random_state", "B71", 42),
        ("A72", "use_pyod_first", "B72", True),
        ("A73", "iqr_k", "B73", 1.5),
        ("A74", "mad_k", "B74", 3.5),
        ("A75", "min_group_size", "B75", 10),
        ("A76", "rule_boost", "B76", 0.25),
        ("A77", "stat_boost_high", "B77", 0.15),
        ("A78", "stat_boost_med", "B78", 0.08),
        ("A79", "min_risk_to_alert", "B79", 0.9),
    ]
    
    for label_cell, label, value_cell, value in stage4_defaults:
        ws_control[label_cell] = label
        ws_control[value_cell] = value
    
    # LOG 시트 생성
    ws_log = wb.create_sheet("LOG")
    ws_log["A1"] = "Timestamp"
    ws_log["B1"] = "Level"
    ws_log["C1"] = "Procedure"
    ws_log["D1"] = "Code"
    ws_log["E1"] = "Message"
    for cell in ["A1", "B1", "C1", "D1", "E1"]:
        ws_log[cell].font = openpyxl.styles.Font(bold=True)
    
    ws_log.column_dimensions["A"].width = 22
    ws_log.column_dimensions["B"].width = 10
    ws_log.column_dimensions["C"].width = 24
    ws_log.column_dimensions["D"].width = 10
    ws_log.column_dimensions["E"].width = 120
    
    # 기존 파일 삭제 (있다면)
    from pathlib import Path
    if Path(output_path).exists():
        try:
            Path(output_path).unlink()
            print(f"[INFO] 기존 파일 삭제: {output_path}")
        except Exception as e:
            print(f"[WARN] 기존 파일 삭제 실패: {e}")
    
    # .xlsx로 먼저 저장 (더 안정적)
    # openpyxl로 .xlsm을 직접 저장하면 Excel이 인식하지 못할 수 있음
    xlsx_path = output_path.replace(".xlsm", ".xlsx")
    try:
        wb.save(xlsx_path)
        print(f"[OK] Excel 템플릿 생성 완료: {xlsx_path}")
        print("[INFO] 파일을 Excel에서 열어서 .xlsm 형식으로 저장하세요:")
        print("   1. Excel에서 파일 열기")
        print("   2. 파일 > 다른 이름으로 저장")
        print("   3. 파일 형식: Excel 매크로 사용 통합 문서 (*.xlsm)")
        print("   4. Alt+F11 (VBE 열기)")
        print("   5. 모듈 추가 → modHVDC_CONTROL.bas 내용 붙여넣기")
        print("   6. CTRL_Install 실행하여 템플릿 완성")
    except Exception as e:
        print(f"[ERROR] 파일 저장 실패: {e}")
        raise
    
    return xlsx_path


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "HVDC_Pipeline_Launcher.xlsm"
    create_excel_template(output)

