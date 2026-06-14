# -*- coding: utf-8 -*-
"""
DataSynchronizer v3.0 - Semantic Header Matching Edition
=========================================================

This is a completely refactored version of the data synchronizer that uses
the new core header matching system. All hardcoded column names have been
replaced with semantic key lookups.

Key Improvements:
- Zero hardcoding of column names
- Automatic adaptation to different Excel formats
- Robust header detection and matching
- Comprehensive error reporting
- Better maintainability

Migration from v2.9:
- All column name strings replaced with semantic keys
- Header finding logic centralized in core modules
- Enhanced validation and error messages

HEADER MANAGEMENT POLICY:
========================
All header/column definitions and matching logic are centralized in @core/.
This file MUST NOT hardcode any column names. All column references must go through:

- get_warehouse_columns() - Get warehouse column list from core registry
- get_site_columns() - Get site column list from core registry
- SemanticMatcher - Match columns by semantic meaning, not exact names
- HVDC_HEADER_REGISTRY - Central registry for all header definitions

DO NOT hardcode column names like "Case No.", "DHL WH", etc.
ALWAYS use semantic keys like "case_number", "dhl_wh" and let core handle matching.

For details, see: @core/README.md and @core/INTEGRATION_GUIDE.md
"""

from __future__ import annotations
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import sys
import os

# Add project root to Python path
# In standalone mode, stage1_standalone.py already sets up paths
# But we still need to handle direct execution of this file
try:
    # Check if already in path (standalone mode)
    import scripts.core
except ImportError:
    # Direct execution: setup paths
    project_root = Path(__file__).parent.parent.parent
    sys.path.insert(0, str(project_root))

# Import the new core header matching system
from scripts.core import (
    SemanticMatcher,
    find_header_by_meaning,
    detect_header_row,
    HVDC_HEADER_REGISTRY,
    HeaderCategory,
    HeaderRegistry,
    STAGE1_BASE_COLS_ORDER,
)
from scripts.core.standard_header_order import reorder_dataframe_columns

# ===== Configuration =====
ORANGE = "FFFFA500"  # Changed date cell (ARGB format) - True ORANGE
YELLOW = "FFFFFF00"  # New row (ARGB format)

# Invalid header patterns to filter out
INVALID_HEADER_PATTERNS = [
    r"^열\d+$",  # 열1, 열2 등
    r"^\d+$",  # 0, 1, 2, 3, 4 등 순수 숫자
    r"^총합계$",  # 총합계
    r"^Unnamed:.*$",  # Unnamed: 0, Unnamed: 1 등
    r"^\.+$",  # ... 등 점만 있는 컬럼
]

# Metadata columns that should not be overwritten during sync
METADATA_COLUMNS = [
    "Source_Sheet",  # Original sheet name - should be preserved
    "Source_Vendor",  # Vendor name (HITACHI/SIEMENS) - should be preserved
]

# Sheet names to exclude (aggregate/summary sheets)
EXCLUDED_SHEET_NAMES = [
    "summary",  # English
    "총합계",  # Korean total
    "total",  # English total
    "aggregate",  # Aggregate data
]

# Sheet name aliases for semantic matching
SHEET_NAME_ALIASES = {
    "case list": ["case list, ril", "case list ril", "caselist", "case_list"],
    "he capacitor": ["he-0214,0252 (capacitor)", "he-0214", "he-0252", "capacitor"],
}

# File name keywords for identifying same file types
FILE_NAME_KEYWORDS = {
    "master": ["case list", "case_list", "caselist", "master"],
    "warehouse": ["hvdc", "hitachi", "warehouse", "he"],
}

# ✅ Corrected: These columns contain WAREHOUSE IN/OUT DATES
# Column names say "location" but actual data is DATES (warehouse entry/exit timestamps)
DATE_SEMANTIC_KEYS = [
    "etd_atd",  # Estimated/Actual Time of Departure
    "eta_ata",  # Estimated/Actual Time of Arrival
    "dhl_wh",  # DHL Warehouse IN date
    "dsv_indoor",  # DSV Indoor Warehouse IN date
    "dsv_al_markaz",  # DSV Al Markaz Warehouse IN date
    "dsv_outdoor",  # DSV Outdoor Warehouse IN date
    "aaa_storage",  # AAA Storage IN date
    "hauler_indoor",  # Hauler Indoor IN date
    "dsv_mzp",  # DSV MZP IN date
    "mosb",  # MOSB IN date
    "shifting",  # Shifting date
    "mir",  # MIR site IN date
    "shu",  # SHU site IN date
    "das",  # DAS site IN date
    "agi",  # AGI site IN date
]

ALWAYS_OVERWRITE_NONDATE = True


def _to_date(val) -> Optional[pd.Timestamp]:
    """Convert various date formats to pandas Timestamp."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        if isinstance(val, pd.Timestamp):
            return val
        return pd.to_datetime(val, errors="coerce")
    except Exception:
        return None


@dataclass
class Change:
    """Record of a single cell change."""

    row_index: int
    column_name: str
    old_value: Any
    new_value: Any
    change_type: str  # "date_update" | "field_update" | "new_record"
    semantic_key: str = ""  # ✅ Phase 1: Semantic key for flexible column mapping
    case_no: str = ""  # ✅ Phase 4: Case No. for correct row lookup after reordering


@dataclass
class ChangeTracker:
    """Tracks all changes made during synchronization."""

    changes: List[Change] = field(default_factory=list)
    new_cases: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    semantic_to_column: Dict[str, str] = field(default_factory=dict)  # ✅ Phase 2: Mapping storage

    def set_column_mapping(self, semantic_key: str, column_name: str):
        """✅ Phase 2: Store semantic key → actual column name mapping"""
        if semantic_key and column_name:
            self.semantic_to_column[semantic_key] = column_name

    def get_column_name(self, semantic_key: str, fallback: str = "") -> str:
        """✅ Phase 2: Get actual column name from semantic key (with fallback)"""
        return self.semantic_to_column.get(semantic_key, fallback)

    def add_change(self, **kw):
        """Add a change record."""
        self.changes.append(
            Change(
                row_index=int(kw.get("row_index", -1)),
                column_name=str(kw.get("column_name", "")),
                old_value=kw.get("old_value"),
                new_value=kw.get("new_value"),
                change_type=str(kw.get("change_type", "field_update")),
                semantic_key=str(kw.get("semantic_key", "")),  # ✅ Phase 2: Include semantic_key
                case_no=str(kw.get("case_no", "")),  # ✅ Phase 4: Include case_no
            )
        )

    def log_new_case(self, case_no: str, row_data: Dict[str, Any], row_index: Optional[int] = None):
        """Log a new case that was appended."""
        self.new_cases[str(case_no)] = dict(row_data or {})
        if row_index is not None:
            self.add_change(
                row_index=row_index,
                column_name="",
                old_value=None,
                new_value=None,
                change_type="new_record",
            )


@dataclass
class SyncResult:
    """Result of a synchronization operation."""

    success: bool
    message: str
    output_path: str
    stats: Dict[str, Any]
    matching_report: Optional[str] = None  # New: matching diagnostics


class DataSynchronizerV30:
    """
    Advanced data synchronizer with semantic header matching.

    This version eliminates all hardcoded column names and uses semantic
    matching to automatically find the correct columns regardless of how
    they're named in the Excel file.

    The synchronizer works in several phases:
    1. Load files and detect header rows
    2. Match all required columns using semantic keys
    3. Validate that required columns were found
    4. Perform synchronization using matched column names
    5. Apply Excel formatting to highlight changes

    Example:
        >>> sync = DataSynchronizerV30()
        >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
        >>> if result.success:
        >>>     print(f"Synchronized to {result.output_path}")
        >>> else:
        >>>     print(f"Error: {result.message}")
    """

    def __init__(self, date_semantic_keys: Optional[List[str]] = None) -> None:
        """
        Initialize the synchronizer.

        Args:
            date_semantic_keys: List of semantic keys for date columns.
                If None, uses the default DATE_SEMANTIC_KEYS.
        """
        # Use semantic keys instead of hardcoded column names
        self.date_semantic_keys = date_semantic_keys or DATE_SEMANTIC_KEYS

        # Initialize the semantic matcher
        self.matcher = SemanticMatcher(min_confidence=0.7, allow_partial=True)

        # Change tracking
        self.change_tracker = ChangeTracker()

        # Column mapping storage (will be populated during matching)
        self.master_columns: Dict[str, str] = {}  # semantic_key -> actual_column
        self.warehouse_columns: Dict[str, str] = {}

        # Debug tracking flag
        self.debug_dhl_wh = True  # Enable DHL WH tracking for debugging

        print("[OK] DataSynchronizer v3.0 initialized with semantic header matching")

    def _dates_equal(self, a, b) -> bool:
        """Check if two dates are equal, ignoring format differences."""
        da = _to_date(a)
        db = _to_date(b)
        if da is None and db is None:
            return True
        if da is None or db is None:
            return False
        if pd.isna(da) or pd.isna(db):
            return pd.isna(da) and pd.isna(db)
        return da.normalize() == db.normalize()

    def _track_dhl_wh(self, df: pd.DataFrame, stage_name: str) -> None:
        """Track DHL WH data through pipeline stages (DEBUG)"""
        if not self.debug_dhl_wh:
            return

        dhl_col = None
        # Try to find DHL WH column
        for col in df.columns:
            if "DHL" in str(col).upper() and "WH" in str(col).upper():
                dhl_col = col
                break

        if dhl_col:
            count = df[dhl_col].notna().sum()
            print(f"[DHL WH TRACK] {stage_name}: '{dhl_col}' = {count}건 데이터")
        else:
            print(f"[DHL WH TRACK] {stage_name}: DHL WH 컬럼 없음")
            print(
                f"[DHL WH TRACK]   사용 가능 컬럼: {[c for c in df.columns if 'DHL' in str(c).upper() or 'WH' in str(c).upper()][:5]}"
            )

    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """
        Check if sheet should be skipped (aggregate/summary sheets)

        Args:
            sheet_name: Name of the sheet

        Returns:
            True if sheet should be skipped
        """
        normalized = sheet_name.strip().lower()
        return normalized in EXCLUDED_SHEET_NAMES

    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """
        시트명을 정규화하여 매칭 키 생성

        Args:
            sheet_name: 원본 시트명

        Returns:
            정규화된 시트명 (소문자, 특수문자/공백 정리)
        """
        normalized = sheet_name.strip().lower()
        # 특수문자 제거 (쉼표, 괄호, 하이픈)
        normalized = normalized.replace(",", "").replace("(", "").replace(")", "")
        normalized = normalized.replace("-", " ").replace("_", " ")
        # 연속된 공백을 하나로
        normalized = " ".join(normalized.split())
        return normalized

    def _find_matching_sheet(self, target_sheet: str, available_sheets: List[str]) -> Optional[str]:
        """
        시트명 semantic matching

        Args:
            target_sheet: 찾으려는 시트명 (Master)
            available_sheets: 사용 가능한 시트명 목록 (Warehouse)

        Returns:
            매칭된 시트명 또는 None
        """
        target_norm = self._normalize_sheet_name(target_sheet)

        # 1. Exact match first
        for sheet in available_sheets:
            if self._normalize_sheet_name(sheet) == target_norm:
                return sheet

        # 2. Alias matching
        for base_name, aliases in SHEET_NAME_ALIASES.items():
            # Check if target matches base or any alias
            if target_norm == base_name or target_norm in aliases:
                # Find warehouse sheet that also matches
                for sheet in available_sheets:
                    sheet_norm = self._normalize_sheet_name(sheet)
                    if sheet_norm == base_name or sheet_norm in aliases:
                        return sheet

        # 3. Partial keyword matching (fallback)
        target_keywords = set(target_norm.split())
        best_match = None
        best_score = 0

        for sheet in available_sheets:
            sheet_norm = self._normalize_sheet_name(sheet)
            sheet_keywords = set(sheet_norm.split())

            # Calculate overlap
            overlap = len(target_keywords & sheet_keywords)
            if overlap > best_score:
                best_score = overlap
                best_match = sheet

        # Only return if significant overlap (at least 2 common words)
        if best_score >= 2:
            return best_match

        return None

    def _normalize_file_name(self, file_name: str) -> str:
        """
        파일명을 정규화

        Args:
            file_name: 원본 파일명 (확장자 포함 가능)

        Returns:
            정규화된 파일명
        """
        # 확장자 제거
        if "." in file_name:
            file_name = file_name.rsplit(".", 1)[0]

        normalized = file_name.strip().lower()
        # 특수문자 제거
        normalized = normalized.replace("_", " ").replace("-", " ")
        normalized = " ".join(normalized.split())
        return normalized

    def _identify_file_type(self, file_path: str) -> str:
        """
        파일 타입 식별 (master 또는 warehouse)

        Args:
            file_path: 파일 경로

        Returns:
            'master', 'warehouse', 또는 'unknown'
        """
        from pathlib import Path

        file_name = Path(file_path).name
        file_norm = self._normalize_file_name(file_name)

        # Check against keywords
        for file_type, keywords in FILE_NAME_KEYWORDS.items():
            for keyword in keywords:
                if keyword in file_norm:
                    return file_type

        return "unknown"

    def _detect_vendor_and_header_row(self, file_path: Path) -> Tuple[str, int]:
        """
        Detect vendor type and appropriate header row based on file name.

        Args:
            file_path: Path to the file

        Returns:
            Tuple of (vendor_type, header_row_index)
        """
        file_norm = file_path.stem.lower()

        if "simense" in file_norm or "sim" in file_norm:
            return ("SIEMENS", 0)  # SIEMENS는 첫 행이 헤더
        else:
            return ("HITACHI", 4)  # HITACHI는 5번째 행이 헤더

    def _load_master_files(self, master_xlsx: str) -> Dict[str, pd.DataFrame]:
        """
        Load Master files (HITACHI + SIEMENS) and merge them.

        Args:
            master_xlsx: Path to Master Excel file (HITACHI)

        Returns:
            Dictionary of sheet_name -> DataFrame
        """
        from pathlib import Path

        master_sheets = {}

        # Load HITACHI Master file
        print(f"[INFO] Loading HITACHI Master file: {Path(master_xlsx).name}")
        hitachi_sheets_data = self._load_file_by_sheets(master_xlsx, "Master")
        # Extract DataFrames from tuples
        hitachi_sheets = {name: df for name, (df, _) in hitachi_sheets_data.items()}

        # ✅ Set Source_Vendor and Source_Sheet for ALL HITACHI data
        for sheet_name, df in hitachi_sheets.items():
            df["Source_Vendor"] = "HITACHI"
            df["Source_Sheet"] = sheet_name
            print(
                f"[HITACHI] Set Source_Vendor='HITACHI', Source_Sheet='{sheet_name}' for {len(df)} rows"
            )

        master_sheets.update(hitachi_sheets)

        # Look for SIEMENS file in the same directory first
        master_dir = Path(master_xlsx).parent
        siemens_files = list(master_dir.glob("*SIMENSE*.xls*")) + list(
            master_dir.glob("*SIM*.xls*")
        )

        # If not found, search in parent directory's subdirectories
        if not siemens_files:
            parent_dir = master_dir.parent  # data/raw/
            print(f"[INFO] Searching for SIEMENS files in subdirectories of {parent_dir}")
            siemens_files = list(parent_dir.glob("*/*SIMENSE*.xls*")) + list(
                parent_dir.glob("*/*SIM*.xls*")
            )

        if siemens_files:
            siemens_file = siemens_files[0]  # Take the first SIEMENS file found
            print(f"[INFO] Found SIEMENS file: {siemens_file.name}")
            print(f"[INFO] Loading SIEMENS Master file...")

            siemens_sheets_data = self._load_file_by_sheets(str(siemens_file), "SIEMENS")
            # Extract DataFrames from tuples
            siemens_sheets = {name: df for name, (df, _) in siemens_sheets_data.items()}

            # ===== SIEMENS 데이터 정제 및 컬럼 통합 =====
            print("[INFO] Cleaning and mapping SIEMENS data...")
            for sheet_name, siemens_df in siemens_sheets.items():
                # 1. Bill of Lading 컬럼 제거
                bill_of_lading_patterns = ["billof", "billoflading", "b/l", "bl", "lading"]
                bl_columns = [
                    col
                    for col in siemens_df.columns
                    if any(
                        pattern in str(col).lower().replace(" ", "").replace("_", "")
                        for pattern in bill_of_lading_patterns
                    )
                ]
                if bl_columns:
                    print(
                        f"[SIEMENS] Dropping Bill of Lading columns from '{sheet_name}': {bl_columns}"
                    )
                    siemens_df.drop(columns=bl_columns, inplace=True, errors="ignore")

                # 2. columns_to_drop 리스트 초기화
                columns_to_drop = []

                # 3. SIEMENS 컬럼을 HITACHI 컬럼명으로 매핑
                column_mapping = {
                    "PackageNo": "Case No.",
                    "PO.No": "EQ No",
                    "HSCode": "HS Code",
                    "No.": "no.",  # 시퀀스 번호 통합
                }

                print(f"[SIEMENS] Mapping columns to HITACHI structure for '{sheet_name}'...")
                for siemens_col, hitachi_col in column_mapping.items():
                    if siemens_col in siemens_df.columns:
                        if hitachi_col in siemens_df.columns:
                            # 기존 HITACHI 컬럼이 있으면 빈 값만 SIEMENS 데이터로 채우기
                            null_count = siemens_df[hitachi_col].isna().sum()
                            if null_count > 0:
                                siemens_df[hitachi_col].fillna(
                                    siemens_df[siemens_col], inplace=True
                                )
                                print(
                                    f"  - Filled {null_count} null values in '{hitachi_col}' with '{siemens_col}' data"
                                )
                        else:
                            # HITACHI 컬럼이 없으면 SIEMENS 컬럼을 rename
                            siemens_df.rename(columns={siemens_col: hitachi_col}, inplace=True)
                            print(f"  - Renamed '{siemens_col}' → '{hitachi_col}'")

                # 4. SIEMENS 전용 컬럼 무조건 제거 (매핑 여부 관계없이)
                siemens_specific_cols = ["No.", "PackageNo", "PO.No", "HSCode", "BillofLading"]
                for siemens_col, hitachi_col in column_mapping.items():
                    # 매핑 후 원본 SIEMENS 컬럼이 남아있으면 제거
                    if (
                        siemens_col in siemens_df.columns
                        and hitachi_col in siemens_df.columns
                        and siemens_col != hitachi_col
                    ):
                        columns_to_drop.append(siemens_col)

                # SIEMENS 전용 컬럼 무조건 제거
                for col in siemens_specific_cols:
                    if col in siemens_df.columns:
                        columns_to_drop.append(col)

                if columns_to_drop:
                    print(
                        f"[SIEMENS] Dropping SIEMENS-specific columns from '{sheet_name}': {columns_to_drop}"
                    )
                    siemens_df.drop(columns=columns_to_drop, inplace=True, errors="ignore")

                # 업데이트된 DataFrame 저장
                siemens_sheets[sheet_name] = siemens_df

            print("[OK] SIEMENS data cleaning and column mapping completed")

            # Merge SIEMENS sheets with HITACHI sheets
            for sheet_name, siemens_df in siemens_sheets.items():
                if sheet_name in master_sheets:
                    # Merge with existing HITACHI sheet
                    print(f"[INFO] Merging SIEMENS sheet '{sheet_name}' with HITACHI sheet")
                    hitachi_df = master_sheets[sheet_name]

                    # Add source vendor and sheet columns (HITACHI already has these, but SIEMENS needs them)
                    # hitachi_df["Source_Vendor"] = "HITACHI"  # Already set above
                    siemens_df["Source_Vendor"] = "SIEMENS"
                    siemens_df["Source_Sheet"] = sheet_name

                    # Merge DataFrames
                    merged_df = pd.concat([hitachi_df, siemens_df], ignore_index=True, sort=False)

                    # CRITICAL: Remove duplicates based on Case No
                    if "Case No." in merged_df.columns:
                        before_dedup = len(merged_df)
                        merged_df = merged_df.drop_duplicates(subset=["Case No."], keep="first")
                        after_dedup = len(merged_df)
                        removed = before_dedup - after_dedup
                        if removed > 0:
                            print(
                                f"[DEDUP] Removed {removed} duplicate Case No entries from merged data"
                            )

                    master_sheets[sheet_name] = merged_df

                    print(
                        f"[OK] Merged '{sheet_name}': HITACHI({len(hitachi_df)}) + SIEMENS({len(siemens_df)}) = {len(merged_df)} rows after dedup"
                    )
                else:
                    # New sheet from SIEMENS
                    print(f"[INFO] Adding new SIEMENS sheet '{sheet_name}'")
                    siemens_df["Source_Vendor"] = "SIEMENS"
                    siemens_df["Source_Sheet"] = sheet_name
                    master_sheets[sheet_name] = siemens_df
        else:
            print(f"[INFO] No SIEMENS file found in {master_dir}")

        return master_sheets

    def _load_file_with_header_detection(
        self, file_path: str, file_label: str
    ) -> Tuple[pd.DataFrame, int]:
        """
        Load all sheets from Excel file with vendor-specific header row detection.

        Args:
            file_path: Path to the Excel file
            file_label: Label for logging (e.g., "Master", "Warehouse")

        Returns:
            Tuple of (merged_dataframe, header_row_index)
        """
        print(f"\n{'='*60}")
        print(f"Loading {file_label} file: {Path(file_path).name}")
        print(f"{'='*60}")

        xl = pd.ExcelFile(file_path, engine="openpyxl")
        all_dfs = []
        header_row = None

        print(f"Found {len(xl.sheet_names)} sheets in file")

        # Detect vendor and appropriate header row
        vendor_type, vendor_header_row = self._detect_vendor_and_header_row(Path(file_path))
        print(f"[INFO] Detected vendor: {vendor_type}, using header row: {vendor_header_row}")

        for sheet_name in xl.sheet_names:
            print(f"\n  Loading sheet: '{sheet_name}'")

            # Skip summary/aggregate sheets
            if self._should_skip_sheet(sheet_name):
                print(f"  [SKIP] Aggregate sheet (not Case data)")
                continue

            # Use vendor-specific header row
            sheet_header_row = vendor_header_row
            confidence = 1.0  # Vendor-specific detection is 100% confident

            if header_row is None:
                header_row = sheet_header_row

            print(
                f"  [OK] Header at row {sheet_header_row} (confidence: {confidence:.0%}) [{vendor_type}]"
            )

            # Load sheet
            df = pd.read_excel(
                xl, sheet_name=sheet_name, header=sheet_header_row, engine="openpyxl"
            )

            if df.empty:
                print(f"  [SKIP] Empty sheet")
                continue

            # Track source sheet (preserve original sheet name)
            df["Source_Sheet"] = sheet_name

            # DEBUG: Track DHL WH in each sheet
            self._track_dhl_wh(df, f"Sheet '{sheet_name}' loaded")

            all_dfs.append(df)
            print(f"  [OK] {len(df)} rows loaded")

        if not all_dfs:
            raise ValueError(f"No valid sheets found in {file_label}")

        # Merge all sheets
        merged_df = pd.concat(all_dfs, ignore_index=True, sort=False)
        print(f"\n[OK] Total: {len(merged_df)} rows from {len(all_dfs)} sheets")

        # DEBUG: Track after merge
        self._track_dhl_wh(merged_df, "After pd.concat (merge)")

        # DEBUG: Check for duplicate column names and WH columns
        if self.debug_dhl_wh:
            from collections import Counter

            col_counts = Counter(merged_df.columns)
            duplicates = {col: count for col, count in col_counts.items() if count > 1}
            if duplicates:
                print(f"  [⚠️WARNING⚠️] Duplicate column names detected!")
                for col, count in duplicates.items():
                    print(f"    - '{col}' appears {count} times")
            else:
                print(f"  [OK] No duplicate column names")

            # Show exact column positions for WH columns
            wh_cols_with_idx = [
                (i, c) for i, c in enumerate(merged_df.columns) if "WH" in str(c).upper()
            ]
            print(f"  [DEBUG] WH columns after concat (position, name):")
            for idx, col in wh_cols_with_idx:
                data_count = merged_df[col].notna().sum()
                print(f"    - [{idx}] '{col}' = {data_count}건")

        # Filter out invalid columns (신규 추가)
        print("\nFiltering invalid columns:")
        merged_df = self._filter_invalid_columns(merged_df)
        self._track_dhl_wh(merged_df, "After filter_invalid_columns")

        # Consolidate incorrectly named warehouse columns
        print("\nConsolidating warehouse columns:")
        merged_df = self._consolidate_warehouse_columns(merged_df)
        self._track_dhl_wh(merged_df, "After consolidate_warehouse_columns")

        # Ensure all location columns exist
        print("\nEnsuring all location columns:")
        merged_df = self._ensure_all_location_columns(merged_df)
        self._track_dhl_wh(merged_df, "After ensure_all_location_columns")

        return merged_df, header_row

    def _load_file_by_sheets(
        self, file_path: str, file_label: str
    ) -> Dict[str, Tuple[pd.DataFrame, int]]:
        """
        Load each sheet from Excel file separately with automatic header row detection.

        Args:
            file_path: Path to the Excel file
            file_label: Label for logging (e.g., "Master", "Warehouse")

        Returns:
            Dict mapping sheet_name to (dataframe, header_row_index)
        """
        print(f"\n{'='*60}")
        print(f"Loading {file_label} file by sheets: {Path(file_path).name}")
        print(f"{'='*60}")

        xl = pd.ExcelFile(file_path, engine="openpyxl")
        sheet_data = {}
        header_row = None

        print(f"Found {len(xl.sheet_names)} sheets in file")

        for sheet_name in xl.sheet_names:
            print(f"\n  Loading sheet: '{sheet_name}'")

            # Skip summary/aggregate sheets
            if self._should_skip_sheet(sheet_name):
                print(f"  [SKIP] Aggregate sheet (not Case data)")
                continue

            # Detect header row for this sheet
            sheet_header_row, confidence = detect_header_row(file_path, sheet_name)

            if header_row is None:
                header_row = sheet_header_row

            print(f"  [OK] Header at row {sheet_header_row} (confidence: {confidence:.0%})")

            # Load sheet
            df = pd.read_excel(
                xl, sheet_name=sheet_name, header=sheet_header_row, engine="openpyxl"
            )

            if df.empty:
                print(f"  [SKIP] Empty sheet")
                continue

            # Track source sheet (preserve original sheet name)
            df["Source_Sheet"] = sheet_name

            # DEBUG: Track DHL WH in each sheet
            self._track_dhl_wh(df, f"Sheet '{sheet_name}' loaded")

            # Process each sheet independently
            df = self._filter_invalid_columns(df)
            df = self._consolidate_warehouse_columns(df)
            df = self._ensure_all_location_columns(df)

            sheet_data[sheet_name] = (df, sheet_header_row)
            print(f"  [OK] {len(df)} rows loaded")

        if not sheet_data:
            raise ValueError(f"No valid sheets found in {file_label}")

        print(f"\n[OK] Loaded {len(sheet_data)} sheets")
        return sheet_data

    def _filter_invalid_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        잘못된 헤더를 가진 컬럼 제거

        Args:
            df: 필터링할 DataFrame

        Returns:
            유효한 컬럼만 포함된 DataFrame
        """
        import re

        valid_columns = []
        removed_columns = []

        for col in df.columns:
            col_str = str(col).strip()

            # 잘못된 패턴 검사
            is_invalid = False
            for pattern in INVALID_HEADER_PATTERNS:
                if re.match(pattern, col_str):
                    is_invalid = True
                    removed_columns.append(col)
                    break

            if not is_invalid:
                valid_columns.append(col)

        if removed_columns:
            print(f"  [CLEANUP] Removed {len(removed_columns)} invalid columns:")
            for col in removed_columns:
                print(f"    - '{col}'")

        return df[valid_columns]

    def _consolidate_warehouse_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Consolidate incorrectly named warehouse columns.

        Some raw data files use incorrect column names for the same warehouse.
        This method merges such columns to ensure data consistency:
        - "DSV WH" → "DSV Indoor" (HE Local sheet uses incorrect name)

        NOTE: DHL WH is a separate, valid warehouse and should NOT be consolidated!

        Args:
            df: DataFrame with potentially incorrect column names

        Returns:
            DataFrame with consolidated columns
        """
        consolidations = {
            "DSV WH": "DSV Indoor",  # HE Local sheet incorrectly names DSV Indoor as DSV WH
        }

        # DEBUG: Check DHL WH before consolidation
        if self.debug_dhl_wh:
            dhl_before = "DHL WH" in df.columns
            all_cols_before = list(df.columns)
            print(f"  [DEBUG] Before consolidation:")
            print(f"    - Total columns: {len(all_cols_before)}")
            print(f"    - DHL WH exists: {dhl_before}")
            if dhl_before:
                dhl_idx = all_cols_before.index("DHL WH")
                print(f"    - DHL WH position: {dhl_idx}")
                print(f"    - DHL WH data count: {df['DHL WH'].notna().sum()}건")

            # Show all warehouse-like columns
            wh_cols = [
                c
                for c in df.columns
                if "WH" in str(c).upper() or "DSV" in str(c).upper() or "DHL" in str(c).upper()
            ]
            print(f"    - WH/DSV/DHL columns: {wh_cols}")

        for wrong_name, correct_name in consolidations.items():
            if wrong_name in df.columns:
                # DEBUG: Track before rename
                if self.debug_dhl_wh:
                    dhl_before_rename = "DHL WH" in df.columns
                    print(f"  [DEBUG] Before '{wrong_name}' → '{correct_name}':")
                    print(f"    - DHL WH exists: {dhl_before_rename}")

                if correct_name in df.columns:
                    # Merge data: use correct_name where it exists, fill with wrong_name otherwise
                    df[correct_name] = df[correct_name].fillna(df[wrong_name])
                    df = df.drop(columns=[wrong_name])
                    print(f"  [OK] Merged '{wrong_name}' → '{correct_name}'")
                else:
                    # Just rename if correct column doesn't exist yet
                    # IMPORTANT: Use columns.str.replace() to avoid affecting other columns
                    new_columns = []
                    renamed = False
                    for col in df.columns:
                        if col == wrong_name and not renamed:
                            # Only rename the FIRST occurrence
                            new_columns.append(correct_name)
                            renamed = True
                            print(
                                f"  [OK] Renamed '{wrong_name}' → '{correct_name}' at position {len(new_columns)-1}"
                            )
                        else:
                            new_columns.append(col)

                    df.columns = new_columns

                # DEBUG: Track after rename
                if self.debug_dhl_wh:
                    dhl_after_rename = "DHL WH" in df.columns
                    print(f"  [DEBUG] After '{wrong_name}' → '{correct_name}':")
                    print(f"    - DHL WH exists: {dhl_after_rename}")
                    if dhl_before_rename and not dhl_after_rename:
                        print(f"  [🚨CRITICAL🚨] DHL WH was deleted by this rename operation!")
                        print(
                            f"    - Operation: rename({{{repr(wrong_name)}: {repr(correct_name)}}})"
                        )
                        print(f"    - DataFrame columns now: {len(df.columns)}")

        # DEBUG: Check DHL WH after consolidation
        if self.debug_dhl_wh:
            dhl_after = "DHL WH" in df.columns
            all_cols_after = list(df.columns)
            print(f"  [DEBUG] After consolidation:")
            print(f"    - Total columns: {len(all_cols_after)}")
            print(f"    - DHL WH exists: {dhl_after}")

            # Show all warehouse-like columns after
            wh_cols_after = [
                c
                for c in df.columns
                if "WH" in str(c).upper() or "DSV" in str(c).upper() or "DHL" in str(c).upper()
            ]
            print(f"    - WH/DSV/DHL columns: {wh_cols_after}")

            if dhl_before and not dhl_after:
                print(f"  [🚨ERROR🚨] DHL WH was LOST during consolidation!")
                print(f"    - Columns lost: {set(all_cols_before) - set(all_cols_after)}")
                print(f"    - Columns gained: {set(all_cols_after) - set(all_cols_before)}")

        return df

    def _ensure_all_location_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ensure all warehouse and site columns exist, but keep Shifting in original position.
        Source_Sheet is metadata and should not be included in ordering logic.

        Args:
            df: DataFrame to check and update

        Returns:
            DataFrame with all location columns present in correct order
        """
        # Get column order from core registry (Single Source of Truth)
        # All warehouse/site column definitions are centrally managed in @core/header_registry.py
        from scripts.core import get_warehouse_columns, get_site_columns

        WAREHOUSE_ORDER = get_warehouse_columns()
        SITE_ORDER = get_site_columns()

        # Combined order: Warehouse first, then Site
        all_locations = WAREHOUSE_ORDER + SITE_ORDER
        location_set = set(all_locations)

        # Check and add missing columns
        missing_cols = []
        for location in all_locations:
            if location not in df.columns:
                df[location] = pd.NaT
                missing_cols.append(location)

        if missing_cols:
            print(f"  [OK] Added {len(missing_cols)} missing location columns:")
            for col in missing_cols:
                print(f"    - {col}")
        else:
            print(f"  [OK] All location columns present")

        # Separate columns into groups using core's standard order

        # Ensure "no." column exists (create if missing)
        if "no." not in df.columns:
            # Create "no." as sequential row number
            df.insert(0, "no.", range(1, len(df) + 1))
            print("  [OK] Created 'no.' column as row index")

        # 1. Base columns in standard order (from core)
        base_cols_in_df = [col for col in STAGE1_BASE_COLS_ORDER if col in df.columns]

        # 2. Extra base columns not in standard order (for dynamic columns)
        # Exclude 'no' (lowercase without dot) as it's likely a duplicate of 'no.'
        extra_base_cols = [
            col
            for col in df.columns
            if col not in STAGE1_BASE_COLS_ORDER
            and col not in location_set
            and col != "Shifting"
            and col != "Source_Sheet"
            and col != "no"  # Exclude 'no' (keep only 'no.')
        ]

        base_cols = base_cols_in_df + extra_base_cols

        # 3. Special columns
        shifting_col = "Shifting" if "Shifting" in df.columns else None
        source_sheet_col = "Source_Sheet" if "Source_Sheet" in df.columns else None

        # Build final column order:
        # base_cols + warehouse_cols + shifting + site_cols + source_sheet
        final_order = (
            base_cols
            + WAREHOUSE_ORDER
            + ([shifting_col] if shifting_col else [])
            + SITE_ORDER
            + ([source_sheet_col] if source_sheet_col else [])
        )

        # Reorder dataframe
        df = df[[c for c in final_order if c in df.columns]]

        print(
            f"  [OK] Column order: base({len(base_cols)}) + warehouses({len(WAREHOUSE_ORDER)}) + Shifting + sites({len(SITE_ORDER)}) + Source_Sheet"
        )
        if base_cols:
            print(f"  [DEBUG] First 5 base columns: {base_cols[:5]}")

        return df

    def _match_and_validate_headers(self, df: pd.DataFrame, file_label: str) -> Dict[str, str]:
        """
        Match semantic keys to actual column names and validate required columns.

        Args:
            df: The DataFrame to match against
            file_label: Label for logging

        Returns:
            Dictionary mapping semantic_key to actual column name

        Raises:
            ValueError: If required columns are not found
        """
        print(f"\nMatching headers for {file_label}...")

        # Define required semantic keys for synchronization
        required_keys = [
            "case_number",  # Must have case number for matching
        ]

        # All keys we want to find (required + date columns + location columns)
        location_headers = HVDC_HEADER_REGISTRY.get_by_category(HeaderCategory.LOCATION)
        location_keys = [h.semantic_key for h in location_headers]
        all_keys = required_keys + self.date_semantic_keys + location_keys

        # Perform semantic matching
        report = self.matcher.match_dataframe(df, all_keys)

        # Print summary
        print(f"  Matched: {report.successful_matches}/{report.total_semantic_keys}")
        print(f"  Success rate: {report.successful_matches/report.total_semantic_keys:.0%}")

        # Validate required columns
        missing_required = []
        for key in required_keys:
            if not report.get_column_name(key):
                missing_required.append(key)

        if missing_required:
            # Print detailed error information
            print(f"\n{'='*60}")
            print(f"ERROR: Missing required columns in {file_label}")
            print(f"{'='*60}")
            report.print_summary()
            raise ValueError(
                f"Required semantic keys not found in {file_label}: {missing_required}"
            )

        # Show what was matched
        print("\n  Key matches:")
        for key in required_keys + self.date_semantic_keys[:3]:  # Show first few
            col = report.get_column_name(key)
            if col:
                print(f"    - {key:20s} → '{col}'")

        if len(self.date_semantic_keys) > 3:
            remaining = len([k for k in self.date_semantic_keys if report.get_column_name(k)])
            print(f"    ... and {remaining} more date columns")

        # Build mapping dictionary
        column_mapping = {
            key: report.get_column_name(key) for key in all_keys if report.get_column_name(key)
        }

        return column_mapping

    def _build_case_index(self, df: pd.DataFrame, case_col: str) -> Dict[str, int]:
        """
        Build an index mapping case numbers to row indices.

        Args:
            df: The DataFrame
            case_col: Name of the case number column

        Returns:
            Dictionary mapping normalized case numbers to row indices
        """
        import re

        idx: Dict[str, int] = {}

        # Normalize case numbers: uppercase, remove special characters
        series = df[case_col].fillna("").astype(str).str.strip().str.upper()
        series = series.apply(lambda x: re.sub(r"[^A-Z0-9]", "", x))

        for i, v in enumerate(series.tolist()):
            if not v:
                continue
            if v not in idx:  # Keep first occurrence
                idx[v] = i

        return idx

    def _apply_master_order_sorting(
        self,
        master: pd.DataFrame,
        warehouse: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Maintain Warehouse original order while updating with Master data.

        HVDC HITACHI 파일의 원본 순서를 유지하면서 Master 데이터로 업데이트합니다.
        신규 케이스만 제일 하단에 추가합니다.

        Args:
            master: Master DataFrame
            warehouse: Warehouse DataFrame
            master_cols: Master column mapping (semantic_key -> column_name)
            wh_cols: Warehouse column mapping

        Returns:
            Tuple of (master, warehouse) - 순서 변경 없음
        """
        print("\nMaintaining Warehouse original order...")

        # Get case columns
        master_case_col = master_cols["case_number"]
        wh_case_col = wh_cols["case_number"]

        # Warehouse 원본 순서 그대로 유지 (정렬하지 않음)
        print(f"  Warehouse original order: {len(warehouse)} rows")

        # Master의 Case No 집합만 추출
        master_case_set = set(master[master_case_col].dropna().tolist())
        print(f"  Master has {len(master_case_set)} cases")

        # Warehouse 순서는 변경하지 않음
        print(f"  First 5 Warehouse Case No.: {warehouse[wh_case_col].head().tolist()}")

        return master, warehouse  # 순서 변경 없음

    def _maintain_warehouse_order(
        self,
        warehouse: pd.DataFrame,
        master: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> pd.DataFrame:
        """
        Warehouse 원본 순서를 유지하고 신규 케이스만 하단에 추가합니다.

        HVDC HITACHI 파일의 원본 순서는 변경하지 않고, Master 데이터로 업데이트만 수행합니다.
        신규 케이스는 제일 하단에 추가됩니다.

        Args:
            warehouse: 업데이트된 Warehouse DataFrame
            master: Master DataFrame
            master_cols: Master column mapping
            wh_cols: Warehouse column mapping

        Returns:
            pd.DataFrame: 원본 순서가 유지된 Warehouse DataFrame (신규 케이스는 하단)
        """
        master_case_col = master_cols["case_number"]
        wh_case_col = wh_cols["case_number"]

        # Master의 Case No 집합
        master_case_set = set(master[master_case_col].dropna().tolist())

        # 1. Warehouse에 이미 있는 케이스 (원본 순서 유지)
        wh_existing_cases = warehouse[warehouse[wh_case_col].isin(master_case_set)].copy()

        # 2. Warehouse에 없는 케이스 (원래부터 Warehouse에만 있던)
        wh_only_cases = warehouse[~warehouse[wh_case_col].isin(master_case_set)].copy()

        # 최종: 기존 Warehouse 순서 + Warehouse 전용 케이스
        # (신규 케이스는 _apply_updates에서 append됨)
        sorted_warehouse = pd.concat([wh_existing_cases, wh_only_cases], ignore_index=True)

        print(f"  Warehouse order maintained: {len(sorted_warehouse)} rows")
        print(f"    - Existing cases: {len(wh_existing_cases)} (original order)")
        print(f"    - Warehouse-only: {len(wh_only_cases)}")
        print(f"  First 5 Warehouse Case No.: {sorted_warehouse[wh_case_col].head().tolist()}")

        return sorted_warehouse

    def _apply_updates(
        self,
        master: pd.DataFrame,
        wh: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Apply updates from Master to Warehouse using matched column names.

        Args:
            master: Master DataFrame
            wh: Warehouse DataFrame
            master_cols: Master column mapping
            wh_cols: Warehouse column mapping

        Returns:
            Tuple of (updated_warehouse, statistics)
        """
        print("\nApplying updates from Master to Warehouse...")

        stats = dict(updates=0, date_updates=0, field_updates=0, appends=0)

        # Build warehouse index by case number
        wh_case_col = wh_cols["case_number"]
        wh_index = self._build_case_index(wh, wh_case_col)

        # Get master case column
        master_case_col = master_cols["case_number"]

        # Find all semantic keys that exist in both files
        common_keys = set(master_cols.keys()) & set(wh_cols.keys())

        # Find master-only keys (columns that exist in Master but not in Warehouse)
        # These need to be added to Warehouse during synchronization
        master_only_keys = set(master_cols.keys()) - set(wh_cols.keys())

        print(f"  Master columns: {list(master_cols.keys())}")
        print(f"  Warehouse columns: {list(wh_cols.keys())}")
        print(f"  Common keys: {list(common_keys)}")
        print(f"  Master-only keys: {list(master_only_keys)}")

        if master_only_keys:
            print(f"  Master-only columns found: {list(master_only_keys)}")
            print(f"  These will be added to Warehouse during sync")

        # ✅ Phase 3: Store semantic key → actual column name mappings
        for semantic_key in common_keys:
            w_col = wh_cols.get(semantic_key)
            if w_col:
                self.change_tracker.set_column_mapping(semantic_key, w_col)

        for semantic_key in master_only_keys:
            m_col = master_cols.get(semantic_key)
            if m_col:
                self.change_tracker.set_column_mapping(semantic_key, m_col)

        print(f"  [Phase 3] Stored {len(self.change_tracker.semantic_to_column)} column mappings")

        # Process each master row
        for mi, mrow in master.iterrows():
            # Get case number
            key = (
                str(mrow[master_case_col]).strip().upper()
                if pd.notna(mrow[master_case_col])
                else ""
            )
            if not key:
                continue

            # Check if case exists in warehouse
            if key not in wh_index:
                # Append new case - START WITH ALL MASTER COLUMNS
                append_row = {}

                # STEP 1: Copy ALL Master columns (matched or not)
                for col in master.columns:
                    append_row[col] = mrow[col]

                # STEP 2: Apply semantic name mapping for common columns
                # This renames Master columns to Warehouse names when they differ
                for semantic_key in common_keys:
                    m_col = master_cols[semantic_key]
                    w_col = wh_cols[semantic_key]

                    # If Warehouse uses different name, rename in append_row
                    if m_col != w_col and m_col in append_row:
                        append_row[w_col] = append_row.pop(m_col)

                # STEP 3: Source_Sheet already copied in Step 1

                # STEP 4: Ensure warehouse has all columns before concat
                for col in append_row.keys():
                    if col not in wh.columns:
                        wh[col] = None  # Initialize for existing rows

                wh = pd.concat([wh, pd.DataFrame([append_row])], ignore_index=True)
                new_index = len(wh) - 1
                stats["appends"] += 1

                # Track new case
                self.change_tracker.log_new_case(
                    case_no=key, row_data=append_row, row_index=new_index
                )
                continue

            # Case exists - update it
            wi = wh_index[key]

            # Update Source_Sheet from Master for existing cases
            # (Source_Sheet is not in common_keys, so we handle it separately)
            if "Source_Sheet" in master.columns and "Source_Sheet" in wh.columns and wi < len(wh):
                # Use Master's Source_Sheet to reflect original sheet name
                old_source = wh.at[wi, "Source_Sheet"]
                new_source = mrow["Source_Sheet"]
                wh.at[wi, "Source_Sheet"] = new_source
                # Track Source_Sheet updates for debugging
                if old_source != new_source:
                    stats["source_sheet_updates"] = stats.get("source_sheet_updates", 0) + 1

            # ✅ Update Source_Vendor from Master for existing cases
            if "Source_Vendor" in master.columns and "Source_Vendor" in wh.columns and wi < len(wh):
                # Use Master's Source_Vendor to reflect vendor
                new_vendor = mrow["Source_Vendor"]
                wh.at[wi, "Source_Vendor"] = new_vendor
                # Track Source_Vendor updates
                stats["source_vendor_updates"] = stats.get("source_vendor_updates", 0) + 1

            # Process each common column
            for semantic_key in common_keys:
                m_col = master_cols[semantic_key]
                w_col = wh_cols[semantic_key]

                # Skip metadata columns - preserve Warehouse's original value
                if w_col in METADATA_COLUMNS:
                    continue

                mval = mrow[m_col]
                wval = wh.at[wi, w_col] if wi < len(wh) else None

                # Check if this is a date column
                is_date = semantic_key in self.date_semantic_keys

                if is_date:
                    # Date column: Master always wins if it has a value
                    if pd.notna(mval):
                        if not self._dates_equal(mval, wval):
                            stats["updates"] += 1
                            stats["date_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,  # w_col is actual Excel header name (e.g., "ETD/ATD")
                                semantic_key=semantic_key,  # ✅ Phase 3
                                case_no=key,  # ✅ Phase 4: Include case_no for row lookup
                                old_value=wval,
                                new_value=mval,
                                change_type="date_update",
                            )
                        else:
                            # Equal logically - ensure consistent format
                            wh.at[wi, w_col] = mval
                else:
                    # Non-date column: Overwrite if Master has value
                    if ALWAYS_OVERWRITE_NONDATE and pd.notna(mval):
                        if (wval is None) or (str(mval) != str(wval)):
                            stats["updates"] += 1
                            stats["field_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,  # w_col is actual Excel header name
                                semantic_key=semantic_key,  # ✅ Phase 3
                                case_no=key,  # ✅ Phase 4: Include case_no
                                old_value=wval,
                                new_value=mval,
                                change_type="field_update",
                            )

            # Process master-only columns (like DHL WH) for existing cases
            for semantic_key in master_only_keys:
                m_col = master_cols[semantic_key]
                mval = mrow[m_col]

                # Add master-only column to warehouse if it doesn't exist
                if m_col not in wh.columns:
                    wh[m_col] = None  # Initialize with None for all rows

                # Update the value for this case
                if pd.notna(mval):
                    old_val = wh.at[wi, m_col] if wi < len(wh) else None
                    if old_val != mval:
                        stats["updates"] += 1
                        stats["field_updates"] += 1
                        wh.at[wi, m_col] = mval
                        self.change_tracker.add_change(
                            row_index=wi,
                            column_name=m_col,
                            semantic_key=semantic_key,  # ✅ Phase 3
                            case_no=key,  # ✅ Phase 4: Include case_no
                            old_value=old_val,
                            new_value=mval,
                            change_type="master_only_update",
                        )

        print(f"  [OK] Updates: {stats['updates']} cells changed")
        print(f"    - Date updates: {stats['date_updates']}")
        print(f"    - Field updates: {stats['field_updates']}")
        print(f"    - New records: {stats['appends']}")
        if "source_sheet_updates" in stats:
            print(f"    - Source_Sheet updates: {stats['source_sheet_updates']}")

        return wh, stats

    def synchronize(
        self, master_xlsx: str, warehouse_xlsx: str, output_path: Optional[str] = None
    ) -> SyncResult:
        """
        Synchronize Master data into Warehouse file.

        This is the main entry point for the synchronization process. It:
        1. Loads both files with automatic header detection
        2. Matches all required columns using semantic keys
        3. Validates that required columns exist
        4. Applies Master data to Warehouse
        5. Maintains Master's row order
        6. Saves the result with color-coded changes

        Args:
            master_xlsx: Path to Master Excel file
            warehouse_xlsx: Path to Warehouse Excel file
            output_path: Path for output file (optional, auto-generated if None)

        Returns:
            SyncResult object with success status, messages, and statistics

        Example:
            >>> sync = DataSynchronizerV30()
            >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
            >>> if result.success:
            >>>     print(f"Success! Output: {result.output_path}")
            >>>     print(f"Changes: {result.stats}")
        """
        try:
            # Phase 1: Load files by sheets (NEW: Sheet-by-sheet processing)
            print("\n" + "=" * 60)
            print("PHASE 1: Loading Files by Sheets")
            print("=" * 60)

            # Load Master files (HITACHI + SIEMENS)
            master_sheets = self._load_master_files(master_xlsx)

            # Check if Master and Warehouse are the same file
            from pathlib import Path

            master_path = Path(master_xlsx).resolve()
            warehouse_path = Path(warehouse_xlsx).resolve()

            if master_path == warehouse_path:
                # Same file: Use Master's merged data as Warehouse baseline
                print("[INFO] Master and Warehouse are the same file - using merged Master data")
                warehouse_sheets = {name: df.copy() for name, df in master_sheets.items()}
                warehouse_sheets_data = {name: (df.copy(), 0) for name, df in master_sheets.items()}
            else:
                # Different files: Load Warehouse separately
                warehouse_sheets_data = self._load_file_by_sheets(warehouse_xlsx, "Warehouse")
                # Extract DataFrames from tuples
                warehouse_sheets = {name: df for name, (df, _) in warehouse_sheets_data.items()}

                # ✅ Add Source_Vendor and Source_Sheet to Warehouse data (initially empty)
                for sheet_name, df in warehouse_sheets.items():
                    if "Source_Vendor" not in df.columns:
                        df["Source_Vendor"] = None  # Will be filled from Master during sync
                    if "Source_Sheet" not in df.columns:
                        df["Source_Sheet"] = sheet_name  # Warehouse's own sheet name
                    print(
                        f"[WAREHOUSE] Initialized metadata for '{sheet_name}': Source_Vendor=None, Source_Sheet='{sheet_name}'"
                    )

            # Phase 2: Process each sheet independently
            print("\n" + "=" * 60)
            print("PHASE 2: Sheet-by-Sheet Synchronization")
            print("=" * 60)

            processed_sheets = {}
            total_stats = {"updates": 0, "appends": 0, "field_updates": 0, "date_updates": 0}

            # Get common sheet names with semantic matching
            common_sheets = {}  # Dict[master_sheet, warehouse_sheet]
            master_only_sheets = []
            warehouse_only_sheets = []

            # Find matching sheets
            matched_warehouse_sheets = set()

            for m_sheet in master_sheets.keys():
                w_sheet = self._find_matching_sheet(m_sheet, list(warehouse_sheets.keys()))
                if w_sheet:
                    common_sheets[m_sheet] = w_sheet
                    matched_warehouse_sheets.add(w_sheet)
                    print(f"  [MATCH] '{m_sheet}' ↔ '{w_sheet}'")
                else:
                    master_only_sheets.append(m_sheet)

            # Find unmatched warehouse sheets
            for w_sheet in warehouse_sheets.keys():
                if w_sheet not in matched_warehouse_sheets:
                    warehouse_only_sheets.append(w_sheet)

            print(f"Common sheets: {len(common_sheets)}")
            print(f"Master-only sheets: {len(master_only_sheets)}")
            print(f"Warehouse-only sheets: {len(warehouse_only_sheets)}")

            # Dictionary to store per-sheet change trackers
            sheet_change_trackers = {}

            # Process common sheets with matched pairs
            for m_sheet_name, w_sheet_name in common_sheets.items():
                print(f"\n--- Synchronizing: '{m_sheet_name}' ↔ '{w_sheet_name}' ---")

                m_df = master_sheets[m_sheet_name]
                w_df = warehouse_sheets[w_sheet_name]
                # Get header rows from original data
                w_header_row = warehouse_sheets_data[w_sheet_name][1]
                m_header_row = 4  # Default for master sheets

                # Use warehouse sheet name for output
                output_sheet_name = w_sheet_name

                # ✅ Create a new change tracker for this sheet
                self.change_tracker = ChangeTracker()

                # Match headers for this sheet
                master_columns = self._match_and_validate_headers(m_df, f"Master-{m_sheet_name}")
                warehouse_columns = self._match_and_validate_headers(
                    w_df, f"Warehouse-{w_sheet_name}"
                )

                # Sort according to Master order
                m_df, w_df = self._apply_master_order_sorting(
                    m_df, w_df, master_columns, warehouse_columns
                )

                # Apply updates
                updated_w_df, stats = self._apply_updates(
                    m_df, w_df, master_columns, warehouse_columns
                )

                # Maintain Master order after updates
                updated_w_df = self._maintain_warehouse_order(
                    updated_w_df, m_df, master_columns, warehouse_columns
                )

                # Remove Source_Sheet column (sheet name is preserved in sheet name)
                if "Source_Sheet" in updated_w_df.columns:
                    updated_w_df = updated_w_df.drop(columns=["Source_Sheet"])

                # ✅ Store the change tracker for this sheet
                sheet_change_trackers[output_sheet_name] = self.change_tracker

                processed_sheets[output_sheet_name] = (updated_w_df, w_header_row)

                # Accumulate stats
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)

                print(
                    f"  [OK] {output_sheet_name}: {len(updated_w_df)} rows, {stats.get('updates', 0)} updates"
                )

            # Process warehouse-only sheets (keep as-is)
            for sheet_name in warehouse_only_sheets:
                print(f"\n--- Keeping warehouse-only sheet: '{sheet_name}' ---")
                w_df = warehouse_sheets[sheet_name]
                # Get header row from warehouse_sheets_data
                w_header_row = warehouse_sheets_data[sheet_name][1]

                # Remove Source_Sheet column
                if "Source_Sheet" in w_df.columns:
                    w_df = w_df.drop(columns=["Source_Sheet"])

                processed_sheets[sheet_name] = (w_df, w_header_row)
                print(f"  [OK] {sheet_name}: {len(w_df)} rows (unchanged)")

            # Process master-only sheets (append to first common sheet or create new)
            for sheet_name in master_only_sheets:
                print(f"\n--- Processing master-only sheet: '{sheet_name}' ---")
                m_df = master_sheets[sheet_name]
                # Use default header row for master sheets
                m_header_row = (
                    4  # Default for HITACHI, SIEMENS uses 0 but we'll use 4 for consistency
                )

                # Remove Source_Sheet column
                if "Source_Sheet" in m_df.columns:
                    m_df = m_df.drop(columns=["Source_Sheet"])

                processed_sheets[sheet_name] = (m_df, m_header_row)
                print(f"  [OK] {sheet_name}: {len(m_df)} rows (master-only)")

            # Phase 3: Save multi-sheet output
            print("\n" + "=" * 60)
            print("PHASE 3: Saving Multi-Sheet Output")
            print("=" * 60)

            # Determine output path
            out = output_path or str(
                Path(warehouse_xlsx).with_name(
                    Path(warehouse_xlsx).stem + ".synced_v3.4_multi.xlsx"
                )
            )

            # Save all sheets with standard 63-column header order
            print(f"  Writing to: {Path(out).name}")
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                for sheet_name, (df, header_row) in processed_sheets.items():
                    # Clean sheet name for Excel (remove invalid characters)
                    clean_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[
                        :31
                    ]  # Excel limit
                    # Apply standard 63-column header order before saving
                    df_reordered = reorder_dataframe_columns(
                        df, is_stage2=False, keep_unlisted=False, use_semantic_matching=True
                    )
                    df_reordered.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                    print(f"    - {clean_sheet_name}: {len(df)} rows, {len(df_reordered.columns)} columns (standard order)")

            print(f"  [OK] Saved {len(processed_sheets)} sheets")

            # Apply color formatting to all sheets
            print(f"  Applying color formatting...")
            for sheet_name, (df, header_row) in processed_sheets.items():
                clean_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[:31]
                # ✅ Use the sheet-specific change tracker
                if sheet_name in sheet_change_trackers:
                    self.change_tracker = sheet_change_trackers[sheet_name]
                    print(f"    - {clean_sheet_name}: {len(self.change_tracker.changes)} changes")
                    self._apply_excel_formatting(out, clean_sheet_name, header_row)
                else:
                    print(f"    - {clean_sheet_name}: No change tracker (skipped)")
            print(f"  [OK] Formatting applied to all sheets")

            # Create merged file (NEW: Single sheet with all data combined)
            print(f"\n[INFO] 합쳐진 단일시트 파일 생성 중...")

            # Combine all sheet DataFrames in specific order
            combined_dfs = []

            # Define sheet order (Case List, RIL first)
            sheet_order = []
            if "Case List, RIL" in processed_sheets:
                sheet_order.append("Case List, RIL")
            if "HE Local" in processed_sheets:
                sheet_order.append("HE Local")
            if "HE-0214,0252 (Capacitor)" in processed_sheets:
                sheet_order.append("HE-0214,0252 (Capacitor)")

            # Add any remaining sheets not in the predefined order
            for sheet_name in processed_sheets.keys():
                if sheet_name not in sheet_order:
                    sheet_order.append(sheet_name)

            # Combine sheets in order
            for sheet_name in sheet_order:
                df, _ = processed_sheets[sheet_name]
                df_copy = df.copy()
                df_copy["Source_Sheet"] = sheet_name  # 출처 기록
                combined_dfs.append(df_copy)

            # Concatenate all sheets
            merged_df = pd.concat(combined_dfs, ignore_index=True, sort=False)
            print(f"  - 합쳐진 데이터: {len(merged_df)}행, {len(merged_df.columns)}컬럼")
            print(f"  - Source_Sheet 컬럼 추가됨")

            # Save merged file with standard 63-column header order
            merged_output_path = Path(out).with_name(
                Path(out).stem.replace("_multi", "_merged") + ".xlsx"
            )
            if merged_output_path == Path(out):
                # If no "_multi" in name, add "_merged" suffix
                merged_output_path = Path(out).with_name(Path(out).stem + "_merged.xlsx")
            # Apply standard 63-column header order before saving
            merged_df_reordered = reorder_dataframe_columns(
                merged_df, is_stage2=False, keep_unlisted=False, use_semantic_matching=True
            )
            with pd.ExcelWriter(merged_output_path, engine="openpyxl") as writer:
                merged_df_reordered.to_excel(writer, sheet_name="Merged Data", index=False)

            print(f"[OK] 합쳐진 파일 저장: {merged_output_path.name} (63개 헤더로 정리됨)")

            # Prepare result
            total_stats["output_file"] = out
            total_stats["merged_file"] = str(merged_output_path)
            total_stats["merged_rows"] = len(merged_df_reordered)
            total_stats["merged_columns"] = len(merged_df_reordered.columns)

            print("\n" + "=" * 60)
            print("[OK] MULTI-SHEET SYNCHRONIZATION COMPLETE")
            print("=" * 60)
            print(f"Multi-sheet output: {out}")
            print(f"Merged output: {merged_output_path}")
            print(f"Sheets processed: {len(processed_sheets)}")
            print(
                f"Total changes: {total_stats['updates']} updates, {total_stats['appends']} new records"
            )

            return SyncResult(
                True,
                "Multi-sheet sync & colorize done.",
                out,
                total_stats,
                matching_report="All headers matched successfully",
            )

        except Exception as e:
            error_msg = f"Synchronization failed: {str(e)}"
            print(f"\n{'='*60}")
            print(f"[ERROR] ERROR: {error_msg}")
            print(f"{'='*60}")

            return SyncResult(
                False,
                error_msg,
                output_path or warehouse_xlsx,
                {},
                matching_report=str(e),
            )

    def _fuzzy_find_column(self, target_col: str, header_map: Dict[str, int]) -> Optional[int]:
        """
        ✅ Phase 4: Find column using normalized matching (Fallback level 3)

        This handles cases where column names differ slightly due to spacing,
        punctuation, or other formatting differences.

        Example: "ETD/ATD" matches "ETD / ATD"
        """
        try:
            from scripts.core.header_normalizer import HeaderNormalizer

            normalizer = HeaderNormalizer()
            target_normalized = normalizer.normalize(target_col)

            for header, idx in header_map.items():
                header_normalized = normalizer.normalize(header)
                if target_normalized == header_normalized:
                    return idx
        except Exception:
            # Fallback failed, return None
            pass

        return None

    def _apply_excel_formatting(self, excel_file: str, sheet_name: str, header_row: int):
        """
        Apply color formatting to the Excel file to highlight changes.

        Args:
            excel_file: Path to the Excel file
            sheet_name: Name of the sheet to format
            header_row: Row index where headers start (0-based from pandas)
        """
        try:
            # Early return if no changes to apply
            if not self.change_tracker.changes:
                print(f"      [DEBUG] No changes to apply for {sheet_name}")
                return

            print(
                f"      [DEBUG] Applying {len(self.change_tracker.changes)} changes to {sheet_name}"
            )
            print(f"      [DEBUG] Target file: {excel_file}")

            # ✅ Load without data_only to preserve formatting
            wb = load_workbook(excel_file, data_only=False)
            if sheet_name not in wb.sheetnames:
                print(f"      [DEBUG] Sheet {sheet_name} not found in {wb.sheetnames}")
                return

            ws = wb[sheet_name]

            # Excel rows are 1-indexed, and we need to account for header
            # FIX: Always use row 1 as header (actual Excel header row)
            excel_header_row = 1
            print(f"      [DEBUG] Excel header row: {excel_header_row} (fixed)")

            # Build header map
            header_map = {}
            case_no_col_idx = None
            for c_idx, cell in enumerate(ws[excel_header_row], start=1):
                if cell.value is None:
                    continue
                header_name = str(cell.value).strip()
                header_map[header_name] = c_idx
                
                # Find Case No. column
                if "Case" in header_name and "No" in header_name:
                    case_no_col_idx = c_idx

            print(
                f"      [DEBUG] Header map size: {len(header_map)}, first 5: {list(header_map.keys())[:5]}"
            )
            print(f"      [DEBUG] Case No. column index: {case_no_col_idx}")
            
            # ✅ Phase 4: Build Case No. → Excel row mapping
            case_to_row = {}
            if case_no_col_idx:
                for row_idx in range(excel_header_row + 1, ws.max_row + 1):
                    case_no_cell = ws.cell(row=row_idx, column=case_no_col_idx)
                    if case_no_cell.value:
                        case_to_row[str(case_no_cell.value).strip()] = row_idx
                print(f"      [DEBUG] Built case_to_row mapping: {len(case_to_row)} cases")

            # Define fills
            orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
            yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

            # ✅ Phase 4: Apply date changes (orange) with 3-level Fallback
            orange_applied = 0
            match_by_semantic = 0
            match_by_exact = 0
            match_by_fuzzy = 0
            orange_cells: List[Tuple[int, int]] = []

            for change in self.change_tracker.changes:
                if change.change_type != "date_update":
                    continue

                # ✅ Phase 4: Use case_no to find correct row after reordering
                if change.case_no and change.case_no in case_to_row:
                    excel_row = case_to_row[change.case_no]
                else:
                    # Fallback to old method if case_no not available
                    excel_row = change.row_index + excel_header_row + 1
                    if change.case_no:
                        print(f"      [WARN] Case No. '{change.case_no}' not found in case_to_row mapping")

                # Level 1: Use semantic key mapping (most accurate)
                actual_col_name = self.change_tracker.get_column_name(
                    change.semantic_key, fallback=change.column_name
                )

                # Level 2: Exact match in header_map
                col_idx = header_map.get(actual_col_name)
                if col_idx:
                    match_by_semantic += 1 if change.semantic_key else 0
                    match_by_exact += 1
                else:
                    # Level 3: Fuzzy matching (last resort)
                    col_idx = self._fuzzy_find_column(actual_col_name, header_map)
                    if col_idx:
                        match_by_fuzzy += 1

                if col_idx:
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = orange_fill
                        orange_applied += 1
                        orange_cells.append((excel_row, col_idx))

            print(f"      [DEBUG] Orange cells applied: {orange_applied}")
            print(f"        - Semantic key matches: {match_by_semantic}")
            print(f"        - Exact matches: {match_by_exact}")
            print(f"        - Fuzzy matches: {match_by_fuzzy}")

            # Apply new records (yellow)
            yellow_applied = 0
            yellow_cells: List[Tuple[int, int]] = []
            for change in self.change_tracker.changes:
                if change.change_type == "new_record":
                    excel_row = change.row_index + excel_header_row + 1

                    # Color all cells with data in this row
                    for cell in ws[excel_row]:
                        if cell.value is not None and str(cell.value).strip():
                            cell.fill = yellow_fill
                            yellow_applied += 1
                            yellow_cells.append((excel_row, cell.col_idx))

            print(f"      [DEBUG] Yellow cells applied: {yellow_applied}")

            # ✅ Save with explicit close to ensure colors persist
            wb.save(excel_file)
            wb.close()
            print(f"      [DEBUG] File saved and closed")

            # ✅ 즉시 검증: 저장된 파일을 다시 읽어 색상 확인
            try:

                def _matches_color(fill: PatternFill, targets: Tuple[str, ...]) -> bool:
                    if not fill:
                        return False
                    color = getattr(fill, "start_color", None)
                    candidates: List[str] = []
                    if color is not None:
                        rgb = str(getattr(color, "rgb", "") or "").upper()
                        if rgb:
                            candidates.append(rgb)
                        indexed = getattr(color, "indexed", None)
                        if isinstance(indexed, str):
                            candidates.append(indexed.upper())
                    fg_color = getattr(fill, "fgColor", None)
                    if fg_color is not None:
                        fg_rgb = str(getattr(fg_color, "rgb", "") or "").upper()
                        if fg_rgb:
                            candidates.append(fg_rgb)
                    return any(
                        any(target in candidate for target in targets) for candidate in candidates
                    )

                wb_verify = load_workbook(excel_file, data_only=False)
                ws_verify = wb_verify[sheet_name]

                verify_orange = 0
                for row_idx, col_idx in orange_cells:
                    cell = ws_verify.cell(row=row_idx, column=col_idx)
                    if _matches_color(cell.fill, ("FFFFA500", "FFA500", "00FFA500")):
                        verify_orange += 1

                verify_yellow = 0
                for row_idx, col_idx in yellow_cells:
                    cell = ws_verify.cell(row=row_idx, column=col_idx)
                    if _matches_color(cell.fill, ("FFFFFF00", "FFFF00", "00FFFF00")):
                        verify_yellow += 1

                print(f"      [VERIFY] Saved file color check:")
                print(f"        - Orange: {verify_orange}")
                print(f"        - Yellow: {verify_yellow}")

                if verify_orange < orange_applied:
                    missing = orange_applied - verify_orange
                    if missing > 0:
                        print(f"        WARNING: Orange color mismatch on {missing} cells!")
                if verify_yellow < yellow_applied:
                    missing = yellow_applied - verify_yellow
                    if missing > 0:
                        print(f"        WARNING: Yellow color mismatch on {missing} cells!")

                wb_verify.close()

            except Exception as verify_error:
                print(f"      [VERIFY] 검증 실패: {verify_error}")

        except Exception as e:
            print(f"  Warning: Formatting failed: {e}")
            import traceback

            traceback.print_exc()


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="DataSynchronizer v3.0 - Semantic Header Matching Edition"
    )
    ap.add_argument("--master", required=True, help="Path to Master Excel file")
    ap.add_argument("--warehouse", required=True, help="Path to Warehouse Excel file")
    ap.add_argument("--out", default="", help="Output path (optional)")
    args = ap.parse_args()

    print("\n" + "=" * 60)
    print("DataSynchronizer v3.0")
    print("Semantic Header Matching Edition")
    print("=" * 60)

    sync = DataSynchronizerV30()
    res = sync.synchronize(args.master, args.warehouse, args.out or None)

    print("\n" + "=" * 60)
    print("FINAL RESULT")
    print("=" * 60)
    print(f"Success: {res.success}")
    print(f"Message: {res.message}")
    print(f"Output:  {res.output_path}")
    print(f"Stats:   {res.stats}")

    if not res.success:
        print(f"\nError details: {res.matching_report}")
        exit(1)
