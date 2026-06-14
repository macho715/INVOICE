#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스팟 실행용 파일 확인
"""
from openpyxl import load_workbook
from pathlib import Path
import sys

project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

def check_spot_files():
    print("=" * 80)
    print("스팟 실행 파일 확인")
    print("=" * 80)
    
    master_file = project_root / "data/raw/HITACHI/Case List(20251030)v3.xlsx"
    warehouse_file = project_root / "data/raw/HITACHI/HVDC WAREHOUSE_HITACHI(HE)20251030v2.xlsx"
    
    print(f"\n[1] Master 파일 (v3):")
    if not master_file.exists():
        print(f"  오류: 파일을 찾을 수 없습니다: {master_file}")
        return
    
    master = load_workbook(master_file, read_only=True)
    print(f"  시트 목록: {master.sheetnames}")
    print(f"  총 {len(master.sheetnames)}개 시트")
    
    for i, sheet_name in enumerate(master.sheetnames[:10]):
        ws = master[sheet_name]
        print(f"    {i+1}. {sheet_name}: {ws.max_row}행")
    
    print(f"\n[2] Warehouse 파일 (v2):")
    if not warehouse_file.exists():
        print(f"  오류: 파일을 찾을 수 없습니다: {warehouse_file}")
        return
    
    warehouse = load_workbook(warehouse_file, read_only=True)
    print(f"  시트 목록: {warehouse.sheetnames}")
    print(f"  총 {len(warehouse.sheetnames)}개 시트")
    
    for i, sheet_name in enumerate(warehouse.sheetnames[:10]):
        ws = warehouse[sheet_name]
        print(f"    {i+1}. {sheet_name}: {ws.max_row}행")
    
    print("\n" + "=" * 80)
    print("스팟 실행 준비 완료")
    print("=" * 80)

if __name__ == "__main__":
    check_spot_files()




