#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스팟 실행 결과 검증
"""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import sys

project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

def verify_spot_run():
    print("=" * 80)
    print("스팟 실행 결과 검증")
    print("=" * 80)
    
    synced_file = project_root / "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx"
    merged_file = project_root / "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4_merged.xlsx"
    
    print(f"\n[1] Stage 1 출력 파일:")
    if synced_file.exists():
        wb = load_workbook(synced_file, read_only=True)
        print(f"  시트: {wb.sheetnames}")
        print(f"  총 {len(wb.sheetnames)}개 시트")
    else:
        print(f"  오류: 파일을 찾을 수 없습니다: {synced_file}")
        return
    
    print(f"\n[2] Merged 파일:")
    if merged_file.exists():
        merged_df = pd.read_excel(merged_file, sheet_name="Merged Data")
        print(f"  총 레코드: {len(merged_df)}건")
        
        if "Source_Sheet" in merged_df.columns:
            print(f"\n  Source_Sheet 분포:")
            print(merged_df["Source_Sheet"].value_counts())
    else:
        print(f"  오류: 파일을 찾을 수 없습니다: {merged_file}")
        return
    
    print(f"\n[3] 최신 날짜 확인:")
    date_cols = ["MIR", "SHU", "DAS", "AGI", "ETD/ATD", "ETA/ATA"]
    for col in date_cols:
        if col in merged_df.columns:
            if merged_df[col].notna().any():
                max_date = merged_df[col].dropna().max()
                count = merged_df[col].notna().sum()
                print(f"  {col}: {max_date} ({count}건)")
            else:
                print(f"  {col}: 데이터 없음")
    
    print("\n" + "=" * 80)
    print("검증 완료")
    print("=" * 80)

if __name__ == "__main__":
    verify_spot_run()




