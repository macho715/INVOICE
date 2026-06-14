#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stage 1 색상 작업 검증 스크립트

목적: Stage 1 동기화 출력 파일에서 색상 포맷팅이 올바르게 적용되었는지 검증

색상 규칙:
- 주황색 (ORANGE: FFFFA500): 날짜 변경사항
- 노란색 (YELLOW: FFFFFF00): 신규 추가된 행
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, Tuple, List, Optional

# 프로젝트 루트 추가
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

# UTF-8 인코딩 설정 (Windows)
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


# 색상 정의 (다양한 형식 지원)
# 주황색 (ORANGE: FFFFA500) - 날짜 변경
ORANGE_COLORS = ["FFFFA500", "FFA500", "00FFA500", "A500"]
# 노란색 (YELLOW: FFFFFF00) - 신규 행
YELLOW_COLORS = ["FFFFFF00", "FFFF00", "00FFFF00", "FF00"]

# 색상 이름 매핑
COLOR_NAMES = {
    "ORANGE": "주황색 (날짜 변경)",
    "YELLOW": "노란색 (신규 행)",
}


def get_color_rgb(fill) -> Optional[str]:
    """셀의 색상 RGB 값을 추출"""
    if not fill:
        return None
    
    # fill_type이 'none'이면 색상 없음
    if hasattr(fill, 'fill_type') and fill.fill_type == 'none':
        return None
    
    if not isinstance(fill, PatternFill):
        return None
    
    # start_color 확인 (우선순위 1)
    if hasattr(fill, "start_color") and fill.start_color:
        color = fill.start_color
        if hasattr(color, "rgb") and color.rgb:
            rgb_str = str(color.rgb).upper()
            # 'FF' 접두사 처리
            if rgb_str.startswith('FF'):
                return rgb_str
            elif len(rgb_str) == 6:
                return f"FF{rgb_str}"
            return rgb_str
        
        # indexed 색상도 확인
        if hasattr(color, "indexed") and color.indexed is not None:
            return str(color.indexed)
    
    # fgColor 확인 (우선순위 2 - 하위 호환성)
    if hasattr(fill, "fgColor") and fill.fgColor:
        fg_color = fill.fgColor
        if hasattr(fg_color, "rgb") and fg_color.rgb:
            rgb_str = str(fg_color.rgb).upper()
            if rgb_str.startswith('FF'):
                return rgb_str
            elif len(rgb_str) == 6:
                return f"FF{rgb_str}"
            return rgb_str
    
    return None


def matches_color(fill, target_colors: List[str]) -> bool:
    """셀 색상이 목표 색상과 일치하는지 확인"""
    rgb = get_color_rgb(fill)
    if not rgb:
        return False
    
    # 부분 매칭 (색상 코드 형식이 다양할 수 있음)
    for target in target_colors:
        if target in rgb or rgb in target:
            return True
    
    return False


def verify_sheet_colors(ws, sheet_name: str, header_row: int = 1) -> Dict:
    """시트의 색상 셀을 검증"""
    print(f"\n  시트: {sheet_name}")
    print(f"    - 총 {ws.max_row}행, {ws.max_column}열")
    
    # 헤더 행 제외
    data_start_row = header_row + 1
    
    orange_cells = []
    yellow_cells = []
    other_colored_cells = []
    total_colored = 0
    
    # 모든 데이터 셀 검사
    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            fill = cell.fill
            if not fill or not isinstance(fill, PatternFill):
                continue
            
            # fill_type이 'none'이거나 색상이 없는 경우 건너뛰기
            if hasattr(fill, 'fill_type') and fill.fill_type == 'none':
                continue
            
            rgb = get_color_rgb(fill)
            if not rgb:
                continue
            
            # 기본 색상(검정, 흰색)은 제외
            if rgb in ["00000000", "FFFFFFFF", "FFFFFF", "000000", None, "NONE"]:
                continue
            
            # 색상이 적용된 셀 카운트 (값이 없어도 색상만 있으면 포함)
            total_colored += 1
            
            if matches_color(fill, ORANGE_COLORS):
                orange_cells.append((row_idx, col_idx))
            elif matches_color(fill, YELLOW_COLORS):
                yellow_cells.append((row_idx, col_idx))
            else:
                other_colored_cells.append((rgb, row_idx, col_idx))
    
    result = {
        "sheet_name": sheet_name,
        "total_rows": ws.max_row,
        "total_columns": ws.max_column,
        "orange_cells": len(orange_cells),
        "yellow_cells": len(yellow_cells),
        "other_colored_cells": len(other_colored_cells),
        "total_colored": total_colored,
        "orange_details": orange_cells[:10],  # 처음 10개만 저장
        "yellow_details": yellow_cells[:10],
        "other_colors": list(set([rgb for rgb, _, _ in other_colored_cells[:20]]))
    }
    
    print(f"    - 주황색 셀: {result['orange_cells']}개")
    print(f"    - 노란색 셀: {result['yellow_cells']}개")
    print(f"    - 기타 색상 셀: {result['other_colored_cells']}개")
    print(f"    - 총 색상 적용 셀: {result['total_colored']}개")
    
    if result['other_colored_cells'] > 0:
        print(f"    - 기타 색상 RGB: {', '.join(result['other_colors'][:5])}")
    
    return result


def verify_stage1_colors(file_path: Path) -> Dict:
    """Stage 1 출력 파일의 색상 검증"""
    print("=" * 80)
    print("Stage 1 색상 작업 검증")
    print("=" * 80)
    
    if not file_path.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {file_path}")
        return {"success": False, "error": "File not found"}
    
    print(f"\n파일: {file_path.name}")
    print(f"경로: {file_path}")
    
    try:
        # Excel 파일 로드 (색상 정보 포함)
        print("\nExcel 파일 로드 중...")
        wb = load_workbook(file_path, data_only=False)
        
        print(f"\n발견된 시트: {len(wb.sheetnames)}개")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # 각 시트 검증
        sheet_results = []
        total_orange = 0
        total_yellow = 0
        total_other = 0
        total_colored = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = verify_sheet_colors(ws, sheet_name)
            sheet_results.append(result)
            
            total_orange += result["orange_cells"]
            total_yellow += result["yellow_cells"]
            total_other += result["other_colored_cells"]
            total_colored += result["total_colored"]
        
        wb.close()
        
        # 종합 결과
        print("\n" + "=" * 80)
        print("종합 결과")
        print("=" * 80)
        print(f"\n전체 시트 통계:")
        print(f"  - 주황색 셀 (날짜 변경): {total_orange}개")
        print(f"  - 노란색 셀 (신규 행): {total_yellow}개")
        print(f"  - 기타 색상 셀: {total_other}개")
        print(f"  - 총 색상 적용 셀: {total_colored}개")
        
        # 검증 결과 판정
        success = True
        warnings = []
        
        if total_orange == 0 and total_yellow == 0:
            if total_colored == 0:
                warnings.append("색상이 전혀 적용되지 않았습니다. 변경사항이 없었거나 색상 적용 로직에 문제가 있을 수 있습니다.")
            else:
                warnings.append(f"주황색/노란색이 없고 기타 색상만 {total_other}개 있습니다.")
        
        if total_colored > 0:
            orange_ratio = (total_orange / total_colored * 100) if total_colored > 0 else 0
            yellow_ratio = (total_yellow / total_colored * 100) if total_colored > 0 else 0
            print(f"\n색상 분포:")
            print(f"  - 주황색 비율: {orange_ratio:.1f}%")
            print(f"  - 노란색 비율: {yellow_ratio:.1f}%")
        
        print("\n" + "=" * 80)
        if warnings:
            print("경고:")
            for warning in warnings:
                print(f"  ⚠ {warning}")
            success = False
        else:
            print("✓ 색상 검증 완료")
            if total_colored > 0:
                print(f"✓ 색상이 정상적으로 적용되었습니다.")
        
        return {
            "success": success,
            "file_path": str(file_path),
            "sheets": sheet_results,
            "summary": {
                "total_orange": total_orange,
                "total_yellow": total_yellow,
                "total_other": total_other,
                "total_colored": total_colored,
            },
            "warnings": warnings
        }
        
    except Exception as e:
        print(f"\n[ERROR] 검증 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


def main():
    """메인 함수"""
    # 기본 파일 경로
    default_file = PROJECT_ROOT / "data" / "processed" / "synced" / "HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx"
    
    # 명령줄 인자 처리
    if len(sys.argv) > 1:
        file_path = Path(sys.argv[1])
        if not file_path.is_absolute():
            file_path = PROJECT_ROOT / file_path
    else:
        file_path = default_file
    
    # 파일 존재 확인
    if not file_path.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {file_path}")
        print(f"\n사용법: python {Path(__file__).name} [파일경로]")
        print(f"기본 경로: {default_file}")
        sys.exit(1)
    
    # 검증 실행
    result = verify_stage1_colors(file_path)
    
    # 종료 코드
    sys.exit(0 if result.get("success", False) else 1)


if __name__ == "__main__":
    main()

