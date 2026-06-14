#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""디버그용 색상 확인 스크립트"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# UTF-8 인코딩 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

file_path = Path('data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx')

print(f"파일 로드: {file_path}")
wb = load_workbook(file_path, data_only=False)

ws = wb['Case List, RIL']
print(f"\n시트: {ws.title}")
print(f"총 행: {ws.max_row}, 총 열: {ws.max_column}")

# 처음 10행, 처음 10열에서 색상 확인
print("\n=== 색상 검사 (처음 10행 10열) ===")
found_colors = 0
for row_idx in range(2, min(12, ws.max_row + 1)):
    for col_idx in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        
        if cell.fill:
            fill = cell.fill
            fill_type = getattr(fill, 'fill_type', 'unknown')
            
            if fill_type != 'none':
                start_color_rgb = getattr(fill.start_color, 'rgb', None)
                start_color_indexed = getattr(fill.start_color, 'indexed', None)
                start_color_theme = getattr(fill.start_color, 'theme', None)
                
                print(f"\n셀 {cell.coordinate} (행{row_idx}, 열{col_idx}):")
                print(f"  - fill_type: {fill_type}")
                print(f"  - start_color.rgb: {start_color_rgb}")
                print(f"  - start_color.indexed: {start_color_indexed}")
                print(f"  - start_color.theme: {start_color_theme}")
                print(f"  - cell.value: {cell.value}")
                
                # 모든 속성 확인
                print(f"  - fill 속성들: {dir(fill)}")
                print(f"  - start_color 속성들: {dir(fill.start_color)}")
                
                found_colors += 1

if found_colors == 0:
    print("\n⚠️ 색상이 적용된 셀을 찾지 못했습니다.")
    print("\n추가 확인: 셀의 기본 fill 상태 확인")
    sample_cell = ws.cell(row=2, column=1)
    print(f"샘플 셀 (B2) fill: {sample_cell.fill}")
    print(f"샘플 셀 fill_type: {getattr(sample_cell.fill, 'fill_type', 'N/A') if sample_cell.fill else 'None'}")

print(f"\n총 색상 셀 발견: {found_colors}개")

