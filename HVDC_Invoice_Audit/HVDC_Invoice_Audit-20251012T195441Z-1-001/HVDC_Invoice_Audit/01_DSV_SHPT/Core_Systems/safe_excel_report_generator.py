#!/usr/bin/env python3
"""
Safe Excel Report Generator
===========================

여러 시트를 포함한 종합 Excel 보고서 생성
Excel 포맷팅 코드 절대 사용 금지 (에러 방지)
"""

import pandas as pd
from openpyxl import Workbook
from pathlib import Path
from typing import Dict, List, Any
import logging

logger = logging.getLogger(__name__)


class SafeExcelReportGenerator:
    """
    안전한 Excel 종합 보고서 생성기
    
    Features:
    - 5개 시트 생성 (Main_Data, Executive_Dashboard, PDF_Integration_Summary, Gate_Analysis_11_14, Supporting_Docs_Mapping)
    - 포맷팅 코드 절대 없음 (열너비, 폰트, 셀병합 등 일체 금지)
    - 순수 데이터만 출력
    - VBA + PDF 통합 결과 완전 반영
    - 언더스코어 네이밍 (참조: SHPT_SEPT_2025_INTEGRATED_REPORT_20251013_224919.xlsx)
    """
    
    # Main_Data 표준 컬럼 (40개 - 참조 파일 기준)
    MAIN_DATA_COLUMNS = [
        's_no', 'sheet_name', 'description', 'rate_source', 'unit_rate', 
        'quantity', 'total_usd', 'REV RATE', 'REV TOTAL', 'DIFFERENCE',
        'status', 'flag', 'delta_pct', 'cg_band', 'charge_group',
        'issues', 'tolerance', 'ref_rate_usd', 'doc_aed',
        'gate_status', 'gate_score', 'gate_fails',
        'supporting_docs_list', 'evidence_count', 'evidence_types',
        'pdf_validation_enabled', 'pdf_parsed_files', 'evidence_types_parsed',
        'cross_doc_status',
        'gate_11_status', 'gate_11_score',
        'gate_12_status', 'gate_12_score',
        'gate_13_status', 'gate_13_score',
        'gate_14_status', 'gate_14_score',
        'demurrage_risk_level', 'demurrage_days_overdue', 'demurrage_estimated_cost'
    ]

    def __init__(self):
        """초기화"""
        self.workbook = None
        self.audit_data = None

    def create_multi_sheet_report(
        self, audit_data: Dict[str, Any], output_path: str
    ) -> bool:
        """
        여러 시트를 포함한 Excel 종합 보고서 생성
        
        Args:
            audit_data: 감사 데이터 (JSON 형식)
            output_path: 출력 파일 경로
            
        Returns:
            bool: 성공 여부
        """
        try:
            self.audit_data = audit_data
            self.workbook = Workbook()

            # 기본 시트 제거
            if "Sheet" in self.workbook.sheetnames:
                default_sheet = self.workbook["Sheet"]
                self.workbook.remove(default_sheet)

            # 5개 시트 생성 (참조 파일 구조)
            logger.info("Creating 5 sheets...")

            self._create_main_data()
            self._create_executive_dashboard()
            self._create_pdf_integration_summary()
            self._create_gate_analysis_11_14()
            self._create_supporting_docs_mapping()

            # 파일 저장
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            self.workbook.save(output_path)

            logger.info(
                f"✅ Excel report created: {output_path} ({len(self.workbook.sheetnames)} sheets)"
            )

            return True

        except Exception as e:
            logger.error(f"❌ Failed to create Excel report: {e}")
            return False

    def _create_executive_dashboard(self):
        """Executive_Dashboard 시트 생성"""
        ws = self.workbook.create_sheet("Executive_Dashboard")

        stats = self.audit_data.get("statistics", {})
        gate = stats.get("gate_validation", {})
        vba = self.audit_data.get("vba_integration", {})

        # 헤더
        ws.append(["9월 SHPT 인보이스 검증 요약", ""])
        ws.append(["", ""])  # 빈 줄

        # 전체 통계
        ws.append(["항목", "값"])
        ws.append(["총 항목 수", stats.get("total_items", 0)])
        ws.append(["PASS", stats.get("pass_items", 0)])
        ws.append(["FAIL", stats.get("fail_items", 0)])
        ws.append(["REVIEW", stats.get("review_items", 0)])
        ws.append(["PASS 비율", str(stats.get("pass_rate", "0%"))])
        
        total_amount = stats.get('total_amount_usd', 0)
        ws.append(["총 금액 (USD)", f"${total_amount:,.2f}" if isinstance(total_amount, (int, float)) else str(total_amount)])
        ws.append(["", ""])

        # Gate 검증
        ws.append(["Gate 검증 결과", ""])
        ws.append(["Gate PASS 항목", gate.get("gate_pass_items", 0)])
        ws.append(["Gate 통과율", str(gate.get("gate_pass_rate", "0%"))])
        ws.append(["평균 Gate Score", gate.get("avg_gate_score", 0)])
        ws.append(["", ""])

        # VBA 통합
        ws.append(["VBA 통합 상태", ""])
        ws.append(["VBA 통합 활성화", "Yes" if vba.get("enabled", False) else "No"])
        ws.append(["MasterData Rows", vba.get("master_data_rows", 0)])
        ws.append(["계산 불일치", vba.get("calculation_mismatches", 0)])

        logger.info("✅ Executive_Dashboard sheet created")

    def _create_main_data(self):
        """Main_Data 시트 생성 - 전체 상세 데이터 (40개 컬럼)"""
        ws = self.workbook.create_sheet("Main_Data")

        items = self.audit_data.get("items", [])

        if not items:
            ws.append(["데이터 없음"])
            logger.warning("⚠️ Main_Data sheet created: 0 items")
            return

        # DataFrame으로 변환
        try:
            df = pd.DataFrame(items)
            
            if df.empty:
                ws.append(["데이터 없음"])
                logger.warning("⚠️ Main_Data sheet created: Empty DataFrame")
                return

            # 컬럼 순서 정렬 (없는 컬럼은 빈 값으로)
            for col in self.MAIN_DATA_COLUMNS:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[self.MAIN_DATA_COLUMNS]
            
            # 헤더
            ws.append(self.MAIN_DATA_COLUMNS)

            # 데이터 (빈 리스트 변환 방지)
            for _, row in df.iterrows():
                # 각 값을 안전하게 변환
                row_data = []
                for val in row.values:
                    # 리스트/딕셔너리는 문자열로 변환
                    if isinstance(val, (list, dict)):
                        row_data.append(str(val))
                    else:
                        row_data.append(val)
                
                ws.append(row_data)

            logger.info(f"✅ Main_Data sheet created: {len(df)} rows, {len(self.MAIN_DATA_COLUMNS)} columns")
            
        except Exception as e:
            logger.error(f"❌ Main_Data sheet creation failed: {e}")
            ws.append(["데이터 변환 실패", str(e)])

    def _create_gate_analysis_11_14(self):
        """Gate_Analysis_11_14 시트 생성 (Gate 11-14 상세 분석)"""
        ws = self.workbook.create_sheet("Gate_Analysis_11_14")

        items = self.audit_data.get("items", [])

        # Gate 11-14 상세 데이터 추출
        ws.append(["S/No", "Description", "Gate-11 Status", "Gate-11 Score", 
                   "Gate-12 Status", "Gate-12 Score", "Gate-13 Status", "Gate-13 Score",
                   "Gate-14 Status", "Gate-14 Score"])

        if not items:
            ws.append(["데이터 없음", "", "", "", "", "", "", "", "", ""])
            return

        for item in items:
            # 안전한 딕셔너리 접근
            if not isinstance(item, dict):
                continue
                
            gate_data = [
                item.get("s_no", ""),
                str(item.get("description", ""))[:50],
                str(item.get("gate_11_status", "N/A")),
                item.get("gate_11_score", ""),
                str(item.get("gate_12_status", "N/A")),
                item.get("gate_12_score", ""),
                str(item.get("gate_13_status", "N/A")),
                item.get("gate_13_score", ""),
                str(item.get("gate_14_status", "N/A")),
                item.get("gate_14_score", ""),
            ]
            ws.append(gate_data)

        logger.info(f"✅ Gate_Analysis_11_14 sheet created")

    def _create_pdf_integration_summary(self):
        """PDF_Integration_Summary 시트 생성"""
        ws = self.workbook.create_sheet("PDF_Integration_Summary")
        items = self.audit_data.get("items", [])

        # 헤더
        ws.append(["S/No", "Description", "PDF Enabled", "Parsed Files", 
                   "Cross-doc Status", "Cross-doc Issues"])

        if not items:
            ws.append(["데이터 없음", "", "", "", "", ""])
            return

        for item in items:
            if not isinstance(item, dict):
                continue
                
            # PDF validation 데이터 추출
            pdf_val = item.get("pdf_validation", {}) if isinstance(item.get("pdf_validation"), dict) else {}
            
            ws.append([
                item.get("s_no", ""),
                str(item.get("description", ""))[:50],
                "Yes" if pdf_val.get("enabled", False) else "No",
                pdf_val.get("parsed_files", 0),
                str(pdf_val.get("cross_doc_status", "N/A")),
                pdf_val.get("cross_doc_issues", 0)
            ])

        logger.info(f"✅ PDF_Integration_Summary sheet created")

    def _create_supporting_docs_mapping(self):
        """Supporting_Docs_Mapping 시트 생성"""
        ws = self.workbook.create_sheet("Supporting_Docs_Mapping")
        items = self.audit_data.get("items", [])

        # 헤더
        ws.append(["S/No", "Sheet Name", "Description", "Supporting Docs List",
                   "Evidence Count", "Evidence Types", "PDF Parsed Files", "Evidence Types Parsed"])

        if not items:
            ws.append(["데이터 없음", "", "", "", "", "", "", ""])
            return

        for item in items:
            if not isinstance(item, dict):
                continue
                
            ws.append([
                item.get("s_no", ""),
                str(item.get("sheet_name", "")),
                str(item.get("description", ""))[:50],
                str(item.get("supporting_docs_list", "")),
                item.get("evidence_count", ""),
                str(item.get("evidence_types", "")),
                item.get("pdf_parsed_files", ""),
                str(item.get("evidence_types_parsed", ""))
            ])

        logger.info(f"✅ Supporting_Docs_Mapping sheet created")


def create_final_excel_report(json_data: Dict, output_path: str) -> str:
    """
    최종 Excel 보고서 생성 (편의 함수)
    
    Args:
        json_data: 감사 결과 JSON 데이터
        output_path: 출력 파일 경로
        
    Returns:
        str: 결과 메시지
    """
    generator = SafeExcelReportGenerator()

    if generator.create_multi_sheet_report(json_data, output_path):
        return f"✅ Excel 보고서 생성 완료: {output_path}"
    else:
        return f"❌ Excel 보고서 생성 실패"


if __name__ == "__main__":
    # 테스트 실행
    import json

    json_file = Path(
        "../Results/Sept_2025/JSON/shpt_sept_2025_enhanced_result_20251014_060321.json"
    )

    if json_file.exists():
        with open(json_file, encoding="utf-8") as f:
            data = json.load(f)

        output = Path("../Results/Sept_2025/Reports/FINAL_EXCEL_REPORT_20251014.xlsx")

        result = create_final_excel_report(data, str(output))
        print(result)
    else:
        print(f"❌ JSON 파일 없음: {json_file}")

