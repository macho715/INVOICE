#!/usr/bin/env python3
"""
VBA 결과 읽기 모듈
================

VBA 실행 결과(MasterData)를 읽고 Python 검증과 통합
Excel 포맷팅 코드 절대 없음 (에러 방지)
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List, Optional
import logging

logger = logging.getLogger(__name__)


class VBAResultReader:
    """
    VBA 실행 결과 읽기
    
    Features:
    - MasterData 시트 로드 (data_only=True로 값만)
    - REV RATE, REV TOTAL, DIFFERENCE 추출
    - Formula 정보 추출
    - Excel 포맷팅 코드 일체 없음
    """

    def __init__(self):
        """초기화"""
        self.logger = logging.getLogger("VBAResultReader")

    def read_master_data(self, invoice_path: str) -> pd.DataFrame:
        """
        MasterData 시트 읽기 (VBA 실행 결과)
        
        Args:
            invoice_path: INVOICE 파일 경로
            
        Returns:
            pd.DataFrame: MasterData 데이터
        """
        try:
            invoice_file = Path(invoice_path)
            
            if not invoice_file.exists():
                self.logger.error(f"Invoice file not found: {invoice_path}")
                return pd.DataFrame()

            # data_only=True: 수식 대신 계산된 값만 로드
            wb = load_workbook(invoice_path, data_only=True, read_only=True)

            if "MasterData" not in wb.sheetnames:
                self.logger.warning("MasterData sheet not found in invoice file")
                return pd.DataFrame()

            ws = wb["MasterData"]

            # 데이터 읽기 (첫 행은 헤더)
            data = list(ws.values)
            
            if not data:
                return pd.DataFrame()
            
            headers = data[0]
            rows = data[1:]

            df = pd.DataFrame(rows, columns=headers)

            self.logger.info(f"✅ MasterData loaded: {len(df)} rows, {len(df.columns)} columns")

            return df

        except Exception as e:
            self.logger.error(f"Failed to read MasterData: {e}")
            return pd.DataFrame()

    def extract_vba_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        VBA 계산 컬럼만 추출
        
        Args:
            df: MasterData DataFrame
            
        Returns:
            pd.DataFrame: VBA 계산 컬럼만 포함
        """
        # VBA 파이프라인이 생성하는 주요 컬럼
        vba_columns = ["S/No", "REV RATE", "REV TOTAL", "DIFFERENCE", "Formula"]

        # 실제 존재하는 컬럼만 선택
        available = [col for col in vba_columns if col in df.columns]

        if not available:
            self.logger.warning("No VBA calculation columns found")
            return pd.DataFrame()

        result = df[available].copy()

        self.logger.info(f"✅ Extracted VBA columns: {available}")

        return result


def integrate_vba_results(
    python_df: pd.DataFrame, vba_df: pd.DataFrame
) -> pd.DataFrame:
    """
    Python 검증 결과 + VBA 계산 결과 병합
    
    Args:
        python_df: Python 감사 시스템 결과
        vba_df: VBA MasterData 결과
        
    Returns:
        pd.DataFrame: 병합된 결과
    """
    # S/No 기준 병합
    # python_df는 's_no', vba_df는 'S/No' 사용
    python_df_copy = python_df.copy()

    # 컬럼명 정규화
    if "s_no" in python_df_copy.columns and "S/No" not in python_df_copy.columns:
        python_df_copy["S/No"] = python_df_copy["s_no"]

    # S/No 타입 변환 (object → numeric, 통일)
    if "S/No" in python_df_copy.columns:
        python_df_copy["S/No"] = pd.to_numeric(
            python_df_copy["S/No"], errors="coerce"
        )

    # VBA DataFrame의 S/No도 numeric으로 변환
    vba_df_copy = vba_df.copy()
    if "S/No" in vba_df_copy.columns:
        vba_df_copy["S/No"] = pd.to_numeric(vba_df_copy["S/No"], errors="coerce")

    # VBA 계산 컬럼 선택
    vba_calc_columns = ["S/No", "REV RATE", "REV TOTAL", "DIFFERENCE"]
    available_vba_cols = [col for col in vba_calc_columns if col in vba_df_copy.columns]

    if not available_vba_cols:
        logger.warning("No VBA calculation columns to merge")
        return python_df_copy

    # 병합 (left join - Python 결과 기준)
    merged = pd.merge(
        python_df_copy,
        vba_df_copy[available_vba_cols],
        on="S/No",
        how="left",
        suffixes=("", "_vba"),
    )

    logger.info(
        f"✅ Merged: {len(python_df_copy)} Python items + {len(vba_df_copy)} VBA items = {len(merged)} merged items"
    )

    return merged


def detect_calculation_mismatch(
    merged_df: pd.DataFrame, tolerance: float = 0.01
) -> List[Dict]:
    """
    VBA와 Python 계산 불일치 감지
    
    Args:
        merged_df: 병합된 DataFrame
        tolerance: 허용 오차 (1% 기본)
        
    Returns:
        List[Dict]: 불일치 항목 목록
    """
    mismatches = []

    # total_usd (Python) vs REV TOTAL (VBA) 비교
    if "total_usd" in merged_df.columns and "REV TOTAL" in merged_df.columns:
        for idx, row in merged_df.iterrows():
            python_total = row.get("total_usd", 0)
            vba_total = row.get("REV TOTAL", 0)

            if pd.notna(python_total) and pd.notna(vba_total):
                # 차이 계산
                if vba_total != 0:
                    delta = abs(python_total - vba_total) / vba_total
                    if delta > tolerance:
                        mismatches.append(
                            {
                                "s_no": row.get("S/No", row.get("s_no", idx)),
                                "python_total": python_total,
                                "vba_total": vba_total,
                                "delta_pct": delta * 100,
                            }
                        )

    logger.info(f"Detected {len(mismatches)} calculation mismatches (tolerance: {tolerance*100}%)")

    return mismatches

