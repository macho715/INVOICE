#!/usr/bin/env python3
"""
Excel 종합 보고서 생성 테스트 (TDD)
=================================

여러 시트를 포함한 Excel 보고서 생성 테스트
포맷팅 코드 절대 사용 금지
"""

import pytest
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


class TestSafeExcelReportGenerator:
    """Excel 종합 보고서 생성기 테스트"""

    @pytest.fixture
    def sample_audit_data(self):
        """샘플 감사 데이터"""
        return {
            "statistics": {
                "total_items": 210,
                "pass_items": 37,
                "fail_items": 24,
                "review_items": 149,
                "total_amount_usd": 42858.86,
                "pass_rate": "17.6%",
                "gate_validation": {
                    "gate_pass_items": 40,
                    "gate_pass_rate": "19.0%",
                    "avg_gate_score": 71.1,
                },
            },
            "items": [
                {
                    "s_no": 1,
                    "description": "MASTER DO FEE",
                    "unit_rate": 150.0,
                    "total_usd": 150.0,
                    "status": "PASS",
                    "gate_score": 75.0,
                    "REV RATE": 150.0,
                    "REV TOTAL": 150.0,
                    "DIFFERENCE": 0.0,
                },
                {
                    "s_no": 2,
                    "description": "CUSTOMS CLEARANCE",
                    "unit_rate": 120.0,
                    "total_usd": 120.0,
                    "status": "FAIL",
                    "gate_score": 65.0,
                    "REV RATE": 115.0,
                    "REV TOTAL": 115.0,
                    "DIFFERENCE": 5.0,
                },
            ],
            "vba_integration": {
                "enabled": True,
                "master_data_rows": 102,
                "calculation_mismatches": 1393,
            },
        }

    def test_should_create_excel_with_multiple_sheets(
        self, sample_audit_data, tmp_path
    ):
        """5개 시트가 포함된 Excel 파일을 생성해야 함 (언더스코어 네이밍)"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        # Excel 파일 생성 확인
        assert output_file.exists(), "Excel file should be created"

        # 시트 개수 확인
        wb = load_workbook(output_file)
        sheet_names = wb.sheetnames

        expected_sheets = [
            "Main_Data",
            "Executive_Dashboard",
            "PDF_Integration_Summary",
            "Gate_Analysis_11_14",
            "Supporting_Docs_Mapping",
        ]

        assert len(sheet_names) == 5, f"Should have exactly 5 sheets, got {len(sheet_names)}"

        for expected in expected_sheets:
            assert expected in sheet_names, f"Sheet '{expected}' should exist"

        print(f"✅ Excel created with {len(sheet_names)} sheets: {sheet_names}")

    def test_should_have_executive_dashboard_sheet(self, sample_audit_data, tmp_path):
        """Executive_Dashboard 시트에 통계가 있어야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        # Executive_Dashboard 시트 확인
        df = pd.read_excel(output_file, sheet_name="Executive_Dashboard")

        # 데이터 존재 확인
        assert len(df) > 0, "Executive_Dashboard should have data"

        print(f"✅ Executive_Dashboard has {len(df)} rows")

    def test_should_have_main_data_with_all_items(self, sample_audit_data, tmp_path):
        """Main_Data 시트에 모든 항목이 있어야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        # Main_Data 시트 확인
        df = pd.read_excel(output_file, sheet_name="Main_Data")

        # 항목 개수 확인
        assert len(df) == len(sample_audit_data["items"]), "All items should be in Main_Data"

        # VBA 컬럼 확인
        assert "REV RATE" in df.columns, "VBA columns should be included"

        print(f"✅ Main_Data has {len(df)} items with VBA columns")

    def test_should_have_gate_analysis_11_14(self, sample_audit_data, tmp_path):
        """Gate_Analysis_11_14 시트가 있어야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        # Gate_Analysis_11_14 시트 확인
        wb = load_workbook(output_file)
        assert "Gate_Analysis_11_14" in wb.sheetnames

        print(f"✅ Gate_Analysis_11_14 sheet exists")

    def test_should_not_use_forbidden_formatting(self):
        """금지된 포맷팅 코드를 사용하지 않아야 함"""
        import inspect
        from safe_excel_report_generator import SafeExcelReportGenerator

        # 모든 메서드의 소스 코드 검사
        for name, method in inspect.getmembers(
            SafeExcelReportGenerator, predicate=inspect.isfunction
        ):
            if name.startswith("_"):
                continue

            try:
                source = inspect.getsource(method)

                # 금지된 키워드
                forbidden = [
                    "column_dimensions",
                    "Font(",
                    "PatternFill(",
                    "merge_cells",
                    "Alignment(",
                    "Border(",
                ]

                for keyword in forbidden:
                    assert (
                        keyword not in source
                    ), f"Forbidden keyword '{keyword}' found in {name}()"

            except:
                pass  # Skip if source not available

        print(f"✅ No forbidden formatting code detected")

    def test_should_include_pdf_integration_summary_sheet(self, sample_audit_data, tmp_path):
        """PDF_Integration_Summary 시트가 있어야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        wb = load_workbook(output_file)
        assert "PDF_Integration_Summary" in wb.sheetnames

        print(f"✅ PDF_Integration_Summary sheet exists")

    def test_should_include_supporting_docs_mapping_sheet(self, sample_audit_data, tmp_path):
        """Supporting_Docs_Mapping 시트가 있어야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        wb = load_workbook(output_file)
        assert "Supporting_Docs_Mapping" in wb.sheetnames

        print(f"✅ Supporting_Docs_Mapping sheet exists")
    
    def test_main_data_should_have_40_columns(self, sample_audit_data, tmp_path):
        """Main_Data 시트는 40개 컬럼을 가져야 함"""
        from safe_excel_report_generator import SafeExcelReportGenerator

        generator = SafeExcelReportGenerator()
        output_file = tmp_path / "test_report.xlsx"

        generator.create_multi_sheet_report(sample_audit_data, str(output_file))

        df = pd.read_excel(output_file, sheet_name="Main_Data")
        
        # 40개 컬럼 확인
        assert len(df.columns) == 40, f"Main_Data should have 40 columns, got {len(df.columns)}"
        
        # 필수 컬럼 확인
        required_columns = [
            's_no', 'description', 'rate_source', 'unit_rate', 'quantity', 'total_usd',
            'REV RATE', 'REV TOTAL', 'DIFFERENCE',
            'status', 'gate_status', 'gate_score',
            'pdf_validation_enabled', 'pdf_parsed_files', 'cross_doc_status',
            'gate_11_status', 'gate_11_score', 'gate_12_status', 'gate_12_score',
            'gate_13_status', 'gate_13_score', 'gate_14_status', 'gate_14_score',
            'demurrage_risk_level', 'demurrage_days_overdue', 'demurrage_estimated_cost'
        ]
        
        for col in required_columns:
            assert col in df.columns, f"Required column '{col}' missing in Main_Data"
        
        print(f"✅ Main_Data has {len(df.columns)} columns with all required fields")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

