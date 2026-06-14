#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check color application in Stage 1 and Stage 4 files"""

import sys
import io
import openpyxl
from pathlib import Path
from collections import Counter

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def check_stage1_colors():
    """Check Stage 1 synced file colors"""
    print("=" * 80)
    print("Stage 1 색상 검증 (동기화 파일)")
    print("=" * 80)
    
    file_path = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx")
    
    if not file_path.exists():
        print(f"ERROR: 파일을 찾을 수 없습니다: {file_path}")
        return False
    
    print(f"파일: {file_path.name}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"시트: {ws.title}")
    print(f"데이터: {ws.max_row}행 x {ws.max_column}열")
    print()
    
    # 색상 수집
    color_counts = Counter()
    colored_cells = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(200, ws.max_row)), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color_rgb = str(cell.fill.fgColor.rgb)
                if color_rgb not in ['00000000', 'None']:
                    color_counts[color_rgb] += 1
                    if len(colored_cells) < 10:  # 처음 10개만 기록
                        colored_cells.append({
                            'row': row_idx,
                            'col': col_idx,
                            'color': color_rgb,
                            'value': cell.value
                        })
    
    print("색상 적용 결과:")
    if color_counts:
        for color, count in color_counts.most_common():
            color_name = "알 수 없음"
            if color == "FFC000" or color == "FFFFC000":
                color_name = "주황색 (날짜 변경)"
            elif color == "FFFF00" or color == "FFFFFF00":
                color_name = "노란색 (신규 행)"
            print(f"  {color_name} ({color}): {count}개 셀")
        
        print()
        print(f"총 색상 적용 셀: {sum(color_counts.values())}개")
        print()
        
        if colored_cells:
            print("색상 적용 예시 (처음 5개):")
            for item in colored_cells[:5]:
                print(f"  행{item['row']}, 열{item['col']}: {item['color']} - {item['value']}")
        
        print()
        print("✅ Stage 1 색상 작업 완료됨")
        return True
    else:
        print("  ❌ 색상이 적용되지 않음")
        return False

def check_stage4_colors():
    """Check Stage 4 anomaly report colors"""
    print()
    print("=" * 80)
    print("Stage 4 색상 검증 (이상치 보고서)")
    print("=" * 80)
    
    file_path = Path("data/anomaly/HVDC_anomaly_report.xlsx")
    
    if not file_path.exists():
        print(f"ERROR: 파일을 찾을 수 없습니다: {file_path}")
        return False
    
    print(f"파일: {file_path.name}")
    wb = openpyxl.load_workbook(file_path)
    
    print(f"전체 시트: {', '.join(wb.sheetnames)}")
    print()
    
    # Check if color legend sheet exists
    if "색상 범례" in wb.sheetnames or "Color Legend" in wb.sheetnames:
        print("✅ 색상 범례 시트 존재")
    else:
        print("❌ 색상 범례 시트 없음")
    
    # Check colored cells in main sheet
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "통합" in sheet_name or "HITACHI" in sheet_name or sheet_name == wb.sheetnames[0]:
            target_sheet = wb[sheet_name]
            break
    
    if not target_sheet:
        target_sheet = wb.active
    
    print(f"검증 대상 시트: {target_sheet.title}")
    print(f"데이터: {target_sheet.max_row}행 x {target_sheet.max_column}열")
    print()
    
    # 색상 수집
    color_counts = Counter()
    colored_rows = set()
    
    for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2, max_row=min(200, target_sheet.max_row)), start=2):
        row_has_color = False
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color_rgb = str(cell.fill.fgColor.rgb)
                if color_rgb not in ['00000000', 'None']:
                    color_counts[color_rgb] += 1
                    row_has_color = True
        if row_has_color:
            colored_rows.add(row_idx)
    
    print("색상 적용 결과:")
    if color_counts:
        for color, count in color_counts.most_common():
            color_name = "알 수 없음"
            if color == "FFFF0000":
                color_name = "빨간색 (시간 역전)"
            elif color == "FFFFC000" or color == "FFC000":
                color_name = "주황색 (ML 이상치-높음)"
            elif color == "FFFFFF00" or color == "FFFF00":
                color_name = "노란색 (ML 이상치-보통)"
            elif color == "FFCC99FF":
                color_name = "보라색 (데이터 품질)"
            print(f"  {color_name} ({color}): {count}개 셀")
        
        print()
        print(f"총 색상 적용 셀: {sum(color_counts.values())}개")
        print(f"색상 적용 행: {len(colored_rows)}개")
        print()
        print("✅ Stage 4 색상 작업 완료됨")
        return True
    else:
        print("  ❌ 색상이 적용되지 않음")
        return False

def main():
    """Main function"""
    stage1_ok = check_stage1_colors()
    stage4_ok = check_stage4_colors()
    
    print()
    print("=" * 80)
    print("최종 결과")
    print("=" * 80)
    print(f"Stage 1 색상 작업: {'✅ 완료' if stage1_ok else '❌ 미완료'}")
    print(f"Stage 4 색상 작업: {'✅ 완료' if stage4_ok else '❌ 미완료'}")
    print()
    
    if stage1_ok and stage4_ok:
        print("🎉 모든 색상 작업이 정상적으로 완료되었습니다!")
    else:
        print("⚠️ 일부 색상 작업이 완료되지 않았습니다.")

if __name__ == "__main__":
    main()

