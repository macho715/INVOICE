#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check color application in Stage 1 and Stage 4 files - CORRECTED VERSION"""

import sys
import io
import openpyxl
from pathlib import Path
from collections import Counter
import glob

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def check_stage1_colors():
    """Check Stage 1 synced file colors"""
    print("=" * 80)
    print("Stage 1 색상 검증 (동기화 파일)")
    print("=" * 80)
    
    file_path = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx")
    
    if not file_path.exists():
        print(f"ERROR: 파일을 찾을 수 없습니다: {file_path}")
        return False
    
    print(f"파일: {file_path.name}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"시트: {ws.title}")
    print(f"데이터: {ws.max_row}행 x {ws.max_column}열")
    print()
    
    # 색상 수집 (더 많은 행 검사)
    color_counts = Counter()
    colored_cells = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color_rgb = str(cell.fill.fgColor.rgb)
                if color_rgb not in ['00000000', 'None']:
                    color_counts[color_rgb] += 1
                    if len(colored_cells) < 20:  # 처음 20개 기록
                        colored_cells.append({
                            'row': row_idx,
                            'col': col_idx,
                            'color': color_rgb,
                            'value': str(cell.value)[:30] if cell.value else ''
                        })
    
    print("색상 적용 결과:")
    if color_counts:
        for color, count in color_counts.most_common():
            color_name = "알 수 없음"
            if color in ["FFC000", "FFFFC000"]:
                color_name = "🟠 주황색 (날짜 변경)"
            elif color in ["FFFF00", "FFFFFF00"]:
                color_name = "🟡 노란색 (신규 행)"
            print(f"  {color_name} ({color}): {count}개 셀")
        
        print()
        print(f"총 색상 적용 셀: {sum(color_counts.values())}개")
        print()
        
        if colored_cells:
            print("색상 적용 예시 (처음 10개):")
            for item in colored_cells[:10]:
                print(f"  행{item['row']}, 열{item['col']}: {item['color']} - {item['value']}")
        
        print()
        print("✅ Stage 1 색상 작업 완료됨")
        return True
    else:
        print("  ❌ 색상이 적용되지 않음")
        return False

def check_stage4_colors():
    """Check Stage 4 colors in the actual report file (NOT anomaly_report.xlsx)"""
    print()
    print("=" * 80)
    print("Stage 4 색상 검증 (종합 보고서 파일)")
    print("=" * 80)
    
    # Find the latest report file
    report_dir = Path("data/processed/reports")
    if not report_dir.exists():
        print(f"ERROR: 보고서 디렉토리를 찾을 수 없습니다: {report_dir}")
        return False
    
    # Look for report files
    report_files = list(report_dir.glob("HVDC_입고로직_종합리포트_*.xlsx"))
    if not report_files:
        print(f"ERROR: 보고서 파일을 찾을 수 없습니다")
        return False
    
    # Get the latest file (not backup)
    report_files = [f for f in report_files if "backup" not in f.name.lower()]
    if not report_files:
        print("ERROR: 백업이 아닌 보고서 파일을 찾을 수 없습니다")
        return False
    
    file_path = max(report_files, key=lambda p: p.stat().st_mtime)
    
    print(f"파일: {file_path.name}")
    wb = openpyxl.load_workbook(file_path)
    
    print(f"전체 시트: {', '.join(wb.sheetnames[:5])}...")
    print()
    
    # Stage 4는 "통합_원본데이터_Fixed" 시트에 색상 적용
    target_sheet_name = None
    for name in wb.sheetnames:
        if "통합_원본데이터" in name and "Fixed" in name:
            target_sheet_name = name
            break
    
    if not target_sheet_name:
        # Try index-based
        if len(wb.sheetnames) > 11:
            target_sheet_name = wb.sheetnames[11]
        else:
            target_sheet_name = wb.sheetnames[0]
    
    ws = wb[target_sheet_name]
    print(f"검증 대상 시트: {target_sheet_name}")
    print(f"데이터: {ws.max_row}행 x {ws.max_column}열")
    print()
    
    # 색상 수집
    color_counts = Counter()
    colored_rows = set()
    colored_cells_sample = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(1000, ws.max_row)), start=2):
        row_has_color = False
        for col_idx, cell in enumerate(row, start=1):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color_rgb = str(cell.fill.fgColor.rgb)
                if color_rgb not in ['00000000', 'None']:
                    color_counts[color_rgb] += 1
                    row_has_color = True
                    if len(colored_cells_sample) < 20:
                        colored_cells_sample.append({
                            'row': row_idx,
                            'col': col_idx,
                            'color': color_rgb,
                            'value': str(cell.value)[:30] if cell.value else ''
                        })
        if row_has_color:
            colored_rows.add(row_idx)
    
    print("색상 적용 결과:")
    if color_counts:
        for color, count in color_counts.most_common():
            color_name = "알 수 없음"
            if color == "FFFF0000":
                color_name = "🔴 빨간색 (시간 역전)"
            elif color in ["FFFFC000", "FFC000"]:
                color_name = "🟠 주황색 (ML 이상치-높음/날짜변경)"
            elif color in ["FFFFFF00", "FFFF00"]:
                color_name = "🟡 노란색 (ML 이상치-보통/신규행)"
            elif color == "FFCC99FF":
                color_name = "🟣 보라색 (데이터 품질)"
            print(f"  {color_name} ({color}): {count}개 셀")
        
        print()
        print(f"총 색상 적용 셀: {sum(color_counts.values())}개")
        print(f"색상 적용 행: {len(colored_rows)}개")
        
        if colored_cells_sample:
            print()
            print("색상 적용 예시 (처음 10개):")
            for item in colored_cells_sample[:10]:
                print(f"  행{item['row']}, 열{item['col']}: {item['color']} - {item['value']}")
        
        print()
        print("✅ Stage 4 색상 작업 완료됨")
        return True
    else:
        print("  ❌ 색상이 적용되지 않음")
        print("  NOTE: Stage 4 --stage4-visualize 옵션으로 실행했는지 확인하세요")
        return False

def main():
    """Main function"""
    stage1_ok = check_stage1_colors()
    stage4_ok = check_stage4_colors()
    
    print()
    print("=" * 80)
    print("최종 결과")
    print("=" * 80)
    print(f"Stage 1 색상 작업: {'✅ 완료' if stage1_ok else '❌ 미완료'}")
    print(f"Stage 4 색상 작업: {'✅ 완료' if stage4_ok else '❌ 미완료'}")
    print()
    
    if stage1_ok and stage4_ok:
        print("🎉 모든 색상 작업이 정상적으로 완료되었습니다!")
    elif not stage1_ok and not stage4_ok:
        print("⚠️ 모든 색상 작업이 미완료 상태입니다.")
        print()
        print("해결 방법:")
        print("1. Stage 1 색상: python run_pipeline.py --stage 1 재실행")
        print("2. Stage 4 색상: python run_pipeline.py --stage 4 --stage4-visualize 실행")
    elif not stage1_ok:
        print("⚠️ Stage 1 색상 작업이 미완료입니다.")
        print("해결: python run_pipeline.py --stage 1 재실행")
    else:
        print("⚠️ Stage 4 색상 작업이 미완료입니다.")
        print("해결: python run_pipeline.py --stage 4 --stage4-visualize 실행")

if __name__ == "__main__":
    main()

