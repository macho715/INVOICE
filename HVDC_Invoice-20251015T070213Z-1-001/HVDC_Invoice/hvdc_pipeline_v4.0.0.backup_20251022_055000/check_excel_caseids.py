#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check Case IDs in Excel report"""

import sys
import io
import openpyxl
from pathlib import Path

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Find the latest report file
report_dir = Path("data/processed/reports")
report_files = [f for f in report_dir.glob("HVDC_입고로직_종합리포트_*.xlsx") if "backup" not in f.name.lower()]
file_path = max(report_files, key=lambda p: p.stat().st_mtime)

print(f"파일: {file_path.name}")
print()

wb = openpyxl.load_workbook(file_path)

# Find "통합_원본데이터_Fixed" sheet
target_sheet = None
for name in wb.sheetnames:
    if "통합_원본데이터" in name and "Fixed" in name:
        target_sheet = name
        break

if not target_sheet:
    target_sheet = wb.sheetnames[11] if len(wb.sheetnames) > 11 else wb.sheetnames[0]

ws = wb[target_sheet]
print(f"검증 시트: {target_sheet}")
print(f"데이터: {ws.max_row}행 x {ws.max_column}열")
print()

# Find Case No. column
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
case_col_idx = None
for c, name in enumerate(header, 1):
    if name and "case" in str(name).lower():
        case_col_idx = c
        print(f"Case No. 컬럼 발견: {c}번 컬럼 (헤더={name})")
        break

if not case_col_idx:
    print("ERROR: Case No. 컬럼을 찾을 수 없습니다!")
    sys.exit(1)

print()

# Sample Case IDs from Excel
print("Excel Case No. 샘플 (처음 20행):")
for r in range(2, min(22, ws.max_row+1)):
    case_val = ws.cell(row=r, column=case_col_idx).value
    print(f"  Row {r}: [{case_val}] (type={type(case_val).__name__})")

# Check if any match with JSON case IDs
import json
with open('data/anomaly/HVDC_anomaly_report.json', encoding='utf-8') as f:
    anomalies = json.load(f)

json_case_ids = {str(a.get('Case_ID')) for a in anomalies}
print()
print(f"JSON에서 추출한 고유 Case_ID: {len(json_case_ids)}개")
print(f"샘플 JSON Case_IDs: {list(json_case_ids)[:10]}")
print()

# Check matches
matches = 0
for r in range(2, min(100, ws.max_row+1)):
    case_val = str(ws.cell(row=r, column=case_col_idx).value)
    if case_val in json_case_ids:
        matches += 1
        if matches <= 5:
            print(f"  매칭 발견! Row {r}: Case No.={case_val}")

print()
print(f"처음 100행 중 매칭된 Case: {matches}개")

if matches == 0:
    print()
    print("⚠️ 매칭 실패! Case ID 형식이 다를 수 있습니다.")
    print("디버깅 정보:")
    excel_sample = str(ws.cell(row=2, column=case_col_idx).value)
    json_sample = list(json_case_ids)[0] if json_case_ids else "N/A"
    print(f"  Excel 샘플: [{excel_sample}] (len={len(excel_sample)})")
    print(f"  JSON 샘플: [{json_sample}] (len={len(json_sample)})")

