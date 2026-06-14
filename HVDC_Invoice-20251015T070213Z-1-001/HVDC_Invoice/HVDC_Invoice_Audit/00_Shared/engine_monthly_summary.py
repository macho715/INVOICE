#!/usr/bin/env python3
"""
월별 요약 시트 통합기 (2024-08 ~ 2025-09)

14개월의 월별 요약 시트(AUG, SEP, OCT, ... SEPT)를 하나로 통합
- 동적 헤더 감지 (하드코딩 없음)
- 컬럼명 동의어 자동 매핑
- JUNE 2025 Batch1+Batch2 모두 처리
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging
from dataclasses import dataclass
from typing import Optional, List, Dict

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class MonthSheetInfo:
    """월별 시트 정보"""
    month_code: str      # '2024-08'
    month_label: str     # 'AUG 2024'
    filename: str        # Excel 파일명
    sheet_name: str      # 시트명 ('AUG', 'FEB', 'JUNE' 등)


# 컬럼명 동의어 사전
COLUMN_SYNONYMS = {
    'sno': ['S/No', 'S.No', 'S NO', 'No.', 'No'],
    'shipment_ref': ['Shpt Ref', 'Shipment Reference', 'Shipment Reference#'],
    'job_number': ['Job #', 'CW1 Job Number', 'Job Number', 'CWI Job Number'],
    'type': ['Type'],
    'bl_number': ['BL #', 'BL Number', 'B/L #'],
    'pol': ['POL', 'Port of Loading'],
    'pod': ['POD', 'Port of Discharge'],
    'mode': ['Mode', 'Transport Mode'],
    'cntr_count': ['No. Of CNTR', 'CNTR Count', 'Container Count'],
    'volume': ['Volume'],
    'quantity': ['Quantity', "Q'ty", 'Qty'],
    'boe': ['BOE', 'Bill of Entry'],
    'boe_date': ['BOE Issued Date', 'BOE Date'],
    'master_do': ['MASTER DO CHARGE', 'MASTER DO / CHARGE'],
    'customs': ['CUSTOMS CLEARANCE CHARGE', 'CUSTOMS / CLEARANCE / CHARGE', 'CUSTOMS / CLEARANCE / CH'],
    'house_do': ['HOUSE DO CHARGE', 'HOUSE / DO / CHARGE (BL EXCHANGE) / DELIVERY ORDER'],
    'port_handling': ['PORT HANDLING CHARGE'],
    'transportation': ['TRANSPORTATION CHARGE', 'TRANSPORTATION / CHARGE', 'TRANSPORTATION / CHARG'],
    'additional': ['ADDITIONAL AMOUNT', 'ADDITIONAL AMOUNT / (Please refer to the detail sheet)'],
    'at_cost': ['AT COST AMOUNT'],
    'grand_total': ['GRAND TOTAL (USD)', 'GRAND TOTAL', 'Total (USD)', 'Total'],
    'remarks': ['Remarks', 'Remark'],
    'difference': ['DIFFERENCE', 'Diff'],
    'delivery_month': ['Delivery Month', 'Delivery Month ']
}


class MonthSheetConsolidator:
    """월별 요약 시트 통합기"""
    
    def __init__(self, shpt_folder: Path):
        self.shpt_folder = Path(shpt_folder)
        self.month_sheets = self._define_month_sheets()
    
    def _define_month_sheets(self) -> List[MonthSheetInfo]:
        """
        15개 월별 시트 정의 (JUNE 2025는 Batch1+Batch2 = 2개)
        
        Returns:
            15개 MonthSheetInfo 리스트
        """
        return [
            MonthSheetInfo('2024-08', 'AUG 2024', 'SCNT SHIPMENT DRAFT INVOICE (AUG 2024).xlsx', 'AUG'),
            MonthSheetInfo('2024-09', 'SEP 2024', 'SCNT SHIPMENT DRAFT INVOICE (SEP 2024).xlsx', 'SEP'),
            MonthSheetInfo('2024-10', 'OCT 2024', 'SCNT SHIPMENT DRAFT INVOICE (OCT 2024).xlsx', 'OCT'),
            MonthSheetInfo('2024-11', 'NOV 2024', 'SCNT SHIPMENT DRAFT INVOICE (NOV 2024).xlsx', 'NOV'),
            MonthSheetInfo('2024-12', 'DEC 2024', 'SCNT SHIPMENT DRAFT INVOICE (DEC 2024).xlsm', 'DEC'),
            MonthSheetInfo('2025-01', 'JAN 2025', 'SCNT SHIPMENT DRAFT INVOICE (JANUARY 2025).xlsx', 'JAN'),
            MonthSheetInfo('2025-02', 'FEB 2025', 'SCNT SHIPMENT DRAFT INVOICE (FEBRUARY 2025)_rev2.xlsm', 'FEB'),
            MonthSheetInfo('2025-03', 'MAR 2025', 'SCNT SHIPMENT DRAFT INVOICE (MARCH 2025).xlsm', 'MAR'),
            MonthSheetInfo('2025-04', 'APR 2025', 'SCNT SHIPMENT DRAFT INVOICE (APRIL 2025).xlsx', 'APR'),
            MonthSheetInfo('2025-05', 'MAY 2025', 'SCNT SHIPMENT DRAFT INVOICE (MAY 2025).xlsm', 'MAY'),
            MonthSheetInfo('2025-06', 'JUN 2025 Batch1', 'SCNT SHIPMENT DRAFT INVOICE (JUNE 2025)_Batch1.xlsm', 'JUNE (2)'),
            MonthSheetInfo('2025-06', 'JUN 2025 Batch2', 'SCNT SHIPMENT DRAFT INVOICE (JUNE 2025)_Batch2.xlsm', 'JUNE'),
            MonthSheetInfo('2025-07', 'JUL 2025', 'SCNT SHIPMENT DRAFT INVOICE (JULY 2025).xlsm', 'July'),
            MonthSheetInfo('2025-08', 'AUG 2025', 'SCNT SHIPMENT DRAFT INVOICE (AUG 2025).xlsm', 'AUG'),
            MonthSheetInfo('2025-09', 'SEPT 2025', 'SCNT SHIPMENT DRAFT INVOICE (SEPT 2025).xlsm', 'SEPT'),
        ]
    
    def consolidate_all_month_sheets(self) -> pd.DataFrame:
        """
        15개 월별 시트 전체 통합
        
        Returns:
            통합된 DataFrame (Source_Month, No, ... 포함)
        """
        logger.info("="*80)
        logger.info("월별 요약 시트 통합 시작")
        logger.info("="*80)
        
        all_data = []
        
        for month_info in self.month_sheets:
            logger.info(f"\n{'='*80}")
            logger.info(f"처리 중: {month_info.month_label} ({month_info.filename})")
            logger.info(f"시트명: {month_info.sheet_name}")
            logger.info(f"{'='*80}")
            
            try:
                df = self._extract_month_sheet(month_info)
                if df is not None and len(df) > 0:
                    all_data.append(df)
                    logger.info(f"✅ {len(df)} rows 추출 완료")
                else:
                    logger.warning(f"⚠️ 데이터 없음")
            except Exception as e:
                logger.error(f"❌ 실패: {e}")
                continue
        
        if not all_data:
            raise ValueError("추출된 데이터가 없습니다")
        
        # 통합
        logger.info(f"\n{'='*80}")
        logger.info("데이터 통합 중...")
        logger.info(f"{'='*80}")
        
        merged_df = pd.concat(all_data, ignore_index=True)
        
        # 컬럼 표준화
        merged_df = self._standardize_output(merged_df)
        
        # No 컬럼 재생성 (전체 순번)
        if 'No' in merged_df.columns:
            merged_df['No'] = range(1, len(merged_df) + 1)
        else:
            merged_df.insert(0, 'No', range(1, len(merged_df) + 1))
        
        logger.info(f"\n{'='*80}")
        logger.info(f"✅ 통합 완료")
        logger.info(f"{'='*80}")
        logger.info(f"총 행 수: {len(merged_df)}")
        logger.info(f"총 컬럼 수: {len(merged_df.columns)}")
        logger.info(f"월별 분포:\n{merged_df['Source_Month'].value_counts().sort_index()}")
        
        return merged_df
    
    def _extract_month_sheet(self, info: MonthSheetInfo) -> Optional[pd.DataFrame]:
        """
        단일 월별 시트에서 데이터 추출 (동적 헤더 감지)
        
        Args:
            info: 월별 시트 정보
        
        Returns:
            추출된 DataFrame (Source_Month 컬럼 포함)
        """
        file_path = self.shpt_folder / info.filename
        
        if not file_path.exists():
            logger.error(f"파일 없음: {file_path}")
            return None
        
        logger.info(f"파일 열기: {info.filename}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if info.sheet_name not in wb.sheetnames:
                logger.error(f"시트 없음: {info.sheet_name}")
                logger.info(f"전체 시트: {', '.join(wb.sheetnames[:10])}...")
                return None
            
            ws = wb[info.sheet_name]
            logger.info(f"시트 크기: {ws.max_row} rows × {ws.max_column} cols")
            
            # 1. 헤더 행 동적 감지
            header_row = self._find_header_row(ws)
            
            if header_row is None:
                logger.error(f"헤더를 찾을 수 없음")
                return None
            
            logger.info(f"헤더 행 감지: Row {header_row}")
            
            # 2. 헤더 읽기
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                value = cell.value
                if value:
                    # 여러 줄 텍스트를 한 줄로
                    header_text = str(value).replace('\n', ' / ').strip()
                    headers.append(header_text)
                else:
                    headers.append(f'Unnamed_{col_idx}')
            
            logger.info(f"헤더 컬럼: {', '.join(headers[:10])}...")
            
            # 3. 데이터 추출
            data_rows = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # S/No 컬럼 확인 (Col 1 또는 Col 2)
                sno_val = None
                for sno_col in [1, 2]:
                    val = ws.cell(row=row_idx, column=sno_col).value
                    if val and str(val).strip():
                        sno_val = val
                        break
                
                # TOTAL 행이면 중단
                if sno_val and 'TOTAL' in str(sno_val).upper():
                    logger.info(f"TOTAL 행 감지 (Row {row_idx}), 데이터 추출 종료")
                    break
                
                # 빈 행 체크
                row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                if not any(v for v in row_values if v is not None and str(v).strip()):
                    continue  # 빈 행 스킵
                
                # 행 데이터 추출
                row_data = []
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                
                data_rows.append(row_data)
            
            if not data_rows:
                logger.warning(f"데이터 행 없음")
                return None
            
            logger.info(f"데이터 행 추출: {len(data_rows)} rows")
            
            # 4. DataFrame 생성
            df = pd.DataFrame(data_rows, columns=headers)
            
            # 5. Source_Month 컬럼 추가
            df.insert(0, 'Source_Month', info.month_label)
            
            # 6. 컬럼명 정규화
            df = self._normalize_column_names(df)
            
            # 7. 중복 컬럼 제거
            df = df.loc[:, ~df.columns.duplicated()]
            
            wb.close()
            
            return df
            
        except Exception as e:
            logger.error(f"추출 실패: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _find_header_row(self, ws) -> Optional[int]:
        """
        헤더 행 동적 감지 (하드코딩 없음)
        
        감지 조건:
        1. 'Delivery Month' 또는 'S/No' 키워드 존재
        2. 추가 키워드 3개 이상 (Job, Ref, BL, Total, Shipment, Type, POL, POD)
        
        Returns:
            헤더 행 번호 (1-based) 또는 None
        """
        for row_idx in range(1, min(16, ws.max_row + 1)):
            # Row 전체 텍스트 수집
            row_values = []
            for col_idx in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    row_values.append(str(val).upper())
            
            if not row_values:
                continue
            
            row_text = ' '.join(row_values)
            
            # 헤더 키워드 확인 (확장)
            has_header_keyword = any(kw in row_text for kw in [
                'S/NO', 'S.NO', 'S NO', 
                'DELIVERY MONTH',  # 2024년 시트들에 있음
                'SHPT REF', 'SHIPMENT REFERENCE'
            ])
            
            # 추가 키워드 카운트
            keywords = ['JOB', 'REF', 'BL', 'TOTAL', 'SHIPMENT', 'TYPE', 'POL', 'POD', 'MODE']
            keyword_count = sum(1 for kw in keywords if kw in row_text)
            
            # 헤더 조건: 헤더 키워드 + 3개 이상 추가 키워드
            if has_header_keyword and keyword_count >= 3:
                return row_idx
        
        return None
    
    def _normalize_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        컬럼명 정규화 (동의어 → 표준명 매핑)
        
        Args:
            df: 원본 DataFrame
        
        Returns:
            컬럼명이 정규화된 DataFrame
        """
        column_mapping = {}
        
        for col in df.columns:
            if col == 'Source_Month':
                continue  # Source_Month는 유지
            
            col_str = str(col).strip()
            
            # 동의어 사전에서 매핑 찾기
            for standard_name, synonyms in COLUMN_SYNONYMS.items():
                for synonym in synonyms:
                    if col_str == synonym or col_str.upper() == synonym.upper():
                        # 표준명으로 매핑
                        if standard_name == 'sno':
                            column_mapping[col] = 'S/No'
                        elif standard_name == 'shipment_ref':
                            column_mapping[col] = 'Shipment Reference'
                        elif standard_name == 'job_number':
                            column_mapping[col] = 'CW1 Job Number'
                        elif standard_name == 'bl_number':
                            column_mapping[col] = 'BL #'
                        elif standard_name == 'grand_total':
                            column_mapping[col] = 'GRAND TOTAL (USD)'
                        elif standard_name == 'master_do':
                            column_mapping[col] = 'MASTER DO CHARGE'
                        elif standard_name == 'customs':
                            column_mapping[col] = 'CUSTOMS CLEARANCE CHARGE'
                        elif standard_name == 'port_handling':
                            column_mapping[col] = 'PORT HANDLING CHARGE'
                        elif standard_name == 'transportation':
                            column_mapping[col] = 'TRANSPORTATION CHARGE'
                        elif standard_name == 'at_cost':
                            column_mapping[col] = 'AT COST AMOUNT'
                        else:
                            column_mapping[col] = col_str
                        break
        
        if column_mapping:
            df = df.rename(columns=column_mapping)
            logger.info(f"컬럼명 정규화: {len(column_mapping)}개 매핑")
        
        return df
    
    def _standardize_output(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        최종 출력 표준화
        
        1. 중복 컬럼 제거
        2. 표준 컬럼 순서 적용
        """
        logger.info("출력 표준화 중...")
        
        # 1. 중복 컬럼 제거
        df = df.loc[:, ~df.columns.duplicated()]
        
        # 2. 표준 컬럼 순서
        standard_columns = [
            'Source_Month',
            'No',
            'S/No',
            'Shipment Reference',
            'CW1 Job Number',
            'Type',
            'BL #',
            'POL',
            'POD',
            'Mode',
            'No. Of CNTR',
            'Volume',
            'Quantity',
            'BOE',
            'BOE Issued Date',
            'MASTER DO CHARGE',
            'CUSTOMS CLEARANCE CHARGE',
            'HOUSE DO CHARGE',
            'PORT HANDLING CHARGE',
            'TRANSPORTATION CHARGE',
            'ADDITIONAL AMOUNT',
            'AT COST AMOUNT',
            'GRAND TOTAL (USD)',
            'Remarks',
            'DIFFERENCE'
        ]
        
        # 3. 실제 존재하는 표준 컬럼만 필터링
        existing_standard = [col for col in standard_columns if col in df.columns]
        
        # 4. 나머지 컬럼들 (표준 컬럼 제외, 알파벳 순)
        remaining_columns = sorted([str(col) for col in df.columns if col not in standard_columns])
        
        # 5. 최종 컬럼 순서
        final_column_order = existing_standard + remaining_columns
        
        # 6. 컬럼 순서 재정렬
        df = df[final_column_order]
        
        logger.info(f"✅ 표준화 완료: {len(existing_standard)} 표준 컬럼 + {len(remaining_columns)} 추가 컬럼")
        
        return df


def main():
    """테스트 실행"""
    shpt_folder = Path(__file__).parent.parent / "01_DSV_SHPT" / "SHPT"
    
    if not shpt_folder.exists():
        logger.error(f"SHPT 폴더 없음: {shpt_folder}")
        return
    
    consolidator = MonthSheetConsolidator(shpt_folder)
    df = consolidator.consolidate_all_month_sheets()
    
    print(f"\n✅ 테스트 완료!")
    print(f"총 행 수: {len(df)}")
    print(f"컬럼: {list(df.columns[:10])}...")


if __name__ == "__main__":
    main()

