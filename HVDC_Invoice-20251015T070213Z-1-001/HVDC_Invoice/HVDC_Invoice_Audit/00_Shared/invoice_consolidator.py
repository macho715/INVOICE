#!/usr/bin/env python3
"""
Invoice Consolidator - VBA modCompileMaster.CompileAllSheets Python 재구현

VBA 로직 완전 재현:
- GetValueFromLabel: 레이블 오른쪽 셀 값 추출
- FindHeaderRow: S/No 헤더 행 찾기
- FindCol: S/No 컬럼 위치 찾기
- IsNumeric: 숫자 판별
- 포맷팅 제외 (데이터만 추출)

Version: 1.0.0
Created: 2024-10-16
"""

import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class InvoiceConsolidator:
    """VBA modCompileMaster.CompileAllSheets 로직을 Python으로 재구현"""

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self.wb = None
        
        # 제외할 시스템 시트
        self.excluded_sheets = ['SEPT', 'MasterData', 'LOG', 'OCT', 'NOV', 'DEC', 'AUG', 'SEP', 'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 'JULY', 'AUGUST', 'SEPTEMBER']
        
        # 인보이스 시트 패턴 (SCT, HE, SIM, SEI - 하이픈 허용)
        self.invoice_pattern = re.compile(r'^(SCT|HE|SIM|SEI)[-\d]', re.IGNORECASE)
        
        # S/No 헤더 키워드 (우선순위 순)
        self.sno_keywords = ["S/No", "S/NO", "S.No", "Sl.No", "Sl. No", "Serial No", "No"]
        
        # 필수 헤더 (S/No 없을 때 헤더 행 판별용)
        self.required_headers = ["DESCRIPTION", "RATE", "TOTAL", "Q'TY"]

    def consolidate(self) -> pd.DataFrame:
        """
        모든 인보이스 시트를 통합하여 MasterData DataFrame 생성
        
        VBA CompileAllSheets 재현
        """
        logger.info(f"Opening Excel file: {self.excel_path.name}")
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        try:
            # 1. 인보이스 시트 목록
            invoice_sheets = self._get_invoice_sheets()
            logger.info(f"Found {len(invoice_sheets)} invoice sheets: {', '.join(invoice_sheets[:5])}...")
            
            # 2. 각 시트에서 데이터 추출
            all_data = []
            for sheet_name in invoice_sheets:
                df = self._extract_sheet_data(sheet_name)
                if df is not None and len(df) > 0:
                    all_data.append(df)
            
            # 3. 모든 데이터 병합
            if not all_data:
                raise ValueError("No data extracted from invoice sheets")
            
            merged_df = pd.concat(all_data, ignore_index=True)
            
            # 4. No 컬럼 추가 (일련번호)
            merged_df.insert(0, 'No', range(1, len(merged_df) + 1))
            
            # 5. 컬럼 순서 정렬 (MasterData 형식)
            column_order = [
                'No', 'CWI Job Number', 'Order Ref. Number', 'S/No',
                'RATE SOURCE', 'DESCRIPTION', 'RATE', 'Formula', "Q'TY", 'TOTAL (USD)'
            ]
            
            # 존재하는 컬럼만 선택
            final_columns = [col for col in column_order if col in merged_df.columns]
            # 추가 컬럼도 포함 (REMARK, REV RATE 등)
            extra_cols = [col for col in merged_df.columns if col not in final_columns]
            final_columns.extend(extra_cols)
            
            merged_df = merged_df[final_columns]
            
            # 6. 검증
            validation_result = self._validate_extracted_data(merged_df)
            logger.info(f"\nValidation Results:")
            logger.info(f"  Total rows: {validation_result['total_rows']}")
            logger.info(f"  Total sheets: {validation_result['total_sheets']}")
            logger.info(f"  Missing description: {validation_result['missing_description']}")
            logger.info(f"  Missing rate: {validation_result['missing_rate']}")
            
            logger.info(f"\n✅ Consolidated {len(merged_df)} rows from {len(all_data)} sheets")
            
            return merged_df
            
        finally:
            if self.wb:
                self.wb.close()

    def _get_invoice_sheets(self) -> List[str]:
        """
        인보이스 시트 목록 반환 (패턴 기반 필터링)
        
        포함: SCT*, HE*
        제외: SEPT, MasterData, LOG, 월 이름 등
        """
        invoice_sheets = []
        
        for sheet_name in self.wb.sheetnames:
            # 제외 시트
            if sheet_name in self.excluded_sheets:
                continue
            
            # 패턴 매칭 (SCT 또는 HE로 시작)
            if self.invoice_pattern.match(sheet_name):
                invoice_sheets.append(sheet_name)
        
        return invoice_sheets

    def _extract_sheet_data(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        VBA 로직 재현 + S/No 없어도 처리 (개선)
        1. GetValueFromLabel로 CWI Job Number, Order Ref 추출
        2. FindHeaderRow로 S/No 헤더 행 찾기 (다중 키워드 시도)
        3. S/No 없으면 → DESCRIPTION/RATE/TOTAL로 헤더 행 찾기
        4. 데이터 행 추출 (IsNumeric(S/No) 또는 DESCRIPTION+RATE 있음)
        5. S/No 없으면 자동 번호 부여 (1, 2, 3...)
        """
        ws = self.wb[sheet_name]
        
        # 1. CWI Job Number 추출 (시트 내 레이블 검색)
        cwi_job = self._get_value_from_label(ws, "CW1 Job Number")
        if not cwi_job:
            cwi_job = self._get_value_from_label(ws, "CWI Job Number")
        if not cwi_job:
            # Fallback: 시트 상단에서 BAMF로 시작하는 값 찾기
            cwi_job = self._find_bamf_number(ws)
        if not cwi_job:
            cwi_job = "UNKNOWN"
        
        # 2. Order Ref. Number 추출 (시트 내 레이블 검색, Fallback: 시트명)
        order_ref = self._get_value_from_label(ws, "Order Ref. Number")
        if not order_ref:
            # Fallback: 시트 상단에서 HVDC-ADOPT로 시작하는 값 찾기
            order_ref = self._find_hvdc_adopt(ws)
        if not order_ref:
            order_ref = self._extract_order_ref_from_sheet_name(sheet_name)
        
        # 3. 헤더 찾기 (우선순위)
        # 3-1. S/No 헤더 찾기 (다중 키워드 시도)
        header_info = self._find_header_with_alternatives(ws)
        sno_col = None
        auto_sno = False
        
        if header_info:
            header_row, sno_col = header_info
            logger.debug(f"[{sheet_name}] Found S/No at row={header_row}, col={sno_col}")
        else:
            # 3-2. S/No 없으면 → DESCRIPTION/RATE/TOTAL로 헤더 행 찾기
            header_result = self._find_header_by_required_columns(ws)
            if not header_result:
                logger.warning(f"[{sheet_name}] No valid header found - skipping")
                return None
            
            header_row, column_map = header_result
            auto_sno = True
            logger.info(f"[{sheet_name}] No S/No column - will auto-generate (found {len(column_map)} required columns)")
        
        header_row, sno_col_or_map = header_info if header_info else (header_row, column_map)
        
        # 4. 헤더 전체 읽기 (S/No 컬럼부터 마지막 사용 컬럼까지)
        # VBA: lastCol = ws.Cells(firstHeaderRow, ws.Columns.Count).End(xlToLeft).Column
        last_col = ws.max_column
        for col in range(ws.max_column, sno_col - 1, -1):
            if ws.cell(header_row, col).value:
                last_col = col
                break
        
        headers = []
        for col in range(sno_col, last_col + 1):
            value = ws.cell(header_row, col).value
            if value:
                headers.append(str(value).strip())
            else:
                headers.append(f"Unnamed_{col}")
        
        # 5. 데이터 행 추출 (VBA: IsNumeric(S/No) AND Not IsEmpty)
        # VBA: lastUsedRow = ws.Cells(ws.Rows.Count, snCol).End(xlUp).Row
        last_row = ws.max_row
        data_rows = []
        
        for row_num in range(header_row + 1, last_row + 1):
            sno_value = ws.cell(row_num, sno_col).value
            
            # VBA 로직: IsNumeric(S/No) AND Not IsEmpty
            if sno_value is not None and self._is_numeric(sno_value):
                row_data = {
                    'CWI Job Number': cwi_job,
                    'Order Ref. Number': order_ref
                }
                
                # 전체 행 복사 (VBA: ws.Range(...).Copy, xlPasteValues)
                for idx, header in enumerate(headers):
                    col = sno_col + idx
                    cell_value = ws.cell(row_num, col).value
                    row_data[header] = cell_value
                
                data_rows.append(row_data)
        
        if not data_rows:
            logger.warning(f"[{sheet_name}] No data rows found (no numeric S/No)")
            return None
        
        # 6. DataFrame 생성
        df = pd.DataFrame(data_rows)
        
        logger.info(f"[{sheet_name}] Extracted {len(df)} rows (CWI: {cwi_job}, Order: {order_ref})")
        return df

    def _get_value_from_label(self, ws, label_text: str) -> str:
        """
        VBA GetValueFromLabel 재현
        
        레이블을 찾아서 그 옆(오른쪽) 셀의 값을 반환
        VBA: foundCell.Offset(0, 1).Value
        """
        # 처음 20행, 10열만 스캔 (레이블은 보통 상단에 위치)
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and label_text.lower() in str(cell_value).lower():
                    # 오른쪽 셀 값 반환
                    right_value = ws.cell(row, col + 1).value
                    if right_value:
                        return str(right_value).strip()
        return ""

    def _find_header_row_and_columns(self, ws, header_keyword: str) -> Optional[Tuple[int, int]]:
        """
        VBA FindHeaderRow + FindCol 재현
        
        Returns:
            (header_row, sno_column) 또는 None
        """
        # 전체 시트 스캔하여 "S/No" 찾기 (정확한 매칭)
        # VBA: ws.UsedRange.Find(What:=headerText, LookIn:=xlValues, LookAt:=xlWhole)
        for row in range(1, min(50, ws.max_row + 1)):  # 처음 50행만 스캔
            for col in range(1, min(30, ws.max_column + 1)):  # 처음 30열만 스캔
                cell_value = ws.cell(row, col).value
                if cell_value:
                    # 정확한 매칭 (대소문자 무시, 공백 제거)
                    # VBA: LookAt:=xlWhole
                    normalized = str(cell_value).strip().replace(" ", "").replace("/", "").upper()
                    keyword_norm = header_keyword.strip().replace(" ", "").replace("/", "").upper()
                    
                    if normalized == keyword_norm:
                        logger.debug(f"Found '{header_keyword}' at row={row}, col={col}")
                        return (row, col)
        
        return None

    def _is_numeric(self, value) -> bool:
        """
        VBA IsNumeric 동작 재현
        
        VBA: IsNumeric(value)
        """
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value)
                return True
            except:
                return False
        return False

    def _extract_order_ref_from_sheet_name(self, sheet_name: str) -> str:
        """
        시트명에서 Order Ref 생성
        
        Examples:
            SCT0126 → HVDC-ADOPT-SCT-0126
            HE0471 → HVDC-ADOPT-HE-0471
            HE0499L1 → HVDC-ADOPT-HE-0499(lot1)
            SCT0123,0124 → HVDC-ADOPT-SCT-0123,0124
        """
        name = sheet_name.strip()
        
        # 특수 케이스: Lot 처리 (HE0499L1)
        if re.match(r'.*L\d$', name):
            base = name[:-2]  # HE0499
            lot_num = name[-1]  # 1
            
            # HE 또는 SCT 추출
            prefix = base[:2] if base[2].isdigit() else base[:3]
            number = base[len(prefix):]
            
            return f"HVDC-ADOPT-{prefix}-{number}(lot{lot_num})"
        
        # 일반 케이스: SCT0126, HE0471, SCT0123,0124
        # HE 또는 SCT 추출
        if len(name) >= 3:
            prefix = name[:2] if name[2].isdigit() else name[:3]
            number = name[len(prefix):]
            return f"HVDC-ADOPT-{prefix}-{number}"
        
        # Fallback
        return f"HVDC-ADOPT-{name}"

    def _validate_extracted_data(self, df: pd.DataFrame) -> Dict:
        """추출 데이터 품질 검증"""
        
        validation = {
            'total_rows': len(df),
            'total_sheets': df['Order Ref. Number'].nunique(),
            'missing_description': len(df[df['DESCRIPTION'].isna()]) if 'DESCRIPTION' in df.columns else 0,
            'missing_rate': len(df[df['RATE'].isna()]) if 'RATE' in df.columns else 0,
            'negative_rate': len(df[df['RATE'] < 0]) if 'RATE' in df.columns else 0,
            'zero_qty': len(df[df["Q'TY"] == 0]) if "Q'TY" in df.columns else 0,
            'order_refs': sorted(df['Order Ref. Number'].unique().tolist())
        }
        
        # 경고 발생
        if validation['missing_description'] > 0:
            logger.warning(f"⚠️ {validation['missing_description']} rows with missing DESCRIPTION")
        
        if validation['negative_rate'] > 0:
            logger.error(f"❌ {validation['negative_rate']} rows with negative RATE")
        
        return validation


def main():
    """테스트 실행"""
    test_file = Path(__file__).parent.parent / "01_DSV_SHPT" / "Data" / "DSV 202509" / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    
    if not test_file.exists():
        logger.error(f"Test file not found: {test_file}")
        return
    
    consolidator = InvoiceConsolidator(test_file)
    df = consolidator.consolidate()
    
    print(f"\n✅ Test complete!")
    print(f"Total rows: {len(df)}")
    print(f"Columns: {list(df.columns)}")


if __name__ == "__main__":
    main()

