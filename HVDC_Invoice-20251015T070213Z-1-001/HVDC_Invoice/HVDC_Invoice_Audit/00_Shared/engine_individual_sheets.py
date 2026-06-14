#!/usr/bin/env python3
"""
Invoice Consolidator v2.0 - S/No 없어도 처리

개선사항:
- 시트명 패턴 확장 (HE-0173, SIM-0005, SEI-0007 지원)
- S/No 없어도 처리 (DESCRIPTION+RATE로 헤더 행 찾기)
- S/No 자동 생성 (1, 2, 3...)
- "No" 컬럼 중복 처리

Version: 2.0.0
Created: 2024-10-16
"""

import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
import unicodedata

import openpyxl
import pandas as pd

try:
    from rapidfuzz import fuzz, process
    _USE_FUZZ = True
except Exception:
    _USE_FUZZ = False

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# --- PATCH: Robust Header Detection (from PATCH_INVOICE.MD) ---

_CANONICAL = [
    "sno", "ratesource", "description", "quantity", "uom", "unitprice", "rate", "amount", "total", "hscode"
]

# 동의어/오타/한글 → 정규 키
_SYNONYMS: Dict[str, str] = {
    # 번호
    "s/no": "sno", "s no": "sno", "no": "sno", "순번": "sno", "일련번호": "sno",
    # RATE SOURCE 추가
    "rate source": "ratesource", "ratesource": "ratesource", "rate soruce": "ratesource",
    # 품명/설명
    "desc": "description", "item desc": "description", "item description": "description",
    "품명": "description", "규격": "description", "설명": "description",
    # 수량
    "qty": "quantity", "q'ty": "quantity", "quantity": "quantity", "수량": "quantity",
    # 단위
    "uom": "uom", "unit": "uom", "단위": "uom",
    # 단가/요율
    "unit price": "unitprice", "unitprice": "unitprice", "단가": "unitprice", "요율": "rate",
    "rate": "rate", "요금": "rate",
    # 금액/합계
    "amount": "amount", "line total": "amount", "금액": "amount",
    "total": "total", "합계": "total", "총액": "total", "total (usd)": "total", "total usd": "total",
    # HS
    "hs": "hscode", "hs code": "hscode", "hscode": "hscode", "hs코드": "hscode", "세번": "hscode",
}

def _norm(s: str) -> str:
    """텍스트 정규화"""
    s = unicodedata.normalize("NFKC", str(s or "")).strip().lower()
    s = re.sub(r"[^\w\s/.-]+", "", s)  # 기호 정리
    s = re.sub(r"\s+", " ", s)
    return s

def _syn_map(token: str) -> str:
    """동의어 매핑"""
    t = token
    # 정확 동의어 매핑
    if t in _SYNONYMS: 
        return _SYNONYMS[t]
    # 부분 포함형(예: 'hs code', 'item description', 'line total')
    for k, v in _SYNONYMS.items():
        if k in t:
            return v
    # 숫자/기호 혼입 (e.g., "Q'TY", "S.NO")
    t2 = t.replace(".", " ").replace("'", "'").replace("'", "")
    if t2 in _SYNONYMS:
        return _SYNONYMS[t2]
    return t

@dataclass
class HeaderDetectionResult:
    header_row_index: int
    two_row_joined: bool
    mapping: Dict[int, str]     # col_index -> canonical
    coverage: float
    reasons: List[str]          # 로그/근거

class HeaderDetector:
    def __init__(self, min_hit:int=3, fuzz_threshold:int=78, scan_rows:int=25):
        self.min_hit = min_hit
        self.fuzz_threshold = fuzz_threshold
        self.scan_rows = scan_rows

    def detect(self, df: pd.DataFrame) -> HeaderDetectionResult:
        """
        단일행/2행 조합 헤더 후보를 0~scan_rows-1 범위에서 탐색.
        1) 퍼지매칭으로 타깃 헤더 커버리지 계산
        2) 폴백: 데이터형 휴리스틱(숫자/통화/HS 패턴/텍스트 비율)
        """
        n_rows = min(len(df), self.scan_rows)
        best = None  # (score, details)
        reasons = []

        def row_text(r: int) -> List[str]:
            return [ _norm(x) for x in list(df.iloc[r].values) ]

        def join_two_rows(r1: int, r2: int) -> List[str]:
            a = row_text(r1); b = row_text(r2)
            out = []
            L = max(len(a), len(b))
            for i in range(L):
                t1 = a[i] if i < len(a) else ""
                t2 = b[i] if i < len(b) else ""
                j = (t1 + " " + t2).strip()
                out.append(j)
            return out

        def fuzzy_match_token(tok: str) -> Tuple[str, int]:
            tok_syn = _syn_map(tok)
            # 정확 매핑되면 높은 점수
            if tok_syn in _CANONICAL:
                return tok_syn, 100
            if not _USE_FUZZ:
                # 간단 유사도(부분 포함)
                for c in _CANONICAL:
                    if c in tok_syn or tok_syn in c:
                        return c, 85
                return "", 0
            # rapidfuzz
            choice, score, _ = process.extractOne(tok_syn, _CANONICAL, scorer=fuzz.token_set_ratio)
            return (choice if score >= self.fuzz_threshold else ""), score

        def score_row(tokens: List[str]) -> Tuple[float, Dict[int, str], Dict[int, int]]:
            mapping, scores = {}, {}
            seen = set()
            for i, t in enumerate(tokens):
                if not t:
                    continue
                canon, sc = fuzzy_match_token(t)
                if canon and canon not in seen:
                    mapping[i] = canon
                    scores[i] = sc
                    seen.add(canon)
            coverage = len(mapping) / max(1, len(_CANONICAL))
            # 보정: 숫자/날짜만 있는 행은 패널티
            pure_numeric_cols = sum(1 for x in tokens if re.fullmatch(r"[\d,.\-/: ]{1,}$", x or ""))
            penalty = 0.15 if pure_numeric_cols >= max(2, len(tokens)//2) else 0.0
            total_score = coverage * 1.0 - penalty
            return total_score, mapping, scores

        candidates: List[Tuple[float, int, bool, Dict[int,str], Dict[int,int]]] = []

        # 1) 단일행 후보
        for r in range(n_rows):
            toks = row_text(r)
            s, m, scs = score_row(toks)
            candidates.append((s, r, False, m, scs))

        # 2) 2행 결합 후보 (머지/두 줄 헤더)
        for r in range(n_rows-1):
            toks = join_two_rows(r, r+1)
            s, m, scs = score_row(toks)
            # 2행 조합은 소폭 가산
            s += 0.05
            candidates.append((s, r, True, m, scs))

        # 최고 점수 후보 선택 (동점 시 더 위 행 우선, 매핑 더 많은 쪽 우선)
        candidates.sort(key=lambda x: (x[0], len(x[3]) ), reverse=True)
        for s, r, two, m, scs in candidates:
            if len(m) >= self.min_hit:
                best = (s, r, two, m, scs)
                reasons.append(f"[primary] picked row {r}{'+'+str(r+1) if two else ''} "
                               f"with {len(m)} canonical hits (score={s:.3f})")
                break

        # 3) 폴백: 휴리스틱(데이터형)로 유추
        if best is None:
            # 상단 5행 내에서 가장 '텍스트 비율' 높은 행을 헤더로
            textiness = []
            for r in range(min(n_rows, 8)):
                toks = row_text(r)
                text_cnt = sum(1 for t in toks if re.search(r"[A-Za-z가-힣]", t or ""))
                textiness.append((text_cnt / max(1, len(toks)), r))
            textiness.sort(reverse=True)
            r = textiness[0][1] if textiness else 0
            toks = row_text(r)
            # 휴리스틱 매핑
            mapping = {}
            for i, t in enumerate(toks):
                tn = _syn_map(t)
                if tn in _CANONICAL:
                    mapping[i] = tn
                elif re.search(r"hs.?code|^\d{6,12}$", t):
                    mapping[i] = "hscode"
                elif re.search(r"(qty|q.?ty|수량)", t):
                    mapping[i] = "quantity"
                elif re.search(r"(unit.?price|단가|rate|요율)", t):
                    mapping[i] = "unitprice"
                elif re.search(r"(amount|total|합계|총액)", t):
                    mapping[i] = "amount"
                elif re.search(r"(uom|단위)", t):
                    mapping[i] = "uom"
                elif re.search(r"(desc|품명|설명|규격|item)", t):
                    mapping[i] = "description"
            best = (0.33, r, False, mapping, {})
            reasons.append(f"[fallback] heuristic header at row {r} with {len(mapping)} fields")

        score, r, two, mapping, scs = best

        # 근거 로그 보강(상위 매칭 점수 상위 5개)
        top5 = sorted(scs.items(), key=lambda kv: kv[1], reverse=True)[:5]
        for ci, sc in top5:
            reasons.append(f"  - col#{ci} '{_norm(df.iloc[r].values[ci])}' -> {mapping[ci]} ({sc})")

        cov = len(mapping)/len(_CANONICAL)
        return HeaderDetectionResult(
            header_row_index=r,
            two_row_joined=two,
            mapping=mapping,
            coverage=cov,
            reasons=reasons
        )

def apply_header(df: pd.DataFrame, hdr: HeaderDetectionResult) -> pd.DataFrame:
    """헤더 적용: 선택된 행(또는 2행 조합)을 컬럼명으로 만들고, 그 위 행들은 드랍."""
    
    def make_unique_columns(cols: List[str]) -> List[str]:
        """중복 컬럼명 처리 (Unnamed_N 추가)"""
        seen = {}
        result = []
        for col in cols:
            if col in seen:
                seen[col] += 1
                result.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                result.append(col)
        return result
    
    if hdr.two_row_joined:
        r = hdr.header_row_index
        joined = []
        a = df.iloc[r].astype(str).map(_norm).tolist()
        b = df.iloc[r+1].astype(str).map(_norm).tolist()
        L = max(len(a), len(b))
        for i in range(L):
            t1 = a[i] if i < len(a) else ""
            t2 = b[i] if i < len(b) else ""
            joined.append((t1 + " " + t2).strip())
        new_cols = []
        for i, raw in enumerate(joined):
            if i in hdr.mapping:
                new_cols.append(hdr.mapping[i])
            else:
                new_cols.append(_syn_map(raw) if raw else f"Unnamed_{i}")
        new_cols = make_unique_columns(new_cols)
        out = df.iloc[r+2:].copy()
        out.columns = new_cols[:len(out.columns)]
        return out.reset_index(drop=True)
    else:
        r = hdr.header_row_index
        cols_raw = df.iloc[r].astype(str).map(_norm).tolist()
        new_cols = []
        for i, raw in enumerate(cols_raw):
            if i in hdr.mapping:
                new_cols.append(hdr.mapping[i])
            else:
                new_cols.append(_syn_map(raw) if raw else f"Unnamed_{i}")
        new_cols = make_unique_columns(new_cols)
        out = df.iloc[r+1:].copy()
        out.columns = new_cols[:len(out.columns)]
        return out.reset_index(drop=True)

# --- END PATCH ---


class InvoiceConsolidator:
    """VBA modCompileMaster 로직 + S/No 없어도 처리"""

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self.wb = None
        
        # 제외할 시스템 시트 (월 이름 포함)
        month_names = ['OCT', 'NOV', 'DEC', 'AUG', 'SEP', 'SEPT', 'FEB', 'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 'JULY', 'AUGUST', 'SEPTEMBER']
        system_sheets = ['SEPT', 'MasterData', 'LOG', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 
                        'InvoiceData', 'SUMMARY', 'Novatech', 
                        'Renamed_InvoiceData', 'Renamed_SUMMARY', 'Renamed_Novatech',
                        'Renamed_InvoiceData (2)', 'Renamed_HVDC-SCT-RE-0003']
        self.excluded_sheets = system_sheets + month_names
        
        # 인보이스 시트 패턴 (하이픈 허용, SIM/SEI/ZEN 추가, Renamed_ 접두사 허용)
        self.invoice_pattern = re.compile(r'^(Renamed_)?(SCT|HE|SIM|SEI|ZEN)[-\d]', re.IGNORECASE)
        
        # S/No 헤더 키워드 (우선순위 순)
        self.sno_keywords = ["S/No", "S/NO", "S.No", "Sl.No", "Serial No"]
        
        # 필수 헤더 (S/No 없을 때 헤더 행 판별용)
        self.required_headers = ["DESCRIPTION", "RATE", "TOTAL", "Q'TY"]

    def consolidate(self) -> pd.DataFrame:
        """모든 인보이스 시트를 통합"""
        logger.info(f"Opening Excel file: {self.excel_path.name}")
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        try:
            # 1. 인보이스 시트 목록
            invoice_sheets = self._get_invoice_sheets()
            logger.info(f"Found {len(invoice_sheets)} invoice sheets: {', '.join(invoice_sheets[:5])}...")
            
            # 2. 각 시트에서 데이터 추출
            all_data = []
            for sheet_name in invoice_sheets:
                df = self._extract_sheet_data(sheet_name)
                if df is not None and len(df) > 0:
                    all_data.append(df)
            
            # 3. 모든 데이터 병합
            if not all_data:
                raise ValueError("No data extracted from invoice sheets")
            
            merged_df = pd.concat(all_data, ignore_index=True)
            
            # 4. "No" 컬럼 처리 (중복 시 제거)
            if 'No' in merged_df.columns:
                logger.info("Removing existing 'No' column (will regenerate)")
                merged_df = merged_df.drop(columns=['No'])
            
            # 5. No 컬럼 추가 (일련번호)
            merged_df.insert(0, 'No', range(1, len(merged_df) + 1))
            
            # 6. 컬럼 순서 정렬
            column_order = [
                'No', 'CWI Job Number', 'Order Ref. Number', 'S/No',
                'RATE SOURCE', 'DESCRIPTION', 'RATE', 'Formula', "Q'TY", 'TOTAL (USD)'
            ]
            
            # 존재하는 컬럼만 선택
            final_columns = [col for col in column_order if col in merged_df.columns]
            # 추가 컬럼도 포함
            extra_cols = [col for col in merged_df.columns if col not in final_columns]
            final_columns.extend(extra_cols)
            
            merged_df = merged_df[final_columns]
            
            # 7. 검증
            validation_result = self._validate_extracted_data(merged_df)
            logger.info(f"\nValidation Results:")
            logger.info(f"  Total rows: {validation_result['total_rows']}")
            logger.info(f"  Total sheets: {validation_result['total_sheets']}")
            logger.info(f"  Missing description: {validation_result['missing_description']}")
            logger.info(f"  Missing rate: {validation_result['missing_rate']}")
            
            logger.info(f"\n✅ Consolidated {len(merged_df)} rows from {len(all_data)} sheets")
            
            return merged_df
            
        finally:
            if self.wb:
                self.wb.close()

    def _get_invoice_sheets(self) -> List[str]:
        """인보이스 시트 목록 반환 (패턴 기반)"""
        invoice_sheets = []
        
        for sheet_name in self.wb.sheetnames:
            # 제외 시트
            if sheet_name in self.excluded_sheets:
                continue
            
            # 패턴 매칭 (SCT/HE/SIM/SEI, 하이픈 허용)
            if self.invoice_pattern.match(sheet_name):
                invoice_sheets.append(sheet_name)
        
        return invoice_sheets

    def _extract_sheet_data(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """시트에서 데이터 추출 (S/No 없어도 처리)"""
        ws = self.wb[sheet_name]
        
        # 1. CWI Job Number 추출
        cwi_job = self._get_value_from_label(ws, "CW1 Job Number")
        if not cwi_job:
            cwi_job = self._get_value_from_label(ws, "CWI Job Number")
        if not cwi_job:
            cwi_job = self._find_bamf_number(ws)
        if not cwi_job:
            cwi_job = "UNKNOWN"
        
        # 2. Order Ref 추출
        order_ref = self._get_value_from_label(ws, "Order Ref. Number")
        if not order_ref:
            order_ref = self._find_hvdc_adopt(ws)
        if not order_ref:
            order_ref = self._extract_order_ref_from_sheet_name(sheet_name)
        
        # 3. 헤더 행 찾기
        # 3-1. S/No 있는 경우 (기존 로직 우선)
        header_info_with_sno = self._find_header_with_alternatives(ws)
        
        if header_info_with_sno:
            # S/No 있음 → 기존 로직
            return self._extract_with_sno(ws, sheet_name, cwi_job, order_ref, header_info_with_sno)
        
        # 3-2. Robust Header Detection (PATCH) - Fuzzy 매칭 기반
        try:
            df_raw = pd.DataFrame(ws.values)
            detector = HeaderDetector(min_hit=3, fuzz_threshold=78, scan_rows=25)
            hdr = detector.detect(df_raw)
            
            if len(hdr.mapping) >= 3:  # 최소 3개 헤더 필요
                logger.info(f"[{sheet_name}] Robust detection: row {hdr.header_row_index}, {len(hdr.mapping)} headers")
                for reason in hdr.reasons[:3]:  # 상위 3개 근거만 로그
                    logger.debug(f"  {reason}")
                
                # apply_header로 DataFrame 생성
                df_clean = apply_header(df_raw, hdr)
                
                # 필수 컬럼 매핑 (canonical -> 실제 컬럼명)
                rename_map = {
                    'ratesource': 'RATE SOURCE',
                    'description': 'DESCRIPTION',
                    'rate': 'RATE',
                    'quantity': "Q'TY",
                    'total': 'TOTAL (USD)',
                    'sno': 'S/No'
                }
                df_clean = df_clean.rename(columns=rename_map)
                
                # CWI Job Number, Order Ref 추가
                df_clean.insert(0, 'CWI Job Number', cwi_job)
                df_clean.insert(1, 'Order Ref. Number', order_ref)
                
                # S/No 자동 생성 (없으면)
                if 'S/No' not in df_clean.columns:
                    df_clean.insert(2, 'S/No', range(1, len(df_clean) + 1))
                    logger.info(f"[{sheet_name}] AUTO S/No: 1~{len(df_clean)}")
                
                # 빈 행 제거 (DESCRIPTION 없으면 제외)
                if 'DESCRIPTION' in df_clean.columns:
                    df_clean = df_clean[df_clean['DESCRIPTION'].notna() & (df_clean['DESCRIPTION'] != '')]
                
                logger.info(f"[{sheet_name}] Extracted {len(df_clean)} rows (ROBUST)")
                return df_clean
        except Exception as e:
            logger.warning(f"[{sheet_name}] Robust detection failed: {e}")
        
        # 3-3. 폴백: 기존 로직 (DESCRIPTION/RATE로 헤더 행 찾기)
        return self._extract_without_sno(ws, sheet_name, cwi_job, order_ref)
    
    def _extract_with_sno(self, ws, sheet_name, cwi_job, order_ref, header_info) -> Optional[pd.DataFrame]:
        """S/No 있는 경우 (기존 로직)"""
        header_row, sno_col = header_info
        
        # 마지막 컬럼 찾기
        last_col = ws.max_column
        for col in range(ws.max_column, 1, -1):  # Column 1부터 검사
            if ws.cell(header_row, col).value:
                last_col = col
                break
        
        # 헤더 읽기 (S/No부터 마지막까지, 빈 컬럼 제외)
        headers = []
        header_cols = []
        for col in range(sno_col, last_col + 1):  # S/No부터 시작
            value = ws.cell(header_row, col).value
            if value:
                headers.append(str(value).strip())
                header_cols.append(col)
            else:
                headers.append(f"Unnamed_{col}")
                header_cols.append(col)
        
        # 데이터 행 추출 (IsNumeric(S/No))
        data_rows = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            sno_value = ws.cell(row_num, sno_col).value
            
            if sno_value is not None and self._is_numeric(sno_value):
                row_data = {
                    'CWI Job Number': cwi_job,
                    'Order Ref. Number': order_ref
                }
                
                for idx, header in enumerate(headers):
                    col = header_cols[idx]  # 실제 컬럼 번호 사용
                    cell_value = ws.cell(row_num, col).value
                    row_data[header] = cell_value
                
                data_rows.append(row_data)
        
        if not data_rows:
            return None
        
        df = pd.DataFrame(data_rows)
        logger.info(f"[{sheet_name}] Extracted {len(df)} rows (WITH S/No)")
        return df
    
    def _extract_without_sno(self, ws, sheet_name, cwi_job, order_ref) -> Optional[pd.DataFrame]:
        """S/No 없는 경우 (자동 번호 부여)"""
        # 헤더 행 찾기 (DESCRIPTION/RATE/TOTAL 기준)
        header_result = self._find_header_by_required_columns(ws)
        if not header_result:
            logger.warning(f"[{sheet_name}] No valid header found")
            return None
        
        header_row, column_map = header_result
        
        # 헤더 읽기 (column_map 기준으로 동적 감지)
        # column_map에는 {'RATE SOURCE': 2, 'DESCRIPTION': 3, ...} 형태로 실제 컬럼 번호가 있음
        headers = []
        header_cols = []
        
        # column_map의 최소 컬럼부터 시작 (빈 컬럼 건너뛰기)
        min_col = min(column_map.values()) if column_map else 1
        max_col = ws.max_column
        
        # 실제 헤더가 있는 컬럼만 읽기
        for col in range(min_col, max_col + 1):
            value = ws.cell(header_row, col).value
            if value and str(value).strip():
                headers.append(str(value).strip())
                header_cols.append(col)
        
        # 데이터 행 추출 (DESCRIPTION + RATE 있으면 데이터)
        desc_col = column_map.get('DESCRIPTION')
        rate_col = column_map.get('RATE')
        
        if not desc_col or not rate_col:
            logger.warning(f"[{sheet_name}] DESCRIPTION or RATE column not found")
            return None
        
        data_rows = []
        auto_sno = 1
        
        for row_num in range(header_row + 1, ws.max_row + 1):
            desc_value = ws.cell(row_num, desc_col).value
            rate_value = ws.cell(row_num, rate_col).value
            
            # DESCRIPTION 있고 RATE가 숫자면 데이터 행
            if desc_value and str(desc_value).strip() and self._is_numeric(rate_value):
                row_data = {
                    'CWI Job Number': cwi_job,
                    'Order Ref. Number': order_ref,
                    'S/No': auto_sno  # 자동 번호
                }
                
                # 모든 헤더 컬럼 복사
                for idx, col in enumerate(header_cols):
                    header = headers[idx]
                    cell_value = ws.cell(row_num, col).value
                    row_data[header] = cell_value
                
                data_rows.append(row_data)
                auto_sno += 1
        
        if not data_rows:
            return None
        
        df = pd.DataFrame(data_rows)
        logger.info(f"[{sheet_name}] Extracted {len(df)} rows (AUTO S/No: 1~{auto_sno-1})")
        return df

    def _find_header_with_alternatives(self, ws) -> Optional[Tuple[int, int]]:
        """다중 S/No 키워드로 헤더 찾기"""
        for keyword in self.sno_keywords:
            result = self._find_header_row_and_columns(ws, keyword)
            if result:
                logger.debug(f"Found S/No variant '{keyword}' at row={result[0]}, col={result[1]}")
                return result
        return None
    
    def _find_header_by_required_columns(self, ws) -> Optional[Tuple[int, Dict[str, int]]]:
        """
        DESCRIPTION, RATE, TOTAL로 헤더 행 찾기
        
        Returns:
            (header_row, {'DESCRIPTION': col, 'RATE': col, ...})
        """
        for row in range(1, min(50, ws.max_row + 1)):
            found_headers = {}
            
            for col in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if not cell_value:
                    continue
                
                cell_upper = str(cell_value).upper().strip()
                
                # 필수 헤더 매칭
                if 'DESCRIPTION' in cell_upper:
                    found_headers['DESCRIPTION'] = col
                elif 'RATE' == cell_upper or (cell_upper.startswith('RATE') and 'SOURCE' not in cell_upper):
                    found_headers['RATE'] = col
                elif 'TOTAL' in cell_upper:
                    found_headers['TOTAL'] = col
                elif "Q'TY" in cell_upper or 'QTY' in cell_upper:
                    found_headers["Q'TY"] = col
                elif 'RATE SOURCE' in cell_upper or 'RATE SORUCE' in cell_upper:  # 오타 포함
                    found_headers['RATE SOURCE'] = col
            
            # 3개 이상 발견되면 헤더 행
            if len(found_headers) >= 3:
                logger.info(f"Found header row at row={row} (no S/No, {len(found_headers)} columns: {list(found_headers.keys())})")
                return (row, found_headers)
        
        return None

    def _find_bamf_number(self, ws) -> str:
        """시트 상단에서 BAMF 번호 찾기"""
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and 'BAMF' in str(cell_value).upper():
                    return str(cell_value).strip()
        return ""
    
    def _find_hvdc_adopt(self, ws) -> str:
        """시트 상단에서 HVDC-ADOPT 번호 찾기"""
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and 'HVDC-ADOPT' in str(cell_value).upper():
                    return str(cell_value).strip()
        return ""

    def _get_value_from_label(self, ws, label_text: str) -> str:
        """VBA GetValueFromLabel 재현"""
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and label_text.lower() in str(cell_value).lower():
                    right_value = ws.cell(row, col + 1).value
                    if right_value:
                        return str(right_value).strip()
        return ""

    def _find_header_row_and_columns(self, ws, header_keyword: str) -> Optional[Tuple[int, int]]:
        """VBA FindHeaderRow + FindCol 재현"""
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value:
                    normalized = str(cell_value).strip().replace(" ", "").replace("/", "").upper()
                    keyword_norm = header_keyword.strip().replace(" ", "").replace("/", "").upper()
                    
                    if normalized == keyword_norm:
                        return (row, col)
        
        return None

    def _is_numeric(self, value) -> bool:
        """VBA IsNumeric 재현"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value)
                return True
            except:
                return False
        return False

    def _extract_order_ref_from_sheet_name(self, sheet_name: str) -> str:
        """시트명 → Order Ref 변환"""
        name = sheet_name.strip()
        
        # 하이픈 제거 (HE-0173 → HE0173)
        name_clean = name.replace('-', '')
        
        # Lot 케이스
        if re.match(r'.*L\d$', name_clean):
            base = name_clean[:-2]
            lot_num = name_clean[-1]
            prefix = base[:2] if base[2].isdigit() else base[:3]
            number = base[len(prefix):]
            return f"HVDC-ADOPT-{prefix}-{number}(LOT{lot_num})"
        
        # 일반 케이스
        if len(name_clean) >= 3:
            prefix = name_clean[:2] if name_clean[2].isdigit() else name_clean[:3]
            number = name_clean[len(prefix):]
            return f"HVDC-ADOPT-{prefix}-{number}"
        
        return f"HVDC-ADOPT-{name}"

    def _validate_extracted_data(self, df: pd.DataFrame) -> Dict:
        """추출 데이터 검증"""
        validation = {
            'total_rows': len(df),
            'total_sheets': df['Order Ref. Number'].nunique(),
            'missing_description': len(df[df['DESCRIPTION'].isna()]) if 'DESCRIPTION' in df.columns else 0,
            'missing_rate': len(df[df['RATE'].isna()]) if 'RATE' in df.columns else 0,
            'negative_rate': len(df[pd.to_numeric(df['RATE'], errors='coerce') < 0]) if 'RATE' in df.columns else 0,
            'order_refs': sorted([str(x) for x in df['Order Ref. Number'].unique().tolist()])
        }
        
        if validation['missing_description'] > 0:
            logger.warning(f"⚠️ {validation['missing_description']} rows with missing DESCRIPTION")
        if validation['negative_rate'] > 0:
            logger.error(f"❌ {validation['negative_rate']} rows with negative RATE")
        
        return validation
    
    def _normalize_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        컬럼명 정규화 (오타 수정, 통일)
        
        매핑:
        - RATE SORUCE → RATE SOURCE
        - Q'TY → Q'TY (통일)
        - TOTAL, TOTAL(USD), TOTAL (USD) → TOTAL (USD)
        - Line Total (USD), Line\u202fTotal (USD) → Line Total (USD)
        """
        column_mapping = {
            'RATE SORUCE': 'RATE SOURCE',
            'TOTAL(USD)': 'TOTAL (USD)',
            'TOTAL': 'TOTAL (USD)',
            'TOTAL(AED)': 'TOTAL (AED)',
            'Line\u202fTotal (USD)': 'Line Total (USD)',
            'Q\'TY': 'Q\'TY',  # 통일된 형식
            'Qty': 'Q\'TY',
            'QTY': 'Q\'TY',
        }
        
        # 컬럼명 매핑 적용
        df = df.rename(columns=column_mapping)
        
        logger.info(f"컬럼명 정규화 완료: {len(df.columns)} columns")
        return df
    
    def _standardize_column_order(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        이미지 기준 표준 컬럼 순서로 정렬
        
        표준 순서:
        1. Source_File, No, Month (사용자 추가 컬럼)
        2. CWI Job Number, Order Ref. Number, S/No
        3. DESCRIPTION, RATE SOURCE, RATE, Formula
        4. Q'TY, TOTAL (USD), REV RATE, REV TOTAL, DIFFERENCE
        5. 나머지 컬럼들 (알파벳 순)
        """
        # 표준 컬럼 순서 정의
        standard_columns = [
            'Source_File',
            'No',
            'Month',
            'CWI Job Number',
            'Order Ref. Number',
            'S/No',
            'DESCRIPTION',
            'RATE SOURCE',
            'RATE',
            'Formula',
            'Q\'TY',
            'TOTAL (USD)',
            'REV RATE',
            'REV TOTAL',
            'DIFFERENCE'
        ]
        
        # 실제 존재하는 표준 컬럼만 필터링
        existing_standard = [col for col in standard_columns if col in df.columns]
        
        # 나머지 컬럼들 (표준 컬럼 제외, 알파벳 순)
        remaining_columns = sorted([col for col in df.columns if col not in standard_columns])
        
        # 최종 컬럼 순서
        final_column_order = existing_standard + remaining_columns
        
        # 컬럼 순서 재정렬
        df = df[final_column_order]
        
        logger.info(f"컬럼 순서 표준화 완료: {len(existing_standard)} standard + {len(remaining_columns)} additional")
        return df

