#!/usr/bin/env python3
"""
Vendor Distance Request Generator
거리 확인이 필요한 항목들에 대한 벤더 요청 패키지 생성
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main():
    print("=" * 80)
    print("Vendor Distance Request Generator")
    print("=" * 80)

    # Load interpolation results
    results_dir = Path(__file__).parent / "Results" / "Sept_2025"

    # Find latest interpolated file
    csv_files = sorted(
        (results_dir / "CSV").glob("domestic_sept_2025_interpolated_*.csv")
    )
    if not csv_files:
        print("[ERROR] No interpolation results found")
        return

    latest_csv = csv_files[-1]
    print(f"\nLoading: {latest_csv.name}")
    df = pd.read_csv(latest_csv)
    print(f"  Total items: {len(df)}")

    # Filter vendor_needed items
    vendor_needed = df[
        df.get("distance_interpolation_method", "") == "vendor_needed"
    ].copy()

    print(f"  Vendor distance needed: {len(vendor_needed)}")

    if len(vendor_needed) == 0:
        print("[INFO] No vendor requests needed!")
        return

    # Prepare request data
    request_data = []

    for idx, row in vendor_needed.iterrows():
        request_data.append(
            {
                "S/No": len(request_data) + 1,
                "Shipment Reference": row.get("shipment_ref", ""),
                "Origin (As per Invoice)": row.get("origin", ""),
                "Destination (As per Invoice)": row.get("destination", ""),
                "Origin (Normalized)": row.get("origin_norm", ""),
                "Destination (Normalized)": row.get("destination_norm", ""),
                "Vehicle Type": row.get("vehicle", ""),
                "Unit": row.get("unit", ""),
                "Draft Rate (USD)": round(row.get("rate_usd", 0), 2),
                "Distance (km) - PLEASE CONFIRM": "",  # Empty for vendor to fill
                "Distance Source (GPS/ERP/Route Plan)": "",  # Empty for vendor to fill
                "Minimum Fare Applied (if any)": "",  # Empty for vendor to fill
                "Notes/Comments": "",  # Empty for vendor to fill
            }
        )

    request_df = pd.DataFrame(request_data)

    # Create Excel workbook
    output_dir = results_dir / "Reports"
    output_excel = output_dir / "VENDOR_DISTANCE_REQUEST_SEPT2025.xlsx"

    print(f"\nCreating vendor request Excel...")
    wb = Workbook()

    # Sheet 1: Instructions
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    instructions = [
        ["DISTANCE CONFIRMATION REQUEST"],
        ["HVDC Project - Domestic Deliveries (September 2025)"],
        [""],
        ["Dear DSV Team,"],
        [""],
        ["We are finalizing the September 2025 domestic invoice validation."],
        [
            "Please confirm the trip distances (km) for the following items where distance is currently recorded as 0."
        ],
        [""],
        ["For each line in the 'Distance_Request' sheet, please provide:"],
        ["  1) Actual trip distance (km)"],
        [
            "  2) Distance source/method (GPS tracking / ERP system / Route plan / Manual calculation)"
        ],
        ["  3) Minimum fare evidence (if short-run distance applies)"],
        [""],
        ["Once completed, please return this file to our team."],
        [""],
        ["Thank you for your prompt support."],
        [""],
        ["Best regards,"],
        ["HVDC Audit Team"],
    ]

    for row in instructions:
        ws_inst.append(row)

    # Format
    ws_inst["A1"].font = Font(bold=True, size=16, color="366092")
    ws_inst["A2"].font = Font(italic=True, size=12)

    for cell in ws_inst["A"]:
        cell.alignment = Alignment(wrap_text=True)

    ws_inst.column_dimensions["A"].width = 80

    # Sheet 2: Distance Request
    ws_req = wb.create_sheet("Distance_Request")

    # Write data
    from openpyxl.utils.dataframe import dataframe_to_rows

    for r in dataframe_to_rows(request_df, index=False, header=True):
        ws_req.append(r)

    # Format header
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws_req[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border

    # Highlight fillable columns
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Find column indices for fillable columns
    headers = [cell.value for cell in ws_req[1]]
    fillable_cols = [
        headers.index("Distance (km) - PLEASE CONFIRM") + 1,
        headers.index("Distance Source (GPS/ERP/Route Plan)") + 1,
        headers.index("Minimum Fare Applied (if any)") + 1,
        headers.index("Notes/Comments") + 1,
    ]

    for row in range(2, len(vendor_needed) + 2):
        for col_idx in fillable_cols:
            cell = ws_req.cell(row=row, column=col_idx)
            cell.fill = yellow_fill

    # Freeze header
    ws_req.freeze_panes = "A2"

    # Auto-adjust column widths
    for column in ws_req.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_req.column_dimensions[column_letter].width = adjusted_width

    # Save
    wb.save(output_excel)
    print(f"[OK] Vendor request saved: {output_excel}")

    # Email template
    print("\n" + "=" * 80)
    print("Email Template (Copy & Paste)")
    print("=" * 80)

    email_template = f"""
Subject: Request for Trip Distance Confirmation - Domestic Deliveries (Sept 2025)

Dear DSV Team,

We are finalizing the September 2025 domestic invoice validation.

Kindly confirm the trip distances (km) for the attached {len(vendor_needed)} lines
where the distance is currently recorded as 0.

For each line, please provide:
1) Origin/Destination (as dispatched)
2) Actual trip distance (km) and method (GPS/ERP/Route plan)
3) Any short-run minimum fare evidence (if applicable)

We will update our reference map accordingly and proceed with validation.

Please fill in the yellow-highlighted columns in the attached Excel file and
return it to us at your earliest convenience.

Attachment: VENDOR_DISTANCE_REQUEST_SEPT2025.xlsx

Thank you for your prompt support.

Best regards,
HVDC Audit Team
"""

    print(email_template)

    # Save email template
    email_file = output_dir / "VENDOR_EMAIL_TEMPLATE.txt"
    with open(email_file, "w", encoding="utf-8") as f:
        f.write(email_template)

    print(f"\n[OK] Email template saved: {email_file}")

    print("\n" + "=" * 80)
    print("Summary")
    print("=" * 80)
    print(f"  Vendor distance requests: {len(vendor_needed)}")
    print(f"  Request file: {output_excel}")
    print(f"  Email template: {email_file}")
    print("=" * 80)


if __name__ == "__main__":
    main()
